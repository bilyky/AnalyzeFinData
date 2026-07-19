"""
Project AETHER: Historical Trade DNA Bootstrapper.

Parses historical game transaction history, pairs BUY/SELL executions,
cross-references historical Excel backups to extract buy-state DNA (PGR, scores, Z-scores),
and pre-populates 'Data/trade_history_dna.json' with your past month of trades.
"""

import json
import os
import re
import datetime
from pathlib import Path
import openpyxl
from aether_logger import get_logger as _get_logger

_log = _get_logger("bootstrap_dna")

BASE_DIR = Path(__file__).resolve().parent
GAME_FILE = BASE_DIR / "Data" / "ai_portfolio_game.json"
BACKUP_DIR = BASE_DIR / "Data" / "Backup"
OUT_FILE = BASE_DIR / "Data" / "trade_history_dna.json"

def find_backup_for_date(date_str: str) -> Path:
    """Find the first Excel backup file for a specific YYYY-MM-DD date."""
    # Backups are in Data/Backup/2026/investment_YYYY-MM-DD_*.xlsx
    pattern = f"investment_{date_str}_*.xlsx"
    matches = list(BACKUP_DIR.glob(f"**/{pattern}"))
    if matches:
        return matches[0]
    return None

def extract_dna_from_backup(backup_path: Path, symbol: str) -> dict:
    """Load historical Excel backup and extract a symbol's ratings/scores."""
    dna = {
        "pgr": "Neutral",
        "s10": 0.0,
        "l60": 0.0,
        "score": 0.0,
        "z_score": 0.0,
        "setup": False,
        "industry": "Unknown"
    }
    if not backup_path or not backup_path.exists():
        return dna
    
    wb = None
    try:
        wb = openpyxl.load_workbook(backup_path, read_only=True, data_only=True)
        # 1. Read Research sheet
        if "Research" in wb.sheetnames:
            ws = wb["Research"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Column index map: row[3] is Ticker, row[6] is PGR, row[20] is Setup, row[24] is Short10, row[25] is Long60
                sym = row[3]
                if sym and sym.strip().upper() == symbol.strip().upper():
                    dna["pgr"] = row[6] or "Neutral"
                    dna["industry"] = row[4] or "Unknown"
                    try:
                        dna["s10"] = float(row[24] or 0.0)
                        dna["l60"] = float(row[25] or 0.0)
                        dna["score"] = round(dna["s10"] + dna["l60"], 1)
                    except (ValueError, TypeError):
                        pass
                    setup_val = str(row[20] or "").strip()
                    dna["setup"] = setup_val in ("1", "OK")
                    break
                    
    except Exception as e:
        _log.warning(f"Failed to parse backup {backup_path.name}", extra={"error": str(e)})
    finally:
        if wb:
            wb.close()
    return dna

def bootstrap():
    _log.info("Initialising historical Trade DNA recovery pass...")
    
    if not GAME_FILE.exists():
        _log.error(f"Game file {GAME_FILE} not found. Cannot bootstrap.")
        return
        
    with open(GAME_FILE, "r", encoding="utf-8") as f:
        state = json.load(f)
        
    history = state.get("history", [])
    _log.info(f"Loaded {len(history)} historical transaction events.")
    
    # Pair up BUY and SELL transactions
    closed_trades = []
    open_buys = {} # symbol -> list of BUY transactions (FIFO order)
    
    for tx in history:
        sym = tx.get("symbol")
        tx_type = tx.get("type")
        tx_date = tx.get("date")
        price = float(tx.get("price", 0.0))
        qty = int(tx.get("qty", 0))
        
        if not sym:
            continue
            
        if tx_type == "BUY":
            if sym not in open_buys:
                open_buys[sym] = []
            open_buys[sym].append(tx)
            
        elif tx_type == "SELL":
            if sym in open_buys and open_buys[sym]:
                # FIFO matching
                buy_tx = open_buys[sym].pop(0)
                
                b_date = datetime.date.fromisoformat(buy_tx["date"])
                s_date = datetime.date.fromisoformat(tx_date)
                holding_days = (s_date - b_date).days
                
                b_price = float(buy_tx["price"])
                pnl_pct = round(((price - b_price) / b_price) * 100, 2) if b_price else 0.0
                
                # Fetch historical Buy DNA from backup sheet
                backup_file = find_backup_for_date(buy_tx["date"])
                if backup_file:
                    _log.info(f"Mapping {sym} closed trade (PnL {pnl_pct:+.1f}%) -> Loading backup for {buy_tx['date']}...")
                else:
                    _log.warning(f"Mapping {sym} closed trade (PnL {pnl_pct:+.1f}%) -> No backup found for {buy_tx['date']}.")
                    
                buy_dna = extract_dna_from_backup(backup_file, sym)
                
                closed_trades.append({
                    "symbol": sym,
                    "buy_date": buy_tx["date"],
                    "sell_date": tx_date,
                    "buy_price": b_price,
                    "sell_price": price,
                    "qty": qty,
                    "pnl_pct": pnl_pct,
                    "holding_days": holding_days,
                    "buy_dna": buy_dna
                })
                
    _log.info(f"Successfully paired and recovered {len(closed_trades)} completed trades!")
    
    # Save the raw DNA ledger
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(closed_trades, f, indent=4)
        
    _log.info(f"Successfully wrote pre-populated trade ledger to: {OUT_FILE}")

if __name__ == "__main__":
    bootstrap()
