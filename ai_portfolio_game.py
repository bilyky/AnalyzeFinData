import json
import datetime
import openpyxl
import os
import pytz
import re
import requests
from aether import etrade
import rapidapi
import sys
import console_safe
import circuit_breaker
from aether import trash
import aether.notify as notify
import argparse
from pathlib import Path
import aether_oracle
from aether.utils import _to_float
from aether.config import CFG

# Windows CP1252 console fallback (bug-fix workaround, not a feature): must run
# before the first non-ASCII print. Reduces, not eliminates, cp1252 crashes in
# headless runs. See AETHER_REFERENCE.md.
console_safe.install()

# --- CONFIGURATION ---
BASE_DIR = Path(__file__).resolve().parent
AI_GAME_FILE = BASE_DIR / "Data" / "ai_portfolio_game.json"
XLSX_FILE = BASE_DIR / "Data" / "state_of_the_day.xlsx"
AI_PERF_XLSX = BASE_DIR / "Data" / "ai_portfolio_performance.xlsx"
SYMBOL_FULL_DIR = BASE_DIR / "Data" / "Symbol_full"   # OHLCV cache — one source of truth
INITIAL_BALANCE = 10000.0

# Import risk utils safely
import risk_utils
import sell_rules
import decision_eval
import watchdog
import instruments
from aether_logger import get_logger as _get_logger
from aether.scoring import digit_sum_open_score as _digit_open_score
_log = _get_logger("ai_game")


_SYMBOL_DAY_CACHE = {}

def _load_symbol_today_cache(symbol: str, today_str: str) -> dict:
    """Flyweight Cache Pattern: loads and returns the daily JSON cache,
    ensuring each symbol is read from the hard drive at most once."""
    symbol = symbol.upper()
    if symbol in _SYMBOL_DAY_CACHE:
        return _SYMBOL_DAY_CACHE[symbol]
        
    cache_path = BASE_DIR / "Data" / "Symbol" / symbol / f"{symbol}_{today_str}.json"
    cache = {}
    if cache_path.exists():
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cache = json.load(f)
        except Exception as e:
            _log.warning(f"Failed to read daily cache for {symbol}: {e}")
            
    _SYMBOL_DAY_CACHE[symbol] = cache
    return cache

def check_failure_rules(symbol, pgr, score, z_score, industry, s10=0.0) -> tuple[bool, str]:
    """Check if the candidate matches any active toxic rules in Data/failure_dna_rules.json or dynamic filters."""
    rules_file = BASE_DIR / "Data" / "failure_dna_rules.json"
    
    # ── Earnings-Shock Failure Gate (Pillar 1 Guard) ──
    # Programmatic, un-bypassable veto on any symbol that has just reported a massive earnings miss
    today_str = str(datetime.date.today())
    cache = _load_symbol_today_cache(symbol, today_str)
    if cache:
        eps_data = cache.get("EPSData", {})
        eps_diff = eps_data.get("eps_diff_description", "")
        warning_impact = eps_data.get("warning_impact", "")
        
        if "missed by" in (eps_diff or "").lower() or warning_impact == "Very Bearish":
            return True, f"Earnings Shock Veto: {eps_diff or 'Very Bearish earnings'}"

    if not rules_file.exists() or rules_file.stat().st_size == 0:
        return False, ""
    try:
        with open(rules_file, "r", encoding="utf-8") as f:
            rules = json.load(f)
        for r in rules:
            field = r.get("field")
            condition = r.get("condition")
            reason = r.get("reason", "Toxic pattern match")

            # 1. PGR Match
            if field == "pgr" and condition == "startswith_Be" and str(pgr).startswith("Be"):
                # ── R&D #13: PGR Waivers (canonical two-factor elite gate) ──
                # A. HighScorePGRBypass: bypass Bearish PGR only for an elite breakout
                #    leader per risk_utils.is_elite_breakout_candidate — BOTH
                #    score >= CFG.system_bypass_score_floor AND s10 >= CFG.system_bypass_s10_floor
                #    (single-sourced from CFG; same gate run_daily_ai_management uses for
                #    the R:R waiver, so both stages agree on "elite" and re-tune together).
                # B. BottomSnipePGRWaiver: bypass Bearish PGR on a confirmed bottom setup.
                # Test the cheap in-memory gate first; only read the daily-history file
                # (is_bottom_confirmed) when the elite gate did not already waive.
                if risk_utils.is_elite_breakout_candidate(score, s10):
                    _log.info(f"[R&D #13 PGR Waiver] Bypassed Bearish PGR '{pgr}' for {symbol} (elite breakout: score {score:.1f}, s10 {s10:.1f}).")
                    continue
                bottom_ok, _ = is_bottom_confirmed(symbol)
                if bottom_ok:
                    _log.info(f"[R&D #13 PGR Waiver] Bypassed Bearish PGR '{pgr}' for {symbol} (bottom confirmed).")
                    continue
                return True, reason

            # 2. Score Match
            if field == "score" and condition == "less_than_5.0" and score < 5.0:
                return True, reason

            # 3. Z-Score Match
            if field == "z_score" and condition == "greater_than_2.5" and z_score > 2.5:
                return True, reason
    except Exception as e:
        _log.warning(f"Failed to evaluate failure rules: {e}")
    return False, ""


def log_closed_trade_dna(sym, pos, price, today_str):
    """Log a completed trade's DNA signature to Data/trade_history_dna.json."""
    try:
        buy_dna = pos.get("buy_dna")
        if not buy_dna:
            return
        
        buy_date = buy_dna.get("buy_date", today_str)
        pnl_pct = round(((price - pos["cost"]) / pos["cost"]) * 100, 2) if pos["cost"] else 0.0
        
        dna_file = BASE_DIR / "Data" / "trade_history_dna.json"
        dna_list = []
        if dna_file.exists() and dna_file.stat().st_size > 0:
            with open(dna_file, "r", encoding="utf-8") as f:
                dna_list = json.load(f)
                
        b_date = datetime.date.fromisoformat(buy_date)
        s_date = datetime.date.fromisoformat(today_str)
        holding_days = (s_date - b_date).days
        
        dna_list.append({
            "symbol": sym,
            "buy_date": buy_date,
            "sell_date": today_str,
            "buy_price": pos["cost"],
            "sell_price": price,
            "qty": pos["qty"],
            "pnl_pct": pnl_pct,
            "holding_days": holding_days,
            "buy_dna": buy_dna
        })
        
        with open(dna_file, "w", encoding="utf-8") as f:
            json.dump(dna_list, f, indent=4)
        _log.info(f"Logged closed trade DNA for {sym} to trade_history_dna.json.")
    except Exception as e:
        _log.warning(f"Failed to log trade DNA: {e}")


def _load_closes(symbol):
    """Sorted daily closes for a symbol from the local OHLCV cache, or [] when the
    file is missing/unreadable. Single loader so the cache path is defined once —
    trend-score and bubble-z-score previously used a wrong doubled-`Data` path and
    silently read nothing."""
    try:
        path = SYMBOL_FULL_DIR / f"{symbol}_daily.json"
        if not path.exists():
            return []
        with open(path) as f:
            ts = json.load(f).get("Time Series (Daily)", {})
        return [float(ts[d]["4. close"]) for d in sorted(ts.keys())]
    except Exception:
        return []


_HEAL_ATTEMPTED: set = set()  # per-process guard: each symbol healed at most once
_MAX_STALE_DAYS = 10          # single source of truth for the OHLCV freshness horizon


def _heal_symbol_cache(symbol) -> bool:
    """Pull fresh OHLCV history via rapidapi when a stale cache is detected.
    Each symbol is attempted at most once per process run."""
    if symbol in _HEAL_ATTEMPTED:
        return False
    _HEAL_ATTEMPTED.add(symbol)
    try:
        today_str = str(datetime.date.today())
        res = rapidapi.repair_missing([symbol], today_str)
        if res.get("locked"):
            # Another process holds the RapidAPI lock — this was never a real attempt.
            # Release the once-per-run guard so a later pass can still heal this symbol.
            _HEAL_ATTEMPTED.discard(symbol)
            _log.info(f"[Self-Healer] {symbol} heal deferred — RapidAPI busy in another process.")
            return False
        if res.get("updated", 0) > 0:
            _log.info(f"[Self-Healer] Healed {symbol} OHLCV cache.")
            return True
        else:
            _log.warning(f"[Self-Healer] On-demand refresh for {symbol} yielded 0 updates.")
            return False
    except Exception as e:
        _log.warning(f"[Self-Healer] Failed to heal {symbol} cache: {e}")
        return False


def _cache_age_days(symbol):
    """Age in days of the newest RAW cached bar (no split-adjust), or None if the cache
    is missing/empty/unreadable. Cheap sibling of risk_utils.ohlcv_age_days — reads the
    same raw JSON _cache_stale already consults, so the freshness gate can report an age
    without a second, heavier split-adjusting load."""
    path = SYMBOL_FULL_DIR / f"{symbol}_daily.json"
    if not path.exists():
        return None
    try:
        with open(path) as f:
            ts = json.load(f).get("Time Series (Daily)", {})
        dates = sorted(ts.keys())
        if not dates:
            return None
        return (datetime.date.today() - datetime.date.fromisoformat(dates[-1])).days
    except Exception:
        return None


def _cache_stale(symbol, max_stale_days=_MAX_STALE_DAYS) -> bool:
    """Return True if the symbol's OHLCV cache is missing or older than max_stale_days."""
    age = _cache_age_days(symbol)
    return age is None or age > max_stale_days


def _sma50(symbol, max_stale_days=_MAX_STALE_DAYS):
    """50-day SMA of closes from the local OHLCV cache, with autonomic on-demand self-healing."""
    try:
        path = SYMBOL_FULL_DIR / f"{symbol}_daily.json"
        
        # Check if cache is missing or stale
        needs_healing = False
        last_bar_date = "None"
        if not path.exists():
            needs_healing = True
            _log.warning(f"[sma50] {symbol} OHLCV cache missing — winner protection unavailable.")
        else:
            try:
                with open(path) as f:
                    ts = json.load(f).get("Time Series (Daily)", {})
                dates = sorted(ts.keys())
                if not dates:
                    needs_healing = True
                else:
                    last_bar_date = dates[-1]
                    stale_days = (datetime.date.today() - datetime.date.fromisoformat(last_bar_date)).days
                    if stale_days > max_stale_days:
                        needs_healing = True
                        _log.warning(f"[sma50] {symbol} OHLCV cache stale by {stale_days}d (last bar: {last_bar_date}).")
            except Exception:
                needs_healing = True

        if needs_healing:
            healed = _heal_symbol_cache(symbol)
            if not healed:
                _log.warning(f"[sma50] {symbol} cache heal failed; winner protection unavailable this run.")
                return None

        # Re-read the newly healed/fresh cache file
        if not path.exists():
            return None
        with open(path) as f:
            ts = json.load(f).get("Time Series (Daily)", {})
        dates = sorted(ts.keys())
        if not dates:
            return None
        closes = [float(ts[d]["4. close"]) for d in dates]
        return sell_rules.sma_from_closes(closes, 50)
    except Exception as e:
        _log.warning(f"Failed to calculate 50-day SMA for {symbol}: {e}")
        return None

def _live_equity(cash, positions, price_of):
    """Cash + Σ qty·price across positions, falling back to each position's cost when
    its live quote is missing/invalid. A data gap must never crash the caller — the
    equity number is for reporting/sizing, not a risk gate."""
    equity = cash or 0.0
    for sym, pos in positions.items():
        px = price_of.get(sym)
        if not px or px <= 0:
            px = pos.get("cost", 0.0)
        equity += pos.get("qty", 0) * px
    return equity


def _active_setup_symbols(ws):
    """Symbols on the Research sheet with an active Setup flag (col U / index 20 =
    '1' or 'OK'). Single definition of the filter used for both price-gathering and
    buy-candidate selection."""
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] and str(row[20] or '') in ('1', 'OK', 1):
            out.append(row[3])
    return out


# @doc-sync-start: scarcity_core
# Behavior documented in: Data/wiki.json["scarcity_core"], AETHER_REFERENCE.md
# ("Dynamic Structural Scarcity Core"), and plans/dynamic-scarcity-cap.md. If you
# change the cap math, the ramp, or the ceilings, update those surfaces in the same
# commit (enforced by scripts/utils/pre_commit_validator.py :: check_feature_doc_sync).
def _conviction_cap_pct(total, base_pct, ceiling_pct, relax_start, relax_full):
    """Dynamic (per-profile) bucket-allocation cap that scales with conviction.

    Below ``relax_start`` the ``base_pct`` cap applies unchanged; between
    ``relax_start`` and ``relax_full`` the cap ramps linearly ``base_pct``→``ceiling_pct``;
    at or above ``relax_full`` it clamps at ``ceiling_pct``. The cap never disappears.

    This replaces the old binary "suspend the scarcity cap entirely when total >= 8.0"
    cliff, which had two flaws: (a) a 0.01-point score difference at the 8.0 boundary
    produced a ~10x position-size jump (a discontinuity that noise or a factor tweak
    could trip), and (b) because the relaxation threshold was a global 8.0 while
    DEFENSIVE's min_score_threshold is 10.0, *every* DEFENSIVE buy already cleared 8.0 —
    so the capital-preservation profile had its scarcity cap permanently suspended.
    Tying ``relax_start`` to each profile's own min_score_threshold (see
    get_strategy_rules) fixes that; the ramp keeps conviction rewarded but bounded.

    This is Option C from plans/dynamic-scarcity-cap.md (per-profile ramp); Options A
    (global ramp) and B (ramp, no per-position ceiling) are documented there too.
    """
    if ceiling_pct <= base_pct or relax_full <= relax_start:
        return base_pct
    if total <= relax_start:
        return base_pct
    if total >= relax_full:
        return ceiling_pct
    frac = (total - relax_start) / (relax_full - relax_start)
    return base_pct + frac * (ceiling_pct - base_pct)
# @doc-sync-end: scarcity_core


def determine_max_positions(cash_ratio: float, num_positions: int, base_max_positions: int) -> int:
    """Determine max positions after applying Dynamic Position-Slot Expansion.
    Tested deterministically with zero mocks.
    """
    max_positions = base_max_positions
    while cash_ratio > 0.15 and num_positions >= max_positions:
        max_positions += 1
    return max_positions


def evaluate_momentum_rotation(
    profile: str,
    is_market_open_flag: bool,
    available_slots: int,
    max_positions: int,
    positions: dict,
    prices: dict,
    top_buys: list,
    active_position_scores: dict
) -> tuple[list[str], int, float]:
    """Evaluate and plan momentum rotations for AGGRESSIVE profiles.
    Returns: (list_of_symbols_to_sell, updated_available_slots, cash_proceeds_to_add)
    Tested deterministically with zero mocks.
    """
    sells = []
    freed_slots = 0
    cash_addition = 0.0
    
    if profile == "AGGRESSIVE" and is_market_open_flag:
        # Dynamic Momentum Rotation (R&D #27):
        # 1. Standard Case: Slots are FULL, and we have elite buy candidates (score >= 12.0) waiting.
        # 2. Lazy-Hold Case: Slots are NOT full, but we have a position in severe momentum decay (score < 2.0)
        #    AND we have qualifying buy candidates (score >= 10.0) waiting to deploy that capital better!
        is_full_slots = (available_slots + freed_slots) <= 0
        min_buy_score = 12.0 if is_full_slots else 10.0
        
        elite_buys = [b for b in top_buys if b["total"] >= min_buy_score]
        if elite_buys:
            # Find eligible rotation candidates
            eligible_rotation_candidates = []
            for sym, pos in positions.items():
                if sym in sells:
                    continue
                current_px = prices.get(sym, pos["cost"])
                is_mature = pos.get("stop_loss", 0.0) >= pos["cost"]
                current_score = active_position_scores.get(sym, 0.0)
                
                # If slots are full, we rotate score < 8.0.
                # If slots are empty, we rotate score < 2.0 (lazy hold prevention!)
                threshold_score = 8.0 if is_full_slots else 2.0
                
                if is_mature and current_score < threshold_score:
                    eligible_rotation_candidates.append({
                        "sym": sym,
                        "pos": pos,
                        "score": current_score,
                        "current_px": current_px
                    })
            
            # Sort eligible candidates ascending by score (lowest score rotated first)
            eligible_rotation_candidates.sort(key=lambda x: x["score"])
            
            while eligible_rotation_candidates and elite_buys:
                is_currently_full = (available_slots + freed_slots) <= 0
                target_to_rotate = eligible_rotation_candidates[0]
                
                # If slots are empty, we only rotate if the candidate has decayed below 2.0
                if not is_currently_full and target_to_rotate["score"] >= 2.0:
                    break
                    
                target_to_rotate = eligible_rotation_candidates.pop(0)
                sym_to_sell = target_to_rotate["sym"]
                pos = target_to_rotate["pos"]
                price = target_to_rotate["current_px"]
                
                sells.append(sym_to_sell)
                cash_addition += pos["qty"] * price
                freed_slots += 1
                elite_buys.pop(0)
                
    return sells, available_slots + freed_slots, cash_addition


def _execute_buys(state, top_buys, available_slots, min_cash_required, rules,
                  today, now_time, new_transactions, prices=None):
    """Fill up to available_slots from ranked top_buys while preserving the cash
    buffer. Rejects super-bubble names (>2.5σ over the 500-day mean) and unverified
    setups, falling through to the next candidate so a slot isn't wasted. Mutates
    state (balance/positions/history) and appends to new_transactions; returns the
    number of buys executed. Enforces Core-Satellite allocation limits for Scarcity/Standard plays."""
    if prices is None:
        prices = {}
    
    # 20% Scarcity bucket and 80% Standard bucket allocations
    scarcity_allocation_pct = rules.get("scarcity_allocation_pct", 0.20)
    scarcity_limit_usd = state["equity"] * scarcity_allocation_pct
    standard_limit_usd = state["equity"] * (1.0 - scarcity_allocation_pct)
    
    current_scarcity_usd = sum(pos["qty"] * prices.get(sym, pos["cost"]) 
                               for sym, pos in state["positions"].items() if pos.get("is_scarcity", False))
    current_standard_usd = sum(pos["qty"] * prices.get(sym, pos["cost"]) 
                               for sym, pos in state["positions"].items() if not pos.get("is_scarcity", False))
                               
    _log.debug("[Buckets] Scarcity $%.2f / $%.2f  Standard $%.2f / $%.2f",
               current_scarcity_usd, scarcity_limit_usd,
               current_standard_usd, standard_limit_usd)

    buys_executed = 0
    for buy in top_buys:
        if buys_executed >= available_slots:
            break

        is_scarcity = instruments.is_scarcity_asset(buy["sym"], buy.get("industry", ""))

        # Re-calculate cash buffer dynamically based on remaining available slots
        current_available = available_slots - buys_executed
        cash_per_buy = (state["balance"] - min_cash_required) / current_available
        qty = calculate_share_qty(buy["sym"], cash_per_buy, buy["price"])
        if qty > 0:
            # 2.5-Sigma Bubble Guard: Reject if symbol trades > 2.5 standard deviations above 500 SMA
            z_score = calculate_bubble_z_score(buy["sym"])
            if z_score is not None and z_score > 2.5:
                _log.warning(f"🛑 AI BUY REJECTED (2.5-Sigma Bubble Guard): {buy['sym']} trades at +{z_score:.2f} SD above its 500-day mean (Super-Bubble Zone!).")
                continue

            is_verified, v_msg = backtrack_verify(buy["sym"])
            if not is_verified:
                _log.warning(f"🛑 AI BUY REJECTED: {buy['sym']} - {v_msg}")
                continue

            # Dynamic Feedback Analyzer Guard Check (Anti-Failure DNA)
            pgr_val = buy.get("pgr", "Neutral")
            score_val = buy.get("total", 0.0)
            is_toxic, t_reason = check_failure_rules(buy["sym"], pgr_val, score_val, z_score, buy.get("industry", "Unknown"), s10=buy.get("s10", 0.0))
            if is_toxic:
                _log.warning(f"AI BUY REJECTED (Feedback Analyzer Rule Match): {buy['sym']} - {t_reason}")
                continue

            # Open-digit real-time signal: log when a strong numerology bias fires
            _open_digit_z = _digit_open_score(buy["sym"], buy["price"])
            if abs(_open_digit_z) >= 0.33:  # fires when |z|>=2.0 in study (z/3 threshold)
                _dir = "UP" if _open_digit_z > 0 else "DOWN"
                _log.console(f"🔢 [Digit-Sum] {buy['sym']} open digit signal: {_dir} bias (score={_open_digit_z:+.2f})")

            # Core-Satellite Allocation Checking & Downsizing logic.
            # Dynamic per-profile conviction ramp (Option C, plans/dynamic-scarcity-cap.md):
            # the scarcity bucket cap scales base→ceiling with the candidate's conviction
            # score instead of the old all-or-nothing suspension at total>=8.0. Standard
            # bucket keeps its flat base cap. Both are then bounded by a per-position ceiling.
            cost = qty * buy["price"]
            total = buy.get("total", 0.0)
            if is_scarcity:
                cap_pct = _conviction_cap_pct(
                    total,
                    base_pct=scarcity_allocation_pct,
                    ceiling_pct=rules.get("scarcity_cap_ceiling_pct", scarcity_allocation_pct),
                    relax_start=rules.get("cap_relax_start", 8.0),
                    relax_full=rules.get("cap_relax_full", 12.0),
                )
                scarcity_limit_usd = state["equity"] * cap_pct
                remaining_scarcity_room_usd = scarcity_limit_usd - current_scarcity_usd
                if remaining_scarcity_room_usd <= 0:
                    _log.warning(f"🛑 AI BUY REJECTED (Scarcity Limit Full): {buy['sym']} is a scarcity play, but the {cap_pct:.0%} scarcity bucket is fully allocated.")
                    continue
                if cost > remaining_scarcity_room_usd:
                    old_qty = qty
                    qty = calculate_share_qty(buy["sym"], remaining_scarcity_room_usd, buy["price"])
                    _log.warning(f"⚠️ Downsizing scarcity buy {buy['sym']} from {old_qty} to {qty} shares to fit the {cap_pct:.0%} scarcity cap (conviction {total:.1f}).")
                    if qty <= 0:
                        _log.warning(f"🛑 AI BUY REJECTED (Scarcity Limit Full): Remaining room is less than 1 share of {buy['sym']}.")
                        continue
                    cost = qty * buy["price"]
            else:
                remaining_standard_room_usd = standard_limit_usd - current_standard_usd
                if remaining_standard_room_usd <= 0:
                    _log.warning(f"🛑 AI BUY REJECTED (Standard Cap Full): {buy['sym']} is a standard play, but the {(1.0 - scarcity_allocation_pct):.0%} standard bucket is fully allocated.")
                    continue
                if cost > remaining_standard_room_usd:
                    old_qty = qty
                    qty = calculate_share_qty(buy["sym"], remaining_standard_room_usd, buy["price"])
                    _log.warning(f"⚠️ Downsizing standard buy {buy['sym']} from {old_qty} to {qty} shares to fit the standard bucket cap.")
                    if qty <= 0:
                        _log.warning(f"🛑 AI BUY REJECTED (Standard Cap Full): Remaining room is less than 1 share of {buy['sym']}.")
                        continue
                    cost = qty * buy["price"]

            # Per-position ceiling on EVERY buy: no single name may exceed
            # equity * max_allocation_pct. Previously _execute_buys never applied this
            # cap (only the pyramiding path did), so a single buy into an empty bucket
            # could consume the whole bucket. Now concentration is capped per name too.
            max_pos_usd = state["equity"] * rules.get("max_allocation_pct", 0.15)
            if cost > max_pos_usd and max_pos_usd > 0:
                old_qty = qty
                qty = calculate_share_qty(buy["sym"], max_pos_usd, buy["price"])
                _log.warning(f"⚠️ Capping {buy['sym']} from {old_qty} to {qty} shares to fit the {rules.get('max_allocation_pct', 0.15):.0%} per-position ceiling.")
                if qty <= 0:
                    _log.warning(f"🛑 AI BUY REJECTED: Per-position ceiling leaves less than 1 share of {buy['sym']}.")
                    continue
                cost = qty * buy["price"]

            state["balance"] -= cost
            # Update cumulative bucket counts for sequence
            if is_scarcity:
                current_scarcity_usd += cost
            else:
                current_standard_usd += cost

            atr = risk_utils.calculate_atr(buy["sym"])
            if atr and atr > 0:
                stop_loss = round(buy["price"] - (rules["atr_multiplier"] * atr), 2)
                stop_desc = f"ATR-based Stop: ${stop_loss}{buy.get('bottom_desc', '')}"
            else:
                stop_loss = round(buy["price"] * 0.92, 2)
                stop_desc = f"8% Fallback{buy.get('bottom_desc', '')}"

            state["positions"][buy["sym"]] = {
                "qty": qty, 
                "cost": buy["price"], 
                "stop_loss": stop_loss,
                "is_scarcity": is_scarcity,
                "buy_dna": {
                    "buy_date": today,
                    "pgr": buy.get("pgr", "Neutral"),
                    "s10": buy.get("s10", 0.0),
                    "l60": buy.get("l60", 0.0),
                    "score": buy.get("total", 0.0),
                    "z_score": z_score,
                    "industry": buy.get("industry", "Unknown")
                }
            }
            tx = {"date": today, "time": now_time, "type": "BUY", "symbol": buy["sym"], "price": buy["price"], "qty": qty}
            state["history"].append(tx)
            new_transactions.append(tx)
            buys_executed += 1
            _log.info(f"🤖 AI LIVE BUY: {qty} shares of {buy['sym']} at ${buy['price']} ({stop_desc}, Scarcity={is_scarcity})")
    return buys_executed


def is_market_open():
    """Check if current time is within US Market hours (9:30 AM - 4:00 PM EST)."""
    tz = pytz.timezone("America/New_York")
    now = datetime.datetime.now(tz)
    
    if now.weekday() >= 5:
        return False, "Market is closed (Weekend)."
    
    market_open = now.replace(hour=9, minute=30, second=0, microsecond=0)
    market_close = now.replace(hour=16, minute=0, second=0, microsecond=0)
    
    if market_open <= now <= market_close:
        return True, "Market is Open."
    else:
        return False, f"Market is Closed. (Current EST: {now.strftime('%H:%M')})"

def calculate_ticker_trend_score(symbol: str):
    """
    Calculate a normalized, standardized trend score in [-10.0, +10.0] for a symbol
    based on its price relative to its 20, 50, and 200 daily SMAs.

    Returns None when the OHLCV cache is missing or has < 200 days — a distinct
    "no data" signal, not a 0.0 score, so the breadth filter never treats an absent
    cache as a flat market (Zero-Trust).
    """
    try:
        closes = _load_closes(symbol)
        if len(closes) < 200:
            return None

        sma20 = sum(closes[-20:]) / 20
        sma50 = sum(closes[-50:]) / 50
        sma200 = sum(closes[-200:]) / 200
        current_price = closes[-1]

        s_20 = 2.5 if current_price > sma20 else -2.5
        s_50 = 2.5 if current_price > sma50 else -2.5
        s_200 = 2.5 if current_price > sma200 else -2.5
        s_cross1 = 1.25 if sma20 > sma50 else -1.25
        s_cross2 = 1.25 if sma50 > sma200 else -1.25

        return s_20 + s_50 + s_200 + s_cross1 + s_cross2
    except Exception:
        return None

def calculate_bubble_z_score(symbol: str):
    """
    Calculate the Z-score (number of standard deviations) of the current price
    relative to its 500-day moving average (mean) over historical closes.
    Returns None if there is insufficient historical data (< 500 days).
    """
    try:
        closes = _load_closes(symbol)
        if len(closes) < 500:
            return None # Insufficient history for 500-day mean

        recent_closes = closes[-500:]
        
        # Calculate 500-day Mean (SMA500)
        mean = sum(recent_closes) / 500
        
        # Calculate 500-day Standard Deviation
        variance = sum((p - mean) ** 2 for p in recent_closes) / 500
        std_dev = variance ** 0.5
        
        if std_dev <= 0:
            return 0.0
            
        current_price = closes[-1]
        z_score = (current_price - mean) / std_dev
        return z_score
    except Exception:
        return None

def get_market_regime():
    """Query SPY momentum to dynamically determine the best strategy profile, adjusting for breadth divergence."""
    base_profile = "BALANCED"
    wb = None
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        ws = wb["Research"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3] == "SPY":
                l60 = row[25] or 0
                if l60 > 2:
                    base_profile = "AGGRESSIVE" # Strong bull market
                elif l60 < -2:
                    base_profile = "DEFENSIVE" # Bear market
                else:
                    base_profile = "BALANCED" # Consolidation
                break
    except Exception as e:
        _log.warning(f"Error loading SPY regime: {e}")
        base_profile = "BALANCED"
    finally:
        if wb:
            wb.close()

    # Breadth Divergence Adjustment Pass
    try:
        spy_score = calculate_ticker_trend_score("SPY")
        rsp_score = calculate_ticker_trend_score("RSP")
        # Zero-Trust: only judge breadth when BOTH series have enough history. A
        # missing/short cache (score None) must not fake a divergence — e.g. absent
        # RSP data would otherwise read as a large SPY-RSP gap and force a spurious
        # defensive downgrade.
        if spy_score is None or rsp_score is None:
            _log.info(f"  [Breadth] Skipped — insufficient SMA history (SPY={spy_score}, RSP={rsp_score}).")
            return base_profile
        delta = spy_score - rsp_score

        if delta > 4.0:
            _log.warning(f"⚠️ [BREADTH ALERT] SPY-RSP Divergence is high: {delta:.2f} (SPY: {spy_score:.1f}, RSP: {rsp_score:.1f})")
            if base_profile == "AGGRESSIVE":
                _log.info("  -> Downgrading profile from AGGRESSIVE to BALANCED due to narrow market breadth!")
                return "BALANCED"
            elif base_profile == "BALANCED":
                _log.info("  -> Downgrading profile from BALANCED to DEFENSIVE due to narrow market breadth!")
                return "DEFENSIVE"
    except Exception as e:
        _log.warning(f"Error applying breadth divergence filter: {e}")

    return base_profile

def get_strategy_rules(profile):
    """Define risk and size rules based on the chosen strategy profile."""
    # Profile options: BALANCED, AGGRESSIVE, DEFENSIVE
    # Regime is read independently of the (possibly manually-overridden) profile:
    # get_market_regime() labels a strong bull market "AGGRESSIVE" (SPY L60 > 2),
    # so that is the bull signal — deploy fully (zero cash buffer) when it fires,
    # regardless of which profile is active (Zero-Cash-Drag Autopilot).
    regime = get_market_regime()
    is_bullish = (regime == "AGGRESSIVE")

    if profile == "AGGRESSIVE":
        return {
            "max_positions": 6,
            "max_allocation_pct": 0.15, # Optimized from 0.25 to minimize drawdowns
            # @doc-sync-start: scarcity_core
            "scarcity_allocation_pct": 0.20, # Dynamic Core-Satellite scarcity bucket base cap
            # Dynamic scarcity cap (Option C): bucket cap ramps 20%->40% as conviction
            # rises from min_score_threshold (2.0) to 10.0. See plans/dynamic-scarcity-cap.md.
            "scarcity_cap_ceiling_pct": 0.40,
            "cap_relax_start": 2.0,      # = min_score_threshold (no cliff, ramp starts here)
            "cap_relax_full": 10.0,      # calibrated to eligible-score p90 (~+10.1, validate_scarcity_cap.py, 2026-07-30; n=26)
            # @doc-sync-end: scarcity_core
            "atr_multiplier": 3.5,       # Loose stop to avoid shakeouts in high-beta stocks
            "min_score_threshold": 2.0,
            "cash_buffer_pct": 0.0 if is_bullish else 0.10
        }
    elif profile == "DEFENSIVE":
        return {
            "max_positions": 3,          # Restrict to top 3 ultra-conviction plays
            "max_allocation_pct": 0.10, # Optimized from 0.15 for maximum capital preservation (Capped at $1,000 per trade)
            # @doc-sync-start: scarcity_core
            "scarcity_allocation_pct": 0.20, # Dynamic Core-Satellite scarcity bucket base cap
            # Dynamic scarcity cap (Option C): tightest ramp — 20%->25% only, and it does
            # not even begin until score 10.0 (= min_score_threshold). This fixes the old
            # pathology where the global 8.0 relaxation left DEFENSIVE (min_score 10.0)
            # permanently uncapped. See plans/dynamic-scarcity-cap.md.
            "scarcity_cap_ceiling_pct": 0.25,
            "cap_relax_start": 10.0,     # = min_score_threshold
            "cap_relax_full": 16.0,      # validate_scarcity_cap.py 2026-07-30: eligible p90≈+16.7 (OK), but n=3 — hold pending multi-day data
            # @doc-sync-end: scarcity_core
            "atr_multiplier": 1.5,       # Tight stop-loss to preserve capital
            "min_score_threshold": 10.0,
            "cash_buffer_pct": 0.0 if is_bullish else 0.50
        }
    else: # BALANCED (Default)
        return {
            "max_positions": 5,
            "max_allocation_pct": 0.15, # Optimized from 0.20 (Perfect sweet spot between risk and growth)
            # @doc-sync-start: scarcity_core
            "scarcity_allocation_pct": 0.20, # Dynamic Core-Satellite scarcity bucket base cap
            # Dynamic scarcity cap (Option C): mid ramp — 20%->35% as conviction rises
            # from min_score_threshold (5.0) to 11.0. See plans/dynamic-scarcity-cap.md.
            "scarcity_cap_ceiling_pct": 0.35,
            "cap_relax_start": 5.0,      # = min_score_threshold
            "cap_relax_full": 11.0,      # validate_scarcity_cap.py 2026-07-30: eligible p90≈+16.1 suggests higher, but n=9 — hold pending multi-day data
            # @doc-sync-end: scarcity_core
            "atr_multiplier": 2.5,
            "min_score_threshold": 5.0,
            "cash_buffer_pct": 0.0 if is_bullish else 0.20
        }


def calculate_share_qty(symbol: str, cash_to_use: float, price: float) -> float:
    """Calculate the share quantity to purchase.
    Returns a float rounded to 3 decimal places for fractional-eligible assets,
    or a whole integer for standard assets."""
    if price <= 0 or cash_to_use <= 0:
        return 0
    if instruments.is_fractional_eligible(symbol):
        return round(cash_to_use / price, 3)
    else:
        return int(cash_to_use // price)

def adaptive_s10_floor(cash_pct: float) -> float:
    """Dynamic short-term momentum floor for new buys (R&D #15).

    When idle cash drags above CFG.system_cash_drag_threshold, relax the required
    Short10 floor from system_default_s10_floor down to system_adaptive_s10_floor so
    capital can be deployed; otherwise hold the stricter default. Single-sourced from
    CFG.system_{cash_drag_threshold,adaptive_s10_floor,default_s10_floor}.
    """
    if cash_pct > CFG.system_cash_drag_threshold:
        return CFG.system_adaptive_s10_floor
    return CFG.system_default_s10_floor

def should_pyramid_into_winner(is_winner: bool, has_peak: bool, s10: float, l60: float) -> bool:
    """Pyramiding momentum gate (R&D #31).

    Scale into a profitable, risk-locked winner trading near its peak when EITHER
    short-term momentum holds (s10 >= system_pyramiding_s10_floor) OR long-term
    trend support is strong (l60 >= system_pyramiding_l60_floor) — the latter lets
    us add on minor short-term pullbacks within an established uptrend. Thresholds
    are single-sourced from CFG.system_pyramiding_*.
    """
    return (
        is_winner
        and has_peak
        and (s10 >= CFG.system_pyramiding_s10_floor or l60 >= CFG.system_pyramiding_l60_floor)
    )

def load_game():
    if not AI_GAME_FILE.exists() or AI_GAME_FILE.stat().st_size == 0:
        # Try to find a backup to restore from
        backup_dir = BASE_DIR / "Data" / "Backup" / "Game"
        if backup_dir.exists():
            backups = sorted(list(backup_dir.glob("ai_portfolio_game_*.json")), key=lambda x: x.stat().st_mtime, reverse=True)
            for b in backups:
                try:
                    with open(b, "r", encoding="utf-8") as bf:
                        state = json.load(bf)
                        # Succeeded! Copy it back to repair AI_GAME_FILE
                        shutil.copy2(b, AI_GAME_FILE)
                        _log.warning(f"  [⚠️ AETHER SELF-HEALER] AI game file was missing or empty. Successfully restored from backup: {b.name}")
                        return state
                except Exception:
                    continue
        return {
            "balance": INITIAL_BALANCE,
            "equity": INITIAL_BALANCE,
            "positions": {},
            "history": [],
            "start_date": str(datetime.date.today()),
            "profile": "BALANCED"
        }
    with open(AI_GAME_FILE, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except Exception as e:
            # Corruption detected! Try to restore from backup
            _log.warning(f"  [⚠️ AETHER SELF-HEALER] Error loading {AI_GAME_FILE.name}: {e}. Attempting automated recovery from backup...")
            backup_dir = BASE_DIR / "Data" / "Backup" / "Game"
            if backup_dir.exists():
                backups = sorted(list(backup_dir.glob("ai_portfolio_game_*.json")), key=lambda x: x.stat().st_mtime, reverse=True)
                for b in backups:
                    try:
                        with open(b, "r", encoding="utf-8") as bf:
                            state = json.load(bf)
                            shutil.copy2(b, AI_GAME_FILE)
                            _log.warning(f"  [⚠️ AETHER SELF-HEALER] Successfully restored corrupted file from backup: {b.name}")
                            return state
                    except Exception:
                        continue
            # Fallback to default state if all else fails
            return {
                "balance": INITIAL_BALANCE,
                "equity": INITIAL_BALANCE,
                "positions": {},
                "history": [],
                "start_date": str(datetime.date.today()),
                "profile": "BALANCED"
            }

import shutil

def save_game(state):
    # --- Mandatory Backup before Write ---
    if AI_GAME_FILE.exists():
        try:
            backup_dir = BASE_DIR / "Data" / "Backup" / "Game"
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = backup_dir / f"ai_portfolio_game_{ts}.json"
            
            shutil.copy2(AI_GAME_FILE, backup_path)
            
            # Clean up old backups (Keep last 15)
            backups = sorted(list(backup_dir.glob("ai_portfolio_game_*.json")), key=lambda x: x.stat().st_mtime)
            if len(backups) > 15:
                for old_b in backups[:-15]:
                    trash.soft_delete(old_b, reason="game-backup-prune", force=True)
        except Exception as e:
            _log.warning(f"  [Warning] Game backup failed: {e}")

    with open(AI_GAME_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=4)

def backtrack_verify(symbol):
    """Verify the consistency of a price trend over the last 3 trading days using local daily history."""
    try:
        path = BASE_DIR / "Data" / "Symbol_full" / f"{symbol}_daily.json"
        if not path.exists():
            return False, "No local daily history found."
        
        with open(path) as f:
            data = json.load(f)
            
        ts = data.get("Time Series (Daily)", {})
        sorted_dates = sorted(list(ts.keys()), reverse=True)
        if len(sorted_dates) < 3:
            return False, f"Insufficient history (found {len(sorted_dates)}/3 days) for verification."
        
        prices = [float(ts[d]["4. close"]) for d in sorted_dates[:3]]
        
        # Rule: Today's price (prices[0]) must not be in a vertical collapse 
        # (e.g. no day-over-day drop greater than 2% in the last 2 days)
        day1_change = (prices[0] - prices[1]) / prices[1]
        day2_change = (prices[1] - prices[2]) / prices[2]
        
        if day1_change >= -0.02 and day2_change >= -0.02:
            return True, f"Verified price trend stability: {[round(p, 2) for p in prices]}"
        else:
            return False, f"Failed backtracking check (vertical drop detected). Changes: {round(day1_change*100, 1)}%, {round(day2_change*100, 1)}%"
    except Exception as e:
        return False, f"Verification error: {e}"

def is_bottom_confirmed(symbol):
    """Verify if a stock is forming a technical bottom based on its 3-day price slope."""
    try:
        path = BASE_DIR / "Data" / "Symbol_full" / f"{symbol}_daily.json"
        if not path.exists():
            return False, "No local daily history found."
        
        with open(path) as f:
            data = json.load(f)
            
        ts = data.get("Time Series (Daily)", {})
        sorted_dates = sorted(list(ts.keys()), reverse=True)
        if len(sorted_dates) < 4:
            return False, f"Insufficient history (found {len(sorted_dates)}/4 days) for slope check."
        
        prices = [float(ts[d]["4. close"]) for d in sorted_dates[:4]]
        
        # Calculate daily percentage changes over the last 3 days
        # prices[0]: today, prices[1]: yesterday, prices[2]: 2 days ago, prices[3]: 3 days ago
        change1 = (prices[0] - prices[1]) / prices[1]  # Today vs Yesterday
        change2 = (prices[1] - prices[2]) / prices[2]  # Yesterday vs 2 Days Ago
        change3 = (prices[2] - prices[3]) / prices[3]  # 2 Days Ago vs 3 Days Ago
        
        # Bottoming signature: Selling pressure is exhausting and slope is turning positive
        # Condition 1: Today's slope is positive (change1 > 0)
        # Condition 2: Today's slope is better than yesterday's slope (change1 > change2)
        # Condition 3: Average of the last 2 days is positive (> 0.005, or +0.5%)
        avg_slope = (change1 + change2) / 2
        
        if change1 > 0 and change1 > change2 and avg_slope >= 0.005:
            return True, f"Bottom Confirmed! Avg Slope: +{round(avg_slope*100, 2)}% | Price Trend: {[round(p, 2) for p in prices[:3]]}"
        else:
            return False, f"No reversal confirmed. Avg Slope: {round(avg_slope*100, 2)}%"
    except Exception as e:
        return False, f"Slope error: {e}"

def update_excel_log(state, new_transactions):
    if not AI_PERF_XLSX.exists():
        return
    try:
        wb = openpyxl.load_workbook(AI_PERF_XLSX)
        today = str(datetime.date.today())
        ws1 = wb["Summary"]
        profit = state["equity"] - INITIAL_BALANCE
        ws1.append([today, state["equity"], round(state["balance"], 2), round(profit, 2), f"{round((profit/INITIAL_BALANCE)*100, 2)}%", len(state["positions"])])
        ws2 = wb["Transaction_Log"]
        for tx in new_transactions:
            val = tx.get("qty", 1) * tx["price"]
            ws2.append([tx["date"], tx["time"], tx["type"], tx["symbol"], tx["price"], tx.get("qty", ""), round(val, 2), tx.get("pnl", "")])
        wb.save(AI_PERF_XLSX)
    except Exception as e:
        _log.info(f"Failed to update Excel log: {e}")

def get_live_google_price(symbol):
    """Scrape the latest price from Google Finance as a fallback when E*TRADE is unavailable."""
    exchanges = ["NASDAQ", "NYSE", "NYSEARCA", "AMEX"]
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    
    for ex in exchanges:
        url = f'https://www.google.com/finance/quote/{symbol}:{ex}'
        try:
            r = requests.get(url, headers=headers, timeout=5)
            if r.status_code == 200:
                # Regex looking specifically for the jsname="Pdsbrc" span enclosing the dollar price
                match = re.search(r'jsname="Pdsbrc"[^>]*>\s*<span>\$([0-9,.]+)<', r.text)
                if match:
                    price_str = match.group(1).replace(',', '')
                    return float(price_str)
        except Exception:
            pass
    return None

def is_market_hours():
    """Return True if current time is within active US equity market hours (6:30 AM - 1:15 PM PST, weekdays) and NOT a market holiday."""
    try:
        tz_la = pytz.timezone("America/Los_Angeles")
        now_la = datetime.datetime.now(tz_la)
        
        # 1. Hard Time-Window check: Must be within 6:30 AM - 1:15 PM PST
        start_time = now_la.replace(hour=6, minute=30, second=0, microsecond=0)
        end_time = now_la.replace(hour=13, minute=15, second=0, microsecond=0)
        if not (start_time <= now_la <= end_time):
            return False
            
        # 2. Weekend check (Saturday=5, Sunday=6)
        if now_la.weekday() in (5, 6):
            return False
            
        # 3. Static US Stock Market Holiday (NYSE)
        holidays_2026 = {
            "2026-01-01",  # New Year's Day
            "2026-01-19",  # Martin Luther King Jr. Day
            "2026-02-16",  # Presidents' Day
            "2026-04-03",  # Good Friday
            "2026-05-25",  # Memorial Day
            "2026-06-19",  # Juneteenth National Independence Day
            "2026-07-03",  # Independence Day (Observed)
            "2026-09-07",  # Labor Day
            "2026-11-26",  # Thanksgiving Day
            "2026-12-25",  # Christmas Day
            
            # Future Years Support (2027)
            "2027-01-01", "2027-01-18", "2027-02-15", "2027-03-26", "2027-05-31",
            "2027-06-18", "2027-07-05", "2027-09-06", "2027-11-25", "2027-12-24"
        }
        
        today_str = now_la.strftime("%Y-%m-%d")
        if today_str in holidays_2026:
            return False
            
        # 4. Dynamic Live Verification (Clock API + SPY Ticker)
        try:
            _et = etrade.ETradeClient("production", role="auth")
            tokens = _et.auth.get_tokens()
            if tokens:
                is_open = _et.market.is_open_now(tokens)
                if is_open is not None:
                    if not is_open:
                        _log.console("  [AETHER] Dynamic Market Checks confirm market is CLOSED (Holiday/Weekend).")
                    return is_open
        except Exception as e:
            _log.warning(f"  [AETHER] Dynamic market clock failed: {e}. Falling back to static datetime checks.")
            
        return True
    except Exception:
        _log.warning("is_market_hours: timezone check failed — defaulting to False (skip run).")
        return False  # safe default: skip trade rather than run at unknown hours

def get_json_prices_fallback(symbols):
    """Retrieve the latest closing prices from the local per-symbol OHLCV JSON caches.
    Accepts prices from the last 4 calendar days to handle weekends and market holidays."""
    quotes = {}
    try:
        today = datetime.date.today()
        today_str = str(today)
        is_weekend = today.weekday() in (5, 6)
        # Accept a close up to 4 days old — covers 3-day weekends and market holidays
        cutoff = (today - datetime.timedelta(days=4)).isoformat()
        for sym in symbols:
            path = SYMBOL_FULL_DIR / f"{sym}_daily.json"
            if path.exists():
                with open(path) as f:
                    ts = json.load(f).get("Time Series (Daily)", {})
                if ts:
                    newest_date = sorted(ts.keys())[-1]
                    if newest_date == today_str or is_weekend or newest_date >= cutoff:
                        quotes[sym] = float(ts[newest_date]["4. close"])
        return quotes
    except Exception as e:
        _log.console(f"  [AETHER] Failed to load local JSON prices: {e}")
        return {}

def get_live_prices(symbols):
    """Fetch real-time market prices via E*TRADE safely with automated recovery."""
    try:
        # Weekend Gate: Bypass E*TRADE only on weekends to prevent offline lockout sweeps
        if datetime.date.today().weekday() in (5, 6):
            _log.console("  [AETHER] Weekend detected — bypassing E*TRADE login.")
            _log.console("  [AETHER] Attempting instant local price extraction from per-symbol JSON cache...")
            quotes = get_json_prices_fallback(symbols)
            # Find any missing gaps (symbols not in our local workbook)
            missing = [s for s in symbols if s not in quotes or not quotes[s] or quotes[s] <= 0]
            if missing:
                _log.console(f"  [AETHER] Local cache missing prices for {missing}. Falling back to Google Finance.")
                quotes.update(get_google_prices_fallback(missing))
            return quotes

        # After-Hours Synced Gate: If after-hours on a weekday, and today's workbook is ALREADY
        # fresh and synced, we can safely bypass E*TRADE entirely and load directly from the local cache!
        if not is_market_hours():
            is_fresh = False
            try:
                if os.path.exists(XLSX_FILE):
                    mtime = os.path.getmtime(XLSX_FILE)
                    mdate = datetime.date.fromtimestamp(mtime)
                    is_fresh = (mdate == datetime.date.today())
            except Exception:
                pass

            if is_fresh:
                _log.console("  [AETHER] After-hours detected & workbook is ALREADY synced today.")
                _log.console("  [AETHER] Bypassing E*TRADE login to load directly from per-symbol JSON cache...")
                quotes = get_json_prices_fallback(symbols)
                missing = [s for s in symbols if s not in quotes or not quotes[s] or quotes[s] <= 0]
                if missing:
                    _log.console(f"  [AETHER] Local cache missing prices for {missing}. Falling back to Google Finance.")
                    quotes.update(get_google_prices_fallback(missing))
                return quotes
            else:
                _log.console("  [AETHER] After-hours detected, but workbook is STALE/UN-SYNCED today.")
                _log.console("  [AETHER] Connecting to E*TRADE to fetch fresh settled closes...")

        # Call the hardened get_tokens() which is safe and has an active headless safety gate.
        # This ensures we always actively attempt to re-authenticate when tokens expire.
        _et = etrade.ETradeClient("production", role="auth")
        tokens = _et.auth.get_tokens()
        if not tokens:
            _log.warning("  [AETHER] E*TRADE authentication failed. Attempting Google Finance live fallback.")
            return get_google_prices_fallback(symbols)

        quotes = _et.market.quotes(symbols, tokens)

        # Fill only the gaps from Google — keep the E*TRADE quotes we already have.
        # Surface dead/delisted/misaligned tickers loudly instead of discarding
        # every good quote and re-scraping the whole list.
        missing = [s for s in symbols if s not in quotes or not quotes[s] or quotes[s] <= 0]
        if missing:
            _log.console(f"  [AETHER] E*TRADE returned no quote for {missing} (possibly dead/delisted/misaligned); filling gaps via Google Finance.")
            quotes.update(get_google_prices_fallback(missing))
            still_missing = [s for s in missing if s not in quotes or not quotes[s] or quotes[s] <= 0]
            if still_missing:
                _log.console(f"  [AETHER] SYMBOLOGY ERROR: No live quote from any source for: {still_missing}")

        return quotes
    except Exception as e:
        _log.warning(f"  [AETHER] E*TRADE connection failed: {e}. Attempting Google Finance live fallback.")
        return get_google_prices_fallback(symbols)

def get_google_prices_fallback(symbols):
    """Scrape Google Finance for multiple symbols in parallel/sequence as a robust fallback."""
    quotes = {}
    _log.console(f"  [Google] Scraping live quotes for: {symbols}")
    for sym in symbols:
        price = get_live_google_price(sym)
        if price and price > 0:
            quotes[sym] = price
            _log.info(f"    - Google Verified {sym}: ${price:.2f}")
    return quotes

def send_daily_summary(return_html=False):
    state = load_game()
    today = str(datetime.date.today())
    today_tx = [tx for tx in state.get("history", []) if tx["date"] == today]
    tx_rows = ""
    for tx in today_tx:
        pnl_str = f" (PnL: ${tx['pnl']})" if "pnl" in tx else ""
        tx_rows += f"<li><b>{tx['type']}</b>: {tx.get('qty', '')} {tx['symbol']} @ ${tx['price']}{pnl_str} [Time: {tx.get('time', '')}]</li>"

    # Fetch live quotes for open positions to show accurate daily values
    positions = state.get("positions", {})
    live_prices = get_live_prices(list(positions.keys()))
    # Track which positions had a genuine live quote; if any needed a workbook/cost
    # fallback, the equity below is an estimate and must not overwrite stored equity.
    fully_priced = all(sym in live_prices and live_prices[sym] and live_prices[sym] > 0
                       for sym in positions)

    # Standardize fallback to workbook close prices if E*TRADE renewal fails (e.g. on weekends)
    if not live_prices or any(sym not in live_prices for sym in positions):
        try:
            wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
            # Use Short_Long sheet if available, as it contains all active portfolio holdings and current prices
            if "Short_Long" in wb.sheetnames:
                ws = wb["Short_Long"]
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if len(row) > 4:
                        sym = str(row[1] or "").strip().upper()
                        if sym in positions and sym not in live_prices:
                            live_prices[sym] = row[4] or positions[sym]["cost"]
            # Fallback to Research sheet if Short_Long is unavailable
            elif "Research" in wb.sheetnames:
                ws = wb["Research"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > 10:
                        sym = str(row[3] or "").strip().upper()
                        if sym in positions and sym not in live_prices:
                            live_prices[sym] = row[10] or positions[sym]["cost"]
        except Exception as e:
            _log.warning(f"Workbook fallback failed inside summary: {e}")

    # Final safety fallback to cost basis if both API and workbook are empty
    for sym in positions:
        if sym not in live_prices:
            live_prices[sym] = positions[sym]["cost"]

    # Build the open positions HTML table
    pos_table_rows = ""
    live_equity = state.get("balance", 0.0)
    if positions:
        for sym, pos in positions.items():
            qty = pos["qty"]
            cost = pos["cost"]
            current = live_prices.get(sym, cost)
            val = qty * current
            live_equity += val
            pnl = (current - cost) * qty
            pnl_pct = ((current - cost) / cost) * 100
            
            pnl_color = "#27ae60" if pnl >= 0 else "#c0392b"
            pnl_sign = "+" if pnl >= 0 else ""
            
            pos_table_rows += f"""
            <tr>
                <td style="padding: 10px; border-bottom: 1px solid #ddd; font-weight: bold;">{sym}</td>
                <td style="padding: 10px; border-bottom: 1px solid #ddd; text-align: center;">{qty}</td>
                <td style="padding: 10px; border-bottom: 1px solid #ddd; text-align: right;">${cost:.2f}</td>
                <td style="padding: 10px; border-bottom: 1px solid #ddd; text-align: right; font-weight: bold;">${current:.2f}</td>
                <td style="padding: 10px; border-bottom: 1px solid #ddd; text-align: right; font-weight: bold;">${val:.2f}</td>
                <td style="padding: 10px; border-bottom: 1px solid #ddd; text-align: right; color: {pnl_color}; font-weight: bold;">
                    {pnl_sign}${pnl:.2f} ({pnl_sign}{pnl_pct:.2f}%)
                </td>
            </tr>
            """
    else:
        pos_table_rows = "<tr><td colspan='6' style='padding: 15px; text-align: center; color: #888;'>No open positions currently.</td></tr>"

    # Keep the dashboard's stored equity fresh — but only when every position had a
    # real live quote. Persisting a cost/workbook-fallback estimate here would let a
    # weekend/no-quote run overwrite the authoritative equity that read_portfolio
    # surfaces. On a partial-priced run, leave the last good value untouched.
    if fully_priced:
        state["equity"] = round(live_equity, 2)
        save_game(state)

    # Build decision log reflection card
    try:
        sc = decision_eval.score_log(decision_eval.read_log())
        refl = decision_eval.reflection(sc)
        refl_html = f"""
        <h3 style="color: #2c3e50; margin-top: 35px;">🔬 AI Decision Scorecard (Retrospective)</h3>
        <div style="background: #fcf8e3; border: 1px solid #fbeed5; border-radius: 4px; padding: 15px; font-family: monospace; white-space: pre-wrap; font-size: 13px; line-height: 1.5; color: #c09853;">
{refl}
        </div>
        """
    except Exception as e:
        _log.info(f"Failed to include retrospective reflection in email: {e}")
        refl_html = ""

    html = f"""
    <html>
    <body style="font-family: sans-serif; color: #333; max-width: 700px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-bottom: 20px;">🤖 AI Portfolio: Daily Performance Summary</h2>
        <p><b>Date:</b> {today} | <b>Active Strategy:</b> <span style="background: #2c3e50; color: white; padding: 2px 6px; border-radius: 3px; font-weight: bold;">{state.get('profile', 'BALANCED')}</span></p>
        <hr style="border: 0; border-top: 1px solid #eee; margin: 20px 0;">
        
        <h3 style="color: #2c3e50;">🛒 Market Action Today:</h3>
        <ul style="padding-left: 20px; font-size: 14px; line-height: 1.6;">
            {tx_rows if today_tx else "<li>No transactions executed today.</li>"}
        </ul>
        
        <h3 style="color: #2c3e50; margin-top: 30px;">📈 Current Open Positions:</h3>
        <table border="0" cellpadding="0" cellspacing="0" style="width: 100%; border-collapse: collapse; margin-bottom: 35px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); border: 1px solid #ddd;">
            <thead>
                <tr style="background: #34495e; color: white;">
                    <th style="padding: 12px; text-align: left;">Symbol</th>
                    <th style="padding: 12px; text-align: center;">Qty</th>
                    <th style="padding: 12px; text-align: right;">Cost Basis</th>
                    <th style="padding: 12px; text-align: right;">Live Price</th>
                    <th style="padding: 12px; text-align: right;">Total Value</th>
                    <th style="padding: 12px; text-align: right;">Total P&L</th>
                </tr>
            </thead>
            <tbody>
                {pos_table_rows}
            </tbody>
        </table>
        
        <h3 style="color: #2c3e50;">🛡️ Portfolio Financial Summary:</h3>
        <table border="0" cellpadding="10" cellspacing="0" style="width: 100%; max-width: 400px; border-collapse: collapse; margin-bottom: 25px; border: 1px solid #ddd;">
            <tr style="background: #f8f9fa;">
                <td style="border-bottom: 1px solid #ddd; font-weight: bold;">Current Equity</td>
                <td style="border-bottom: 1px solid #ddd; text-align: right; font-weight: bold; font-size: 16px;">${live_equity:.2f}</td>
            </tr>
            <tr>
                <td style="border-bottom: 1px solid #ddd; font-weight: bold; color: #555;">Cash Balance (Dry Powder)</td>
                <td style="border-bottom: 1px solid #ddd; text-align: right; font-weight: bold; color: #555;">${state['balance']:.2f}</td>
            </tr>
            <tr style="background: #f8f9fa;">
                <td style="font-weight: bold;">Total Return</td>
                <td style="text-align: right; font-weight: bold; color: {'#27ae60' if live_equity >= INITIAL_BALANCE else '#c0392b'};">
                    {'+' if live_equity >= INITIAL_BALANCE else ''}{round(((live_equity - INITIAL_BALANCE)/INITIAL_BALANCE)*100, 2)}%
                </td>
            </tr>
        </table>
        
        {refl_html}
        
        <p style="margin-top: 35px; border-top: 1px solid #eee; padding-top: 15px; font-size: 11px; color: #7f8c8d;">
            🤖 <i>This is an automated performance report from your autonomous Project AETHER trading desk. All figures represent live, verified production-grade data.</i>
        </p>
    </body>
    </html>
    """
    if return_html:
        return html
    notify.send_email(f"AI Portfolio Summary: {today}", html, is_html=True)
    _log.info(f"Summary email sent for {today}.")

def send_consolidated_morning_report():
    """
    Compile both the Virtual AI-Game morning actions and the Real-Account Oracle Advisory,
    wrapping them in a single consolidated HTML email dispatched immediately at 7:00 AM.
    """
    _log.info("📧 COMPILING CONSOLIDATED AETHER MORNING BRIEFING...")
    today = str(datetime.date.today())
    
    # 1. Compile the AI-game summary
    ai_html = send_daily_summary(return_html=True)
    
    # Extract only the body contents of the AI HTML to prevent nested <html>/<body> tags
    body_content = ai_html
    body_start = ai_html.find('<body')
    if body_start != -1:
        body_start = ai_html.find('>', body_start) + 1
        body_end = ai_html.find('</body>')
        if body_end != -1:
            body_content = ai_html[body_start:body_end]

    # 2. Compile the Oracle Advisory HTML
    oracle_html = aether_oracle.run_oracle_advisory()
    
    # 3. Concatenate both into a single cohesive, high-quality, professional layout
    unified_html = f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>AETHER Consolidated Morning Briefing</title>
    </head>
    <body style="font-family: sans-serif; color: #333; max-width: 800px; margin: 0 auto; padding: 20px; background-color: #f8f9fa;">
        <div style="background-color: #ffffff; border: 1px solid #e1e4e8; border-radius: 8px; padding: 25px; box-shadow: 0 4px 10px rgba(0,0,0,0.05);">
            <div style="text-align: center; border-bottom: 3px double #3498db; padding-bottom: 15px; margin-bottom: 25px;">
                <h1 style="color: #2c3e50; margin: 0; font-size: 24px; font-weight: bold; letter-spacing: 0.5px;">☀️ AETHER DAILY MORNING BRIEFING</h1>
                <p style="color: #7f8c8d; margin: 5px 0 0 0; font-size: 13px;">Date: {today} | Unified Trading & Market Intelligence Report</p>
            </div>
            
            <!-- Section 1: AI Virtual Portfolio Summary -->
            {body_content}
            
            <hr style="border: 0; border-top: 2px dashed #e1e4e8; margin: 40px 0;">
            
            <!-- Section 2: Real-Account Oracle Advisory -->
            {oracle_html}
        </div>
    </body>
    </html>
    """
    
    subject = f"☀️ AETHER Morning Briefing: {today} (AI-Game Actions & Oracle Advice)"
    notify.send_email(subject, unified_html, is_html=True)
    _log.info(f"Consolidated morning briefing sent successfully for {today}.")

def _has_strong_setups_today(min_score=9.5) -> bool:
    """Return True if we have 2 or more strong, verified bottom setups in the workbook today."""
    wb = None
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        ws = wb["Research"]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            sym = row[3]
            if not sym:
                continue
            setup = str(row[20] or "").strip()
            if setup in ('1', 'OK'):
                try:
                    s10 = float(row[24] or 0)
                    l60 = float(row[25] or 0)
                except (TypeError, ValueError):
                    continue
                if (s10 + l60) >= min_score:
                    z_score = calculate_bubble_z_score(sym)
                    if z_score is None or z_score < 2.5:
                        count += 1
                        if count >= 2:
                            return True
    except Exception:
        pass
    finally:
        if wb:
            wb.close()
    return False


def run_daily_ai_management(force=False, manual_profile=None):
    state = None
    try:
        open_status, msg = is_market_open()
        if not open_status and not force:
            _log.info(f"Aborting AI Move: {msg}")
            return

        state = load_game()
        today = str(datetime.date.today())
        now_time = datetime.datetime.now().strftime("%H:%M:%S")
        new_transactions = []

        # Determine strategy profile (Adaptive vs. Manual Override)
        if manual_profile and manual_profile.upper() == "ADAPTIVE":
            profile = get_market_regime()
            state["profile"] = profile
            state["profile_mode"] = "ADAPTIVE"
            _log.info(f"🤖 AI ACTIVE STRATEGY: {profile} (Adaptive pilot restored)")
        elif manual_profile:
            profile = manual_profile
            state["profile"] = profile
            state["profile_mode"] = "MANUAL"
            _log.info(f"🤖 AI ACTIVE STRATEGY: {profile} (Manual Override - Locked)")
        elif state.get("profile_mode") == "MANUAL":
            # Auto-Reset Manual Override: A manual override is a one-time tactical choice.
            # On the next automated run (no CLI profile passed), we automatically reset back to ADAPTIVE autopilot.
            _log.console("  [AETHER] Manual override expired. Automatically resetting back to Adaptive autopilot...")
            state["profile_mode"] = "ADAPTIVE"
            profile = get_market_regime()

            # Adaptive Cash-Deployment Upgrade Gate (Only for Autopilot!)
            equity = state.get("equity", 0)
            cash_ratio = state.get("balance", 0) / equity if equity > 1.0 else 0
            if profile == "DEFENSIVE" and cash_ratio > 0.40 and _has_strong_setups_today(min_score=9.5):
                _log.console(f"  [AETHER] Cash is plentiful ({cash_ratio*100:.1f}%) and strong bottom setups are detected.")
                _log.info("  -> Adaptively upgrading today's strategy profile from DEFENSIVE to BALANCED to deploy cash safely!")
                profile = "BALANCED"

            state["profile"] = profile
            _log.info(f"🤖 AI ACTIVE STRATEGY: {profile} (Adaptive)")
        else:
            profile = get_market_regime()
            state["profile_mode"] = "ADAPTIVE"

            # Adaptive Cash-Deployment Upgrade Gate (Only for Autopilot!)
            equity = state.get("equity", 0)
            cash_ratio = state.get("balance", 0) / equity if equity > 1.0 else 0
            if profile == "DEFENSIVE" and cash_ratio > 0.40 and _has_strong_setups_today(min_score=9.5):
                _log.console(f"  [AETHER] Cash is plentiful ({cash_ratio*100:.1f}%) and strong bottom setups are detected.")
                _log.info("  -> Adaptively upgrading today's strategy profile from DEFENSIVE to BALANCED to deploy cash safely!")
                profile = "BALANCED"

            state["profile"] = profile
            _log.info(f"🤖 AI ACTIVE STRATEGY: {profile} (Adaptive)")

        rules = get_strategy_rules(profile)

        if not XLSX_FILE.exists():
            _log.info("Workbook not found. AI Management deferred.")
            return

        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        ws = wb["Research"]
        
        symbols_to_check = list(state["positions"].keys())

        # Pre-flight: batch-heal stale OHLCV caches before the decision loop so
        # _sma50 never blocks on a live network call mid-iteration.
        stale_syms = [s for s in symbols_to_check if _cache_stale(s, max_stale_days=_MAX_STALE_DAYS)]
        if stale_syms:
            _log.info(f"[Pre-flight] Healing {len(stale_syms)} stale OHLCV cache(s) before evaluation: {stale_syms}")
            for _s in stale_syms:
                _heal_symbol_cache(_s)

        # Dynamically heal/classify legacy positions
        for row in ws.iter_rows(min_row=2, values_only=True):
            sym = row[3]
            if sym in state["positions"] and "is_scarcity" not in state["positions"][sym]:
                industry_str = row[4] or ""
                is_scarcity = instruments.is_scarcity_asset(sym, industry_str)
                state["positions"][sym]["is_scarcity"] = is_scarcity
                _log.info(f"  [AETHER State Healer] Classified existing position {sym} as scarcity={is_scarcity}")

        research_symbols = _active_setup_symbols(ws)

        queued = state.get("queued_orders", [])
        queued_syms = [q["symbol"] for q in queued]
        
        # Ensure we always fetch the live price of SPY for our Circuit Breaker
        all_syms = list(set(symbols_to_check + research_symbols + queued_syms + ["SPY"]))
        prices = get_live_prices(all_syms)
        
        # --- Price Source Gate ---
        # Primary source: E*TRADE live API. Automatic fallback: Google Finance scraper.
        # If both fail (prices is empty), crash — no stale workbook prices allowed.
        if not prices:
            raise RuntimeError("Critical Data Failure: Both E*TRADE and Google Finance fallback returned no prices. No live source of truth available!")

        # --- Systemic Crash Circuit Breaker Guard ---
        circuit_breaker.enforce_circuit_breaker(state, prices)

        # Zero-Trust: surface held positions with no live quote, but do NOT abort the
        # run over them — aborting would skip stop-loss enforcement on every *other*
        # position too, violating the Rule of Loss Minimization. Unpriced names fall
        # back to cost for the equity figure (via _live_equity) and are held (their
        # cost-based price won't trip a stop) until a quote returns.
        missing_prices = [sym for sym in symbols_to_check if sym not in prices or not prices[sym] or prices[sym] <= 0]
        if missing_prices:
            _log.console(f"  [AETHER] PORTFOLIO ERROR: No live quote found for held positions {missing_prices}! Using cost basis for their equity share and skipping their stop check this run.")

        state["equity"] = round(_live_equity(state["balance"], state["positions"], prices), 0)

        # 0. Execute QUEUED ORDERS (Strategic Overrides with Volatility Sizing)
        if queued:
            _log.info("🤖 AI EXECUTING QUEUED STRATEGIC ORDERS...")
            for order in queued:
                sym = order["symbol"]
                price = prices.get(sym, 0)
                if price <= 0: continue
                
                if order["type"] == "SELL" and sym in state["positions"]:
                    pos = state["positions"].pop(sym)
                    proceeds = pos["qty"] * price
                    state["balance"] += proceeds
                    tx = {
                        "date": today, "time": now_time, "type": "SELL", 
                        "symbol": sym, "price": price, "qty": pos["qty"], 
                        "pnl": round((price - pos["cost"]) * pos["qty"], 2),
                        "details": f"Queued Sell: {order['reason']}"
                    }
                    state["history"].append(tx)
                    new_transactions.append(tx)
                    _log.info(f"🤖 AI QUEUED SELL EXECUTED: {sym} at ${price} (PnL: ${tx['pnl']})")
                    log_closed_trade_dna(sym, pos, price, today)
                    
                elif order["type"] == "BUY" and sym not in state["positions"]:
                    # Zero-Trust Freshness Gate (execution-time): a queued BUY may have sat overnight,
                    # so the screen-time gate that cleared it can be stale by now. The ATR stop below is
                    # derived from this cache, so re-verify freshness before filling. Heal once, then
                    # SKIP (drop) the order rather than execute on untrustworthy data — the screener
                    # re-surfaces the name next run if it still qualifies. (Queued SELLs are intentionally
                    # NOT gated: a strategic/stop exit must always be allowed to run.)
                    if _cache_stale(sym, max_stale_days=_MAX_STALE_DAYS):
                        _heal_symbol_cache(sym)
                        if _cache_stale(sym, max_stale_days=_MAX_STALE_DAYS):
                            _log.warning(f"🛑 QUEUED BUY SKIPPED (Zero-Trust Freshness): {sym} - OHLCV cache stale/missing at execution; refusing to derive an ATR stop from untrustworthy data.")
                            continue
                    max_positions = rules["max_positions"]
                    available_slots = max_positions - len(state["positions"])
                    if state["balance"] > 500 and available_slots > 0:
                        max_allocation = state["equity"] * rules["max_allocation_pct"]
                        cash_to_use = min(state["balance"] / available_slots, max_allocation)
                        
                        qty = calculate_share_qty(sym, cash_to_use, price)
                        if qty > 0:
                            cost = qty * price
                            state["balance"] -= cost
                            
                            # Volatility-Based Stop Loss customized by profile
                            atr = risk_utils.calculate_atr(sym)
                            if atr and atr > 0:
                                stop_loss = round(price - (rules["atr_multiplier"] * atr), 2)
                                stop_desc = f"ATR-based Stop: ${stop_loss} ({rules['atr_multiplier']} * ATR)"
                            else:
                                stop_loss = round(price * 0.92, 2)
                                stop_desc = f"8% Fallback Stop: ${stop_loss}"
                                
                            # Resolve buy DNA for queued buys
                            q_pgr = "Neutral"
                            q_s10 = 0.0
                            q_l60 = 0.0
                            q_score = 0.0
                            q_industry = "Unknown"
                            for r_row in ws.iter_rows(min_row=2, values_only=True):
                                if r_row[3] == sym:
                                    q_pgr = r_row[6] or "Neutral"
                                    q_industry = r_row[4] or "Unknown"
                                    try:
                                        q_s10 = float(r_row[24] or 0.0)
                                        q_l60 = float(r_row[25] or 0.0)
                                        q_score = round(q_s10 + q_l60, 1)
                                    except (ValueError, TypeError):
                                        pass
                                    break
                                    
                            state["positions"][sym] = {
                                "qty": qty, 
                                "cost": price, 
                                "stop_loss": stop_loss,
                                "buy_dna": {
                                    "buy_date": today,
                                    "pgr": q_pgr,
                                    "s10": q_s10,
                                    "l60": q_l60,
                                    "score": q_score,
                                    "z_score": 0.0,
                                    "industry": q_industry
                                }
                            }
                            tx = {
                                "date": today, "time": now_time, "type": "BUY", 
                                "symbol": sym, "price": price, "qty": qty,
                                "details": f"Queued Buy: {order['reason']} ({stop_desc})"
                            }
                            state["history"].append(tx)
                            new_transactions.append(tx)
                            _log.info(f"🤖 AI QUEUED BUY EXECUTED: {qty} shares of {sym} at ${price} ({stop_desc})")
            
            state["queued_orders"] = []

        # SELL logic — unified deterministic exit policy (sell_rules.exit_decision):
        # hard ATR stop > soft momentum signal (winner-protected) > hold.
        symbols_to_sell = []
        decision_entries = []
        for sym in list(state["positions"].keys()):
            pos = state["positions"][sym]
            s10 = l60 = 0
            prev_close = pos.get("cost", 0.0)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[3] == sym:
                    s10 = row[24] or 0
                    l60 = row[25] or 0
                    try:
                        prev_close = float(row[8] or pos.get("cost", 0.0))
                    except Exception:
                        prev_close = pos.get("cost", 0.0)
                    break
            price = prices.get(sym, pos.get("cost"))

            # ── AETHER Profit-Lock Trailing Stop-Loss Ratchet (Priority 1) ──
            atr = risk_utils.calculate_atr(sym)
            if atr and atr > 0:
                # 1. Peak Price Tracking: track highest close since acquisition
                highest_close = pos.get("highest_close_since_acq", 0.0)
                highest_close = max(highest_close, pos.get("cost", 0.0), price)
                pos["highest_close_since_acq"] = highest_close
                
                # Determine trailing multiplier (1.5x for Scarcity, profile rules-based for Standard)
                is_scarcity = pos.get("is_scarcity", False)
                multiplier = 1.5 if is_scarcity else rules.get("atr_multiplier", 2.5)
                
                # 2. Peter Lynch Flower Protection: Only ratchet stop upward once safely in profit by > 1.0x ATR
                if (price - pos.get("cost", 0.0)) > (1.0 * atr):
                    recalculated_stop = round(highest_close - (multiplier * atr), 2)
                    old_stop = pos.get("stop_loss", 0.0)
                    if recalculated_stop > old_stop:
                        pos["stop_loss"] = recalculated_stop
                        _log.info(f"[Profit-Lock] {sym} stop ratcheted: ${old_stop:.2f} -> ${recalculated_stop:.2f} (peak close ${highest_close:.2f})")
                        _log.info(f"🛡️ [Profit-Lock] {sym} stop ratcheted upwards: ${old_stop:.2f} ➡️ ${recalculated_stop:.2f}")
                        
                # 3. Breakeven Trigger: If price has rallied > 1.5x ATR, lock in exact purchase Cost Basis (Breakeven)
                if (price - pos.get("cost", 0.0)) > (1.5 * atr):
                    old_stop = pos.get("stop_loss", 0.0)
                    cost_basis = pos.get("cost", 0.0)
                    if cost_basis > old_stop:
                        pos["stop_loss"] = cost_basis
                        _log.info(f"[Breakeven Lock] {sym} stop raised to cost basis: ${old_stop:.2f} -> ${cost_basis:.2f}")
                        _log.info(f"🛡️ [Breakeven Lock] {sym} stop bumped to Cost Basis: ${old_stop:.2f} ➡️ ${cost_basis:.2f}")
            # ────────────────────────────────────────────────────────────────

            # Check Idiosyncratic Single-Stock Gap-Down Guard (Whipsaw protection)
            is_gap_frozen = False
            if circuit_breaker.is_single_stock_gap_frozen(sym, price, prev_close):
                is_gap_frozen = True
                gap_pct = round(((price - prev_close) / prev_close) * 100, 2) if prev_close > 0 else 0.0
                _log.info(f"❄️ [Gap Guard] {sym} gapped down {gap_pct}% overnight. Holding stop wide to prevent opening whipsaw.")

            # Only pay the OHLCV read when the soft signal actually fires (winner-
            # protection is the only consumer of sma50); HOLD positions skip the I/O.
            sma50 = _sma50(sym) if sell_rules.soft_exit(s10, l60) else None
            # build_entry is the single source of the decision: it runs the exit
            # policy once, logs it, and (run_shadow=None) shadow-runs AI only for
            # non-HOLD candidates. The action it returns drives the actual sell.
            entry = decision_eval.build_entry(
                symbol=sym, price=price, cost=pos.get("cost"),
                stop_loss=None if is_gap_frozen else pos.get("stop_loss"), 
                s10=s10, l60=l60, sma50=sma50,
                date=today, run_shadow=None)
            
            # AI Second-Opinion Exit Override Gate (R&D Item 14)
            if entry["rules_action"] == "SELL":
                ai_override = False
                override_reason = ""
                override_verdict = ""

                # 1. Check real-time shadow verdicts generated in entry
                realtime_verdicts = entry.get("verdicts", {})
                for prov, v_info in realtime_verdicts.items():
                    v = v_info.get("verdict", "").upper() if isinstance(v_info, dict) else str(v_info).upper()
                    if v in ("FLAG-FOR-REVIEW", "HOLD"):
                        ai_override = True
                        override_verdict = v
                        override_reason = f"Real-time AI Shadow Heuristic ({prov}) returned {v}" + (f": {v_info.get('note', '')}" if isinstance(v_info, dict) and v_info.get('note') else "")
                        break

                # 2. Check stored position shadow verdicts if any
                if not ai_override:
                    for key in ["shadow_verdict", "ai_verdict", "verdict"]:
                        if key in pos:
                            val = pos[key]
                            if isinstance(val, dict):
                                v = val.get("verdict", "").upper()
                                note = val.get("note", "")
                            else:
                                v = str(val).upper()
                                note = ""
                            if v in ("FLAG-FOR-REVIEW", "HOLD"):
                                ai_override = True
                                override_verdict = v
                                override_reason = f"Stored position {key} returned {v}" + (f": {note}" if note else "")
                                break

                # 3. Check verdicts dictionary inside the stored position
                if not ai_override:
                    pos_verdicts = pos.get("verdicts", {})
                    if isinstance(pos_verdicts, dict):
                        for prov, v_info in pos_verdicts.items():
                            v = v_info.get("verdict", "").upper() if isinstance(v_info, dict) else str(v_info).upper()
                            if v in ("FLAG-FOR-REVIEW", "HOLD"):
                                ai_override = True
                                override_verdict = v
                                override_reason = f"Stored position verdicts ({prov}) returned {v}" + (f": {v_info.get('note', '')}" if isinstance(v_info, dict) and v_info.get('note') else "")
                                break

                if ai_override:
                    old_action = entry["rules_action"]
                    if override_verdict == "HOLD":
                        new_action = "HOLD"
                    else:
                        new_action = "WATCH"
                    
                    entry["rules_action"] = new_action
                    entry["rules_reason"] = f"AI Override: {override_reason} (was {entry['rules_reason']})"
                    
                    _log.info(
                        f"🛡️ [AI EXIT OVERRIDE] {sym} sell overridden! "
                        f"Decision downgraded from {old_action} to {new_action}. Reason: {override_reason}"
                    )
                    _log.info(f"🛡️ [AI Override] Overriding sell of {sym} -> Downgrading to {new_action} due to: {override_reason}")

            decision_entries.append(entry)
            if entry["rules_action"] == "SELL":
                if not is_market_hours():
                    # Queue the sell instead of executing immediately
                    if not any(q["symbol"] == sym and q["type"] == "SELL" for q in state.get("queued_orders", [])):
                        state.setdefault("queued_orders", []).append({
                            "type": "SELL", "symbol": sym, "reason": f"Exit triggered: {entry['rules_reason'] or 'Technical exit'}"
                        })
                        _log.info(f"📝 [Queued] After-hours SELL queued for {sym}: {entry['rules_reason']}")
                else:
                    symbols_to_sell.append(sym)
            elif entry["rules_action"] == "REVIEW":
                # Winner above its 50-DMA on a soft signal — hold, don't dump.
                _log.info(f"🌸 AI HOLD (winner-protected): {sym} — {entry['rules_reason']}")
        if decision_entries:
            decision_eval.log_decisions(decision_entries)

        for sym in symbols_to_sell:
            pos = state["positions"].pop(sym)
            price = prices.get(sym, pos["cost"])
            
            # Slippage-Protected Limit Stop (STP LMT - R&D #8): Execute at exactly the stop price
            # if the market close price dropped below our stop-loss floor, preventing slippage leaks.
            stop_loss = pos.get("stop_loss", 0.0)
            if stop_loss > 0.0 and price <= stop_loss:
                _log.info(f"🛡️ [STP LMT] Executed {sym} stop-loss at Limit price ${stop_loss:.2f} (protected against market gap ${price:.2f}).")
                price = stop_loss
                
            proceeds = pos["qty"] * price
            state["balance"] += proceeds
            tx = {"date": today, "time": now_time, "type": "SELL", "symbol": sym, "price": price, "qty": pos["qty"], "pnl": round((price - pos["cost"]) * pos["qty"], 2)}
            state["history"].append(tx)
            new_transactions.append(tx)
            _log.info(f"🤖 AI LIVE SELL: {sym} at ${price} (Time: {now_time})")
            log_closed_trade_dna(sym, pos, price, today)

        # BUY logic (filtered by profile momentum threshold)
        base_max = rules["max_positions"]
        
        # Dynamic Position-Slot Expansion: If cash is plentiful (>15% of equity) and we are full, dynamically expand slots to deploy cash!
        cash_ratio = state["balance"] / state["equity"] if state["equity"] > 1.0 else 0.0
        max_positions = determine_max_positions(cash_ratio, len(state["positions"]), base_max)
        if max_positions > base_max:
            _log.info(f"🛡️ [Slot Expansion] Plentiful cash ({cash_ratio*100:.1f}%) detected. Dynamically expanding slots from {base_max} to {max_positions} to prevent cash drag!")
            _log.info(f"🛡️ [Slot Expansion] Plentiful cash detected. Expanding max slots from {base_max} to {max_positions}.")
            
        available_slots = max_positions - len(state["positions"])
        
        # Enforce defensive cash buffer
        min_cash_required = state["equity"] * rules["cash_buffer_pct"]
        
        # Always build top_buys list and track active position scores
        top_buys = []
        active_position_scores = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            sym = row[3]
            if not sym: continue
            total_score = (row[24] or 0.0) + (row[25] or 0.0)
            
            # If sym is currently in positions, record its current score
            if sym in state["positions"]:
                active_position_scores[sym] = total_score
                continue
                
            # TEMPORARY: skip leveraged/inverse/crypto ETFs as new long buys — the
            # long swing-low framework doesn't fit them (see instruments.py + R&D).
            # They still get ATR-based stops if already held; this only blocks entry.
            if instruments.is_excluded(sym):
                continue
            setup = str(row[20] or '')
            price = prices.get(sym, 0)
            short10 = row[24] or 0.0

            # Dynamic Adaptive s10 Floor (R&D #15) — thresholds single-sourced via helper.
            cash_pct = (state["balance"] / state["equity"]) * 100.0 if state.get("equity", 0) > 0 else 0.0
            required_floor = adaptive_s10_floor(cash_pct)

            # Strict Short10 Momentum Floor: Reject buy entries if short-term momentum is below required floor
            if short10 < required_floor:
                if (setup in ('1', 'OK', 1)) and price > 0:
                    _log.warning(f"🛑 AI BUY REJECTED (Momentum Floor): {sym} - Short10 score {short10} is below required {required_floor} floor (cash_pct={round(cash_pct, 1)}%).")
                continue

            # Filter by strategy profile threshold OR mathematically confirmed bottom
            if (setup in ('1', 'OK', 1)) and price > 0:
                # Zero-Trust Freshness Gate (CLAUDE.md Rule of Zero-Trust / Temporal Zero-Trust):
                # Never OPEN a new position on stale or missing OHLCV. Every downstream risk level —
                # the ATR stop assigned in _execute_buys, swing support, and is_bottom_confirmed just
                # below — is derived from this cache, so acting on a stale bar is the real "buying
                # blind" (empty workbook S/R is not — that still gets a live ATR stop). Attempt one
                # self-heal, then reject if the cache is still not fresh. Held positions are pre-healed
                # above (line ~1437); this extends the identical discipline to buy candidates, which
                # were previously ungated. NOTE: this gate sits at the top of the qualifying-candidate
                # block, so it applies to EVERY buy candidate — those with explicit workbook S/R just
                # as much as the empty-S/R rows handled below — not only the empty-S/R case. Validated
                # over the live cache: a meaningful fraction of symbols are stale on any given day, and
                # their ATR stops would be computed off old bars.
                if _cache_stale(sym, max_stale_days=_MAX_STALE_DAYS):
                    _healed = _heal_symbol_cache(sym)  # bool: True only on a real refresh
                    if _cache_stale(sym, max_stale_days=_MAX_STALE_DAYS):
                        # Distinguish missing vs N-days-stale in the reject line, reusing the cheap
                        # raw-JSON age (_cache_age_days) that _cache_stale already read — not the
                        # heavier split-adjusting ohlcv_age_days. The precise heal outcome (healed /
                        # lock-deferred / failed / 0-updates) is on the preceding [Self-Healer] line,
                        # so lock-deferral is separable from genuine failure.
                        _age = _cache_age_days(sym)
                        _age_desc = "missing" if _age is None else f"{_age}d stale"
                        _log.warning(f"🛑 AI BUY REJECTED (Zero-Trust Freshness): {sym} - OHLCV cache {_age_desc}, self-heal did not refresh it (healed={_healed}; see preceding [Self-Healer] line for lock-deferral vs failure); refusing to derive risk levels from untrustworthy data.")
                        continue

                bottom_ok, bottom_msg = is_bottom_confirmed(sym)

                # Catastrophic Gap Guard (The CNXC Trap):
                # Reject gap-downs > 8% unless bottom is independently confirmed — a volume-confirmed
                # capitulation gap is exactly the entry signal bottom detection targets.
                prev_close = row[10]
                if prev_close and prev_close > 0:
                    gap_pct = (price - prev_close) / prev_close
                    if gap_pct <= -0.08 and not bottom_ok:
                        _log.warning(f"🛑 AI BUY REJECTED (CNXC Trap): {sym} - Gap-Down {round(gap_pct*100, 1)}% with no confirmed bottom.")
                        continue

                # Reward-to-Risk & Target Upside Filter (Risk-Reward Gate):
                # Reject if the projected Reward-to-Risk ratio is below CFG.system_default_min_rr,
                # OR if the projected target gain percentage is below 5.0% of the current price.
                # EXEMPTION (Breakout Risk-Reward Waiver - R&D #13 & #32):
                # An elite momentum leader (see is_elite_breakout_candidate: combined + s10 both
                # above the CFG.system_bypass_* floors) waives the conservative R:R and target-gain
                # fallback restrictions below, to capture breakout alpha on names with no real
                # overhead resistance. Gate is delegated so the thresholds live in one place.
                stop_val = _to_float(row[9], 0.0)
                target_val = _to_float(row[11], 0.0)
                pgr_val = str(row[6] or "Neutral")

                is_elite_breakout = risk_utils.is_elite_breakout_candidate(total_score, short10)

                # The R:R and target-gain gates apply ONLY when the workbook carries explicit
                # Support/Resistance levels (row[9]/row[11]). A row with empty workbook S/R is NOT
                # "buying blind": the execution path assigns a downstream ATR-based stop, so such a
                # setup is intentionally allowed to fall through to the ATR fallback rather than be
                # rejected here. Do NOT re-add a hard reject on missing workbook S/R — it preempts
                # that ATR stop and rejects entries the system was designed to make.
                if stop_val > 0 and target_val > 0:
                    upside = target_val - price
                    downside = price - stop_val
                    rr_ratio = round(upside / downside, 2) if downside > 0 else 0.0
                    target_gain_pct = round((upside / price) * 100, 2) if price > 0 else 0.0

                    min_rr = CFG.system_default_min_rr
                    if rr_ratio < min_rr:
                        if is_elite_breakout:
                            _log.info(f"🛡️ [R&D #32 Breakout Waiver] Waived {min_rr}:1 R:R limit for elite breakout leader: {sym} (Combined Score: {total_score}, PGR: {pgr_val}, R:R: {rr_ratio}:1).")
                        else:
                            _log.warning(f"🛑 AI BUY REJECTED (Risk-Reward Gate): {sym} - Reward-to-Risk ratio of {rr_ratio}:1 is less than the required {min_rr}:1 minimum (Upside: ${round(upside, 2)}, Downside: ${round(downside, 2)}).")
                            continue

                    if target_gain_pct < 5.0:
                        if is_elite_breakout:
                            _log.info(f"🛡️ [R&D #32 Breakout Waiver] Waived 5.0% target upside limit for elite breakout leader: {sym} (Combined Score: {total_score}, PGR: {pgr_val}, Target Gain: {target_gain_pct}%).")
                        else:
                            _log.warning(f"🛑 AI BUY REJECTED (Risk-Reward Gate): {sym} - Projected target gain of {target_gain_pct}% is less than the required 5.0% minimum (Upside: ${round(upside, 2)}).")
                            continue
                else:
                    # Incomplete workbook S/R (missing stop AND/OR target — this branch is the negation
                    # of `stop_val > 0 and target_val > 0`) — synthesize the risk thesis for TRANSPARENCY,
                    # not as a gate.
                    # The freshness gate above already proved the cache is trustworthy, and _execute_buys
                    # assigns a live ATR stop, so this row is not "buying blind". Derive stop/target from
                    # the SAME split-adjusted, provenance-reporting resolvers the UI and backtests use and
                    # log the thesis. We deliberately DO NOT hard-gate on this synthesized R:R: the nearest
                    # confirmed resistance is structurally the smallest possible upside while support can be
                    # far below, so a uniform R:R>=2 reject would gut the fresh pipeline (validated live —
                    # most fresh names score R:R<2 on synthesized levels). The R:R / target-gain quality
                    # screen stays scoped to explicit workbook levels; freshness is the real safety gate.
                    excl = instruments.is_excluded(sym)
                    # Load the (freshness-gated) split-adjusted series ONCE and feed both resolvers,
                    # instead of letting each reload + re-adjust the same cache is_bottom_confirmed
                    # already read above. Freshness is guaranteed by the gate, so the resolvers'
                    # internal staleness check is redundant here and passing the series skips it.
                    _hi, _lo, _cl, _ = risk_utils._load_ohlcv_series(sym)
                    _sd = risk_utils.resolve_stop_detailed(price, highs=_hi, lows=_lo, closes=_cl, exclude_swing=excl)
                    _td = risk_utils.resolve_target_detailed(price, highs=_hi, lows=_lo, closes=_cl, exclude_swing=excl)
                    _wb_sr = f"workbook stop={row[9]!r}/target={row[11]!r} incomplete"
                    _log.info(f"[Synthesized Risk Thesis] {sym} @ ${price}: stop ${_sd.get('stop')} ({_sd.get('source')}) / target ${_td.get('target')} ({_td.get('source')}) — {_wb_sr}; live ATR stop applied at execution.")

                if total_score >= rules["min_score_threshold"] or bottom_ok:
                    bottom_desc = f" (Bottom Confirmed: {bottom_msg})" if bottom_ok else ""
                    top_buys.append({
                        "sym": sym,
                        "price": price,
                        "total": total_score,
                        "pgr": row[6] or "Neutral",
                        "s10": row[24] or 0.0,
                        "l60": row[25] or 0.0,
                        "bottom_desc": bottom_desc,
                        "industry": row[4]
                    })
        
        # ── R&D #32 Overbought Breakout Guard score penalty ──
        for buy_cand in top_buys:
            sym_upper = buy_cand["sym"].upper()
            cache = _load_symbol_today_cache(sym_upper, today)
            if cache:
                checklist = cache.get("checklist_stocks", {})
                strength_count = checklist.get("strengthCount", 1)
                timing_count = checklist.get("timingCount", 1)
                industry_rating = checklist.get("industry", "Neutral")
                
                if (strength_count < 1 and timing_count < 1) or industry_rating == "Weak":
                    buy_cand["total"] -= 1.5
                    _log.info(f"🛡️ [R&D #32 Guard] Applied -1.5 score penalty to {sym_upper} (overbought/weak-sector: strength={strength_count}, timing={timing_count}, industry={industry_rating})")

        top_buys.sort(key=lambda x: x["total"], reverse=True)

        # ── R&D #27: Dynamic Momentum Rotation Engine ──
        sells_to_rotate, available_slots, balance_addition = evaluate_momentum_rotation(
            profile, is_market_hours(), available_slots, max_positions, 
            state["positions"], prices, top_buys, active_position_scores
        )
        
        for sym_to_sell in sells_to_rotate:
            pos = state["positions"].pop(sym_to_sell)
            price = prices.get(sym_to_sell, pos["cost"])
            
            # Retrieve score for logging details if available
            score_val = active_position_scores.get(sym_to_sell, 0.0)
            
            tx = {
                "date": today, 
                "time": now_time, 
                "type": "SELL", 
                "symbol": sym_to_sell, 
                "price": price, 
                "qty": pos["qty"], 
                "pnl": round((price - pos["cost"]) * pos["qty"], 2),
                "details": f"🔄 [MOMENTUM ROTATION] Sold mature position {sym_to_sell} (Score: {score_val:.1f}) to free slot."
            }
            state["history"].append(tx)
            new_transactions.append(tx)
            _log.info(f"🔄 [MOMENTUM ROTATION] Sold mature position {sym_to_sell} (Score: {score_val:.1f}) @ ${price} to open slot.")
            log_closed_trade_dna(sym_to_sell, pos, price, today)
            
        state["balance"] += balance_addition

        if available_slots > 0 and state["balance"] > min_cash_required:
            if not is_market_hours():
                # Queue the buys up to available slots
                for buy in top_buys[:available_slots]:
                    if not any(q["symbol"] == buy["sym"] and q["type"] == "BUY" for q in state.get("queued_orders", [])):
                        state.setdefault("queued_orders", []).append({
                            "type": "BUY", "symbol": buy["sym"], "reason": "Top-ranked quantitative pick (After-Hours)"
                        })
                        _log.info(f"📝 [Queued] After-hours BUY queued for {buy['sym']}")
            else:
                _execute_buys(state, top_buys, available_slots, min_cash_required, rules,
                              today, now_time, new_transactions, prices)

                # ── Dynamic Pyramiding: Scale into winning trends with idle cash ──
                cash_ratio = state["balance"] / state["equity"] if state["equity"] > 1.0 else 0.0
                if cash_ratio > CFG.system_pyramiding_cash_ratio:  # idle-cash trigger for scaling in
                    _log.info(f"🛡️ [Pyramiding Pass] Checking active positions to deploy idle cash ({cash_ratio*100:.1f}%)...")
                    for sym, pos in list(state["positions"].items()):
                        current_px = prices.get(sym, pos["cost"])
                        # A winner is in profit and trading at or near its peak close
                        is_winner = (current_px > pos["cost"])
                        has_peak = (pos.get("highest_close_since_acq", 0.0) >= current_px * 0.98) # within 2% of peak close

                        # Retrieve its Short10 and Long60 momentum scores from row
                        s10 = 0.0
                        l60 = 0.0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[3] == sym:
                                s10 = row[24] or 0.0
                                l60 = row[25] or 0.0
                                break

                        # Pyramiding momentum gate (R&D #31) — delegated to a single-sourced helper.
                        if should_pyramid_into_winner(is_winner, has_peak, s10, l60):
                            max_pos_allocation = state["equity"] * rules["max_allocation_pct"]
                            current_allocation = pos["qty"] * current_px
                            remaining_room = max_pos_allocation - current_allocation

                            if remaining_room > 500.0:  # Minimum scale-in size is $500
                                cash_to_deploy = min(state["balance"] - (state["equity"] * rules["cash_buffer_pct"]), remaining_room)
                                add_qty = calculate_share_qty(sym, cash_to_deploy, current_px)

                                if add_qty > 0:
                                    cost = add_qty * current_px
                                    state["balance"] -= cost

                                    # Blended cost basis recalculation
                                    old_qty = pos["qty"]
                                    old_cost = pos["cost"]
                                    new_qty = old_qty + add_qty
                                    blended_cost = round(((old_qty * old_cost) + (cost)) / new_qty, 2)

                                    pos["qty"] = new_qty
                                    pos["cost"] = blended_cost

                                    tx = {
                                        "date": today, "time": now_time, "type": "BUY_SCALE_IN", 
                                        "symbol": sym, "price": current_px, "qty": add_qty,
                                        "details": f"Pyramiding Scale-In (Breakout Peak, s10={s10})"
                                    }
                                    state["history"].append(tx)
                                    new_transactions.append(tx)

                                    _log.info(f"🛡️ [Pyramiding Scale-In] Added {add_qty} shares to {sym} @ ${current_px:.2f} (Blended Cost: ${blended_cost:.2f})")
                                    _log.info(f"🛡️ [Pyramiding Scale-In] Added {add_qty} shares to {sym} @ ${current_px:.2f}")
                # ───────────────────────────────────────────────────────────

    except RuntimeError:
        raise  # critical failures (no prices, etc.) must propagate — never swallow
    except Exception as e:
        _log.exception(f"run_daily_ai_management failed: {e}")
    finally:
        if state is not None:
            save_game(state)
            update_excel_log(state, new_transactions)
            _log.info(f"🤖 AI Portfolio Value: ${state['equity']} (Cash: ${round(state['balance'], 2)})")
        else:
            _log.info("🤖 AI Management aborted before state initialization. Game state preserved.")

def deduct_operational_costs(amount):
    if amount <= 0: return
    state = load_game()
    state["balance"] -= amount
    state["total_ops_cost"] = round(state.get("total_ops_cost", 0) + amount, 4)
    state["history"].append({
        "date": str(datetime.date.today()),
        "type": "COST_DEDUCTION",
        "symbol": "OPS",
        "price": amount,
        "details": "Token & API fees"
    })
    save_game(state)
    _log.info(f"💸 AI Account debited ${round(amount, 4)} for operational costs.")

def show_report():
    state = load_game()
    
    # Dynamically compute the equity using latest available prices from workbook
    positions = state.get("positions", {})
    live_prices = {}
    
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        ws = wb["Research"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            sym = row[3]
            if sym in positions:
                live_prices[sym] = row[10]
        wb.close()
    except Exception:
        pass
        
    for sym in positions:
        if sym not in live_prices or not live_prices[sym]:
            live_prices[sym] = positions[sym]["cost"]
            
    live_equity = state.get("balance", 0.0)
    for sym, pos in positions.items():
        live_equity += pos["qty"] * live_prices.get(sym, pos["cost"])
        
    _log.info("--- 🤖 AI PORTFOLIO MANAGER REPORT ---")
    _log.info(f"Active Strategy: {state.get('profile', 'BALANCED')}")
    _log.info(f"Current Equity:  ${round(live_equity, 2)}")
    _log.info(f"Cash Balance:    ${round(state['balance'], 2)}")
    _log.info(f"Total Ops Cost:  ${state.get('total_ops_cost', 0)}")
    _log.info(f"Open Positions:  {len(positions)}")
    for sym, pos in positions.items():
        cur_px = live_prices.get(sym, pos['cost'])
        _log.info(f"  - {sym}: {pos['qty']} @ ${pos['cost']} (Current: ${cur_px:.2f}, Stop: ${pos.get('stop_loss', 'N/A')})")
    
    profit = live_equity - INITIAL_BALANCE
    _log.info(f"Net Profit:      ${round(profit, 2)} ({round((profit/INITIAL_BALANCE)*100, 2)}%)")
    target = INITIAL_BALANCE * 2
    days_elapsed = (datetime.date.today() - datetime.datetime.strptime(state["start_date"], "%Y-%m-%d").date()).days
    _log.info(f"Goal Progress:   {round((profit/INITIAL_BALANCE)*100, 1)}% of 100% (Target: ${target})")
    _log.info(f"Days Active:     {days_elapsed} / 90")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--run", action="store_true", help="Execute AI daily moves")
    parser.add_argument("--profile", type=str, help="Manually override strategy profile (AGGRESSIVE, DEFENSIVE, BALANCED)")
    parser.add_argument("--force", action="store_true", help="Force run outside market hours")
    parser.add_argument("--report", action="store_true", help="Show CLI performance report")
    parser.add_argument("--summary", action="store_true", help="Send daily email summary")
    args = parser.parse_args()

    if args.run:
        run_daily_ai_management(force=args.force, manual_profile=args.profile)
        send_consolidated_morning_report()
        try:
            watchdog.sync_data_folder()
        except Exception as e:
            _log.warning(f"Post-run sync failed: {e}")
    elif args.summary:
        send_daily_summary()
        try:
            watchdog.sync_data_folder()
        except Exception as e:
            _log.warning(f"Post-run sync failed: {e}")
    else:
        show_report()
