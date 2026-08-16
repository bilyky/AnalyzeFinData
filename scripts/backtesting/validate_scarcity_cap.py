"""
Historical / distributional validation of the dynamic scarcity-bucket cap (Option C,
see plans/dynamic-scarcity-cap.md).

WHAT THIS IS
------------
This is NOT a portfolio-P&L backtest — the repo has no position-sizing/P&L simulator,
only factor-level forward-return studies (backtest.py / backtest_ratings.py). Instead
this script validates the *cap mechanism* against the REAL conviction-score distribution:

  * It imports the SHIPPED functions `_conviction_cap_pct` and `get_strategy_rules`
    from ai_portfolio_game (no re-implementation, no mocks), so it tests the exact code
    that runs in production.
  * It draws the real `total = Short10 + Long60` distribution from the live workbook's
    Research sheet — the same columns (`row[24] + row[25]`) `_execute_buys` reads. This
    is a single trading day's cross-section of the full ~485-symbol universe: a real,
    trustworthy snapshot of the quantity that gates the cap (not a reconstructed guess).

It then quantifies three things the Option-C change was meant to fix:
  1. The old binary cliff at total>=8.0 vs the new smooth ramp (discontinuity removed).
  2. The DEFENSIVE pathology: under the old global 8.0 suspension, every DEFENSIVE-
     eligible buy (min_score 10.0 >= 8.0) ran with the scarcity cap fully suspended.
  3. Concentration: max scarcity-bucket exposure, old vs new, per profile.

And it suggests calibrated `cap_relax_full` endpoints from the real score percentiles.

Read-only: opens the workbook data_only, writes nothing. Run:
    python scripts/backtesting/validate_scarcity_cap.py
"""
import os
import sys


sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))

import openpyxl

import ai_portfolio_game as game


# The old (pre-Option-C) behavior we are validating against: a binary cliff.
OLD_RELAX_THRESHOLD = 8.0          # global "Adaptive Cap Relaxation" trigger
OLD_BASE_CAP = 0.20                # scarcity base cap when not relaxed
SAMPLE_EQUITY = 100_000.0          # for translating cap % -> $ concentration

PROFILES = ("AGGRESSIVE", "BALANCED", "DEFENSIVE")


def load_totals(xlsx_path):
    """Real total = Short10(col24) + Long60(col25) for every scored Research row.
    Returns (all_totals, active_totals) where active = rows with an active Setup flag
    (col20 in {1,'1','OK'}) — the ones actually eligible to be bought."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Research"] if "Research" in wb.sheetnames else wb.active
    all_t, active_t = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= 25:
            continue
        s10, l60 = row[24], row[25]
        if not isinstance(s10, (int, float)) or not isinstance(l60, (int, float)):
            continue
        total = float(s10) + float(l60)
        all_t.append(total)
        if str(row[20] or "") in ("1", "OK") or row[20] == 1:
            active_t.append(total)
    return all_t, active_t


def pctile(sorted_vals, q):
    if not sorted_vals:
        return float("nan")
    k = (len(sorted_vals) - 1) * q
    lo = int(k)
    hi = min(lo + 1, len(sorted_vals) - 1)
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (k - lo)


def describe(label, vals):
    if not vals:
        return
    sorted(vals)


def old_cap(total):
    """Old binary model: base 20% below 8.0, else the cap was SUSPENDED (no bucket
    limit at all). Returns None to denote 'suspended / unbounded'."""
    return None if total >= OLD_RELAX_THRESHOLD else OLD_BASE_CAP


def new_cap(total, rules):
    """Shipped Option-C ramp — calls the real production helper."""
    return game._conviction_cap_pct(
        total,
        base_pct=rules.get("scarcity_allocation_pct", 0.20),
        ceiling_pct=rules.get("scarcity_cap_ceiling_pct", 0.20),
        relax_start=rules.get("cap_relax_start", 8.0),
        relax_full=rules.get("cap_relax_full", 12.0),
    )


def fmt_cap(c):
    return "SUSPENDED" if c is None else f"{c*100:5.1f}%"


def main():
    xlsx = game.XLSX_FILE
    if not os.path.exists(xlsx):
        return 1


    all_t, active_t = load_totals(xlsx)
    describe("full universe ", all_t)
    describe("active setups ", active_t)

    for prof in PROFILES:
        rules = game.get_strategy_rules(prof)
        min_score = rules.get("min_score_threshold", 0.0)
        base = rules.get("scarcity_allocation_pct", 0.20)
        rules.get("scarcity_cap_ceiling_pct", base)
        rules.get("cap_relax_start", 8.0)
        r_full = rules.get("cap_relax_full", 12.0)
        rules.get("max_allocation_pct", 0.15)

        eligible = sorted(t for t in active_t if t >= min_score)
        if not eligible:
            probes = [min_score, min_score + 2, r_full, r_full + 2]
        else:
            probes = [eligible[0], pctile(eligible, 0.5), eligible[-1]]

        # DEFENSIVE pathology quantification
        old_uncapped = sum(1 for t in eligible if old_cap(t) is None)
        (old_uncapped / len(eligible) * 100) if eligible else float("nan")
        new_caps = [new_cap(t, rules) for t in eligible]
        if new_caps:
            pass

        # sample points: cap % and $ concentration in scarcity bucket
        for t in probes:
            oc = old_cap(t)
            nc = new_cap(t, rules)
            "unbounded" if oc is None else f"${oc*SAMPLE_EQUITY:,.0f}"
            f"${nc*SAMPLE_EQUITY:,.0f}"

        # calibration hint from the real distribution
        if eligible:
            suggest_full = round(pctile(eligible, 0.90), 1)
            "OK" if abs(suggest_full - r_full) < 1.5 else "consider retune"

    for prof in PROFILES:
        rules = game.get_strategy_rules(prof)
        rules.get("scarcity_cap_ceiling_pct", rules.get("scarcity_allocation_pct", 0.20))
        rules.get("max_allocation_pct", 0.15)
    return 0


if __name__ == "__main__":
    sys.exit(main())
