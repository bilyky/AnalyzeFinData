import os
import sys


sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import datetime
import json
import os
from pathlib import Path


# --- CONFIGURATION ---
BASE_DIR = Path(__file__).resolve().parent
GAME_STATE_FILE = BASE_DIR / "Data" / "game_state.json"
PERFORMANCE_LOG = BASE_DIR / "Data" / "performance_log.json"

def load_json(path):
    if not path.exists():
        return {}
    with open(path, "r") as f:
        try:
            return json.load(f)
        except:
            return {}

def save_json(path, data):
    with open(path, "w") as f:
        json.dump(data, f, indent=4)

def pick_stock(symbol):
    """Pick a stock for today's challenge."""
    state = load_json(GAME_STATE_FILE)
    today = str(datetime.date.today())
    
    # Ensure symbols are tracked in performance_log first
    perf = load_json(PERFORMANCE_LOG)
    if today not in perf:
        return

    valid_symbols = [p["symbol"].upper() for p in perf[today]]
    if symbol.upper() not in valid_symbols:
        pass

    state["current_pick"] = {
        "date": today,
        "symbol": symbol.upper(),
        "entry_price": None # Will be filled from perf_log if available
    }
    
    # Try to find entry price
    for p in perf[today]:
        if p["symbol"].upper() == symbol.upper():
            state["current_pick"]["entry_price"] = p["entry_price"]
            break
            
    save_json(GAME_STATE_FILE, state)

def evaluate_challenge():
    """Check how yesterday's pick performed."""
    state = load_json(GAME_STATE_FILE)
    pick = state.get("current_pick")
    if not pick:
        return

    today_date = datetime.date.today()
    pick_date = datetime.datetime.strptime(pick["date"], "%Y-%m-%d").date()
    
    if pick_date >= today_date:
        return

    # To evaluate, we need today's price for the picked symbol
    # In a real game, this would be called AFTER main.py updates the workbook
    from openpyxl import load_workbook
    XLSX_FILE = BASE_DIR / "Data" / "state_of_the_day.xlsx"
    
    if not XLSX_FILE.exists():
        return

    wb = load_workbook(XLSX_FILE, data_only=True)
    ws = wb["Research"]
    exit_price = None
    spy_change = 0.0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == pick["symbol"]:
            exit_price = row[10]
        if row[3] == "SPY":
            spy_change = row[15] or 0.0 # Today's Change %

    if exit_price and pick["entry_price"]:
        pct_change = ((exit_price - pick["entry_price"]) / pick["entry_price"]) * 100
        alpha = pct_change - spy_change
        
        # Update Leaderboard
        history = state.get("history", [])
        history.append({
            "date": pick["date"],
            "symbol": pick["symbol"],
            "return": round(pct_change, 2),
            "alpha": round(alpha, 2)
        })
        state["history"] = history
        state["total_alpha"] = round(state.get("total_alpha", 0) + alpha, 2)
        state["current_pick"] = None # Reset
        
        save_json(GAME_STATE_FILE, state)
    else:
        pass

def show_status():
    state = load_json(GAME_STATE_FILE)
    pick = state.get("current_pick")
    if pick:
        pass
    else:
        pass
    
    
    history = state.get("history", [])
    if history:
        for _h in history[-5:]:
            pass

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--pick", type=str, help="Pick a symbol for today")
    parser.add_argument("--evaluate", action="store_true", help="Evaluate the pending challenge")
    parser.add_argument("--status", action="store_true", help="Show leaderboard")
    args = parser.parse_args()

    if args.pick:
        pick_stock(args.pick)
    elif args.evaluate:
        evaluate_challenge()
    elif args.status:
        show_status()
    else:
        show_status()
