"""
Project AETHER Intraday Stop-Breach Monitor.
Performs 30-minute interval audits of held E*TRADE positions against their ATR-based stop-loss levels.
Includes state-persistence (cures 30-min duplicate email spam) and dynamic, AI-powered qualitative analyses.
"""
import datetime
import json
import os
import sys
from pathlib import Path

import openpyxl


_BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_BASE_DIR))

import ai_client
import notify
from aether_logger import get_logger
from ai_portfolio_game import get_live_prices
from data_api import _SL  # canonical Short_Long column map (single source of truth)


_log = get_logger("monitor")

# Anchor state/workbook paths to the project root, not the process CWD — the
# scheduled task may launch from anywhere, and CWD-relative paths would silently
# yield "no positions found" and skip every cycle.
XLSX_FILE = _BASE_DIR / "Data" / "state_of_the_day.xlsx"
STATE_FILE = _BASE_DIR / "Data" / "intraday_monitor_state.json"

# Re-alert an existing breach only when it materially deteriorates: the price
# drops at least this fraction below the last-alerted price, or the stop moves.
# Without this, ordinary sub-percent ticks flip "changed" on every 30-min run and
# re-send the alert — the duplicate-email spam this monitor is meant to cure.
_MATERIAL_DROP_PCT = 0.01

def load_state() -> dict:
    """Load the last active stop-breach state from disk."""
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except Exception as e:
            _log.warning(f"Failed to load monitor state: {e}")
    return {"last_breached": {}}

def save_state(state: dict):
    """Save the active stop-breach state to disk atomically (temp file + replace),
    so a crash or overlapping run can't leave a truncated JSON that resets state."""
    try:
        STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        tmp = STATE_FILE.with_name(STATE_FILE.name + ".tmp")
        tmp.write_text(json.dumps(state, indent=2), encoding="utf-8")
        os.replace(tmp, STATE_FILE)
    except Exception as e:
        _log.warning(f"Failed to save monitor state: {e}")

def get_monitored_positions() -> list:
    """Load positions and stops from Short_Long sheet."""
    positions = []
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, data_only=True, read_only=True)
        if "Short_Long" not in wb.sheetnames:
            return []
        
        ws = wb["Short_Long"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            sym  = row[_SL["sym"]]  if len(row) > _SL["sym"]  else None
            stop = row[_SL["stop"]] if len(row) > _SL["stop"] else None
            if (isinstance(sym, str) and sym.strip() and sym.strip() != "Symb"
                    and isinstance(stop, (int, float)) and stop > 0):
                positions.append({"symbol": sym.strip().upper(), "stop": stop})
    except Exception as e:
        _log.error(f"Error loading positions from workbook: {e}")
    return positions

def generate_ai_analysis(sym: str, price: float, stop: float) -> str:
    """Generate a highly specific, professional quantitative analysis for the stop breach."""
    if not ai_client.primary():
        return "ATR Stop-Loss Floor Breached. Technical momentum is currently negative, representing an immediate capital preservation exit risk."
        
    system_prompt = (
        "You are AETHER, an expert hedge-fund quantitative analyst.\n"
        "Provide a highly concise, professional 2-sentence analysis and reasoning of why this stock has hit its stop-loss, "
        "what the primary technical risk factor is, and what action the trader should take immediately. "
        "Be concise, direct, and authoritative."
    )
    user_prompt = f"Asset {sym} has breached its ATR stop-loss floor. Current Price: ${price:.2f} | Stop-Loss level: ${stop:.2f}."
    try:
        raw_reply = ai_client.evaluate(system_prompt, user_prompt, max_tokens=150)
        return raw_reply.strip()
    except Exception as e:
        _log.warning(f"Failed calling AI analyzer for {sym}: {e}")
        return "ATR Stop-Loss Floor Breached. Technical momentum is currently negative, representing an immediate capital preservation exit risk."

def monitor():
    _log.console(f"[{datetime.datetime.now()}] Starting Intraday Stop Monitor...")
    monitored = get_monitored_positions()
    if not monitored:
        _log.info("No positions with valid stops found to monitor.")
        return

    _log.console(f"Monitoring {len(monitored)} positions.")

    symbols = [p["symbol"] for p in monitored]
    quotes = get_live_prices(symbols)

    current_breaches = {}
    for p in monitored:
        sym = p["symbol"]
        stop = p["stop"]
        last_price = quotes.get(sym)
        if last_price and last_price > 0:
            if last_price <= stop:
                current_breaches[sym] = {"stop": stop, "price": last_price}
        else:
            _log.warning(f"Could not fetch live price for {sym}")

    # Load previous state
    state = load_state()
    last_breached = state.get("last_breached", {})

    # Determine if something changed compared to the previous run
    new_breaches = [s for s in current_breaches if s not in last_breached]
    cleared_breaches = [s for s in last_breached if s not in current_breaches]
    
    # Re-alert an existing breach only when it materially worsens: the stop level
    # moved (e.g. a trailing/ratchet adjustment) or the price made a new low at
    # least _MATERIAL_DROP_PCT below what we last alerted on. A small tick or an
    # upward bounce toward the stop must NOT re-trigger the email.
    modified_breaches = []
    for s in current_breaches:
        if s in last_breached:
            prev = last_breached[s]
            cur = current_breaches[s]
            stop_moved = abs(cur["stop"] - prev.get("stop", cur["stop"])) > 1e-9
            dropped_further = cur["price"] < prev["price"] * (1 - _MATERIAL_DROP_PCT)
            if stop_moved or dropped_further:
                modified_breaches.append(s)

    has_changed = bool(new_breaches or cleared_breaches or modified_breaches)

    if current_breaches:
        if not has_changed:
            _log.info(f"Stop breach state unchanged ({len(current_breaches)} active breaches). Bypassing duplicate email dispatch.")
            return

        # Prepare rich qualitative report with AI analysis for each item
        diff_msgs = []
        if new_breaches:
            diff_msgs.append(f"🔴 NEW BREACHES DETECTED: {', '.join(new_breaches)}")
        if cleared_breaches:
            diff_msgs.append(f"🟢 BREACHES CLEARED/EXITED: {', '.join(cleared_breaches)}")
        if modified_breaches:
            diff_msgs.append(f"🔄 EXISTING BREACH PRICES UPDATED: {', '.join(modified_breaches)}")

        _log.warning("Stop breach state changed! Dispatching rich, AI-analyzed alert email...")
        
        subject = f"AETHER ALERT: Stop Breach Detected ({len(current_breaches)} Position{'s' if len(current_breaches) > 1 else ''})"
        
        body_parts = [
            "The following stop-loss breaches have been detected in your active E*TRADE portfolio.\n",
            "📊 DIFFERENCE SUMMARY:",
            "\n".join(f"  * {m}" for m in diff_msgs),
            "\n" + "=" * 80,
            "🔬 QUALITATIVE ANALYSIS & RISK REASONING FOR ACTIONABLE BREACHES:",
            "=" * 80 + "\n"
        ]

        actionable_symbols = sorted(list(set(new_breaches + modified_breaches)))
        
        for sym in actionable_symbols:
            data = current_breaches[sym]
            price = data["price"]
            stop = data["stop"]
            
            _log.console(f"Running live AI risk analysis and reasoning for {sym}...")
            ai_reasoning = generate_ai_analysis(sym, price, stop)
            
            body_parts.append(
                f"📈 {sym} — BREACH UPDATE!"
                f"\n  * Current Price:   ${price:.2f}"
                f"\n  * Stop-Loss Level: ${stop:.2f}"
                f"\n  * Technical Delta: {round(((price - stop)/stop)*100, 2):+.2f}%"
                f"\n  * AI Risk Analysis & Reasoning:"
                f"\n    {ai_reasoning}"
                f"\n\n" + "-" * 80
            )

        # Include a quiet list of other active, unchanged breaches
        unchanged_breaches = [s for s in current_breaches if s not in actionable_symbols]
        if unchanged_breaches:
            body_parts.append(
                "ℹ️ OTHER UNCHANGED ACTIVE BREACHES (Already Alerted):\n" +
                ", ".join(f"{s} (${current_breaches[s]['price']:.2f})" for s in sorted(unchanged_breaches)) + "\n"
            )

        body_parts.append(
            "\nAction required: These positions are trading below their capital-preservation stop-loss levels. "
            "Please review and execute these exits manually in your broker immediately to prevent further drawdown."
        )

        notify.send_email(subject, "\n".join(body_parts))
        _log.info(f"Sent rich consolidated alert email for {len(current_breaches)} breaches.")
    else:
        if cleared_breaches:
            # If breaches exist in memory but are now cleared (e.g. they were sold!)
            _log.info(f"All stop breaches cleared (bypassed email notification to avoid noise): {', '.join(cleared_breaches)}")
        else:
            _log.info("No stop breaches detected.")

    # Save the current state for the next run
    save_state({"last_breached": current_breaches})

if __name__ == "__main__":
    monitor()
