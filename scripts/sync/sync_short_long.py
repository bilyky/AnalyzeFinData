"""Standalone entry point: sync Short_Long sheet with live E*TRADE positions.

Run:
    python sync_short_long.py              # production account
    python sync_short_long.py --sandbox    # sandbox account

Does NOT require a full check_from_xls run. Uses cached production tokens
(or runs browser auth if not cached). Picks scores come from the last
Research sheet data already in the workbook.
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))



import argparse
import datetime
import json
import os
import openpyxl

from aether import etrade
from aether.logger import get_logger
from workbook_write import update_short_long_scores, fix_comment_shape_ids, backup_xlsx
from powergauge import XLSX_FILE, SRC_XLSX, OHLCV_DIR

_log = get_logger("sync_short_long")


def _read_picks_from_research(wb) -> dict:
    """Build a picks_lookup dict from the current Research sheet data."""
    if "Research" not in wb.sheetnames:
        return {}
    ws     = wb["Research"]
    lookup = {}

    def _g(row, idx, default=None):
        return row[idx] if len(row) > idx else default

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        sym = str(_g(row, 3) or "").strip().upper()   # col D
        if not sym:
            continue
        lookup[sym] = {
            "symbol":   sym,
            "industry": str(_g(row, 4) or "").strip(),    # col E
            "pgr":      _g(row, 6),                        # col G
            "br":       _g(row, 21),                       # col V
            "short10":  _g(row, 24),                       # col Y
            "long60":   _g(row, 25),                       # col Z
            "ob_os":    str(_g(row, 19) or "").strip(),    # col T
            "money_fl": str(_g(row, 18) or "").strip(),    # col S
            "lt_trend": str(_g(row, 17) or "").strip(),    # col R
            "stop":     _g(row, 9),                        # col J
            "target":   _g(row, 11),                       # col L
        }
    return lookup


def main():
    parser = argparse.ArgumentParser(description="Sync Short_Long sheet with E*TRADE")
    parser.add_argument("--sandbox", action="store_true", help="Use sandbox environment")
    args = parser.parse_args()
    env = "sandbox" if args.sandbox else "production"

    _log.console("[sync_short_long] env=%s", env)

    # ── Get tokens ───────────────────────────────────────────────────────────
    tokens = etrade.get_tokens(env)
    if not tokens:
        # get_tokens() fails soft (-> None). Without this guard the None flows into
        # fetch_positions and blows up mid-sync with an opaque error; fail fast instead.
        _log.console("[sync_short_long] E*TRADE tokens unavailable "
                     "(run scripts/diagnostics/test_etrade.py to re-authenticate); aborting sync.")
        raise SystemExit(1)

    # ── Fetch positions + quotes ──────────────────────────────────────────────
    _log.console("Fetching E*TRADE positions...")
    positions = etrade.fetch_positions(tokens, env)
    _log.console("  %d open positions found.", len(positions))

    syms   = list({p["symbol"] for p in positions})
    quotes = etrade.fetch_quotes(tokens, syms, env)
    _log.console("  %d live quotes fetched.", len(quotes))

    # ── Load workbook + build picks lookup from Research sheet ───────────────
    try:
        wb = openpyxl.load_workbook(SRC_XLSX)
    except Exception as e:
        _log.error("Failed to load source %s: %s", SRC_XLSX, e)
        _log.console("  Attempting to load existing output %s instead...", XLSX_FILE)
        wb = openpyxl.load_workbook(XLSX_FILE)

    picks_lookup = _read_picks_from_research(wb)
    _log.console("  %d symbols in Research sheet for score lookup.", len(picks_lookup))

    # ── Load OHLCV for streak computation ────────────────────────────────────
    ohlcv_cache = {}
    for sym in syms:
        path = os.path.join(OHLCV_DIR, f"{sym}_daily.json")
        try:
            with open(path) as f:
                ohlcv_cache[sym] = json.load(f).get('Time Series (Daily)')
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            pass

    # ── Sync Short_Long sheet ────────────────────────────────────────────────
    orig_backup = backup_xlsx(XLSX_FILE)
    update_short_long_scores(wb, picks_lookup, quotes, positions, ohlcv_cache)

    while True:
        try:
            wb.save(XLSX_FILE)
            fix_comment_shape_ids(XLSX_FILE,
                                  original_xlsx=orig_backup,
                                  touched_sheet_names={"Short_Long"})
            _log.console("Saved -> %s", XLSX_FILE)
            break
        except PermissionError:
            input(f"Close {XLSX_FILE} in Excel, then press Enter to retry...")


if __name__ == "__main__":
    main()
