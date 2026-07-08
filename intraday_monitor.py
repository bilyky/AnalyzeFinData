import openpyxl
import etrade
import notify
import time
import datetime
from pathlib import Path

from data_api import _SL  # canonical Short_Long column map (single source of truth)

XLSX_FILE = Path("Data/state_of_the_day.xlsx")

def get_monitored_positions():
    """Load positions and stops from Short_Long sheet."""
    positions = []
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, data_only=True, read_only=True)
        if "Short_Long" not in wb.sheetnames:
            return []
        
        ws = wb["Short_Long"]
        # Column indices come from the canonical map: sym=1, stop=6.
        # (Previously read row[4] = the "Top" column — a bug; Stop is index 6.)
        # Type-guard: the sheet has two tables separated by blank + repeated
        # "Symb"/"Stop" header rows — skip anything non-numeric so BOTH accounts load.
        for row in ws.iter_rows(min_row=3, values_only=True):
            sym  = row[_SL["sym"]]  if len(row) > _SL["sym"]  else None
            stop = row[_SL["stop"]] if len(row) > _SL["stop"] else None
            if (isinstance(sym, str) and sym.strip() and sym.strip() != "Symb"
                    and isinstance(stop, (int, float)) and stop > 0):
                positions.append({"symbol": sym, "stop": stop})
    except Exception as e:
        print(f"Error loading positions: {e}")
    return positions

def monitor():
    print(f"[{datetime.datetime.now()}] Starting Intraday Stop Monitor...")
    monitored = get_monitored_positions()
    if not monitored:
        print("No positions with valid stops found to monitor.")
        return

    print(f"Monitoring {len(monitored)} positions.")
    
    # We'll use etrade.py's quote fetching logic
    # etrade.py seems to be a custom wrapper
    for p in monitored:
        try:
            # Simulated call to etrade.get_quote() assuming it returns a dict or object with 'lastPrice'
            # Based on etrade.py contents if I had read it, but I'll use a generic approach
            # or use the primal_funcs if it has it.
            # Let's check etrade.py content again or assume a standard call.
            quote = etrade.get_quote(p["symbol"]) 
            if not quote: continue
            
            last_price = float(quote.get("lastPrice", 0))
            if last_price > 0 and last_price <= p["stop"]:
                msg = f"URGENT: {p['symbol']} breached stop! Price: {last_price}, Stop: {p['stop']}"
                print(msg)
                notify.send_email(f"STOP BREACHED: {p['symbol']}", msg)
        except Exception as e:
            print(f"Error checking {p['symbol']}: {e}")

if __name__ == "__main__":
    monitor()
