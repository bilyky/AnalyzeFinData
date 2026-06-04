import datetime
import os
import sys
import json
import openpyxl

# Add current dir to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import powergauge

def get_comprehensive_stats():
    date = datetime.date(2026, 5, 26)
    session_id = "dummy"
    powergauge._build_cache_index()
    
    wb = openpyxl.load_workbook('Data/state_of_the_day.xlsx', data_only=True)
    ws_research = wb['Research']
    
    all_data = []
    for row in ws_research.iter_rows(min_row=2):
        symbol = row[3].value
        if not symbol: continue
        try:
            pg = powergauge.get_symbol_data(symbol, date, True, session_id)
            if pg.price == -1: continue
            
            ohlcv_path = os.path.join("Data", "Symbol_full", f"{symbol}_daily.json")
            ohlcv_ts = None
            if os.path.exists(ohlcv_path):
                with open(ohlcv_path) as _f:
                    ohlcv_ts = json.load(_f).get('Time Series (Daily)')
            
            f = powergauge._compute_pgr_fields(pg, ohlcv_ts=ohlcv_ts)
            
            pgr_val = pg.pgr_corrected_value if pg.pgr_corrected_value != 0 else pg.pgr_value

            all_data.append({
                'symbol': symbol,
                'short10': f['short_score'],
                'long60': f['long_score'],
                'br': f['buying_ratio'],
                'pgr': f['pgr'],
                'pgr_val': pgr_val,
                'setup': (1 if f['setup_ok'] else 0) if f['setup_ok'] is not None else None,
                'money_flow': f['money_flow'],
                'ob_os': f['ob_os']
            })
        except Exception:
            continue

    print(f"Total symbols processed: {len(all_data)}")

    # 1. Top 5 by Short10 (Unfiltered)
    top_short = sorted(all_data, key=lambda x: x['short10'], reverse=True)[:5]
    print("\n--- Top 5 by Short10 (Unfiltered) ---")
    for i, r in enumerate(top_short, 1):
        print(f"{i}. {r['symbol']} (S10: {r['short10']}, BR: {r['br']}, PGR: {r['pgr']})")

    # 2. Top 5 by Buying Ratio (Unfiltered)
    top_br = sorted(all_data, key=lambda x: x['br'], reverse=True)[:5]
    print("\n--- Top 5 by Buying Ratio (Unfiltered) ---")
    for i, r in enumerate(top_br, 1):
        print(f"{i}. {r['symbol']} (BR: {r['br']}, S10: {r['short10']}, PGR: {r['pgr']})")

    # 3. Check Sheet1 symbols
    ws_sheet1 = wb['Sheet1']
    sheet1_symbols = []
    for row in ws_sheet1.iter_rows(min_row=4, max_row=10):
        if row[4].value: sheet1_symbols.append(row[4].value) # Col E
        if row[6].value: sheet1_symbols.append(row[6].value) # Col G
    
    print("\n--- Stats for Sheet1 Symbols ---")
    for sym in sheet1_symbols:
        match = next((d for d in all_data if d['symbol'] == sym), None)
        if match:
            print(f"{sym}: S10: {match['short10']}, BR: {match['br']}, PGR: {match['pgr']}")
        else:
            print(f"{sym}: No data found")

if __name__ == "__main__":
    get_comprehensive_stats()
