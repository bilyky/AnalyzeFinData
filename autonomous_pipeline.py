import atexit
import datetime
import html
import json
import os
import subprocess
import sys
from pathlib import Path

import openpyxl

import console_safe
import notify
import rapidapi
import watchdog
from aether_logger import get_logger as _get_logger
from config import CFG
from run_history import load_symbols


_pipeline_log = _get_logger("pipeline")

# Windows CP1252 console fallback (bug-fix workaround, not a feature): must run
# before the first non-ASCII print. Reduces, not eliminates, cp1252 crashes in
# headless runs. See AETHER_REFERENCE.md.
console_safe.install()

import external_intel

# Custom modules
import performance_tracker
from scripts.diagnostics.preflight_validator import run_preflight_diagnostics
from workbook_read import (
    get_market_regime,
    get_replacement_pairs,
    get_reserves_data,
    get_top_5_picks,
)


# --- CONFIGURATION ---
BASE_DIR = Path(__file__).resolve().parent
SRC_XLSX  = BASE_DIR / "state_of_the_day.xlsx"
XLSX_FILE = BASE_DIR / "Data" / "state_of_the_day.xlsx"
LOG_FILE_PATH = BASE_DIR / "Data" / "autonomous_run.log"

def log(msg):
    """Pipeline log — routes through the AETHER logger (txt + jsonl + stdout)
    and keeps appending to the legacy LOG_FILE_PATH for backward compatibility
    with the dashboard's Pipeline Log viewer."""
    _pipeline_log.info(msg)
    try:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(LOG_FILE_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {msg}\n")
    except Exception as e:
        _pipeline_log.warning("Failed to write to legacy log file", extra={"error": str(e)})

def verify_data_freshness():
    if not XLSX_FILE.exists():
        return False, "File does not exist."
    
    mtime = datetime.datetime.fromtimestamp(XLSX_FILE.stat().st_mtime)
    today = datetime.date.today()
    if mtime.date() < today:
        return False, f"Data is stale. Last updated: {mtime.strftime('%Y-%m-%d %H:%M')}"
    
    return True, f"Data is fresh ({mtime.strftime('%Y-%m-%d %H:%M')})"

def validate_sheets():
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        required_sheets = ["Research", "Picks", "Replacements"]
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                return False, f"Missing sheet: {sheet}"
            
            ws = wb[sheet]
            if ws.max_row < 2:
                return False, f"Sheet {sheet} appears empty."
        
        return True, "All required sheets validated and rendered."
    except Exception as e:
        return False, f"Validation error: {e}"


def check_earnings(symbol):
    return "Check Required"

def get_reasoning(symbol, pgr, s10, l60, industry):
    """Retrieve live LLM reasoning via the GitHub Models API (gpt-4o-mini) with local heuristics fallback."""
    # If in report-only/cached mode, bypass slow sequential AI calls to prevent subprocess deadlocks and timeouts
    if "--report-only" in sys.argv or "--cached" in sys.argv:
        return f"<b>Technical Setup Active:</b> Setup OK with total score: {s10+l60:.1f}.<br>🚨 <b>Local Fallback:</b> Running in report-only/cached mode."
    try:
        # Call the live LLM reasoning engine we just implemented
        return external_intel.get_ai_reasoning(symbol, industry, pgr, s10, l60)
    except Exception as e:
        log(f"Warning: Live AI reasoning failed for {symbol}: {e}")
        # Standard Fallback
        return f"<b>Technical Setup Active:</b> Setup OK with total score: {s10+l60:.1f}.<br>🚨 <b>Devil's Advocate:</b> High market volatility could override technical momentum."

def cleanup_orphaned_processes():
    """Ensure no Excel-locking processes are hung."""
    try:
        if sys.platform == "win32":
            subprocess.run(["powershell", "Get-Process | Where-Object { $_.Name -match 'excel|python' -and $_.CommandLine -match 'main.py|daily_task.py' } | Stop-Process -Force"], capture_output=True)
    except:
        pass

def format_html_report(status_msg, picks, replacements, intel_ideas):
    today = datetime.date.today()
    regime, color = get_market_regime()
    
    # 0. Format External Ideas + Structural Intel
    # All email-sourced strings (from/subject) and AI-extracted text (event/impact/rd_topics)
    # are HTML-escaped before insertion to prevent injection into the email report.
    def _e(v): return html.escape(str(v or ""), quote=True)

    intel_section = """
    <div style="background: #fafbfc; border: 1px solid #e1e4e6; border-radius: 6px; padding: 15px 20px; margin-bottom: 30px;">
        <h3 style="color: #2c3e50; margin: 0 0 5px 0; font-size: 15px;">📰 External Intelligence Feed</h3>
        <p style="font-size: 12px; color: #7f8c8d; margin: 0; font-style: italic;">No new newsletter intelligence or external catalysts parsed for today's session. System evaluated purely on indicators.</p>
    </div>
    """
    if intel_ideas:
        idea_list = ""
        for i in intel_ideas:
            if i.get("symbol") and i.get("sentiment"):
                c = "#27ae60" if i["sentiment"] == "BUY" else ("#c0392b" if i["sentiment"] == "SELL" else "#f39c12")
                badge = f'<span style="background:{c};color:white;padding:2px 6px;border-radius:3px;font-weight:bold;font-size:11px;">{_e(i["sentiment"])}</span>'
                idea_list += f"""
                <li style="margin-bottom:12px;border-bottom:1px dashed #eee;padding-bottom:10px;">
                    <b>Source:</b> {_e(i['from'])}<br>
                    <b>Topic:</b> {_e(i['subject'])}<br>
                    <b>Decision:</b> {badge} <b>{_e(i['symbol'])}</b>
                    {f"<br><b>Thesis:</b> <i>{_e(i['thesis'])}</i>" if i.get('thesis') else ''}
                </li>"""

        # Structural intel: aggregate catalysts, missing symbols, R&D across all emails.
        # Catalysts older than 15 days are historical filler, not action signals — drop them.
        _cutoff = (datetime.date.today() - datetime.timedelta(days=15)).isoformat()
        all_catalysts, all_missing, all_rd = [], [], []
        for i in intel_ideas:
            if not isinstance(i, dict):
                continue
            iv = i.get("intel") or {}
            if isinstance(iv, dict):
                for c in iv.get("dated_catalysts", []):
                    if not isinstance(c, dict):
                        continue
                    d = str(c.get("date") or "")
                    # Keep: future/near dates (>= cutoff) OR year-only entries like "2026"
                    # Drop: specific past dates older than 15 days
                    try:
                        if len(d) >= 10 and d[:10] < _cutoff:
                            continue
                    except Exception:
                        pass
                    all_catalysts.append(c)
                
                missing_syms = iv.get("missing_symbols", [])
                if isinstance(missing_syms, list):
                    all_missing.extend(missing_syms)
                elif isinstance(missing_syms, dict):
                    all_missing.append(missing_syms)
                
                rd_topics_list = iv.get("rd_topics", [])
                if isinstance(rd_topics_list, list):
                    all_rd.extend(rd_topics_list)
                elif isinstance(rd_topics_list, str):
                    all_rd.append(rd_topics_list)

        structural = ""
        if all_catalysts:
            rows = "".join(
                f"<tr><td style='padding:4px 8px;color:#555;'>{_e(c.get('date','?'))}</td>"
                f"<td style='padding:4px 8px;'>{_e(c.get('event',''))}</td>"
                f"<td style='padding:4px 8px;color:#777;font-style:italic;'>{_e(c.get('impact',''))}</td></tr>"
                for c in all_catalysts)
            structural += f"""
            <p style="font-weight:bold;margin:12px 0 4px;color:#555;font-size:11px;text-transform:uppercase;">
                Dated Catalysts</p>
            <table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:10px;">
                <thead><tr style="background:#f5f5f5;">
                    <th style="padding:4px 8px;text-align:left;">Date</th>
                    <th style="padding:4px 8px;text-align:left;">Event</th>
                    <th style="padding:4px 8px;text-align:left;">Why it matters</th>
                </tr></thead><tbody>{rows}</tbody></table>"""

        if all_missing:
            seen = set()
            badges = ""
            for m in all_missing:
                if not isinstance(m, dict):
                    # Handle robustly when elements are strings instead of dict objects
                    sym = str(m).strip().upper()
                    tip = ""
                else:
                    if m.get("in_universe"):
                        continue
                    sym = m.get("symbol", "")
                    tip = _e(m.get("reason", ""))
                
                if sym and sym not in seen:
                    seen.add(sym)
                    badges += (f'<span title="{tip}" style="display:inline-block;margin:2px 4px;'
                               f'padding:2px 8px;background:#fff3e0;border:1px solid #ffb74d;'
                               f'border-radius:3px;font-size:12px;cursor:help;"><b>{_e(sym)}</b></span>')
            structural += f"""
            <p style="font-weight:bold;margin:12px 0 4px;color:#555;font-size:11px;text-transform:uppercase;">
                Not in Our Watchlist (hover for reason)</p>
            <div style="margin-bottom:10px;">{badges}</div>"""

        if all_rd:
            items = "".join(f"<li style='margin-bottom:4px;font-size:12px;color:#555;'>{_e(r)}</li>" for r in all_rd[:4])
            structural += f"""
            <p style="font-weight:bold;margin:12px 0 4px;color:#555;font-size:11px;text-transform:uppercase;">
                R&D Topics Implied</p>
            <ul style="margin:0;padding-left:18px;">{items}</ul>"""

        intel_section = f"""
        <div style="background:#fff8e1;border-left:5px solid #ffc107;padding:15px;margin-bottom:30px;border-radius:4px;">
            <h4 style="margin:0;color:#ffa000;text-transform:uppercase;font-size:11px;padding-bottom:5px;border-bottom:1px solid #ffe082;">
                External Intelligence ({len(intel_ideas)} emails scanned)</h4>
            {f'<ul style="margin:10px 0 0;font-size:13px;padding-left:20px;list-style:none;">{idea_list}</ul>' if idea_list else ''}
            {structural}
        </div>"""
    
    picks_rows = ""
    for i, p in enumerate(picks, 1):
        earnings = check_earnings(p['Symbol'])
        reasoning = get_reasoning(p['Symbol'], p['PGR'], p['S10'], p['L60'], p['Industry'])
        
        pattern_display = p.get('Patterns') or ''
        pattern_cell = (f'<span style="font-size:12px; color:#8e44ad; font-weight:bold;">{pattern_display}</span>'
                        if pattern_display else '<span style="color:#aaa; font-size:11px;">—</span>')
        picks_rows += f"""
        <tr style="border-bottom: 1px solid #ddd;">
            <td style="padding: 10px;">{i}</td>
            <td style="padding: 10px;"><b>{p['Symbol']}</b><br><small>{p['Industry']}</small></td>
            <td style="padding: 10px;">{p['PGR']}</td>
            <td style="padding: 10px;">{p['S10']:.1f} / {p['L60']:.1f}</td>
            <td style="padding: 10px;"><b>{p['Total']:.1f}</b></td>
            <td style="padding: 10px; font-size: 11px;">{reasoning}</td>
            <td style="padding: 10px;">Stop: ${p['Stop']}<br>Target: ${p['Target']}</td>
            <td style="padding: 10px; background-color: #e8f4fd;">
                ATR: <b>{p['Shares_ATR']}</b> | Stop: <b>{p['Shares_Stop']}</b>
            </td>
            <td style="padding: 10px;">{pattern_cell}</td>
            <td style="padding: 10px; color: {'#e74c3c' if 'Required' in earnings else '#333'};">{earnings}</td>
        </tr>
        """

    replacement_rows = ""
    for pair in replacements:
        reasoning = f"Rotating from {pair['Sell']} (Weakest) to {pair['Buy']} (Strongest institutional accumulation)."
        replacement_rows += f"""
        <tr style="border-bottom: 1px solid #ddd; font-size: 12px;">
            <td style="padding: 8px; color: #c0392b;"><b>{pair['Sell']}</b> ({pair['Sell_Score']:.1f})<br><small>{pair['Sell_Status']}</small></td>
            <td style="padding: 8px; text-align: center;">➡️</td>
            <td style="padding: 8px; color: #27ae60;"><b>{pair['Buy']}</b> ({pair['Buy_Score']:.1f})<br><small>PGR: {pair['Buy_PGR']}</small></td>
            <td style="padding: 8px; font-size: 11px; color: #555;">{reasoning}</td>
        </tr>
        """

    # 0.5. Format A-Reserves Audit
    reserves_data = get_reserves_data()
    reserves_rows = ""
    for r in reserves_data:
        reasoning = get_reasoning(r['Symbol'], r['PGR'], r['S10'], r['L60'], r['Industry'])
        color_total = '#27ae60' if r['Total'] >= 0 else '#c0392b'
        reserves_rows += f"""
        <tr style="border-bottom: 1px solid #ddd; font-size: 13px;">
            <td style="padding: 10px; font-weight: bold; color: #2c3e50;">{r['Symbol']}</td>
            <td style="padding: 10px;">{r['Industry']}</td>
            <td style="padding: 10px; font-weight: bold;">{r['PGR']}</td>
            <td style="padding: 10px;">{r['S10']:.1f} / {r['L60']:.1f}</td>
            <td style="padding: 10px; font-weight: bold; color: {color_total};">{r['Total']:.1f}</td>
            <td style="padding: 10px; font-size: 11px;">{reasoning}</td>
        </tr>
        """

    # NOTE: this local MUST NOT be named `html` — the nested _e() helper above
    # calls html.escape(), and a local `html` here would shadow the module for
    # the whole function, making _e() raise NameError once intel is non-empty.
    html_doc = f"""
    <html>
    <body style="font-family: 'Segoe UI', Tahoma, sans-serif; color: #333; line-height: 1.6; max-width: 1100px; margin: auto;">
        <div style="background: #2c3e50; color: white; padding: 20px; border-radius: 8px 8px 0 0;">
            <h1 style="margin: 0; font-size: 24px;">Daily Trading Intelligence</h1>
            <p style="margin: 5px 0 0 0; opacity: 0.8;">Autonomous Market Analysis | {today}</p>
        </div>

        <div style="padding: 20px; border: 1px solid #eee; border-top: none;">
            {intel_section}
            <div style="display: flex; gap: 20px; margin-bottom: 30px;">
                <div style="flex: 1; background: #f8f9fa; padding: 15px; border-left: 5px solid {color}; border-radius: 4px;">
                    <h4 style="margin: 0; color: #7f8c8d; text-transform: uppercase; font-size: 11px;">Market Regime</h4>
                    <p style="margin: 5px 0 0 0; font-weight: bold; font-size: 18px; color: {color};">{regime}</p>
                </div>
                <div style="flex: 1; background: #f8f9fa; padding: 15px; border-left: 5px solid #3498db; border-radius: 4px;">
                    <h4 style="margin: 0; color: #7f8c8d; text-transform: uppercase; font-size: 11px;">System Health</h4>
                    <p style="margin: 5px 0 0 0; font-weight: bold;">{status_msg}</p>
                </div>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #eee; padding-bottom: 5px;">Top High-Probability Setups</h3>
            <table style="border-collapse: collapse; width: 100%; font-size: 13px; margin-bottom: 40px;">
                <thead>
                    <tr style="background-color: #34495e; color: white; text-align: left;">
                        <th style="padding: 12px;">Rank</th><th style="padding: 12px;">Symbol</th><th style="padding: 12px;">PGR</th>
                        <th style="padding: 12px;">S10/L60</th><th style="padding: 12px;">Total</th><th style="padding: 12px;">Reasoning & Ruthless Audit</th>
                        <th style="padding: 12px;">Levels</th><th style="padding: 12px;">Shares</th><th style="padding: 12px;">Patterns</th><th style="padding: 12px;">Earnings</th>
                    </tr>
                </thead>
                <tbody>{picks_rows if picks else '<tr><td colspan="10">No candidates found today.</td></tr>'}</tbody>
            </table>

            <h3 style="color: #2c3e50; border-bottom: 2px solid #eee; padding-bottom: 5px;">Portfolio Rotation Strategy</h3>
            <table style="border-collapse: collapse; width: 80%; font-size: 13px; margin-bottom: 40px;">
                <thead>
                    <tr style="background-color: #f2f2f2; text-align: left;">
                        <th style="padding: 10px;">SELL (Weakest)</th><th style="padding: 10px;"></th><th style="padding: 10px;">BUY (Strongest)</th><th style="padding: 10px;">Rotation Rationale</th>
                    </tr>
                </thead>
                <tbody>{replacement_rows if replacements else '<tr><td colspan="4">No replacement pairs identified.</td></tr>'}</tbody>
            </table>

            <h3 style="color: #2c3e50; border-bottom: 2px solid #eee; padding-bottom: 5px; margin-top: 45px;">🛡️ A-Reserves Sentinel & AI Risk Audit</h3>
            <table style="border-collapse: collapse; width: 100%; font-size: 13px;">
                <thead>
                    <tr style="background-color: #2c3e50; color: white; text-align: left;">
                        <th style="padding: 10px;">Symbol</th><th style="padding: 10px;">Industry / Theme</th><th style="padding: 10px;">PGR</th>
                        <th style="padding: 10px;">S10/L60</th><th style="padding: 10px;">Total Score</th><th style="padding: 10px;">AI Ruthless Audit & Strategic Catalyst</th>
                    </tr>
                </thead>
                <tbody>{reserves_rows}</tbody>
            </table>
            
            <div style="margin-top: 40px; border-top: 1px solid #eee; padding-top: 20px; font-size: 11px; color: #95a5a6; text-align: center;">
                <p><b>Risk Management:</b> ATR-based sizing (2*ATR volatility) vs. Stop-based gap. <b>AI Manager:</b> Live tracking at Data/ai_portfolio_performance.xlsx</p>
                <p>Generated by Project AETHER Professional Desk | 5:30 AM PST Execution</p>
            </div>
        </div>
    </body>
    </html>
    """
    return html_doc

def main():
    cleanup_orphaned_processes()
    
    # ── Pillar 3: Single-Instance Pipeline Lock (Cross-Process Overlap Guard) ──
    lock_path = BASE_DIR / "Data" / "pipeline_run.lock"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    
    if lock_path.exists():
        try:
            with open(lock_path, "r", encoding="utf-8") as lf:
                old_pid = int(lf.read().strip())
        except Exception:
            old_pid = 0
            
        if old_pid > 0:
            # Check if the process is actively running
            res = subprocess.run(["tasklist", "/FI", f"PID eq {old_pid}", "/FO", "CSV"], capture_output=True, text=True, errors="replace")
            if str(old_pid) in res.stdout:
                log(f"🛑 [Overlap Guard] Active pipeline process (PID {old_pid}) is already running! Exiting immediately to prevent race conditions or duplicate dispatches.")
                sys.exit(0)
                
    # Write current PID to lock file
    try:
        with open(lock_path, "w", encoding="utf-8") as lf:
            lf.write(str(os.getpid()))
    except Exception:
        pass
        
    # Register automatic lock cleanup on exit
    def _cleanup_pipeline_lock():
        try:
            if lock_path.exists():
                os.remove(lock_path)
        except Exception:
            pass
    atexit.register(_cleanup_pipeline_lock)

    no_email = "--no-email" in sys.argv[1:]
    no_history = "--no-history" in sys.argv[1:]
    report_only = "--report-only" in sys.argv[1:] or "--cached" in sys.argv[1:]

    # ── Pillar 1: Centralized Pre-Flight Diagnostics (R&D #21) ──
    # Actively test connections and fail-loud early before doing any write operations
    if not report_only:
        preflight_passed = run_preflight_diagnostics()
        if not preflight_passed:
            log("❌ [Pre-flight Failure] One or more critical system gateways are offline! Aborting daily pipeline run to prevent corrupt states.")
            try:
                notify.send_email(
                    "🚨 [CRITICAL AETHER ERROR] Pre-Flight Connection Diagnostics Failed!",
                    "The 5:30 AM morning pipeline failed its pre-flight diagnostic checklist.\n\n"
                    "One or more external API or email gateways are offline. The pipeline has safely and "
                    "defensively aborted to prevent duplicate run or file-locking leaks.\n\n"
                    "Please run 'python scripts/diagnostics/preflight_validator.py' manually to isolate the offline connection."
                )
            except Exception as e:
                log(f"Warning: Failed to send pre-flight failure alert email: {e}")
            sys.exit(1)

    log("Starting Daily Trading Pipeline...")
    
    # ── Configuration Placeholder Audit ──
    if getattr(CFG, "has_placeholders", False):
        log("Warning: Active placeholders detected in configuration! Dispatching health alert email...")
        details_str = "\n".join(f"- {ph}" for ph in CFG.placeholder_details)
        msg = f"AETHER Configuration Health Alert!\n\nActive placeholders were detected in your configuration:\n\n{details_str}\n\nPlease update your config.json or environment variables immediately to resolve this."
        try:
            if not no_email:
                notify.send_email("ALERT: AETHER Configuration Placeholders Detected", msg)
        except Exception as e:
            log(f"ERROR: Failed to send configuration health alert email: {e}")

    intel_ideas = []
    cache_path = BASE_DIR / "Data" / "intel_ideas_cache.json"

    if report_only:
        log("⏩ Running in --report-only / --cached mode. Skipping email intelligence, history backfills, and workbook refreshes...")
        if cache_path.exists():
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    intel_ideas = json.load(f)
                if intel_ideas:
                    log(f"Loaded {len(intel_ideas)} cached ideas from {cache_path.name}.")
            except Exception as e:
                log(f"Warning: Failed to load cached email ideas: {e}")
    else:
        # 0. Gather External Intel (Emails & News)
        log("Gathering external intelligence...")
        try:
            intel_ideas = external_intel.fetch_idea_emails()
            if intel_ideas:
                log(f"Fetched {len(intel_ideas)} ideas from email.")
                try:
                    cache_path.parent.mkdir(parents=True, exist_ok=True)
                    with open(cache_path, "w", encoding="utf-8") as f:
                        json.dump(intel_ideas, f, indent=2)
                except Exception as e:
                    log(f"Warning: Failed to cache email ideas: {e}")
            else:
                _pipeline_log.warning("External intel fetch returned 0 ideas — mailbox auth may have failed. Check aether.jsonl for details.")
                log("Warning: 0 ideas from email scan — see log for details.")
        except Exception as e:
            _pipeline_log.error(f"Could not fetch external intel: {e}")
            log(f"ERROR: Could not fetch external intel: {e}")

    if not report_only:
        # 1. Sync history (Backfill cache for deltas)
        if no_history:
            log("Skipping history backfill via --no-history...")
        else:
            log("Backfilling 5-day history (run_history.py)...")
            try:
                script_path = str(BASE_DIR / "run_history.py")
                subprocess.run(
                    [sys.executable, script_path, "5"], 
                    check=True, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE, 
                    encoding="utf-8", errors="replace",
                    timeout=600  # Fail fast after 10 minutes!
                )
                log("History backfilled.")
            except subprocess.TimeoutExpired:
                log("Warning: run_history.py timed out after 600s (Playwright hang likely). Bypassing backfill...")
            except subprocess.CalledProcessError as e:
                log(f"Warning: run_history.py failed (will continue): {e.stderr}")

        # 2. Execute main.py (writes today's closes into OHLCV JSON via _append_ohlcv_entry)
        log("Refreshing workbook (main.py)...")
        try:
            script_path = str(BASE_DIR / "main.py")
            subprocess.run(
                [sys.executable, script_path], 
                check=True, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE, 
                encoding="utf-8", errors="replace",
                timeout=600  # Fail fast after 10 minutes!
            )
            log("Workbook regenerated.")
        except subprocess.TimeoutExpired:
            # The workbook is the day's source of truth — do NOT silently continue on
            # stale data (Zero-Trust). Alert and abort so nothing downstream trades on it.
            error_msg = "main.py execution timed out after 600s (Playwright hang likely)."
            log(f"ABORT: {error_msg}")
            if not no_email:
                notify.send_email("ALERT: Daily Pipeline Failed", f"Pipeline aborted: {error_msg}")
            return
        except subprocess.CalledProcessError as e:
            error_msg = f"main.py failed: {e.stderr}"
            log(f"ABORT: {error_msg}")
            if not no_email:
                notify.send_email("ALERT: Daily Pipeline Failed",
                                  f"Pipeline aborted during main.py execution.\n\n{error_msg}")
            return

    if not report_only:
        # 2b. OHLCV recovery pass — repair missing/corrupted/stale Symbol_full files via RapidAPI.
        #     Today's closes are already written by main.py (Chaikin). This only touches symbols
        #     with gaps > 30 days. Non-fatal: pipeline continues even if RapidAPI is unavailable.
        log("OHLCV recovery pass (rapidapi.py)...")
        try:
            _ohlcv_syms = load_symbols()
            _today_str = str(datetime.date.today())
            _ohlcv_result = rapidapi.repair_missing(_ohlcv_syms, _today_str)
            log(f"OHLCV: {_ohlcv_result['updated']} recovered, "
                f"{_ohlcv_result['skipped']} already current, "
                f"{len(_ohlcv_result['errors'])} errors")
        except Exception as e:
            log(f"Warning: OHLCV recovery failed (non-fatal, pipeline continues): {e}")

    # 3. Verify data freshness
    fresh, msg = verify_data_freshness()
    log(msg)
    if not fresh:
        if not no_email:
            notify.send_email("ALERT: Daily Pipeline Stale Data", msg)
        return

    # 4. Validate sheets
    valid, v_msg = validate_sheets()
    log(v_msg)
    if not valid:
        if not no_email:
            notify.send_email("ALERT: Daily Pipeline Validation Failed", v_msg)
        return

    # 5. Compute top-5 picks & replacements
    log("Computing top-5 picks and replacements...")
    picks = get_top_5_picks()
    replacements = get_replacement_pairs()

    # 6. Log for performance tracking
    if picks:
        log("Logging picks for performance tracking...")
        performance_tracker.log_picks(picks)

    # 7. Send report
    if not no_email:
        log("Drafting and sending HTML report...")
        html = format_html_report(msg, picks, replacements, intel_ideas)
        notify.send_email(f"AETHER Daily Autopilot Trading & Intelligence Report: {datetime.date.today()}", html, is_html=True)
    else:
        log("Drafting HTML report (email disabled via --no-email)...")

    log("Pipeline completed successfully.")

    try:
        watchdog.sync_data_folder()
    except Exception as e:
        log(f"Post-pipeline sync failed: {e}")


if __name__ == "__main__":
    main()
