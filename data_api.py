"""
Pure data-reading layer for the AETHER web dashboard.
No HTTP, no FastAPI — only reads from files and calls existing modules.
All functions are safe to call from async FastAPI route handlers.
"""

import json
import logging
import re
import subprocess
import sys
import threading
import time
import glob
import pytz
from datetime import datetime, date, timedelta
from pathlib import Path
from scripts.backtesting import backtest_levels
from aether import decision_eval as _decision_eval
from aether.config import CFG as _cfg
import ai_portfolio_game
import instruments
import risk_utils
import sell_rules
import openpyxl
import etrade
from workbook_read import (
    get_top_5_picks as _ap_picks,
    get_market_regime as _ap_regime,
    get_replacement_pairs as _ap_replacements,
    get_reserves_data as _ap_reserves,
)
import powergauge as _pg

_log = logging.getLogger("aether.data_api")


def _live_prices(symbols: list[str]) -> dict[str, float]:
    """Canonical live-price fetch — the single source of truth for the
    E*TRADE/JSON/Google fallback ladder lives in ai_portfolio_game."""
    return ai_portfolio_game.get_live_prices(symbols)


def _google_prices(symbols: list[str]) -> dict[str, float]:
    """Canonical Google-Finance fallback (see _live_prices)."""
    return ai_portfolio_game.get_google_prices_fallback(symbols)

_DIR      = Path(__file__).resolve().parent
_DATA_DIR = _DIR / "Data"
_XLSX     = _DATA_DIR / "state_of_the_day.xlsx"
_GAME     = _DATA_DIR / "ai_portfolio_game.json"
_LOG      = _DATA_DIR / "autonomous_run.log"
_PERF     = _DATA_DIR / "performance_log.json"

# ── Digit-sum study index (loaded once, keyed by symbol) ──────────────────────
_digit_study_index: dict | None = None  # {symbol: [row, ...]}
_digit_study_lock = threading.Lock()

def _get_digit_study(sym: str) -> list:
    """Return all study rows for the given symbol (both integer and full-cents variants).
    Loaded once from Data/digit_sum_study.json and Data/digit_sum_full_study.json."""
    global _digit_study_index
    with _digit_study_lock:
        if _digit_study_index is None:
            idx: dict = {}
            for path in [_DATA_DIR / "digit_sum_study.json",
                         _DATA_DIR / "digit_sum_full_study.json"]:
                if path.exists():
                    try:
                        with open(path, "r", encoding="utf-8") as f:
                            for r in json.load(f):
                                idx.setdefault(r["symbol"], []).append(r)
                    except Exception:
                        pass
            _digit_study_index = idx  # lock held for entire build; no duplicate work
    return _digit_study_index.get(sym, [])

# ── Simple in-process TTL cache ───────────────────────────────────────────────

_cache: dict = {}
_cache_lock = threading.Lock()


def _cached(key: str, ttl: float, fn):
    now = time.monotonic()
    with _cache_lock:
        entry = _cache.get(key)
        if entry and now - entry["ts"] < ttl:
            return entry["val"]
    val = fn()
    with _cache_lock:
        _cache[key] = {"ts": now, "val": val}
    return val


def _load_latest_close_from_cache(sym: str) -> float | None:
    """Retrieve the latest closing price directly from the local per-symbol OHLCV JSON cache on disk."""
    try:
        sym = (sym or "").strip().upper()
        path = _DATA_DIR / "Symbol_full" / f"{sym}_daily.json"
        if not path.exists():
            return None
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        ts = data.get("Time Series (Daily)", {})
        if not ts:
            return None
        newest_date = sorted(ts.keys())[-1]
        return float(ts[newest_date]["4. close"])
    except Exception:
        return None


class PricingDiscrepancyError(ValueError):
    """Raised when a live price deviates from the settled OHLCV cache beyond tolerance."""


_PRICE_CHECK_EXEMPT = frozenset({"931CVR013", "BBB"})
_PRICE_CACHE_MAX_STALE_DAYS = 10  # cache older than this is not authoritative


def is_active_nyse_market_hours() -> bool:
    """Check if NYSE is currently open (9:30 AM - 4:00 PM Eastern, weekdays)."""
    try:
        tz_ny = pytz.timezone("America/New_York")
        now_ny = datetime.now(tz_ny)
        if now_ny.weekday() in (5, 6):
            return False
        start_time = now_ny.replace(hour=9, minute=30, second=0, microsecond=0)
        end_time = now_ny.replace(hour=16, minute=0, second=0, microsecond=0)
        return start_time <= now_ny <= end_time
    except Exception:
        return False


def _get_chaikin_price(sym: str) -> float | None:
    """Read the newest cached Chaikin closing price for a symbol on disk.

    Deliberately re-reads on each call rather than memoizing: verify_price_integrity
    fires only on held positions / account holdings (dozens per run, not the full
    universe), and the sole long-lived caller (server.py) must see fresh Chaikin
    files as they land on disk — a process-lifetime memo would freeze prices and
    defeat the drift detection this feeds."""
    try:
        symbol_dir = _DATA_DIR / "Symbol" / sym
        if not symbol_dir.exists():
            return None
        json_files = sorted(list(symbol_dir.glob(f"{sym}_*.json")), reverse=True)
        if not json_files:
            return None

        newest_file = json_files[0]
        with open(newest_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        meta = data.get("metaInfo")
        if isinstance(meta, list) and len(meta) > 0:
            return float(meta[0].get("Last", 0.0))
        elif isinstance(meta, dict):
            return float(meta.get("Last", 0.0))

        checklist = data.get("checklist_stocks")
        if isinstance(checklist, dict):
            return float(checklist.get("lastPrice", 0.0))
    except Exception:
        pass
    return None


def verify_price_integrity(symbol: str, price: float, source: str) -> None:
    """Raise PricingDiscrepancyError when a price deviates from our fresh OHLCV cache
    or Chaikin PG cache beyond tolerance. Skips stale caches (>10 days old).

    Cross-verification degrades gracefully by available sources: the active price
    (from `source`, e.g. E*TRADE) is always checked against the RapidAPI Symbol_full
    cache (2-way); when a fresh Chaikin PG price is also available it escalates to a
    3-way check (active vs Chaikin, and Symbol_full vs Chaikin). Error messages are
    tagged "2-Way"/"3-Way" to reflect which comparison actually fired."""
    sym = (symbol or "").strip().upper()
    if sym in _PRICE_CHECK_EXEMPT or instruments.is_excluded(sym):
        return
    if price is None or price <= 0:
        raise PricingDiscrepancyError(
            f"[PRICING DISCREPANCY] {sym} has invalid price: {price} (Source: {source})"
        )

    # Skip cache reconciliation if NYSE is currently open (volatile active market hours)
    if is_active_nyse_market_hours():
        return

    path = _DATA_DIR / "Symbol_full" / f"{sym}_daily.json"
    if not path.exists():
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            ts = json.load(f).get("Time Series (Daily)", {})
        if not ts:
            return
        newest_date = sorted(ts.keys())[-1]
        stale_days = (date.today() - date.fromisoformat(newest_date)).days
        if stale_days > _PRICE_CACHE_MAX_STALE_DAYS:
            return  # cache is stale — not authoritative, skip check
        cache_close = float(ts[newest_date]["4. close"])
        cache_open = float(ts[newest_date].get("1. open", 0.0))
        cache_high = float(ts[newest_date].get("2. high", 0.0))
        cache_low = float(ts[newest_date].get("3. low", 0.0))
    except Exception as e:
        _log.warning(f"[PRICING] {sym}: could not read OHLCV cache for integrity check: {e}")
        return

    if cache_close <= 0:
        return

    # ── 0. Detect Flat Placeholders (OHLC Validation) ──
    # Weekday prices must have an active trading range; if they are perfectly flat,
    # it indicates a fake placeholder written by powergauge._append_ohlcv_entry.
    is_weekday = date.fromisoformat(newest_date).weekday() not in (5, 6)
    if is_weekday and cache_open == cache_high == cache_low == cache_close:
        # Ignore flat placeholder if it is for today's date, as it's a temporary entry
        # before the official after-hours RapidAPI sync succeeds.
        if newest_date == date.today().isoformat():
            _log.debug(f"[PRICING] {sym}: today's temporary flat placeholder tolerated.")
        else:
            raise PricingDiscrepancyError(
                f"🛑 [PRICING DISCREPANCY] Flat placeholder detected for {sym}!\n"
                f"  Date:                            {newest_date}\n"
                f"  Cache OHLC (Symbol_full):       ${cache_close:.4f} (Open == High == Low == Close)\n"
                f"  Action required: This is a stale placeholder. Force-refresh Symbol_full history immediately!"
            )

    # ── 1. Compare Active Price (Source) vs. RapidAPI (Symbol_full) ──
    diff_rap = abs(price - cache_close)
    if diff_rap > 0.05:
        pct_diff = (diff_rap / cache_close) * 100
        if pct_diff > 3.0:
            raise PricingDiscrepancyError(
                f"🛑 [PRICING DISCREPANCY] 2-Way Pricing discrepancy detected for {sym}!\n"
                f"  Active Price (Source: {source}): ${price:.4f}\n"
                f"  Cache Price (Symbol_full):       ${cache_close:.4f}\n"
                f"  Difference:                      ${diff_rap:.4f} ({pct_diff:.2f}% - Max tolerance: 3.0%)\n"
                f"  Action required: Re-sync Data/Symbol_full or Excel workbook immediately!"
            )

    # ── 2. Compare Active Price (Source) and RapidAPI vs. Chaikin PowerGauge (PG) ──
    chaikin_close = _get_chaikin_price(sym)
    if chaikin_close is not None and chaikin_close > 0:
        # Check Active Price vs Chaikin
        diff_pg1 = abs(price - chaikin_close)
        if diff_pg1 > 0.05:
            pct_diff1 = (diff_pg1 / chaikin_close) * 100
            if pct_diff1 > 3.0:
                raise PricingDiscrepancyError(
                    f"🛑 [PRICING DISCREPANCY] 3-Way Pricing discrepancy detected for {sym}!\n"
                    f"  Active Price (Source: {source}): ${price:.4f}\n"
                    f"  Chaikin PG Price (Cached):       ${chaikin_close:.4f}\n"
                    f"  Difference:                      ${diff_pg1:.4f} ({pct_diff1:.2f}% - Max tolerance: 3.0%)\n"
                    f"  Action required: Re-authenticate Chaikin or check Data/Symbol cache!"
                )
                
        # Check RapidAPI vs Chaikin
        diff_pg2 = abs(cache_close - chaikin_close)
        if diff_pg2 > 0.05:
            pct_diff2 = (diff_pg2 / chaikin_close) * 100
            if pct_diff2 > 3.0:
                raise PricingDiscrepancyError(
                    f"🛑 [PRICING DISCREPANCY] 3-Way Pricing discrepancy detected for {sym}!\n"
                    f"  Cache Price (Symbol_full):       ${cache_close:.4f}\n"
                    f"  Chaikin PG Price (Cached):       ${chaikin_close:.4f}\n"
                    f"  Difference:                      ${diff_pg2:.4f} ({pct_diff2:.2f}% - Max tolerance: 3.0%)\n"
                    f"  Action required: Re-sync Data/Symbol_full or Chaikin cache immediately!"
                )


# ── Portfolio ─────────────────────────────────────────────────────────────────

def read_portfolio() -> dict:
    """Read ai_portfolio_game.json and compute position-level P&L."""
    try:
        with open(_GAME, encoding="utf-8") as f:
            state = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"balance": 0, "equity": 0, "return_pct": 0, "positions": [],
                "profile": "UNKNOWN", "open_positions": 0, "max_positions": 5,
                "start_date": "", "total_return": 0}

    positions = state.get("positions", {})
    balance   = state.get("balance", 0)
    history   = state.get("history", [])

    # Derive initial balance from history or default
    initial = 10000.0
    pos_list = []
    total_value = balance
    for sym, pos in positions.items():
        cost  = pos.get("cost", 0)
        qty   = pos.get("qty", 0)
        stop  = pos.get("stop_loss", 0)
        # Compute days held from history
        days_held = 0
        for tx in reversed(history):
            if tx.get("symbol") == sym and tx.get("type") == "BUY":
                try:
                    d = datetime.fromisoformat(tx["date"]).date()
                    days_held = (date.today() - d).days
                except Exception:
                    pass
                break

        # Load actual correct close from JSON cache instead of defaulting to cost
        current = _load_latest_close_from_cache(sym) or cost
        try:
            verify_price_integrity(sym, current, "Game Portfolio")
        except PricingDiscrepancyError as e:
            _log.warning(f"[PRICING] Game position {sym} has discrepancy: {e}")
        pnl = round((current - cost) * qty, 2)
        pnl_pct = round((current - cost) / cost * 100, 2) if cost > 0 else 0.0
        total_value += qty * current

        pos_list.append({
            "symbol":        sym,
            "qty":           qty,
            "cost":          round(cost, 2),
            "current_price": round(current, 2),
            "pnl":           pnl,
            "pnl_pct":       pnl_pct,
            "stop_loss":     round(stop, 2),
            "days_held":     days_held,
        })

    equity = round(total_value, 2)
    return_pct = round((equity - initial) / initial * 100, 2) if initial else 0

    return {
        "balance":        round(balance, 2),
        "equity":         round(equity, 2),
        "return_pct":     return_pct,
        "total_return":   round(equity - initial, 2),
        "profile":        state.get("profile", "BALANCED"),
        "positions":      pos_list,
        "open_positions": len(pos_list),
        "max_positions":  5,
        "start_date":     state.get("start_date", ""),
    }


# ── Picks & replacements ──────────────────────────────────────────────────────

def read_picks() -> dict:
    """Cached 60s — reads state_of_the_day.xlsx Research sheet."""
    def _load():
        try:
            picks = _ap_picks()
            regime, color = _ap_regime()
            return {"market_regime": regime, "regime_color": color, "picks": picks}
        except Exception as e:
            return {"market_regime": "Unknown", "regime_color": "#7f8c8d",
                    "picks": [], "error": str(e)}
    return _cached("picks", 60.0, _load)


def read_replacements() -> dict:
    """Cached 60s — reads Replacements sheet."""
    def _load():
        try:
            pairs = _ap_replacements()
            return {"pairs": pairs}
        except Exception as e:
            return {"pairs": [], "error": str(e)}
    return _cached("replacements", 60.0, _load)


def read_reserves() -> dict:
    """Cached 60s — reads A-Reserves from game state + Research sheet scores."""
    def _load():
        try:
            data = _ap_reserves()
            return {"reserves": data}
        except Exception as e:
            return {"reserves": [], "error": str(e)}
    return _cached("reserves", 60.0, _load)


# ── Research sheet (full screener output) ──────────────────────────────────────

# Research sheet column layout (0-based tuple index), matching powergauge.py writes.
_RESEARCH = {"sym": 3, "industry": 4, "prev_pgr": 5, "pgr": 6, "ind_strength": 7,
             "stop": 9, "price": 10, "target": 11, "risk_ratio": 12,
             "lt_trend": 17, "money_flow": 18, "obos": 19, "setup": 20,
             "buying_ratio": 21, "seasonality": 22, "winpct": 23,
             "s10": 24, "l60": 25, "patterns": 26}


def read_research() -> dict:
    """All screened symbols from the Research sheet with their full computed fields
    (PGR, S10/L60, setup flag, buying ratio, money flow, OB/OS, win%, patterns …)
    plus a summary block. Cached 60s."""
    def _load():
        if not _XLSX.exists():
            return {"rows": [], "summary": {}, "error": "state_of_the_day.xlsx not found"}
        rows = []
        stale_stops = 0
        max_stale_age = 0
        support_misses = 0
        target_misses = 0
        wb = None
        try:
            wb = openpyxl.load_workbook(_XLSX, read_only=True, data_only=True)
            ws = wb["Research"]
            maxi = max(_RESEARCH.values())
            for r in ws.iter_rows(min_row=2, values_only=True):
                if len(r) <= maxi:
                    continue
                sym = r[_RESEARCH["sym"]]
                if not sym or not isinstance(sym, str) or sym.strip().upper() == "SYMB":
                    continue
                g = lambda k: r[_RESEARCH[k]]
                s10, l60 = _f(g("s10")), _f(g("l60"))
                setup_raw = g("setup")
                win = _f(g("winpct"))
                price = _f(g("price"))
                sheet_stop = _f(g("stop"))
                # OHLCV-authoritative stop: when the cache is fresh, derive the stop
                # from it (confirmed swing-low -> shallow low -> ATR -> 8%), ignoring
                # a possibly-stale sheet value. Fall back to the sheet only when the
                # cache can't produce one (missing file / no price).
                # TEMPORARY: leveraged/inverse/crypto ETFs skip the long swing-low
                # method and use ATR levels (instruments.py) — still fully protected.
                excl = instruments.is_excluded(sym.strip())
                d = risk_utils.resolve_stop_detailed(price, symbol=sym.strip(), exclude_swing=excl)
                stop_source = d["source"]
                if d["stop"] is not None:
                    stop = d["stop"]
                else:
                    stop = sheet_stop if (sheet_stop and sheet_stop > 0) else None
                    if stop is not None:
                        stop_source = "sheet"
                if d["stale"]:
                    stale_stops += 1
                    max_stale_age = max(max_stale_age, d["age"] or 0)
                elif not excl and d["source"] in ("atr", "pct"):
                    support_misses += 1   # fresh data but no real support/swing found

                # Target — mirror of the stop: nearest confirmed swing-high resistance
                # (OHLCV-authoritative), sheet fallback only when the cache can't help.
                sheet_target = _f(g("target"))
                t = risk_utils.resolve_target_detailed(price, symbol=sym.strip(), exclude_swing=excl)
                target_source = t["source"]
                if t["target"] is not None:
                    target = t["target"]
                else:
                    target = sheet_target if (sheet_target and sheet_target > 0) else None
                    if target is not None:
                        target_source = "sheet"
                if not excl and not t["stale"] and t["source"] in ("atr", "pct"):
                    target_misses += 1    # fresh data but no real resistance above

                # Recompute R:R from the resolved levels so the row is self-consistent.
                if stop and target and price and price > stop:
                    risk_ratio = round((target - price) / (price - stop), 2)
                else:
                    risk_ratio = _f(g("risk_ratio"))
                rows.append({
                    "symbol": sym.strip(),
                    "fractional": instruments.is_fractional_eligible(sym.strip()),
                    "industry": g("industry"),
                    "pgr": g("pgr"), "prev_pgr": g("prev_pgr"),
                    # These four are categorical text ratings (e.g. Weak/Neutral/Wait),
                    # not numbers — pass through raw.
                    "industry_strength": g("ind_strength"),
                    "lt_trend": g("lt_trend"), "money_flow": g("money_flow"),
                    "obos": g("obos"),
                    "price": price, "stop": stop, "stop_source": stop_source,
                    "target": target, "target_source": target_source,
                    "risk_ratio": risk_ratio, "instrument": instruments.classify(sym.strip()),
                    "setup": str(setup_raw) in ("1", "OK") or setup_raw == 1,
                    "buying_ratio": _f(g("buying_ratio")),
                    "seasonality": _f(g("seasonality")),
                    "win_pct": round(win * 100, 1) if win is not None else None,
                    "s10": s10, "l60": l60,
                    "combined": round((s10 or 0) + (l60 or 0), 1),
                    "status": sell_rules.status_label(l60),
                    "patterns": g("patterns") or "",
                })
        except Exception as e:
            return {"rows": [], "summary": {}, "error": str(e)}
        finally:
            if wb:
                wb.close()

        setups = sum(1 for x in rows if x["setup"])
        bullish = sum(1 for x in rows if x["combined"] > 0)
        combos = [x["combined"] for x in rows]
        summary = {
            "total": len(rows),
            "setups": setups,
            "bullish": bullish,
            "bearish": sum(1 for x in rows if x["combined"] < 0),
            "avg_combined": round(sum(combos) / len(combos), 2) if combos else 0.0,
            "stale_stops": stale_stops,
            "ohlcv_max_age_days": max_stale_age if stale_stops else 0,
            "support_misses": support_misses,
            "target_misses": target_misses,
        }
        # Alert on data gaps that weaken the stop.
        if stale_stops:
            _log.warning(f"OHLCV STALE: {stale_stops}/{len(rows)} symbols have caches "
                         f"older than {risk_utils.STALE_STOP_DAYS}d (oldest {max_stale_age}d) — "
                         f"their stops fell back to 8% off the live price. Refresh Data/Symbol_full.")
        if support_misses:
            _log.warning(f"SUPPORT MISS: {support_misses}/{len(rows)} symbols have fresh "
                         f"data but no confirmed swing-low support — stop used an ATR/8% fallback.")
        if target_misses:
            _log.warning(f"TARGET MISS: {target_misses}/{len(rows)} symbols have fresh "
                         f"data but no overhead resistance — target used an ATR/8% projection.")
        try:
            regime, color = _ap_regime()
            summary["market_regime"], summary["regime_color"] = regime, color
        except Exception:
            summary["market_regime"], summary["regime_color"] = "Unknown", "#7f8c8d"
        return {"rows": rows, "summary": summary}

    return _cached("research", 60.0, _load)


# ── Accounts (2 real from Short_Long sheet + 1 AI game) ────────────────────────

# Short_Long column layout (0-based), matching workbook_write.update_short_long_scores.
_SL = {"sym": 1, "qty": 2, "buy": 3, "top": 4, "target": 5, "stop": 6, "buy_date": 10,
       "s10": 16, "l60": 17, "winpct": 18, "status": 19, "in_profit": 22,
       "streak_green": 20, "streak_red": 21}


def _to_date_str(v):
    """Normalize a Short_Long date cell (Excel serial int, datetime, date, or str)
    to 'YYYY-MM-DD', or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        try:
            return (date(1899, 12, 30) + timedelta(days=int(v))).isoformat()
        except Exception:
            return None
    s = str(v).strip()
    return s[:10] if s else None
# The two real E*TRADE accounts (last-4 IDs), top table first. Sourced from config
# (PII — never hardcode). Falls back to generic T1/T2 labels if unset.
def _real_acct_ids():
    try:
        return _cfg.accounts_real or []
    except Exception:
        return []


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _get_streak_from_cache(sym: str) -> int:
    """Calculate the current consecutive Green/Red closing price streak dynamically from local Price History Closes (Symbol_full).
    Filters out weekend dates and counts consecutive upward (Green) or downward (Red) closing price changes."""
    try:
        sym = (sym or "").strip().upper()
        path = _DATA_DIR / "Symbol_full" / f"{sym}_daily.json"
        if not path.exists():
            return None
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        ts = data.get("Time Series (Daily)", {})
        
        # Filter out weekend dates (Saturdays and Sundays) to get actual trading days
        dates = []
        for d_str in ts.keys():
            try:
                dt = date.fromisoformat(d_str)
                if dt.weekday() not in (5, 6):
                    dates.append(d_str)
            except Exception:
                pass
                
        dates = sorted(dates, reverse=True)  # Newest first
        if len(dates) < 2:
            return 0

        first_close = float(ts[dates[0]]["4. close"])
        second_close = float(ts[dates[1]]["4. close"])

        streak = 0
        if first_close > second_close:
            # Green streak (consecutive positive daily changes)
            for i in range(len(dates) - 1):
                cur = float(ts[dates[i]]["4. close"])
                prev = float(ts[dates[i+1]]["4. close"])
                if cur > prev:
                    streak += 1
                else:
                    break
        elif first_close < second_close:
            # Red streak (consecutive negative daily changes)
            for i in range(len(dates) - 1):
                cur = float(ts[dates[i]]["4. close"])
                prev = float(ts[dates[i+1]]["4. close"])
                if cur < prev:
                    streak -= 1
                else:
                    break
        return streak
    except Exception:
        return None


def read_accounts() -> dict:
    """Return live E*TRADE account holdings and balances from the broker,
    with an automatic fallback to the local Excel sheet if offline."""
    def _load():
        accounts = []
        scores = {}
        _streak_cache: dict = {}

        def _streak(sym: str):
            if sym not in _streak_cache:
                _streak_cache[sym] = _get_streak_from_cache(sym)
            return _streak_cache[sym]
        sl_decorations = {}
        env = "production"
        broker_status = "live"   # "live" | "offline" | "reconnecting"

        # ── PRIMARY: Live E*TRADE Broker Feed ──────────────────────────────────
        in_unittest = any("unittest" in arg for arg in sys.argv)
        
        # Load Research scores and Short_Long decorations unconditionally —
        # needed whether E*TRADE is online or offline (fallback path uses them too).
        if not in_unittest:
            try:
                wb = openpyxl.load_workbook(_XLSX, data_only=True, read_only=True)
                try:
                    ws = wb["Research"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        sym = row[3]
                        if sym:
                            scores[sym] = {
                                "s10": row[24], "l60": row[25], "status": row[19],
                                "streak": _streak(sym),
                            }

                    ws_sl = wb["Short_Long"]
                    sl_rows = list(ws_sl.iter_rows(values_only=True))
                    hdrs = [i for i, r in enumerate(sl_rows)
                            if len(r) > 1 and str(r[1]).strip() == "Symb"]

                    if hdrs:
                        h = hdrs[0]
                        started = False
                        for r in sl_rows[h + 1:]:
                            sym = r[_SL["sym"]] if len(r) > _SL["sym"] else None
                            sym = str(sym).strip().upper() if sym else ""
                            if not sym:
                                if started:
                                    break
                                continue
                            started = True
                            buy = _f(r[_SL["buy"]])
                            buy_date = _to_date_str(r[_SL["buy_date"]] if len(r) > _SL["buy_date"] else None)
                            excl = instruments.is_excluded(sym)
                            stop = _f(r[_SL["stop"]])
                            target = _f(r[_SL["target"]])
                            stop_src = target_src = "sheet"
                            if buy:
                                sd = risk_utils.resolve_stop_detailed(buy, symbol=sym, as_of=buy_date, exclude_swing=excl)
                                if sd["stop"] is not None:
                                    stop, stop_src = sd["stop"], sd["source"]
                                td = risk_utils.resolve_target_detailed(buy, symbol=sym, as_of=buy_date, exclude_swing=excl)
                                if td["target"] is not None:
                                    target, target_src = td["target"], td["source"]
                            red_streak = _f(r[_SL["streak_red"]]) if len(r) > _SL["streak_red"] else None
                            green_streak = _f(r[_SL["streak_green"]]) if len(r) > _SL["streak_green"] else None
                            if green_streak and green_streak > 0:
                                sl_streak = int(green_streak)
                            elif red_streak and red_streak > 0:
                                sl_streak = -int(red_streak)
                            else:
                                sl_streak = None

                            sl_decorations[sym] = {
                                "stop": stop, "stop_source": stop_src,
                                "target": target, "target_source": target_src,
                                "s10":     _f(r[_SL["s10"]]      if len(r) > _SL["s10"]      else None),
                                "l60":     _f(r[_SL["l60"]]      if len(r) > _SL["l60"]      else None),
                                "status":  str(r[_SL["status"]]  if len(r) > _SL["status"]   else "") or "",
                                "win_pct": r[_SL["winpct"]]      if len(r) > _SL["winpct"]   else None,
                                "buy_date": buy_date,
                                "in_profit": str(r[_SL["in_profit"]] if len(r) > _SL["in_profit"] else "") or "",
                                "streak": sl_streak,
                            }
                finally:
                    wb.close()
            except Exception as ex:
                _log.warning(f"Error loading Excel decorations: {ex}")

        if not in_unittest:
            try:
                acct_list = []
                raw_positions = []
                tokens = etrade.get_tokens(env)
                if tokens:
                    accts_api = etrade.get_accounts(tokens, env)
                    resp = accts_api.list_accounts(resp_format="json")
                    acct_list = resp.get("AccountListResponse", {}).get("Accounts", {}).get("Account", [])
                    if isinstance(acct_list, dict):
                        acct_list = [acct_list]
                    raw_positions = etrade.fetch_positions(tokens, env)

                for acct in acct_list:
                    desc = acct.get("accountDesc", "Brokerage")
                    acct_id = acct.get("accountId", "")
                    acct_key = acct.get("accountIdKey", "")
                    
                    # Fetch live balance from E*TRADE
                    bal_resp = accts_api.get_account_balance(acct_key, resp_format="json")
                    comp = bal_resp.get("BalanceResponse", {}).get("Computed", {})
                    rt_vals = comp.get("RealTimeValues", {})
                    
                    val = float(rt_vals.get("totalAccountValue", 0.0))
                    cash = float(comp.get("netCash", 0.0)) or float(comp.get("cashBalance", 0.0))
                    
                    # Skip empty real accounts to keep the UI clean
                    if val <= 0 and cash <= 0:
                        continue
                        
                    # Filter positions belonging to this specific account last-4 digits
                    last4 = acct_id[-4:] if len(acct_id) >= 4 else acct_id
                    holdings = []
                    for p in raw_positions:
                        if p.get("account_last4") == last4:
                            sym = p["symbol"]
                            sc = scores.get(sym, {})
                            dec = sl_decorations.get(sym.strip().upper(), {})

                            s10 = _f(sc.get("s10")) if sc.get("s10") is not None else _f(dec.get("s10"))
                            l60 = _f(sc.get("l60")) if sc.get("l60") is not None else _f(dec.get("l60"))
                            status = sc.get("status") or dec.get("status", "")

                            stop = dec.get("stop")
                            stop_source = dec.get("stop_source", "E*TRADE")
                            target = dec.get("target")
                            target_source = dec.get("target_source", "E*TRADE")
                            
                            buy = p.get("cost", 0.0)
                            current = p.get("price", 0.0)
                            try:
                                verify_price_integrity(sym, current, f"E*TRADE Account {last4}")
                            except PricingDiscrepancyError:
                                corrected = _load_latest_close_from_cache(sym)
                                if corrected and corrected > 0:
                                    _log.warning(f"[PRICING] {sym}: overriding broker ${current:.2f} with cache ${corrected:.2f}")
                                    current = corrected
                            qty = p.get("qty", 0)
                            
                            pnl = 0.0
                            pnl_pct = 0.0
                            if buy and buy > 0:
                                pnl = round((current - buy) * qty, 2)
                                pnl_pct = round((current - buy) / buy * 100, 2)
                            
                            holdings.append({
                                "symbol": sym,
                                "fractional": instruments.is_fractional_eligible(sym),
                                "qty": qty,
                                "buy": buy,
                                "current": current,
                                "pnl": pnl,
                                "pnl_pct": pnl_pct,
                                "stop": stop,
                                "stop_source": stop_source,
                                "target": target,
                                "target_source": target_source,
                                "s10": s10,
                                "l60": l60,
                                "total": round((s10 or 0) + (l60 or 0), 1) if s10 is not None or l60 is not None else None,
                                "status": status,
                                "streak": _streak(sym),
                                "buy_date": dec.get("buy_date", ""),
                                "win_pct": dec.get("win_pct"),
                                "in_profit": dec.get("in_profit", ""),
                                "instrument": instruments.classify(sym)
                            })
                            
                    accounts.append({
                        "id": last4,
                        "label": f"Real · {desc} (...{last4})",
                        "type": "real",
                        "balance": cash,
                        "equity": val,
                        "holdings": holdings,
                        "count": len(holdings)
                    })
            except ValueError as ve:
                _log.warning(f"Live broker feed failed: {ve}. Falling back to Excel.")
                broker_status = "offline"
            except Exception as e:
                _log.warning(f"Live broker feed failed: {e}. Falling back to Excel.")
                broker_status = "offline"

        # ── FALLBACK: Parse merged Short_Long sheet if API fails ────────────────
        if not accounts:
            real_ids = _real_acct_ids()
            try:
                wb = openpyxl.load_workbook(_XLSX, data_only=True, read_only=True)
                try:
                    rows = list(wb["Short_Long"].iter_rows(values_only=True))
                finally:
                    wb.close()

                hdrs = [i for i, r in enumerate(rows)
                        if len(r) > 1 and str(r[1]).strip() == "Symb"]

                for tbl_idx, h in enumerate(hdrs[:2]):
                    holdings, started = [], False
                    for r in rows[h + 1:]:
                        sym = r[_SL["sym"]] if len(r) > _SL["sym"] else None
                        sym = str(sym).strip() if sym else ""
                        if not sym:
                            if started:
                                break          # blank row after data → end of table
                            continue            # skip leading blanks between header and data
                        started = True
                        buy = _f(r[_SL["buy"]]); top = _f(r[_SL["top"]]); qty = _f(r[_SL["qty"]])
                        # Prefer Research sheet scores (always fresher) over Short_Long columns
                        _rsc = scores.get(sym.upper(), scores.get(sym, {}))
                        s10 = _f(_rsc.get("s10")) if _rsc.get("s10") is not None else _f(r[_SL["s10"]])
                        l60 = _f(_rsc.get("l60")) if _rsc.get("l60") is not None else _f(r[_SL["l60"]])
                        buy_date = _to_date_str(r[_SL["buy_date"]] if len(r) > _SL["buy_date"] else None)
                        pnl = pnl_pct = None
                        if buy and top and qty:
                            pnl = round((top - buy) * qty, 2)
                            pnl_pct = round((top - buy) / buy * 100, 2)
                        
                        red_streak = _f(r[_SL["streak_red"]]) if len(r) > _SL["streak_red"] else None
                        green_streak = _f(r[_SL["streak_green"]]) if len(r) > _SL["streak_green"] else None
                        if green_streak and green_streak > 0:
                            sl_streak = int(green_streak)
                        elif red_streak and red_streak > 0:
                            sl_streak = -int(red_streak)
                        else:
                            sl_streak = None

                        excl = instruments.is_excluded(sym)
                        stop = _f(r[_SL["stop"]]); target = _f(r[_SL["target"]])
                        stop_source = target_source = "sheet"
                        if buy:
                            sd = risk_utils.resolve_stop_detailed(buy, symbol=sym, as_of=buy_date,
                                                                 exclude_swing=excl)
                            if sd["stop"] is not None:
                                stop, stop_source = sd["stop"], sd["source"]
                            td = risk_utils.resolve_target_detailed(buy, symbol=sym, as_of=buy_date,
                                                                    exclude_swing=excl)
                            if td["target"] is not None:
                                target, target_source = td["target"], td["source"]
                        try:
                            verify_price_integrity(sym, top, "Excel Short_Long fallback")
                        except PricingDiscrepancyError:
                            corrected = _load_latest_close_from_cache(sym)
                            if corrected and corrected > 0:
                                _log.warning(f"[PRICING] {sym}: overriding sheet ${top:.2f} with cache ${corrected:.2f}")
                                top = corrected
                                if buy and qty:
                                    pnl = round((top - buy) * qty, 2)
                                    pnl_pct = round((top - buy) / buy * 100, 2)
                        holdings.append({
                            "symbol":    sym,
                            "fractional": instruments.is_fractional_eligible(sym),
                            "qty":       qty,
                            "buy":       buy,
                            "buy_date":  buy_date,
                            "current":   top,
                            "target":    target,
                            "target_source": target_source,
                            "stop":      stop,
                            "stop_source": stop_source,
                            "instrument": instruments.classify(sym),
                            "s10":       s10,
                            "l60":       l60,
                            "total":     round((s10 or 0) + (l60 or 0), 1),
                            "win_pct":   r[_SL["winpct"]],
                            "status":    _rsc.get("status") or str(r[_SL["status"]] or ""),
                            "streak":    _streak(sym),
                            "in_profit": str(r[_SL["in_profit"]] or ""),
                            "pnl":       pnl,
                            "pnl_pct":   pnl_pct,
                        })
                    acct_id = real_ids[tbl_idx] if tbl_idx < len(real_ids) else f"T{tbl_idx+1}"
                    accounts.append({
                        "id":       acct_id,
                        "label":    f"Real · {acct_id}",
                        "type":     "real",
                        "holdings": holdings,
                        "count":    len(holdings),
                    })
            except FileNotFoundError:
                pass
            except Exception as e:
                accounts.append({"id": "real", "label": "Real accounts", "type": "real",
                                 "holdings": [], "count": 0, "error": str(e)})

        # ── AI game account ─────────────────────────────────────────────────
        pf = read_portfolio()
        game_holdings = []
        
        # Fetch live prices for the game symbols safely
        game_symbols = [p["symbol"] for p in pf.get("positions", [])]
        game_prices = {}
        try:
            game_prices = _live_prices(game_symbols)
        except Exception:
            pass
            
        for g_pos in pf.get("positions", []):
            sym = g_pos["symbol"]
            qty = g_pos["qty"]
            buy = g_pos["cost"]
            days_held = g_pos.get("days_held", 0)
            
            # Look up stop and target from Excel or use game fallback
            dec = sl_decorations.get(sym.upper(), {})
            stop = dec.get("stop") or g_pos.get("stop_loss", 0.0)
            stop_source = dec.get("stop_source") or "Game"
            target = dec.get("target") or round(buy * 1.15, 2)
            target_source = dec.get("target_source") or "Game"
            
            # Retrieve live price and compute P&L
            current = game_prices.get(sym) or buy
            pnl = round((current - buy) * qty, 2)
            pnl_pct = round((current - buy) / buy * 100, 2) if buy > 0 else 0.0
            
            sc = scores.get(sym, {})
            s10 = _f(sc.get("s10"))
            l60 = _f(sc.get("l60"))
            
            game_holdings.append({
                "symbol": sym,
                "fractional": instruments.is_fractional_eligible(sym),
                "qty": qty,
                "buy": buy,
                "current": current,
                "pnl": pnl,
                "pnl_pct": pnl_pct,
                "stop": stop,
                "stop_source": stop_source,
                "target": target,
                "target_source": target_source,
                "days_held": days_held,
                "s10": s10,
                "l60": l60,
                "total": round((s10 or 0) + (l60 or 0), 1) if s10 is not None or l60 is not None else None,
                "status": sc.get("status", ""),
                "streak": sc.get("streak") if sc.get("streak") is not None else _streak(sym),
                "instrument": instruments.classify(sym)
            })

        accounts.append({
            "id":       "game",
            "label":    "AI Game",
            "type":     "game",
            "balance":  pf["balance"],
            "equity":   pf["equity"],
            "return_pct": pf["return_pct"],
            "profile":  pf["profile"],
            "holdings": game_holdings,
            "count":    pf["open_positions"],
        })

        return {"accounts": accounts, "broker_status": broker_status}
    return _cached("accounts", 30.0, _load)


# ── Transaction history ───────────────────────────────────────────────────────

def read_history(limit: int = 50, offset: int = 0) -> dict:
    """Read transaction log from ai_portfolio_game.json history array."""
    try:
        with open(_GAME, encoding="utf-8") as f:
            state = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"total": 0, "transactions": [], "win_rate": 0, "total_pnl": 0}

    history = list(reversed(state.get("history", [])))  # newest first
    total   = len(history)
    page    = history[offset: offset + limit]

    sells   = [t for t in history if t.get("type") == "SELL" and t.get("pnl") is not None]
    wins    = [t for t in sells if (t.get("pnl") or 0) > 0]
    win_rate = round(len(wins) / len(sells) * 100, 1) if sells else 0
    total_pnl = round(sum(t.get("pnl") or 0 for t in sells), 2)

    return {
        "total":        total,
        "transactions": page,
        "win_rate":     win_rate,
        "total_pnl":    total_pnl,
    }


def read_equity_curve() -> list[dict]:
    """Reconstruct daily equity snapshots from transaction history."""
    try:
        with open(_GAME, encoding="utf-8") as f:
            state = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

    history  = sorted(state.get("history", []), key=lambda t: t.get("date", ""))
    balance  = 10000.0
    by_date: dict[str, float] = {}
    for tx in history:
        d = tx.get("date", "")[:10]
        if not d:
            continue
        # Cash-only curve: BUY spends price*qty, SELL returns price*qty — which
        # already equals buy cost + realized pnl (pnl = (price-cost)*qty), so pnl
        # must NOT be added again or profit double-counts. Mirrors the live game's
        # own `balance += proceeds`; open positions are not marked to market here.
        if tx.get("type") == "BUY":
            balance -= (tx.get("price", 0) * tx.get("qty", 0))
        elif tx.get("type") == "SELL":
            balance += (tx.get("price", 0) * tx.get("qty", 0))
        by_date[d] = round(balance, 2)

    return [{"date": d, "balance": v} for d, v in sorted(by_date.items())]


# ── Log tailing ───────────────────────────────────────────────────────────────

def read_log_tail(n_lines: int = 100) -> list[str]:
    """Return last n_lines from autonomous_run.log."""
    if not _LOG.exists():
        return ["[Log file not found]"]
    try:
        with open(_LOG, encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        return [l.rstrip() for l in lines[-n_lines:]]
    except Exception as e:
        return [f"[Error reading log: {e}]"]


# ── System health ─────────────────────────────────────────────────────────────

def _server_needs_restart() -> bool:
    """True when the running server's git commit differs from HEAD on disk —
    i.e. code was updated after the server started and needs a restart."""
    try:
        pid_path = _DATA_DIR / "webserver.pid"
        if not pid_path.exists():
            return False
        # git HEAD on disk
        head = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=str(_DIR), text=True, timeout=3,
        ).strip()
        # commit baked into a stamp file written at server start
        stamp = _DIR / "Data" / "server_commit.txt"
        if not stamp.exists():
            return False
        running = stamp.read_text().strip()
        return running != head
    except Exception:
        return False


def get_system_health() -> dict:
    """Check file freshness and last pipeline status."""
    now = datetime.now()
    today = date.today()

    # Workbook freshness
    data_fresh   = False
    last_refresh = None
    if _XLSX.exists():
        mtime = datetime.fromtimestamp(_XLSX.stat().st_mtime)
        data_fresh   = mtime.date() >= today
        last_refresh = mtime.isoformat(timespec="minutes")

    # Last pipeline run time (parse from log)
    last_pipeline_run  = None
    pipeline_status    = "UNKNOWN"
    if _LOG.exists():
        try:
            with open(_LOG, encoding="utf-8", errors="replace") as f:
                lines = f.readlines()
            for line in reversed(lines):
                if "Starting Daily Trading Pipeline" in line:
                    # [2026-07-01 11:06:53] Starting ...
                    ts = line.strip()[1:20]
                    last_pipeline_run = ts
                    break
            for line in reversed(lines):
                if "Pipeline completed successfully" in line:
                    pipeline_status = "OK"
                    break
                if "Pipeline failed" in line or "ALERT" in line:
                    pipeline_status = "ERROR"
                    break
        except Exception:
            pass

    # Watchdog — check if watchdog.log or similar file was updated today
    watchdog_ok = True  # default optimistic; future: check watchdog log

    # E*TRADE auth state — the shared read-only classifier (probe=False: pure-local, no network,
    # no mutation, since /api/health is polled frequently). /api/health is UNAUTHENTICATED, so
    # embed only the non-sensitive posture (state + whether a human must act); the rich blob
    # (trust marker, breaker counts, token date, and the bootstrap-command hint in `summary`)
    # stays behind the admin-gated GET /api/etrade/status. Never let a broker read break health.
    etrade_auth = None
    try:
        _st = etrade.auth_status("production", probe=False)
        etrade_auth = {
            "state":             _st["state"],
            "needs_manual_auth": _st["needs_manual_auth"],
            "can_auto_reauth":   _st["can_auto_reauth"],
        }
    except Exception:
        etrade_auth = None

    return {
        "data_fresh":           data_fresh,
        "last_refresh":         last_refresh,
        "last_pipeline_run":    last_pipeline_run,
        "pipeline_status":      pipeline_status,
        "watchdog_ok":          watchdog_ok,
        "server_time":          now.isoformat(timespec="seconds"),
        "server_needs_restart": _server_needs_restart(),
        "etrade":               etrade_auth,
    }


# ── Scheduled tasks ───────────────────────────────────────────────────────────

_KNOWN_TASKS = [
    "AETHER_Watchdog",
    "AETHER_StopMonitor",
    "AETHER_DailyDriver",
    "AETHER_PostMarketReporter",
    "AETHER_PostMarketSync",
    "AETHER_RD_Scientist",
]

# Registry of manually-runnable scripts exposed in the dashboard.
# Each entry: id (unique slug), label, description, script, args list,
# admin_only (requires admin token to run), confirm (prompt user first).
MANUAL_TASKS = [
    {
        "id": "pipeline",
        "label": "Morning Pipeline",
        "description": "Full daily screener + workbook refresh + email report (autonomous_pipeline.py).",
        "script": "autonomous_pipeline.py",
        "args": [],
        "admin_only": True,
        "confirm": "Run the full morning pipeline? This fetches fresh data and emails a report.",
        "category": "pipeline",
    },
    {
        "id": "ai_game_run",
        "label": "AI Game — Daily Moves",
        "description": "Run the AI portfolio game's daily buy/sell logic (ai_portfolio_game.py --run).",
        "script": "ai_portfolio_game.py",
        "args": ["--run"],
        "admin_only": True,
        "confirm": "Execute today's AI game moves?",
        "category": "ai_game",
    },
    {
        "id": "ai_game_summary",
        "label": "AI Game — Email Summary",
        "description": "Send the daily performance summary email (ai_portfolio_game.py --summary).",
        "script": "ai_portfolio_game.py",
        "args": ["--summary"],
        "admin_only": True,
        "confirm": "Send the AI game summary email?",
        "category": "ai_game",
    },
    {
        "id": "ai_game_force",
        "label": "AI Game — Force Run (After-Hours)",
        "description": "Force the AI game to run outside market hours for testing (ai_portfolio_game.py --run --force).",
        "script": "ai_portfolio_game.py",
        "args": ["--run", "--force"],
        "admin_only": True,
        "confirm": "Force-run the AI game outside market hours?",
        "category": "ai_game",
    },
    {
        "id": "ohlcv_recovery",
        "label": "OHLCV Recovery",
        "description": "Repair missing/stale OHLCV history for all symbols via RapidAPI (rapidapi.py).",
        "script": "rapidapi.py",
        "args": [],
        "admin_only": True,
        "confirm": "Run OHLCV recovery? This calls the RapidAPI for stale symbols (rate-limited).",
        "category": "data",
    },
    {
        "id": "run_history",
        "label": "History Backfill (5d)",
        "description": "Backfill the last 5 days of price history into the OHLCV cache (run_history.py 5).",
        "script": "run_history.py",
        "args": ["5"],
        "admin_only": True,
        "confirm": "Backfill 5 days of history?",
        "category": "data",
    },
    {
        "id": "backtest_levels",
        "label": "Backtest Levels",
        "description": "Walk-forward accuracy test of support/resistance levels for any symbol.",
        "script": "scripts/backtesting/backtest_levels.py",
        "args": [],
        "input": {"placeholder": "Symbol (e.g. INTC)", "default": "INTC", "arg_position": 0},
        "admin_only": False,
        "confirm": None,
        "category": "research",
    },
    {
        "id": "backtest_levels_all",
        "label": "Backtest Levels — Universe",
        "description": "Walk-forward accuracy across all 500+ cached symbols (backtest_levels.py --all --step 20). Takes ~2 min.",
        "script": "scripts/backtesting/backtest_levels.py",
        "args": ["--all", "--step", "20"],
        "admin_only": False,
        "confirm": "Run full universe backtest? This takes ~2 minutes.",
        "category": "research",
    },
    {
        "id": "decision_eval",
        "label": "Decision Scorecard",
        "description": "Backtrack exit decisions against forward prices and print the selector scorecard (decision_eval.py).",
        "script": "decision_eval.py",
        "args": [],
        "admin_only": False,
        "confirm": None,
        "category": "research",
    },
    {
        "id": "intraday_monitor",
        "label": "Intraday Stop Monitor",
        "description": "Run one pass of the intraday stop-loss monitor and email alerts for breaches (intraday_monitor.py).",
        "script": "intraday_monitor.py",
        "args": [],
        "admin_only": True,
        "confirm": "Run the intraday monitor? This will email alerts for any stop breaches.",
        "category": "monitoring",
    },
    {
        "id": "real_copilot",
        "label": "Real Account Shadow Copilot",
        "description": "Scan real E*TRADE holdings overnight and email trade tickets (real_copilot.py).",
        "script": "real_copilot.py",
        "args": [],
        "admin_only": True,
        "confirm": "Run the real-account copilot? This emails trade recommendations for your real positions.",
        "category": "monitoring",
    },
    {
        "id": "watchdog",
        "label": "Watchdog / Self-Healer",
        "description": "Run one watchdog cycle: check logs, heal tasks, refresh E*TRADE session (watchdog.py).",
        "script": "watchdog.py",
        "args": [],
        "admin_only": True,
        "confirm": "Run the watchdog healer?",
        "category": "system",
    },
    {
        "id": "pattern_discovery",
        "label": "Pattern Discovery",
        "description": "Historical replay: reconstruct scores for a past date, find top winners, identify what the system missed and why. Extracts candidate new scoring factors.",
        "script": "scripts/backtesting/pattern_discovery.py",
        "args": ["--top", "10"],
        "input": {"placeholder": "Replay date (e.g. 2026-03-01)", "default": "2026-03-01", "arg": "--date"},
        "admin_only": False,
        "confirm": None,
        "category": "research",
    },
    {
        "id": "pattern_discovery_validate",
        "label": "Pattern Discovery + Validate",
        "description": "Run pattern discovery AND validate found patterns across Jun 2025 → Feb 2026 to confirm statistical significance. Takes ~30 min.",
        "script": "scripts/backtesting/pattern_discovery.py",
        "args": ["--top", "10", "--validate", "--date-range", "2025-06-01:2026-02-28"],
        "input": {"placeholder": "Replay date (e.g. 2026-03-01)", "default": "2026-03-01", "arg": "--date"},
        "admin_only": False,
        "confirm": "Run full pattern discovery with validation? This takes ~30 minutes.",
        "category": "research",
    },
]


def read_manual_tasks() -> list[dict]:
    """Return the registry of manually-runnable scripts for the dashboard."""
    return MANUAL_TASKS


_ADDRESSED_BUY_SIDE_MISSES = {
    "ETHT": "Addressed",
    "WDAY": "Addressed",
}


def read_scorecard(horizon_days: int = 10) -> dict:
    """Backtracked selector scorecard (rules vs each AI provider) + winner-selling
    misses, from Data/decision_log.jsonl. Cached; empty when no log yet."""
    def _load():
        entries = _decision_eval.read_log()
        sc = {}
        if entries:
            sc = _decision_eval.score_log(entries, horizon_days=horizon_days)
            sc["logged"] = len(entries)
        else:
            sc = {"selectors": {}, "winner_selling_misses": [], "logged": 0}

        # Load latest Buy-Side Missed Winners from Saturday R&D reports
        buy_side_missed = []
        buy_side_date = None
        try:
            pattern = str(_DATA_DIR / "pattern_discovery_*.json")
            files = sorted(glob.glob(pattern), reverse=True)
            if files:
                latest_report_path = files[0]
                with open(latest_report_path, "r", encoding="utf-8") as rf:
                    report = json.load(rf)
                buy_side_date = report.get("replay_date")
                s10_missed = report.get("s10_analysis", {}).get("missed", [])
                for m in s10_missed[:10]: # Top 10 missed buy-side winners
                    sym_upper = str(m.get("symbol") or "").strip().upper()
                    status_val = _ADDRESSED_BUY_SIDE_MISSES.get(sym_upper, "Reviewing")
                    buy_side_missed.append({
                        "symbol": sym_upper,
                        "pgr": m.get("pgr") or "N/A",
                        "score": round(m.get("score", 0.0), 1),
                        "fwd_return_pct": round(m.get("fwd_r10", 0.0), 1),
                        "reasons": m.get("reasons", []),
                        "status": status_val
                    })
        except Exception:
            pass

        sc["buy_side_missed_winners"] = buy_side_missed
        sc["buy_side_replay_date"] = buy_side_date
        return sc
    return _cached(f"scorecard:{horizon_days}", 300.0, _load)


_SYMBOL_RE = re.compile(r'^[A-Z0-9.\-]{1,10}$')

def read_symbol(symbol: str) -> dict:
    """Aggregate all available data for one symbol: research row, 90-day price
    series for charting, backtest accuracy, and account holding if held."""
    sym = (symbol or "").upper().strip()
    if not _SYMBOL_RE.match(sym):
        return {"symbol": sym, "error": f"Invalid symbol: {sym!r}"}
    def _load():
        out: dict = {"symbol": sym}

        # ── Research row ──────────────────────────────────────────────────────
        research = read_research()
        row = next((r for r in research.get("rows", []) if r["symbol"] == sym), None)
        out["research"] = row

        # ── 365-day OHLCV series for the candlestick chart ──────────────────
        path = _DATA_DIR / "Symbol_full" / f"{sym}_daily.json"
        chart = []
        if path.exists():
            try:
                with open(path, "r", encoding="utf-8") as f:
                    ts = json.load(f).get("Time Series (Daily)", {})
                dates = sorted(ts.keys())[-365:]
                chart = [{"date": d,
                          "open":   round(float(ts[d]["1. open"]),  2),
                          "high":   round(float(ts[d]["2. high"]),  2),
                          "low":    round(float(ts[d]["3. low"]),   2),
                          "close":  round(float(ts[d]["4. close"]), 2),
                          "volume": int(float(ts[d].get("5. volume", 0)))} for d in dates]
            except Exception:
                pass
        out["chart"] = chart

        # ── Backtest accuracy ─────────────────────────────────────────────────
        out["backtest"] = backtest_levels.backtest_symbol(sym)

        # ── Account holding (any real account or game) ───────────────────────
        holding = None
        for acct in read_accounts().get("accounts", []):
            for h in acct.get("holdings", []):
                if h.get("symbol") == sym:
                    holding = {"account_id": acct["id"], "account_label": acct["label"], **h}
                    break
            if holding:
                break
        out["holding"] = holding

        out["digit_study"] = _get_digit_study(sym)

        return out
    return _cached(f"symbol:{sym}", 60.0, _load)


def read_backtest(symbol: str, horizon: int = 20) -> dict:
    """Walk-forward accuracy of the support/resistance levels for one symbol
    (backtest_levels.backtest_symbol). Cached 10 min — it scans full history."""
    sym = (symbol or "").upper()
    if not _SYMBOL_RE.match(sym):
        return {"symbol": sym, "error": f"Invalid symbol: {sym!r}"}
    def _load():
        return backtest_levels.backtest_symbol(sym, horizon=horizon)
    return _cached(f"bt:{sym}:{horizon}", 600.0, _load)


def requalify_symbol(symbol: str, cost: float | None = None) -> dict:
    """
    Fetch live Chaikin data for one symbol and return a structured factor dict
    ready for the AI prompt. Also runs the deterministic exit engine.
    Called from the /api/requalify endpoint (Phase 1).
    """
    sym = (symbol or "").upper().strip()
    if not _SYMBOL_RE.match(sym):
        return {"symbol": sym, "error": f"Invalid symbol: {sym!r}", "factors": {}}
    today = date.today()
    result: dict = {"symbol": sym, "error": None}

    # ── 1. Live price ─────────────────────────────────────────────────────────
    price = None
    try:
        tokens = etrade.get_tokens("production")
        if tokens:
            quotes = etrade.fetch_quotes(tokens, [sym], env="production")
            price = quotes.get(sym)
    except Exception:
        pass
    if not price:
        try:
            goog = _google_prices([sym])
            price = goog.get(sym)
        except Exception:
            pass

    # ── 2. Live Chaikin factors ───────────────────────────────────────────────
    factors: dict = {}
    try:
        session = _pg._load_session_from_file()
        if session and session.get("jsessionid"):
            pg = _pg.get_symbol_data(sym, today, prefer_cache=False, session_id=session)
            if pg and pg.price > 0:
                # Load OHLCV for stop/target computation
                ohlcv_path = _DATA_DIR / "Symbol_full" / f"{sym}_daily.json"
                ohlcv_ts: dict = {}
                if ohlcv_path.exists():
                    try:
                        with open(ohlcv_path, "r", encoding="utf-8") as f_ohlcv:
                            ohlcv_ts = json.load(f_ohlcv).get("Time Series (Daily)", {})
                    except Exception:
                        pass
                f = _pg._compute_pgr_fields(pg, ohlcv_ts=ohlcv_ts)
                factors = {
                    "pgr":          f.get("pgr", "N"),
                    "s10":          round(f.get("short_score", 0.0), 1),
                    "l60":          round(f.get("long_score", 0.0), 1),
                    "combined":     round(f.get("short_score", 0.0) + f.get("long_score", 0.0), 1),
                    "stop":         f.get("stop_price", 0.0),
                    "target":       f.get("prev_move_price", 0.0),
                    "buying_ratio": round(f.get("buying_ratio", 0.0), 1),
                    "money_flow":   pg.money_flow or "Neutral",
                    "lt_trend":     pg.lt_trend or "Neutral",
                    "over_bt_sl":   pg.over_bt_sl or "Neutral",
                    "patterns":     f.get("pattern_text", ""),
                    "industry":     pg.industry_name or "",
                    "price":        price or pg.price,
                }
                if price and price > 0:
                    factors["price"] = price
    except Exception as ex:
        result["error"] = f"Chaikin fetch failed: {ex}"

    # Fall back to cached Excel data if live fetch failed
    if not factors:
        try:
            research = read_research()
            row = next((r for r in research.get("rows", []) if r["symbol"] == sym), None)
            if row:
                factors = {
                    "pgr":          row.get("pgr", "N"),
                    "s10":          row.get("s10", 0.0),
                    "l60":          row.get("l60", 0.0),
                    "combined":     (row.get("s10") or 0) + (row.get("l60") or 0),
                    "stop":         row.get("stop"),
                    "target":       row.get("target"),
                    "buying_ratio": row.get("buying_ratio", 0.0),
                    "money_flow":   row.get("money_flow", "Neutral"),
                    "lt_trend":     "Neutral",
                    "over_bt_sl":   "Neutral",
                    "patterns":     row.get("patterns", ""),
                    "industry":     row.get("industry", ""),
                    "price":        price or row.get("price"),
                    "_cached":      True,
                }
                if not result["error"]:
                    result["error"] = "Using cached Excel data (no live Chaikin session)"
        except Exception:
            pass

    result["factors"] = factors

    # ── 3. Deterministic exit engine ──────────────────────────────────────────
    det_action = det_reason = ""
    if cost and factors.get("price") and factors.get("stop"):
        try:
            s10 = factors.get("s10", 0.0) or 0.0
            l60 = factors.get("l60", 0.0) or 0.0
            det_action, det_reason = sell_rules.exit_decision(
                price=factors["price"],
                cost=cost,
                stop_loss=factors["stop"],
                s10=s10,
                l60=l60,
            )
        except Exception:
            pass
    factors["det_action"] = det_action
    factors["det_reason"] = det_reason

    # ── 4. Market regime ─────────────────────────────────────────────────────
    try:
        health = get_system_health()
        result["regime"] = health.get("market_regime", "Unknown")
    except Exception:
        result["regime"] = "Unknown"

    return result


def build_requalify_prompt(sym: str, factors: dict, cost: float | None,
                           regime: str, news: list[str] | None = None) -> str:
    """Build the user-turn prompt for the AI requalify call."""
    price  = factors.get("price", 0.0) or 0.0
    pnl_pct = ""
    if cost and cost > 0 and price > 0:
        pnl_pct = f"{round((price - cost) / cost * 100, 2):+.2f}%"

    lines = [
        f"SYMBOL: {sym}",
        f"ENTRY PRICE: {f'${cost:.2f}' if cost else 'n/a'}",
        f"CURRENT PRICE: ${price:.2f}" + (f"  (PnL: {pnl_pct})" if pnl_pct else ""),
        f"PGR RATING: {factors.get('pgr', 'N')}",
        f"S10 (short-term): {factors.get('s10', 0.0)}",
        f"L60 (long-term):  {factors.get('l60', 0.0)}",
        f"COMBINED SCORE:   {factors.get('combined', 0.0)}",
        f"STOP-LOSS: ${factors.get('stop', 0.0):.2f}" if factors.get("stop") else "STOP-LOSS: n/a",
        f"TARGET:    ${factors.get('target', 0.0):.2f}" if factors.get("target") else "TARGET:    n/a",
        f"BUYING RATIO: {factors.get('buying_ratio', 0.0)}",
        f"MONEY FLOW:   {factors.get('money_flow', 'Neutral')}",
        f"LT TREND:     {factors.get('lt_trend', 'Neutral')}",
        f"OB/OS ZONE:   {factors.get('over_bt_sl', 'Neutral')}",
        f"PATTERNS:     {factors.get('patterns') or 'none detected'}",
        f"INDUSTRY:     {factors.get('industry', '')}",
        f"MARKET REGIME: {regime}",
    ]
    if factors.get("det_action"):
        lines.append(f"DETERMINISTIC ENGINE: {factors['det_action']} — {factors.get('det_reason', '')}")
    if factors.get("_cached"):
        lines.append("NOTE: Chaikin data is from cached Excel sheet (no live session available).")
    if news:
        lines.append("\nRECENT NEWS:")
        for item in news[:5]:
            lines.append(f"  • {item}")
    return "\n".join(lines)


def read_scheduled_tasks() -> list[dict]:
    """Query Windows Task Scheduler for known AETHER tasks."""
    results = []
    try:
        out = subprocess.check_output(
            ["schtasks", "/query", "/fo", "CSV", "/v"],
            encoding="utf-8", errors="replace",
            timeout=10, stderr=subprocess.DEVNULL,
        )
        lines = out.splitlines()
        if not lines:
            return []
        header = [h.strip('"') for h in lines[0].split('","')]

        def col(row_parts, name):
            try:
                return row_parts[header.index(name)].strip('"')
            except (ValueError, IndexError):
                return ""

        seen = set()
        for line in lines[1:]:
            parts = line.split('","')
            task_name = col(parts, "TaskName").lstrip("\\")
            # Handle tasks scheduled in folders (e.g., AETHER_Agents\AETHER_Watchdog)
            base_name = task_name.split("\\")[-1] if "\\" in task_name else task_name
            if base_name not in _KNOWN_TASKS:
                continue
            if base_name in seen:
                continue
            seen.add(base_name)
            results.append({
                "name":     base_name,
                "status":   col(parts, "Status"),
                "last_run": col(parts, "Last Run Time"),
                "next_run": col(parts, "Next Run Time"),
                "last_result": col(parts, "Last Result"),
            })
    except (subprocess.TimeoutExpired, FileNotFoundError, Exception):
        # Not on Windows or schtasks unavailable — return stubs
        results = [{"name": t, "status": "N/A", "last_run": "", "next_run": "", "last_result": ""}
                   for t in _KNOWN_TASKS]
    return results
