"""
Populate Data/Symbol cache for the last N trading days.

For each trading day that is missing cache files, fetches live data from the
Chaikin Analytics API and saves JSON under that date.  Days already fully
cached are skipped entirely.  Run this once to backfill, then run daily.

Usage:
    python run_history.py            # last 14 trading days (default)
    python run_history.py 10         # last 10 trading days

Proxy:
    Behind a corporate proxy — proxy is auto-detected from the HTTP_PROXY /
                               HTTPS_PROXY environment variables.
    Direct connection       — clear the proxy env vars before running:
        Windows:  set HTTP_PROXY=  &&  set HTTPS_PROXY=
        macOS/Linux: unset HTTP_PROXY; unset HTTPS_PROXY

Python deps: requests, openpyxl, playwright (for browser login fallback)
"""

import datetime
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed

import pytz


sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import powergauge


US_HOLIDAYS_2026 = {
    datetime.date(2026, 1, 1),   # New Year's Day
    datetime.date(2026, 1, 19),  # MLK Day
    datetime.date(2026, 2, 16),  # Presidents' Day
    datetime.date(2026, 4, 3),   # Good Friday
    datetime.date(2026, 5, 25),  # Memorial Day
    datetime.date(2026, 7, 3),   # Independence Day (observed)
    datetime.date(2026, 9, 7),   # Labor Day
    datetime.date(2026, 11, 26), # Thanksgiving
    datetime.date(2026, 11, 27), # Day after Thanksgiving
    datetime.date(2026, 12, 25), # Christmas
}


def trading_days(n: int) -> list[datetime.date]:
    days = []
    d = datetime.date.today()
    while len(days) < n:
        if d.weekday() < 5 and d not in US_HOLIDAYS_2026:
            days.append(d)
        d -= datetime.timedelta(days=1)
    return days  # most recent first


def load_symbols() -> list[str]:
    import re

    import openpyxl
    _sym_re = re.compile(r"^[A-Z0-9._\-]+$")
    try:
        wb = openpyxl.load_workbook(powergauge.XLSX_FILE, data_only=True, read_only=True)
    except Exception:
        # Fallback if read_only=True still fails, though unlikely
        wb = openpyxl.load_workbook(powergauge.XLSX_FILE, data_only=True)
    ws = wb['Research']
    syms = []
    for row in ws.iter_rows(min_row=2):
        val = str(row[3].value or '').strip()
        if val and _sym_re.match(val):
            syms.append(val)
    return syms


def days_missing(symbols: list[str], day: datetime.date) -> list[str]:
    symbol_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "Symbol")
    if day < datetime.date.today():
        return [s for s in symbols
                if not (os.path.exists(os.path.join(symbol_dir, s, f"{s}_{day}.json")) or
                        os.path.exists(os.path.join(symbol_dir, f"{s}_{day}.json")))]

    # For today: re-fetch if the file is missing or was written before NYSE 4:05 PM ET
    # (pre-market / intraday writes use the same filename and must be overwritten at EOD).
    try:
        eod_ts = datetime.datetime.combine(
            day, datetime.time(16, 5),
            tzinfo=pytz.timezone("America/New_York")
        ).timestamp()
    except Exception:
        return list(symbols)  # safe fallback: fetch everything

    missing = []
    for s in symbols:
        p1 = os.path.join(symbol_dir, s, f"{s}_{day}.json")
        p2 = os.path.join(symbol_dir, f"{s}_{day}.json")
        path = p1 if os.path.exists(p1) else (p2 if os.path.exists(p2) else None)
        if path is None or os.path.getmtime(path) < eod_ts:
            missing.append(s)
    return missing


def main():
    n_days = int(sys.argv[1]) if len(sys.argv) > 1 else 14

    print("Loading symbols from Research sheet...")
    symbols = load_symbols()
    print(f"  {len(symbols)} symbols")

    days = trading_days(n_days)
    print(f"Trading days to check: {days[-1]} -> {days[0]} ({len(days)} days)\n")

    session_id = powergauge.login()
    sid_str = session_id.get('jsessionid', '') if isinstance(session_id, dict) else str(session_id)
    print(f"Session: {sid_str[:8]}...\n")

    for day in reversed(days):   # oldest first so prevPG chain builds forward
        missing = days_missing(symbols, day)
        if not missing:
            print(f"{day}: all {len(symbols)} symbols cached — skip")
            continue

        total = len(missing)
        print(f"{day}: fetching {total}/{len(symbols)} symbols in parallel...")
        ok = 0
        skip = 0
        errors = 0
        
        workers = int(os.environ.get("CHAIKIN_WORKERS", "10"))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            future_to_sym = {
                pool.submit(
                    powergauge.get_symbol_data,
                    symbol,
                    day,
                    day != datetime.date.today()
                ): symbol
                for symbol in missing
            }
            done = 0
            for future in as_completed(future_to_sym):
                symbol = future_to_sym[future]
                done += 1
                print(f"  [{done}/{total}] {symbol:<8}", end='\r', flush=True)
                try:
                    pg = future.result()
                    if pg.price == -1:
                        skip += 1
                    else:
                        ok += 1
                except Exception as e:
                    print(f"\n  {symbol}: ERROR {e}")
                    errors += 1

        print(f"  done: {ok} fetched, {skip} no-data, {errors} errors          \n")

    print("History backfill complete.")
    print("Run check_from_xls to update the Research sheet with today's data.")


if __name__ == "__main__":
    main()
