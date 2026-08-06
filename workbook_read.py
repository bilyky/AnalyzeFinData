"""
Shared read-only accessors for the daily state workbook (state_of_the_day.xlsx)
and the AI-game JSON.

These are pure readers: they open the workbook (read_only, data_only), pull rows,
and return plain dicts/lists. They have no side effects, send no email, and run no
subprocesses. Extracted here so both the top-level orchestrator (autonomous_pipeline)
and the web data layer (data_api) can share ONE definition instead of the web layer
reaching up into the pipeline module (which drags in email/subprocess machinery on
import). Neither of those modules is a dependency of this one.
"""

import json
import traceback
from pathlib import Path

import openpyxl

import risk_utils
from aether_logger import get_logger as _get_logger

_log = _get_logger("workbook_read")

BASE_DIR = Path(__file__).resolve().parent
XLSX_FILE = BASE_DIR / "Data" / "state_of_the_day.xlsx"
GAME_FILE = BASE_DIR / "Data" / "ai_portfolio_game.json"
ACCOUNT_RISK_USD = 500  # Dollars to lose if a stop is hit (position-sizing basis).


def get_top_5_picks():
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        ws = wb["Research"]

        candidates = []
        seen_symbols = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            sym = row[3]
            if not sym: continue

            # De-duplicate: Ensure each symbol is only processed once
            sym_clean = str(sym).strip().upper()
            if sym_clean in seen_symbols:
                continue
            seen_symbols.add(sym_clean)

            pgr = str(row[6] or "")
            price = row[10] or 0.0
            stop = row[9] or 0.0
            target = row[11] or 0.0
            setup = str(row[20] or "")

            is_setup_ok = (setup == "1" or setup == "OK" or setup == 1)

            win_pct = row[23] or 0.0
            s10 = row[24] or 0.0
            l60 = row[25] or 0.0
            pattern_text = str(row[26] or "").strip() if len(row) > 26 else ""
            industry = str(row[4] or "")

            if is_setup_ok:
                # Enforce strict Risk/Reward filter (Reward must be >= 1.5x of the Risk)
                risk = price - stop
                reward = target - price
                rr_ratio = (reward / risk) if risk > 0 else 0.0

                if rr_ratio < 1.5:
                    continue  # Fail-safe R:R check: reject unfavorable narrow-upside setups

                atr = risk_utils.calculate_atr(sym)
                shares_atr = risk_utils.get_atr_position_size(price, atr, ACCOUNT_RISK_USD)
                shares_stop = risk_utils.get_position_size(price, stop, ACCOUNT_RISK_USD)

                candidates.append({
                    "Symbol": sym,
                    "PGR": pgr,
                    "Price": price,
                    "Stop": stop,
                    "Target": target,
                    "S10": s10,
                    "L60": l60,
                    "WinPct": win_pct,
                    "Total": s10 + l60,
                    "ATR": atr,
                    "Shares_ATR": shares_atr,
                    "Shares_Stop": shares_stop,
                    "Patterns": pattern_text,
                    "Industry": industry
                })

        candidates.sort(key=lambda x: x["Total"], reverse=True)
        return candidates[:5]
    except Exception as e:
        _log.error("Error computing picks", extra={"error": str(e)})
        traceback.print_exc()
        return []


def get_replacement_pairs():
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        if "Replacements" not in wb.sheetnames:
            return []

        ws = wb["Replacements"]
        pairs = []
        for row in ws.iter_rows(min_row=3, max_row=13, values_only=True):
            if row[1] and row[7]:
                pairs.append({
                    "Sell": row[1],
                    "Sell_Score": row[4],
                    "Sell_Status": row[5],
                    "Buy": row[7],
                    "Buy_Score": row[10],
                    "Buy_PGR": row[11]
                })
        return pairs
    except Exception as e:
        _log.error("Error reading replacements", extra={"error": str(e)})
        return []


def get_reserves_data():
    """Extract today's scores for our dynamic A-Reserves (Backup Players) list loaded from the central JSON."""
    reserves_syms = ['EIX', 'AMAT', 'URI', 'VLO', 'RS'] # Standard Fallback
    try:
        if GAME_FILE.exists():
            with open(GAME_FILE, "r", encoding="utf-8") as f:
                state = json.load(f)
                reserves_syms = state.get("reserves", reserves_syms)
    except Exception as e:
        _log.warning("Could not load dynamic reserves from JSON (using fallback)",
                     extra={"error": str(e)})

    reserves_data = []
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        ws = wb["Research"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            sym = row[3]
            if sym in reserves_syms:
                reserves_data.append({
                    "Symbol": sym,
                    "Industry": row[4],
                    "PGR": row[6],
                    "S10": row[24] or 0,
                    "L60": row[25] or 0,
                    "Total": (row[24] or 0) + (row[25] or 0),
                    "Price": row[10] or 0
                })
        # Sort to preserve priority order
        reserves_data.sort(key=lambda x: reserves_syms.index(x["Symbol"]))
    except Exception as e:
        _log.error("Error loading reserves data", extra={"error": str(e)})
    return reserves_data


def get_market_regime():
    """Detect current market regime based on SPY momentum. Returns (label, color)."""
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
        ws = wb["Research"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3] == "SPY":
                l60 = row[25] or 0  # 60-day momentum drives the regime label
                if l60 > 2: return "🚀 BULLISH (Risk-On)", "#2ecc71"
                if l60 < -2: return "⚠️ BEARISH (Risk-Off / Defensive)", "#e74c3c"
                return "⚖️ NEUTRAL (Consolidation)", "#f39c12"
    except Exception as e:
        _log.warning("Could not determine market regime", extra={"error": str(e)})
    return "Unknown", "#7f8c8d"
