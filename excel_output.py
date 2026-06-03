"""
Excel output helpers for the Research and Picks sheets.

Functions here depend only on openpyxl — no PowerGauge or API imports.
Called by check_from_xls() in powergauge.py.
"""

import os
import re
import shutil
import zipfile
import datetime
from io import BytesIO


# ── Column headers + cell-comment text for Research sheet row 1 ─────────────
# Keys are 0-based column indices. PRESERVE cols A-D (0-3) and I (8) and Q (16).
RESEARCH_HEADERS = {
    4:  ("Industry",    "Sector/industry name from Chaikin metaInfo."),
    5:  ("Prev PGR",    "Corrected PGR from the most recent prior cache file.\n1=Be-  2=Be  3=N  4=Bu  5=Bu+"),
    6:  ("PGR",         "Current Corrected Power Gauge Rating.\n1=Be-  2=Be  3=Neutral  4=Bu  5=Bu+"),
    7:  ("Ind Strength","Industry group signal from Chaikin.\nStrong / Weak / NA"),
    9:  ("Stop",        "Stop price = min(3-day lows) x 0.99.\nZeroed when entry filter (col U) fails."),
    10: ("Price",       "Last price from Chaikin API."),
    11: ("Target",      "Resistance target = highest 10-day high above current price.\nZeroed when entry filter (col U) fails."),
    12: ("R/R",         "Risk/Reward ratio = (Target - Price) / (Price - Stop).\nZeroed when entry filter (col U) fails."),
    13: ("Prev Move%",  "% price move since the previous Chaikin cache snapshot."),
    14: ("Prev %",      "Day-change% recorded in the previous Chaikin snapshot."),
    15: ("Change%",     "Today's price change% from Chaikin."),
    17: ("LT Trend",    "Long-term price trend from Chaikin.\nStrong / Neutral / Weak\n\nNote: Weak = recovery play; Strong = already extended."),
    18: ("Money Flow",  "Institutional money flow signal.\nStrong / Neutral / Weak"),
    19: ("OB/OS",       "Overbought / Oversold zone.\nOptimal / Early / Neutral / Wait"),
    20: ("Setup",       "Entry filter: 1 = passed, 0 = failed.\nPass condition: Price > SMA(20) AND Price > Close[3d ago].\nAffects Stop / Target / R/R display only."),
    21: ("BR Score",    "Buying Ratio: composite entry-quality score -10 to +10.\n\nComponents:\n  PGR (1->-2 ... 5->+2)\n  R/R (0->-1, >=0.5->+0.5, >=1->+1, >=2->+1.5, >=3->+2)\n  LT Trend (Weak->+1, Strong->-1)\n  Money Flow (Strong->+0.75, Weak->-0.75)\n  OB/OS (Optimal->+1, Early->+0.25, Wait->-0.25)\n  Industry (Weak->+0.5, Strong->-0.5)\n  PGR Delta (any change->+0.25)\n  Seasonality (-1 to +1)\n\nThresholds: >=4 strong buy | 2-4 moderate | 0-2 weak | -2-0 avoid | <=-2 strong avoid"),
    22: ("Seasonal",    "Week-of-month seasonality score.\n+1.0 strong tailwind  +0.5 mild tailwind\n 0.0 neutral           -0.5 headwind  -1.0 strong headwind\nBlank = less than 3 years of OHLCV data."),
    23: ("Win% 10d",    "Predicted 10-day win% from backtest (238k obs, 466 symbols, 2023-2025).\nBased on Buying Ratio bucket:\n  BR >=  4  -> 64.3%\n  BR 2-4    -> 57.6%\n  BR 0-2    -> 53.1%\n  BR -2-0   -> 50.3%\n  BR <= -2  -> 46.3%"),
    24: ("Short10",     "10-day entry-quality score: -10 to +10.\nWeights (336k obs, 2023-2025, NA-filtered):\n  Rel Volume (4.4%): High->+2.5, Very High->+0.5, Low->-2\n  OB/OS (4.3%): Optimal->+3, Early->+1, Wait->-2\n  Money Flow (3.5%): Strong->+3, Weak->-2\n  Industry Str (3.1%, contrarian): Weak->+2, Strong->-2\n  LT Trend (2.1%, contrarian): Weak->+1.5, Strong->-1.5\n  Seasonality: +-1.0  |  Regime: +-1.0  |  Fibonacci: +-1.0\nRemoved: PGR, PGR Delta, R/R -- all <2% spread."),
    25: ("Long60",      "60-day position-quality score: -10 to +10.\nWeights (336k obs, 2023-2025, NA-filtered):\n  LT Trend (4.5%, contrarian): Weak->+4, Strong->-3\n  Rel Volume (2.8%): High->+2, Low->-1\n  Money Flow (2.5%): Strong->+2.5, Weak->-2\n  Industry Str (2.4%, contrarian): Weak->+2, Strong->-1.5\n  OB/OS (2.3%): Optimal->+1.5, Early->+0.5, Wait->-0.5\n  Seasonality: +-0.5  |  Regime: +-1.5  |  Fibonacci: +-0.5\nRemoved: PGR, PGR Delta, R/R -- all <2% spread at 60d."),
}


def write_research_headers(ws):
    """Write column labels and cell comments to Research sheet row 1."""
    from openpyxl.comments import Comment
    for col_idx, (label, memo) in RESEARCH_HEADERS.items():
        cell = ws.cell(row=1, column=col_idx + 1)
        cell.value = label
        cell.comment = Comment(memo, "PowerGauge")


def write_picks_sheet(wb, picks_data: list, run_date):
    """Create/refresh the Picks sheet with four Top-5 tables (Short10 + Long60, buy + sell)."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK_HDR  = PatternFill("solid", fgColor="2E4057")
    BLUE_HDR  = PatternFill("solid", fgColor="4472C4")
    BUY_FILL  = PatternFill("solid", fgColor="E2EFDA")
    SELL_FILL = PatternFill("solid", fgColor="FCE4D6")
    SETUP_OK  = PatternFill("solid", fgColor="C6EFCE")
    SETUP_NO  = PatternFill("solid", fgColor="D9D9D9")
    WHITE_BOLD  = Font(bold=True, color="FFFFFF")
    BOLD        = Font(bold=True)
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center")
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="BBBBBB"),
        right=Side(style="thin",  color="BBBBBB"),
    )

    if "Picks" in wb.sheetnames:
        del wb["Picks"]
    ws = wb.create_sheet("Picks")
    if "Research" in wb.sheetnames:
        target_idx  = wb.sheetnames.index("Research") + 1
        current_idx = len(wb.sheetnames) - 1
        wb.move_sheet("Picks", offset=target_idx - current_idx)

    col_widths = [6, 9, 30, 8, 7, 7, 10, 10, 10, 7, 9]
    col_labels = ["Rank", "Symbol", "Industry", "Score", "BR", "PGR",
                  "OB/OS", "Money Flow", "LT Trend", "Setup", "Price"]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    regime = picks_data[0]['regime'] if picks_data else "N/A"

    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = f"Top Picks  --  {run_date.strftime('%Y-%m-%d')}    |    Market Regime: {regime}"
    title.fill = DARK_HDR
    title.font = Font(bold=True, color="FFFFFF", size=12)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 22

    def write_table(start_row: int, label: str, rows: list, row_fill):
        ws.merge_cells(f"A{start_row}:K{start_row}")
        sh = ws.cell(start_row, 1)
        sh.value = label
        sh.fill = BLUE_HDR
        sh.font = WHITE_BOLD
        sh.alignment = CENTER
        ws.row_dimensions[start_row].height = 18

        hr = start_row + 1
        for col, lbl in enumerate(col_labels, 1):
            c = ws.cell(hr, col)
            c.value = lbl
            c.font  = BOLD
            c.alignment = CENTER
            c.border = THIN_BORDER
        ws.row_dimensions[hr].height = 15

        for rank, rec in enumerate(rows, 1):
            dr = hr + rank
            vals = [
                rank,
                rec['symbol'],
                rec['industry'],
                rec['score'],
                rec['br'],
                rec['pgr'],
                rec['ob_os'],
                rec['money_fl'],
                rec['lt_trend'],
                "OK" if rec['setup'] == 1 else ("--" if rec['setup'] == 0 else "?"),
                rec['price'],
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(dr, col)
                c.value = val
                c.fill  = row_fill
                c.alignment = CENTER if col != 3 else LEFT
                c.border = THIN_BORDER
                if col == 4:
                    c.font = Font(bold=True, color="375623" if (val or 0) >= 0 else "9C0006")
                if col == 10:
                    c.fill = SETUP_OK if rec['setup'] == 1 else SETUP_NO
            ws.row_dimensions[dr].height = 14

    def top5(data, key, reverse):
        ranked = sorted((r for r in data if r.get(key) is not None),
                        key=lambda r: r[key], reverse=reverse)[:5]
        return [dict(r, score=r[key]) for r in ranked]

    write_table(3,  "TOP 5 BUY  --  Short10 (10-day entry score)",     top5(picks_data, 'short10', True),  BUY_FILL)
    write_table(11, "TOP 5 SELL  --  Short10 (10-day entry score)",    top5(picks_data, 'short10', False), SELL_FILL)
    write_table(19, "TOP 5 BUY  --  Long60  (60-day position score)",  top5(picks_data, 'long60',  True),  BUY_FILL)
    write_table(27, "TOP 5 SELL  --  Long60  (60-day position score)", top5(picks_data, 'long60',  False), SELL_FILL)


def _strip_external_formulas(data: bytes) -> bytes:
    """Replace [N] external-reference formula cells with their cached values.

    When external links are stripped from the workbook, cells that reference
    [N]Sheet!Cell become unresolvable and Excel removes them ("Removed Records:
    Formula").  We fix this by converting those cells to plain inline-string or
    empty cells using the cached value already stored in <v>.
    """
    text = data.decode('utf-8')
    if '[' not in text:
        return data  # fast exit — no external refs in this sheet

    def _replace(m):
        full = m.group(0)
        val_m = re.search(r'<v>([^<]+)</v>', full)
        if val_m and val_m.group(1).strip():
            return (f'<c t="inlineStr"><is>'
                    f'<t xml:space="preserve">{val_m.group(1)}</t>'
                    f'</is></c>')
        # No cached value — emit a minimal empty cell preserving r= and s=
        attrs = ''.join(
            f' {k}="{v}"'
            for k, v in re.findall(r'\b(r|s)="([^"]*)"', full)
        )
        return f'<c{attrs}/>'

    return re.sub(
        r'<c[^>]*>\s*<f>\[\d+\][^<]*</f>.*?</c>',
        _replace,
        text,
        flags=re.DOTALL,
    ).encode('utf-8')


def _xml_escape(s: str) -> str:
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _generate_research_comment_xml(base_shape_id: int = 1025) -> bytes:
    """Generate a complete comment XML for the Research sheet row-1 headers."""
    from openpyxl.utils import get_column_letter
    RPR = ('<rPr><sz val="9"/><color indexed="81"/>'
           '<rFont val="Tahoma"/><charset val="1"/></rPr>')
    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        '<comments xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<authors><author>PowerGauge</author></authors>'
        '<commentList>'
    ]
    for i, (col_idx, (_, memo)) in enumerate(sorted(RESEARCH_HEADERS.items())):
        col_letter = get_column_letter(col_idx + 1)
        shape_id = base_shape_id + i
        text = _xml_escape(memo)
        parts.append(
            f'<comment ref="{col_letter}1" authorId="0" shapeId="{shape_id}">'
            f'<text><r>{RPR}<t xml:space="preserve">{text}</t></r></text>'
            f'</comment>'
        )
    parts.append('</commentList></comments>')
    return ''.join(parts).encode('utf-8')


def _generate_research_vml(base_shape_id: int = 1025) -> bytes:
    """Generate a complete VML drawing for the Research sheet row-1 comment boxes."""
    parts = [
        '<xml xmlns:v="urn:schemas-microsoft-com:vml"'
        ' xmlns:o="urn:schemas-microsoft-com:office:office"'
        ' xmlns:x="urn:schemas-microsoft-com:office:excel">\r\n'
        '<o:shapelayout v:ext="edit">'
        '<o:idmap v:ext="edit" data="1"/>'
        '</o:shapelayout>\r\n'
        '<v:shapetype id="_x0000_t202" coordsize="21600,21600" o:spt="202"'
        ' path="m,l,21600r21600,l21600,xe">'
        '<v:stroke joinstyle="miter"/>'
        '<v:path gradientshapeok="t" o:connecttype="rect"/>'
        '</v:shapetype>\r\n'
    ]
    for i, (col_idx, _) in enumerate(sorted(RESEARCH_HEADERS.items())):
        shape_id = base_shape_id + i
        row = 0
        col = col_idx
        anchor = f'{col + 1}, 15, {row + 1}, 2, {col + 3}, 0, {row + 5}, 3'
        parts.append(
            f'<v:shape id="_x0000_s{shape_id}" type="#_x0000_t202"'
            f' style="position:absolute;margin-left:59.25pt;margin-top:1.5pt;'
            f'width:144pt;height:72pt;z-index:{i + 1};visibility:hidden"'
            f' fillcolor="#ffffe1" o:insetmode="auto">'
            f'<v:fill color2="#ffffe1"/>'
            f'<v:shadow on="t" color="black" obscured="t"/>'
            f'<v:path o:connecttype="none"/>'
            f'<v:textbox style="mso-direction-alt:auto">'
            f'<div style="text-align:left"/>'
            f'</v:textbox>'
            f'<x:ClientData ObjectType="Note">'
            f'<x:MoveWithCells/><x:SizeWithCells/>'
            f'<x:Anchor>{anchor}</x:Anchor>'
            f'<x:AutoFill>False</x:AutoFill>'
            f'<x:Row>{row}</x:Row>'
            f'<x:Column>{col}</x:Column>'
            f'</x:ClientData>'
            f'</v:shape>\r\n'
        )
    parts.append('</xml>')
    return ''.join(parts).encode('utf-8')


def _fix_comment_xml(data: bytes) -> bytes:
    """Add XML declaration and convert plain-text comments to rich-text format.

    openpyxl emits <text><t>plain</t></text>; Excel requires (and generates)
    <text><r><rPr>...</rPr><t xml:space="preserve">plain</t></r></text> plus
    an explicit XML declaration.  Without these, Excel silently removes all
    comments and reports "Removed Records: Comments from /xl/comments/...".
    """
    text = data.decode('utf-8')
    RPR = ('<rPr><sz val="9"/><color indexed="81"/>'
           '<rFont val="Tahoma"/><charset val="1"/></rPr>')

    def _to_rich(m):
        content = m.group(1)
        return f'<text><r>{RPR}<t xml:space="preserve">{content}</t></r></text>'

    # Only convert simple <text><t>…</t></text>; leave already-rich <text><r>… alone
    text = re.sub(r'<text><t>(.*?)</t></text>', _to_rich, text, flags=re.DOTALL)
    result = text.encode('utf-8')
    if not result.startswith(b'<?xml'):
        result = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n' + result
    return result


def _fix_vml_ns(data: bytes) -> bytes:
    """Rename openpyxl's ns0/ns1/ns2 VML prefixes to standard o/v/x and add
    missing <x:Anchor> elements.

    Two openpyxl artefacts:
    1. Non-standard namespace prefixes (ns0/ns1/ns2) — Excel's legacy VML parser
       may reject arbitrary prefixes even though they are semantically correct.
    2. Missing <x:Anchor> — openpyxl omits the anchor box from its generated VML.
       Without it modern Excel cannot position the comment box and silently removes
       the whole comment set, emitting "Removed Records: Comments …".
    """
    text = data.decode('utf-8')

    # ── 1. Namespace prefix rename ──────────────────────────────────────────────
    text = (text
        .replace('xmlns:ns0="urn:schemas-microsoft-com:office:office"',
                 'xmlns:o="urn:schemas-microsoft-com:office:office"')
        .replace('xmlns:ns1="urn:schemas-microsoft-com:vml"',
                 'xmlns:v="urn:schemas-microsoft-com:vml"')
        .replace('xmlns:ns2="urn:schemas-microsoft-com:office:excel"',
                 'xmlns:x="urn:schemas-microsoft-com:office:excel"')
        .replace('ns0:', 'o:').replace('ns1:', 'v:').replace('ns2:', 'x:')
    )

    # ── 2. Add missing <x:Anchor> to comment shapes ────────────────────────────
    def _add_anchor(m):
        cd = m.group(0)
        if '<x:Anchor>' in cd:
            return cd
        row_m = re.search(r'<x:Row>(\d+)</x:Row>', cd)
        col_m = re.search(r'<x:Column>(\d+)</x:Column>', cd)
        if not row_m or not col_m:
            return cd
        row = int(row_m.group(1))
        col = int(col_m.group(1))
        anchor = (f'<x:Anchor>{col + 1}, 15, {row + 1}, 2, '
                  f'{col + 3}, 0, {row + 5}, 3</x:Anchor>')
        return cd.replace('<x:AutoFill>', anchor + '<x:AutoFill>', 1)

    text = re.sub(
        r'<x:ClientData ObjectType="Note">.*?</x:ClientData>',
        _add_anchor,
        text,
        flags=re.DOTALL,
    )

    return text.encode('utf-8')


def _fix_ns_prefix(data: bytes, default_ns_url: str, prefix: str) -> bytes:
    """Restore a namespace prefix that openpyxl flattened into a default xmlns.

    openpyxl saves spreadsheetDrawing files as  xmlns="...spreadsheetDrawing"
    and chart files as  xmlns="...chart"  instead of the expected xdr:/c: prefixes.
    Excel's parser rejects the default-namespace form and removes the shape/chart.

    Strategy: convert  xmlns="<url>"  back to  xmlns:<prefix>="<url>"  and add
    <prefix>: to every element tag that has no namespace prefix (i.e., belongs to
    the default namespace).  Already-prefixed tags (a:xxx, c:xxx, r:xxx …) are left
    untouched because the regex stops matching when it sees a colon in the tag name.
    """
    text = data.decode('utf-8')
    default_decl = f'xmlns="{default_ns_url}"'
    if prefix + ':' in text or default_decl not in text:
        return data  # already fixed or not our target

    if not text.startswith('<?xml'):
        text = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n' + text

    text = text.replace(default_decl, f'xmlns:{prefix}="{default_ns_url}"')

    def _add_prefix(m):
        slash, tag = m.group(1), m.group(2)
        return f'<{slash}{prefix}:{tag}'

    # Match unprefixed tags: the identifier stops at ':' so prefixed tags never match
    text = re.sub(r'<(/?)((?:[a-zA-Z][a-zA-Z0-9_]*))(?=[\s/>])', _add_prefix, text)
    return text.encode('utf-8')


def _fix_drawing_xml(data: bytes) -> bytes:
    return _fix_ns_prefix(
        data,
        'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'xdr',
    )


def _fix_chart_xml(data: bytes) -> bytes:
    return _fix_ns_prefix(
        data,
        'http://schemas.openxmlformats.org/drawingml/2006/chart',
        'c',
    )


def fix_comment_shape_ids(xlsx_path: str, original_xlsx: str = None,
                          touched_sheet_names: set = None):
    """Fix openpyxl-save artefacts in one zip pass.

    1. Untouched-sheet preservation — if original_xlsx is provided, worksheet XML
       files for sheets NOT in touched_sheet_names (plus their associated comment
       and drawing VML files) are restored verbatim from the original to prevent
       openpyxl's load/save cycle from corrupting formulas or comments.
    2. Comment shapeId="0" bug — patches shapeIds from the matching VML for any
       comment XML that was NOT restored from the original.
    3. External-link ghost references — strips xl/externalLinks/* and removes their
       entries from workbook.xml + workbook.xml.rels.
    """
    # ── Build restore map from original ─────────────────────────────────────
    restore_from_orig: dict[str, bytes] = {}

    if original_xlsx:
        try:
            with zipfile.ZipFile(original_xlsx, 'r') as zorig:
                orig_names = set(zorig.namelist())
                name_to_rid: dict[str, str] = {}
                rid_to_target: dict[str, str] = {}

                if 'xl/workbook.xml' in orig_names:
                    _wbx = zorig.read('xl/workbook.xml').decode('utf-8')
                    for chunk in re.findall(r'<sheet\s([^>]*/?>)', _wbx):
                        nm = re.search(r'\bname="([^"]*)"', chunk)
                        ri = re.search(r'\br:id="([^"]*)"', chunk)
                        if nm and ri:
                            name_to_rid[nm.group(1)] = ri.group(1)

                if 'xl/_rels/workbook.xml.rels' in orig_names:
                    _wbr = zorig.read('xl/_rels/workbook.xml.rels').decode('utf-8')
                    for chunk in re.findall(r'<Relationship\s([^>]*/?>)', _wbr):
                        id_m = re.search(r'\bId="([^"]*)"', chunk)
                        tgt  = re.search(r'\bTarget="([^"]*)"', chunk)
                        if id_m and tgt:
                            rid_to_target[id_m.group(1)] = tgt.group(1)

                ts = touched_sheet_names or set()
                touched_xml: set[str] = set()
                for sname, rid in name_to_rid.items():
                    tgt = rid_to_target.get(rid, '')
                    if tgt and sname in ts:
                        xf = ('xl/' + tgt) if not tgt.startswith('/') else tgt.lstrip('/')
                        touched_xml.add(xf)

                for sname, rid in name_to_rid.items():
                    tgt = rid_to_target.get(rid, '')
                    if not tgt:
                        continue
                    xf = ('xl/' + tgt) if not tgt.startswith('/') else tgt.lstrip('/')
                    if xf in touched_xml or xf not in orig_names:
                        continue

                    restore_from_orig[xf] = zorig.read(xf)

                    # Restore associated rels + comment/drawing files
                    parts     = xf.split('/')      # ['xl', 'worksheets', 'sheetN.xml']
                    rels_path = '/'.join(parts[:-1]) + '/_rels/' + parts[-1] + '.rels'
                    if rels_path in orig_names:
                        rels_bytes = zorig.read(rels_path)
                        restore_from_orig[rels_path] = rels_bytes
                        for rel_tgt in re.findall(r'\bTarget="([^"]*)"', rels_bytes.decode('utf-8')):
                            if rel_tgt.startswith('../'):
                                resolved = 'xl/' + rel_tgt[3:]
                            elif rel_tgt.startswith('/'):
                                resolved = rel_tgt.lstrip('/')
                            else:
                                resolved = 'xl/worksheets/' + rel_tgt
                            if resolved in orig_names:
                                restore_from_orig[resolved] = zorig.read(resolved)
        except Exception:
            restore_from_orig.clear()   # on any error, fall back to default behaviour

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        names = zin.namelist()

        # ── 1. shapeId patch (only for comment XMLs not restored from original) ──
        comment_files = sorted(n for n in names if re.match(r'xl/comments/comment\d+\.xml', n))
        vml_files     = sorted(n for n in names if re.match(r'xl/drawings/commentsDrawing\d+\.vml', n))

        modified: dict[str, bytes] = {}

        # ── 0. Always regenerate Research sheet's comment+VML from scratch ────
        # This bypasses all openpyxl/backup-restore uncertainty: our generated
        # content is guaranteed to be valid Excel XML with matching shapeIds.
        try:
            _wb_xml  = zin.read('xl/workbook.xml').decode('utf-8') if 'xl/workbook.xml' in names else ''
            _wb_rels = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8') if 'xl/_rels/workbook.xml.rels' in names else ''
            _research_rid = None
            for _chunk in re.findall(r'<sheet\s[^>]+/?>', _wb_xml):
                _nm = re.search(r'\bname="([^"]*)"', _chunk)
                _ri = re.search(r'\br:id="([^"]*)"', _chunk)
                if _nm and _nm.group(1) == 'Research' and _ri:
                    _research_rid = _ri.group(1)
                    break
            if _research_rid:
                _research_tgt = None
                for _chunk in re.findall(r'<Relationship\s[^>]+/?>', _wb_rels):
                    _id_m = re.search(r'\bId="([^"]*)"', _chunk)
                    _tgt  = re.search(r'\bTarget="([^"]*)"', _chunk)
                    if _id_m and _id_m.group(1) == _research_rid and _tgt:
                        _research_tgt = _tgt.group(1)
                        break
                if _research_tgt:
                    _rxf = ('xl/' + _research_tgt) if not _research_tgt.startswith('/') else _research_tgt.lstrip('/')
                    _rparts = _rxf.split('/')
                    _rels_p = '/'.join(_rparts[:-1]) + '/_rels/' + _rparts[-1] + '.rels'
                    if _rels_p in names:
                        _rrels = zin.read(_rels_p).decode('utf-8')
                        for _rel_tgt in re.findall(r'\bTarget="([^"]*)"', _rrels):
                            if _rel_tgt.startswith('/'):
                                _res = _rel_tgt.lstrip('/')
                            elif _rel_tgt.startswith('../'):
                                _res = 'xl/' + _rel_tgt[3:]
                            else:
                                _res = 'xl/worksheets/' + _rel_tgt
                            if re.match(r'xl/comments/comment\d+\.xml$', _res):
                                modified[_res] = _generate_research_comment_xml()
                                restore_from_orig.pop(_res, None)
                            elif re.match(r'xl/drawings/commentsDrawing\d+\.vml$', _res):
                                modified[_res] = _generate_research_vml()
                                restore_from_orig.pop(_res, None)
                        # Convert absolute paths to relative in the _rels file
                        _patched_rels = re.sub(
                            r'Target="(/xl/comments/(comment\d+\.xml))"',
                            r'Target="../comments/\2"',
                            _rrels,
                        )
                        _patched_rels = re.sub(
                            r'Target="(/xl/drawings/(commentsDrawing\d+\.vml))"',
                            r'Target="../drawings/\2"',
                            _patched_rels,
                        )
                        if _patched_rels != _rrels:
                            modified[_rels_p] = _patched_rels.encode('utf-8')
                            restore_from_orig.pop(_rels_p, None)

                    # Ensure the Research sheet XML has:
                    # 1. xmlns:r at the ROOT <worksheet> element (not inline on legacyDrawing)
                    # 2. <legacyDrawing r:id="anysvml"/> at the end
                    # Excel's worksheet XML parser requires the r: namespace at root level;
                    # an inline xmlns:r on <legacyDrawing> is valid XML but Excel rejects it
                    # and fails to resolve the VML reference, causing "Removed Records: Comments".
                    _R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                    _ws_data = (restore_from_orig.get(_rxf)
                                or (zin.read(_rxf) if _rxf in names else None))
                    if _ws_data:
                        _ws_text = _ws_data.decode('utf-8')
                        # Remove any inline xmlns:r from legacyDrawing
                        _ws_text = re.sub(
                            r'<legacyDrawing[^>]*/?>',
                            '<legacyDrawing r:id="anysvml" />',
                            _ws_text,
                        )
                        # Ensure xmlns:r is on the root <worksheet> element
                        if f'xmlns:r="{_R_NS}"' not in _ws_text:
                            _ws_text = _ws_text.replace(
                                '<worksheet ',
                                f'<worksheet xmlns:r="{_R_NS}" ',
                                1,
                            )
                        # Add legacyDrawing if still missing (e.g. openpyxl stripped it)
                        if '<legacyDrawing' not in _ws_text:
                            _ws_text = _ws_text.replace(
                                '</worksheet>',
                                '<legacyDrawing r:id="anysvml" /></worksheet>',
                            )
                        modified[_rxf] = _ws_text.encode('utf-8')
                        restore_from_orig.pop(_rxf, None)
        except Exception:
            pass  # fall back to existing shapeId patch + _fix_comment_xml/_fix_vml_ns

        # ── 1. shapeId patch (only for comment XMLs not restored from original) ──
        for cf, vf in zip(comment_files, vml_files):
            if cf in restore_from_orig or cf in modified:
                continue    # restored from original or already regenerated
            vml   = zin.read(vf).decode('utf-8')
            spids = re.findall(r'id="_x0000_s(\d+)"', vml)
            if not spids:
                continue
            comments = zin.read(cf).decode('utf-8')
            it = iter(spids)
            patched = re.sub(r'shapeId="\d+"', lambda _: f'shapeId="{next(it)}"', comments)
            modified[cf] = patched.encode('utf-8')

        # ── 2. External-link strip ──────────────────────────────────────────
        ext_link_files = {n for n in names if n.startswith('xl/externalLinks/')}

        if ext_link_files:
            wb_rels_name = 'xl/_rels/workbook.xml.rels'
            if wb_rels_name in names:
                rels_xml = zin.read(wb_rels_name).decode('utf-8')
                rels_xml = re.sub(
                    r'<Relationship[^>]+Type="[^"]*externalLinkPath"[^/]*/?>',
                    '', rels_xml,
                )
                modified[wb_rels_name] = rels_xml.encode('utf-8')

            wb_name = 'xl/workbook.xml'
            if wb_name in names:
                wb_xml = zin.read(wb_name).decode('utf-8')
                wb_xml = re.sub(r'<externalReferences\b[^>]*>.*?</externalReferences>', '',
                                wb_xml, flags=re.DOTALL)
                modified[wb_name] = wb_xml.encode('utf-8')

        # ── Rewrite zip ─────────────────────────────────────────────────────
        skip = ext_link_files
        if not modified and not skip and not restore_from_orig:
            return  # nothing to do

        _is_vml     = re.compile(r'xl/drawings/commentsDrawing\d+\.vml').fullmatch
        _is_ws      = re.compile(r'xl/worksheets/sheet\d+\.xml').fullmatch
        _is_comment = re.compile(r'xl/comments/comment\d+\.xml').fullmatch
        _is_drawing = re.compile(r'xl/drawings/drawing\d+\.xml').fullmatch
        _is_chart   = re.compile(r'xl/charts/chart\d+\.xml').fullmatch

        def _apply_fixes(name, data):
            if _is_vml(name):
                return _fix_vml_ns(data)
            if _is_ws(name):
                return _strip_external_formulas(data)
            if _is_comment(name):
                return _fix_comment_xml(data)
            if _is_drawing(name):
                return _fix_drawing_xml(data)
            if _is_chart(name):
                return _fix_chart_xml(data)
            return data

        buf = BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            written: set[str] = set()
            for name in names:
                if name in skip:
                    continue
                if name in restore_from_orig:
                    data = restore_from_orig[name]
                elif name in modified:
                    data = modified[name]
                else:
                    data = zin.read(name)
                zout.writestr(name, _apply_fixes(name, data))
                written.add(name)
            # Re-add any original files that openpyxl dropped (e.g. rels for
            # unmodified sheets that openpyxl decided not to write).
            for fname, fbytes in restore_from_orig.items():
                if fname not in written:
                    zout.writestr(fname, _apply_fixes(fname, fbytes))

    with open(xlsx_path, 'wb') as f:
        f.write(buf.getvalue())


def update_short_long_scores(wb, picks_lookup: dict, quotes: dict, positions: list[dict]):
    """Sync the Short_Long sheet with E*TRADE positions and write score columns Q-W.

    Two account-based tables separated by exactly 3 blank rows:
    - Account 0053 (ACCT_T1) → top table
    - Account 1315 (ACCT_T2) → bottom table
    - User notes below T2 are never touched.
    """
    import datetime
    from openpyxl.styles import PatternFill, Font, Alignment
    from scoring import predicted_win_pct as _pwp

    if "Short_Long" not in wb.sheetnames:
        return
    ws = wb["Short_Long"]

    ACCT_T1 = "0053"
    ACCT_T2 = "1315"

    # ── Styling constants ────────────────────────────────────────────────────
    GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")
    RED_FILL  = PatternFill("solid", fgColor="FCE4D6")
    YEL_FILL  = PatternFill("solid", fgColor="FFEB9C")
    ORG_FILL  = PatternFill("solid", fgColor="FFCC99")
    LGN_FILL  = PatternFill("solid", fgColor="E2EFDA")
    GRY_FILL  = PatternFill("solid", fgColor="D9D9D9")
    GRN_FONT  = Font(bold=True, color="375623")
    RED_FONT  = Font(bold=True, color="9C0006")
    CTR       = Alignment(horizontal="center", vertical="center")

    STATUS_CFG = [
        (4,    "STRONG HOLD", GRN_FILL),
        (2,    "HOLD",        LGN_FILL),
        (0,    "WATCH",       YEL_FILL),
        (-2,   "REDUCE",      ORG_FILL),
        (-999, "EXIT",        RED_FILL),
    ]

    def _status(score):
        if score is None:
            return "N/A", GRY_FILL
        for thresh, label, fill in STATUS_CFG:
            if score >= thresh:
                return label, fill
        return "EXIT", RED_FILL

    today = datetime.date.today()

    import re as _re
    _TICKER_RE = _re.compile(r'^[A-Z][A-Z0-9.]{0,7}$')

    def _is_valid_sym(sym: str) -> bool:
        return bool(_TICKER_RE.match(sym)) if sym else False

    # Drop CVRs, bonds, zero-price items before any sheet logic runs
    positions = [
        p for p in positions
        if _is_valid_sym(p["symbol"]) and (p.get("price") or 0) > 0
    ]

    # ── 1. Locate header row ─────────────────────────────────────────────────
    HDR_ROW = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        if row[1].value == "Symb":
            HDR_ROW = row[0].row
            break

    def _is_data_row(r):
        v = r[1].value
        return isinstance(v, str) and v.strip() and v.strip().upper() != "SYMB"

    def _find_separator() -> int | None:
        """Return first blank row of the T1/T2 separator (3+ consecutive blanks with data on both sides)."""
        data_seen   = False
        blank_start = None
        blank_count = 0
        for row in ws.iter_rows(min_row=3, max_row=200):
            if _is_data_row(row):
                if data_seen and blank_start is not None and blank_count >= 3:
                    return blank_start
                blank_start = None
                blank_count = 0
                data_seen   = True
            elif all(c.value is None for c in row[:6]):
                if blank_start is None:
                    blank_start = row[0].row
                blank_count += 1
            else:
                blank_start = None
                blank_count = 0
        return None

    def _build_row_maps():
        """Return (t1_rows, t2_rows): {sym: row_num}, split at the 3-blank separator."""
        sep = _find_separator()
        t1: dict[str, int] = {}
        t2: dict[str, int] = {}
        for row in ws.iter_rows(min_row=3, max_row=200):
            if _is_data_row(row):
                sym = row[1].value.strip().upper()
                rn  = row[0].row
                if sep is None or rn < sep:
                    if sym not in t1:
                        t1[sym] = rn
                else:
                    if sym not in t2:
                        t2[sym] = rn
        return t1, t2

    # ── Pre-0. Remove invalid symbols and zero-price rows from the sheet ──────
    _bad_rows = []
    for _row in ws.iter_rows(min_row=3, max_row=300):
        if not _is_data_row(_row):
            continue
        _sym   = (_row[1].value or "").strip().upper()
        _cost  = _row[3].value   # col D  (buy price)
        _price = _row[4].value   # col E  (current price)
        if not _is_valid_sym(_sym) or (not _price and not _cost):
            _bad_rows.append(_row[0].row)
    for _rn in sorted(_bad_rows, reverse=True):
        _sym_val = ws.cell(_rn, 2).value
        ws.delete_rows(_rn)
        print(f"[Short_Long] Removed invalid/priceless row {_rn}: {_sym_val!r}")

    # ── 0. Compact any internal blank rows within T1 ─────────────────────────
    # Identifies T1 rows by matching current E*TRADE T1 symbols; deletes blank
    # rows found strictly between the first and last T1 data row.
    _acct_t1_syms = {
        p["symbol"] for p in positions
        if p.get("account_last4") not in {ACCT_T2}   # unknown accounts → T1
    }
    _t1_data_rows = sorted(
        r[0].row for r in ws.iter_rows(min_row=3, max_row=200)
        if _is_data_row(r) and r[1].value.strip().upper() in _acct_t1_syms
    )
    if len(_t1_data_rows) >= 2:
        _t1_first, _t1_last = _t1_data_rows[0], _t1_data_rows[-1]
        _internal_blanks = [
            r[0].row for r in ws.iter_rows(min_row=_t1_first + 1, max_row=_t1_last - 1)
            if not _is_data_row(r)
        ]
        for _rn in sorted(_internal_blanks, reverse=True):
            ws.delete_rows(_rn)
            print(f"[Short_Long] Compacted internal blank row {_rn} in T1")

    sheet_t1_rows, sheet_t2_rows = _build_row_maps()

    # Split E*TRADE positions by account; unknown accounts fall through to T1
    etrade_t1 = [p for p in positions if p.get("account_last4") == ACCT_T1]
    etrade_t2 = [p for p in positions if p.get("account_last4") == ACCT_T2]
    known     = {ACCT_T1, ACCT_T2}
    etrade_t1 += [p for p in positions if p.get("account_last4") not in known]

    etrade_t1_by_sym = {p["symbol"]: p for p in etrade_t1}
    etrade_t2_by_sym = {p["symbol"]: p for p in etrade_t2}
    etrade_t1_syms   = set(etrade_t1_by_sym)
    etrade_t2_syms   = set(etrade_t2_by_sym)

    to_remove_t1 = set(sheet_t1_rows) - etrade_t1_syms
    to_add_t1    = etrade_t1_syms - set(sheet_t1_rows)
    to_remove_t2 = set(sheet_t2_rows) - etrade_t2_syms
    to_add_t2    = etrade_t2_syms - set(sheet_t2_rows)

    # ── 1b. Detect sell/rebuy per table ──────────────────────────────────────
    for sym in set(sheet_t1_rows) & etrade_t1_syms:
        etrade_date = etrade_t1_by_sym[sym].get("date_acquired")
        if not etrade_date:
            continue
        kval = ws.cell(sheet_t1_rows[sym], 11).value
        if isinstance(kval, datetime.datetime):
            kval = kval.date()
        if isinstance(kval, datetime.date) and etrade_date > kval:
            to_remove_t1.add(sym)
            to_add_t1.add(sym)
            print(f"[Short_Long] T1 {sym}: dateAcquired moved {kval} -> {etrade_date}, replacing record")

    for sym in set(sheet_t2_rows) & etrade_t2_syms:
        etrade_date = etrade_t2_by_sym[sym].get("date_acquired")
        if not etrade_date:
            continue
        kval = ws.cell(sheet_t2_rows[sym], 11).value
        if isinstance(kval, datetime.datetime):
            kval = kval.date()
        if isinstance(kval, datetime.date) and etrade_date > kval:
            to_remove_t2.add(sym)
            to_add_t2.add(sym)
            print(f"[Short_Long] T2 {sym}: dateAcquired moved {kval} -> {etrade_date}, replacing record")

    # ── 2. Remove closed/replaced positions (reverse row order) ─────────────
    rows_to_delete = (
        [sheet_t1_rows[s] for s in to_remove_t1] +
        [sheet_t2_rows[s] for s in to_remove_t2]
    )
    for rn in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(rn)

    sheet_t1_rows, sheet_t2_rows = _build_row_maps()

    # ── 3. Add new positions — T1 first (insertions shift T2 row numbers) ────
    for sym in sorted(to_add_t1):
        pos  = etrade_t1_by_sym[sym]
        pick = picks_lookup.get(sym, {})
        t1_last     = max(sheet_t1_rows.values(), default=2)
        new_row_num = t1_last + 1
        ws.insert_rows(new_row_num)
        ws.cell(new_row_num, 1).value  = None
        ws.cell(new_row_num, 2).value  = sym
        ws.cell(new_row_num, 3).value  = pos["qty"]
        ws.cell(new_row_num, 4).value  = pos["cost"]
        ws.cell(new_row_num, 5).value  = pos["price"]
        ws.cell(new_row_num, 11).value = pos.get("date_acquired") or today
        ws.cell(new_row_num, 12).value = pick.get("industry", "")
        sheet_t1_rows[sym] = new_row_num

    # Rebuild T2 map after T1 insertions may have shifted rows
    _, sheet_t2_rows = _build_row_maps()

    for sym in sorted(to_add_t2):
        pos  = etrade_t2_by_sym[sym]
        pick = picks_lookup.get(sym, {})
        t2_last     = max(sheet_t2_rows.values(), default=max(sheet_t1_rows.values(), default=2) + 3)
        new_row_num = t2_last + 1
        ws.insert_rows(new_row_num)
        ws.cell(new_row_num, 1).value  = None
        ws.cell(new_row_num, 2).value  = sym
        ws.cell(new_row_num, 3).value  = pos["qty"]
        ws.cell(new_row_num, 4).value  = pos["cost"]
        ws.cell(new_row_num, 5).value  = pos["price"]
        ws.cell(new_row_num, 11).value = pos.get("date_acquired") or today
        ws.cell(new_row_num, 12).value = pick.get("industry", "")
        sheet_t2_rows[sym] = new_row_num

    # ── 3b. Ensure exactly 3 blank separator rows between T1 and T2 ──────────
    sheet_t1_rows, sheet_t2_rows = _build_row_maps()
    if sheet_t1_rows and sheet_t2_rows:
        t1_end   = max(sheet_t1_rows.values())
        t2_start = min(sheet_t2_rows.values())
        gap = t2_start - t1_end - 1
        if gap < 3:
            for _ in range(3 - gap):
                ws.insert_rows(t1_end + 1)
        elif gap > 3:
            for _ in range(gap - 3):
                ws.delete_rows(t1_end + 1)

    # ── 3c. Final compact: remove any residual internal blanks within T1 ─────
    sheet_t1_rows, _ = _build_row_maps()
    if len(sheet_t1_rows) >= 2:
        _t1_sorted = sorted(sheet_t1_rows.values())
        _final_blanks = [
            r[0].row for r in ws.iter_rows(
                min_row=_t1_sorted[0] + 1, max_row=_t1_sorted[-1] - 1
            )
            if not _is_data_row(r)
        ]
        for _rn in sorted(_final_blanks, reverse=True):
            ws.delete_rows(_rn)
            print(f"[Short_Long] Final-compact: removed residual blank row {_rn}")

    # ── 4. Renumber col A within each table ──────────────────────────────────
    sheet_t1_rows, sheet_t2_rows = _build_row_maps()

    def _renumber(row_map):
        for rank, rn in enumerate(sorted(row_map.values()), 1):
            ws.cell(rn, 1).value = rank

    _renumber(sheet_t1_rows)
    _renumber(sheet_t2_rows)

    sheet_t1_rows, sheet_t2_rows = _build_row_maps()

    # ── 5. Write header for new score columns ────────────────────────────────
    if HDR_ROW:
        for col, label in enumerate(
            ["Short10", "Long60", "Win%", "Status", "Days G", "Days R", "In Profit"],
            start=17,
        ):
            ws.cell(HDR_ROW, col).value = label

    # ── 6. Update Top (col E) and write score columns Q-W ────────────────────
    # Process each table with its own position lookup so AMZN (in both) is correct
    all_rows = (
        [(sym, rn, etrade_t1_by_sym) for sym, rn in sheet_t1_rows.items()] +
        [(sym, rn, etrade_t2_by_sym) for sym, rn in sheet_t2_rows.items()]
    )
    for sym, rn, pos_lookup in all_rows:
        if sym in quotes:
            ws.cell(rn, 5).value = quotes[sym]

        top  = ws.cell(rn, 5).value
        buy  = ws.cell(rn, 4).value
        kval = ws.cell(rn, 11).value

        pick = picks_lookup.get(sym, {})
        s10  = pick.get("short10")
        l60  = pick.get("long60")
        br   = pick.get("br")
        winp = round(_pwp(br) * 100, 1) if br is not None else None
        status_label, status_fill = _status(l60)

        if isinstance(kval, datetime.datetime):
            kval = kval.date()
        days = (today - kval).days if isinstance(kval, datetime.date) else None

        top_f = float(top or 0)
        buy_f = float(buy or 0)
        in_profit = top_f > buy_f if (top_f and buy_f) else None

        days_green = days if in_profit else 0
        days_red   = days if (in_profit is False) else 0

        # Q: Short10
        c = ws.cell(rn, 17)
        c.value = s10
        c.font  = GRN_FONT if (s10 or 0) >= 0 else RED_FONT
        c.alignment = CTR

        # R: Long60
        c = ws.cell(rn, 18)
        c.value = l60
        c.font  = GRN_FONT if (l60 or 0) >= 0 else RED_FONT
        c.alignment = CTR

        # S: Win%
        c = ws.cell(rn, 19)
        c.value = f"{winp}%" if winp is not None else None
        c.alignment = CTR

        # T: Status
        c = ws.cell(rn, 20)
        c.value = status_label
        c.fill  = status_fill
        c.alignment = CTR

        # U: Days Green
        c = ws.cell(rn, 21)
        c.value = days_green
        c.fill  = GRN_FILL if in_profit else GRY_FILL
        c.alignment = CTR

        # V: Days Red
        c = ws.cell(rn, 22)
        c.value = days_red
        c.fill  = RED_FILL if (in_profit is False) else GRY_FILL
        c.alignment = CTR

        # W: In Profit
        c = ws.cell(rn, 23)
        if in_profit is True:
            c.value = "YES"
            c.fill  = GRN_FILL
            c.font  = GRN_FONT
        elif in_profit is False:
            c.value = "NO"
            c.fill  = RED_FILL
            c.font  = RED_FONT
        else:
            c.value = None
        c.alignment = CTR


def backup_xlsx(xlsx_path: str) -> str | None:
    """Copy xlsx to a timestamped backup in Backup/{year}/. Returns backup path or None."""
    if not os.path.exists(xlsx_path):
        return None
    now = datetime.datetime.now()
    ts  = now.strftime("%Y-%m-%d_%H%M%S")
    dst = os.path.join(os.path.dirname(xlsx_path), "Backup",
                       str(now.year), f"investment_{ts}.xlsx")
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.copy2(xlsx_path, dst)
    print(f"Backup saved to {dst}")
    return dst
