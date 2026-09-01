import datetime
import sys
import re
import requests
import json
import os
import time
import urllib3
import pytz
try:
    from playwright_stealth import Stealth
except ImportError:
    Stealth = None
from aether.notify import send_email
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from utils import _to_float
from aether_logger import get_logger as _get_logger
import risk_utils
import instruments
from config import CFG
try:
    from playwright.sync_api import sync_playwright
except ImportError:
    sync_playwright = None

_pg_log = _get_logger("powergauge")


def is_nyse_market_open() -> bool:
    """Return True if the NYSE is currently open (9:30 AM - 4:00 PM Eastern, weekdays).

    Setting FORCE_MARKET_CLOSED=true forces a CLOSED result. This is a test /
    dry-run seam only, and it can only ever suppress live-market behavior (the
    conservative direction) — it can never force the market falsely OPEN."""
    if os.environ.get("FORCE_MARKET_CLOSED") == "true":
        return False
    try:
        tz_ny = pytz.timezone("America/New_York")
        now_ny = datetime.datetime.now(tz_ny)
        if now_ny.weekday() in (5, 6):
            return False
        # Check market hours (9:30 AM - 4:00 PM Eastern)
        start_time = now_ny.replace(hour=9, minute=30, second=0, microsecond=0)
        end_time = now_ny.replace(hour=16, minute=0, second=0, microsecond=0)
        return start_time <= now_ny <= end_time
    except Exception:
        return False

from workbook_write import (
    write_research_headers      as _write_research_headers,
    write_picks_sheet           as _write_picks_sheet,
    update_short_long_scores    as _update_short_long_scores,
    update_replacements_sheet   as _update_replacements_sheet,
    fix_comment_shape_ids       as _fix_comment_shape_ids,
    backup_xlsx                 as _backup_xlsx,
)
from scoring import (
    REGIME_SYMBOL,
    ohlcv_streak_perc    as _ohlcv_streak_perc,
    ohlcv_streak_count   as _ohlcv_streak_count,
    week_of_month        as _week_of_month,
    compute_seasonality  as _compute_seasonality,
    predicted_win_pct    as _predicted_win_pct,
    market_regime        as _market_regime,
    rel_volume_bucket    as _rel_volume_bucket,
    fibonacci_retracement_score as _fib_score,
    rsi_divergence_score        as _rsi_div_score,
    digit_sum_score             as _digit_sum_score,   # close→next-day only
    short_score          as _short_score_fn,
    long_score           as _long_score_fn,
)
from patterns import (
    candlestick_score    as _cs_score,
    chart_pattern_score  as _cp_score,
    momentum_pattern_score as _mo_score,
    pattern_summary      as _pattern_summary,
    rubber_band_reversal_score as _rbr_score,
)

PGR_STR = ["", "Be-", "Be", "N", "Bu", "Bu+", ""]
_CHAIKIN_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"

# _CHAIKIN_API_KEY is defined once, below, in the "Chaikin API" constants block.


def _pgr_str(v: int) -> str:
    if 0 <= v < len(PGR_STR):
        return PGR_STR[v]
    return ""


def _chaikin_uuid() -> str:
    """Return the Chaikin account email from config (used as the 'uuid' request header)."""
    try:
        from config import CFG
        return CFG.chaikin_email or ""
    except Exception:
        return ""

from aether.token_renewer import TokenRenewer as _TokenRenewer

_SESSION_VALID_TTL   = 300   # seconds to trust a validated session without re-checking
_session_valid_until = 0.0   # monotonic timestamp; avoids HTTP validation on every call

_chaikin_renewer = _TokenRenewer(
    lock_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "chaikin_reauth.lock"),
    renew_fn=lambda: login(interactive=False),
    load_fn=lambda: _load_session_from_file(),
    lock_ttl=180,
    wait_timeout=90,
)


def ensure_valid_session() -> dict:
    """Return a valid Chaikin session, refreshing headlessly if expired.
    Uses TokenRenewer for cross-process safety — see aether/token_renewer.py.
    """
    global _session_valid_until
    session = _load_session_from_file()
    # Fast path: TTL cache avoids HTTP validation probe on every call
    if session and time.monotonic() < _session_valid_until:
        return session
    status = _probe_session(session) if session else "invalid"
    if status == "valid":
        _session_valid_until = time.monotonic() + _SESSION_VALID_TTL
        return session
    if status == "unreachable":
        # Network/proxy fault or transient upstream 5xx — do NOT re-auth. Keep the
        # existing token; a proxy blip must not destroy a good session or spin up a
        # doomed headless browser (see _probe_session's 3-state contract). This runs
        # before the JWT/browser refresh because both make network calls that would
        # also fail when Chaikin is unreachable.
        _pg_log.warning("Chaikin unreachable (network/proxy/5xx) — keeping existing session, skipping re-auth.")
        return session or {}
    # status == "invalid": genuinely expired/rejected — try the cheap refresh path first.
    # API-based JWT Token Refresh (Bypasses Browser/Turnstile completely in 0.2 seconds!)
    if session and session.get("jwttoken"):
        try:
            new_sid = _jwt_to_session_id(session["jwttoken"])
            if new_sid:
                # Mutation Hygiene: Copy dict before modifying to prevent in-place corruption
                test_session = session.copy()
                test_session["jsessionid"] = new_sid
                if _validate_session(test_session):
                    _save_session_to_file(test_session)
                    _session_valid_until = time.monotonic() + _SESSION_VALID_TTL
                    _pg_log.info("Successfully refreshed Chaikin session using saved JWT token (API bypass).")
                    return test_session
                else:
                    _pg_log.warning("JWT exchanged session ID failed active validation check.")
        except Exception as e:
            # Security Best Practice: Sanitize and redact raw JWT token from exception logs to prevent exposure
            err_msg = str(e)
            if "jwtToken=" in err_msg:
                err_msg = re.sub(r"jwtToken=[^&\s]+", "jwtToken=REDACTED", err_msg)
            _pg_log.warning(f"Failed to refresh session using JWT token (will fall back to browser): {err_msg}")

    # Expired — delegate to the cross-process singleton (protected by try-except to send email outside of lock duration)
    try:
        new_session = _chaikin_renewer.ensure(current_token=session)
        if new_session and new_session.get("jsessionid"):
            _session_valid_until = time.monotonic() + _SESSION_VALID_TTL
        return new_session or session or {}
    except EnvironmentError as e:
        # We are now OUTSIDE the cross-process lock! We can safely send the email alert
        # without adding any 30s SMTP latency to other waiting background tasks.
        try:
            session_abs_path = os.path.abspath(SESSION_FILE)
            send_email(
                subject="ALERT: Chaikin Turnstile Block - Manual Auth Required",
                body=f"Chaikin automated session token renewal failed due to browser login timeout/Turnstile challenge.\n\nError: {e}\n\nActions required:\n1. Log in manually at https://app.chaikinanalytics.com in a regular browser.\n2. Extract JSESSIONID from DevTools request headers.\n3. Save JSESSIONID to {session_abs_path}.\n4. Re-run the daily pipeline."
            )
        except Exception as mail_err:
            _pg_log.warning("Failed to send Turnstile block alert email: %s", mail_err)
        raise


# Pre-built index of symbol → sorted list of cached JSON paths.
# None = not yet scanned; {} = scanned but empty directory.
_cache_file_index: dict | None = None


class LazyCacheFileIndex(dict):
    def __init__(self):
        super().__init__()
        self._scanned_symbols = set()
        self._symbol_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "Symbol")
        try:
            if os.path.isdir(self._symbol_dir):
                for d in os.listdir(self._symbol_dir):
                    if os.path.isdir(os.path.join(self._symbol_dir, d)):
                        self[d] = []
        except OSError:
            pass

    def _ensure_symbol(self, sym: str):
        if sym in self._scanned_symbols:
            return
        self._scanned_symbols.add(sym)
        sym_dir = os.path.join(self._symbol_dir, sym)
        paths = []
        if os.path.isdir(sym_dir):
            try:
                for name in os.listdir(sym_dir):
                    if name.endswith('.json'):
                        if name.rsplit('_', 1)[0] == sym:
                            paths.append(os.path.join(sym_dir, name))
            except OSError:
                pass
        self[sym] = sorted(paths)

    def __getitem__(self, item):
        if isinstance(item, str):
            self._ensure_symbol(item)
        return super().__getitem__(item)

    def get(self, key, default=None):
        if isinstance(key, str):
            self._ensure_symbol(key)
        return super().get(key, default)

    def __contains__(self, key):
        if isinstance(key, str):
            self._ensure_symbol(key)
        return super().__contains__(key)


def _build_cache_index():
    """Scan Data/Symbol recursively and build a symbol→[paths] index for find_prev_pf."""
    global _cache_file_index
    if _cache_file_index is not None:
        return
    _cache_file_index = LazyCacheFileIndex()
    _pg_log.info(f"Cache index built: {len(_cache_file_index)} symbols")


SESSION_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "session.json")

_http_session: requests.Session | None = None
_resolved_proxy: str | None = None   # process-level memo: "" (direct) or a proxy URL
_DIRECT_TOKENS = {"", "direct", "none", "off", "no"}


def _proxy_reachable(proxy_url: str) -> bool:
    """Quick TCP check that a proxy host:port accepts connections (~1.5s cap).

    Lets 'auto' mode pick the Intel/E*TRADE proxy only when we are actually on that
    network, and fall back to a direct connection everywhere else.
    """
    try:
        import socket
        from urllib.parse import urlparse
        u = urlparse(proxy_url if "://" in proxy_url else "http://" + proxy_url)
        host, port = u.hostname, (u.port or 911)
        if not host:
            return False
        with socket.create_connection((host, port), timeout=1.5):
            return True
    except Exception:
        return False


def _resolve_proxy() -> str:
    """Resolve the Chaikin proxy at call time (NOT import time), memoized per process.

    Returns a proxy URL to use, or "" for a direct connection. Honors CFG.chaikin_proxy:
    an explicit URL is used verbatim; "" / "direct" force direct; "auto" (default) uses
    CFG.etrade_proxy only when that host is reachable — so the same build works on and
    off the Intel network. Override with env CHAIKIN_PROXY or config chaikin.proxy.
    """
    global _resolved_proxy
    if _resolved_proxy is not None:
        return _resolved_proxy
    mode, candidate = "auto", ""
    try:
        from config import CFG
        mode = (CFG.chaikin_proxy or "").strip()
        candidate = (CFG.etrade_proxy or "").strip()
    except Exception:
        pass
    if mode.lower() == "auto":
        resolved = candidate if (candidate and _proxy_reachable(candidate)) else ""
    elif mode.lower() in _DIRECT_TOKENS:
        resolved = ""
    else:
        resolved = mode   # explicit proxy URL
    _resolved_proxy = resolved
    _pg_log.info(f"Chaikin proxy resolved: {resolved or '(direct)'}")
    return resolved


def _get_http_session() -> requests.Session:
    """Return a shared Session with retry logic and the configured proxy applied."""
    global _http_session
    if _http_session is None:
        _http_session = requests.Session()
        _retry = Retry(total=1, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
        adapter = HTTPAdapter(max_retries=_retry)
        _http_session.mount("https://", adapter)
        _http_session.mount("http://", adapter)
        _http_session.verify = False
        proxy = _resolve_proxy()
        if proxy:
            _http_session.proxies.update({"http": proxy, "https": proxy})
        else:
            # Explicit direct: ignore any stray HTTPS_PROXY/HTTP_PROXY the E*TRADE client
            # may have exported into the environment, so off-Intel runs stay direct.
            _http_session.trust_env = False
    return _http_session


SRC_XLSX  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "state_of_the_day.xlsx")
XLSX_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "state_of_the_day.xlsx")
XLSX_BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "Backup")
OHLCV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "Symbol_full")

# ── Chaikin API ───────────────────────────────────────────────────────────────
# OMNI (/api/*) app key. It is sent as `x-api-key` (alongside `x-app-id: omni`) on
# every /api/* call. It is *reportedly* the OMNI web client key that ships in the
# members.chaikinanalytics.com JS bundle as $rootScope.config.apiKey — i.e. plausibly
# a public client-side key — but that has NOT been verified here.
#
# It is intentionally NOT hardcoded here. This repo is public and the value's
# public-vs-secret status is unverified, so we ship no default in source. Supply it via
# either `chaikin.api_key` in config.json or the CHAIKIN_API_KEY environment variable.
# To obtain the live value, read the `x-api-key` request header from a logged-in OMNI
# session (DevTools → Network → any members-backend.chaikinanalytics.com/api/* call), or
# pull `$rootScope.config.apiKey` from the OMNI web JS bundle. With an empty key the API
# returns HTTP 403 {"code":"SESSION_EXPIRED","message":"Missing required headers"} — the
# message is misleading (it's the missing key, not an expired token), so the session
# probe/fetch will report unreachable/invalid until the key is configured.
try:
    from config import CFG
    _CHAIKIN_API_KEY = CFG.chaikin_api_key or os.environ.get("CHAIKIN_API_KEY") or ""
except Exception:
    _CHAIKIN_API_KEY = os.environ.get("CHAIKIN_API_KEY") or ""
# Concurrent workers for parallel symbol fetch in check_from_xls.
_FETCH_WORKERS = int(os.environ.get("CHAIKIN_WORKERS", "10"))

# ── Symbol validation ─────────────────────────────────────────────────────────
_SYMBOL_RE = re.compile(r"^[A-Z0-9._\-]+$")

# ── OHLCV / entry-filter parameters ──────────────────────────────────────────
# Stop/target now come from risk_utils.detect_support/detect_resistance (shared with
# the dashboard); only the trend/direction entry filter is computed here.
_TREND_SMA_PERIOD     = 20   # SMA period for trend filter
_DIR_CHECK_DAYS       = 3    # "price above N days ago" direction check
SESSION_INSTRUCTIONS = """
Session expired or missing. To get a new session token:
  1. Open https://app.chaikinanalytics.com in your browser and log in.
  2. Press F12 to open DevTools, go to the Network tab.
  3. Click on any API request (e.g. getSymbolData or getChecklistStocks).
  4. In the Request Headers, find the 'Cookie' header.
  5. Copy the value of JSESSIONID (the part after 'JSESSIONID=' and before ';').
  6. Save that value to: {session_file}
Then re-run the script.
""".strip()

class PowerGauge:
    def __init__(self, symbol, date=None):
        self.symbol = symbol
        self.date = date if date is not None else datetime.date.today()
        self.pgr_value = 0
        self.pgr_corrected_value = 0
        self.industry_name = ""
        self.price = 0.0
        self.max_price = 0.0
        self.signals = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        self.percentage = 0.0
        self.change = 0.0
        self.prevPG: PowerGauge | None = None
        self.industry_strength = ""
        self.lt_trend = ""
        self.money_flow = ""
        self.over_bt_sl = ""
        self.relative_strength = ""

    def init_from_json(self, data_json, check_schema=True):
        cl = data_json.get('checklist_stocks') or {}

        # Early-out: invalid symbol reported by API
        if data_json.get('status') == 'invalid symbol':
            self.price = -1
            if check_schema:
                self._check_schema(data_json)
            return

        pgr_list = data_json.get('pgr') or []
        self.pgr_value = (
            pgr_list[0].get('PGR Value', 0) if len(pgr_list) > 0
            else cl.get('rawPgrRating', 0)
        )
        self.pgr_corrected_value = (
            pgr_list[5].get('Corrected PGR Value', 0) if len(pgr_list) > 5
            else cl.get('pgrRating', 0)
        )
        metainfo = data_json.get('metaInfo') or [{}]
        m = metainfo[0]
        industry = (m.get('industry_name') or m.get('etf_group_name') or m.get('industry_logo_name')
                    or (m.get('etf_data') or {}).get('list_name')
                    or m.get('name') or '')
        self.industry_name = industry.replace(',', '')
        if not self.industry_name:
            if _cache_file_index is None:
                _build_cache_index()
            candidates = (_cache_file_index or {}).get(self.symbol, [])
            for path in reversed(candidates):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        prev_data = json.load(f)
                    prev_meta = prev_data.get("metaInfo") or []
                    if prev_meta and isinstance(prev_meta, list) and len(prev_meta) > 0:
                        pm = prev_meta[0]
                        prev_ind = (pm.get('industry_name') or pm.get('etf_group_name') or pm.get('industry_logo_name')
                                    or (pm.get('etf_data') or {}).get('list_name')
                                    or pm.get('name') or '')
                        if prev_ind:
                            self.industry_name = prev_ind.replace(',', '')
                            break
                except Exception:
                    continue
        self.price = m.get('Last') if m.get('Last') is not None else _to_float(cl.get('lastPrice'), -1)
        self.max_price = self.price
        self.signals = m.get('signals')
        self.percentage = m.get('Percentage ') if m.get('Percentage ') is not None else _to_float(cl.get('changePercentage'), 0)
        self.change = m.get('Change') if m.get('Change') is not None else _to_float(cl.get('change'), None)
        self.industry_strength = cl.get('industry')
        self.lt_trend = cl.get('ltTrend')
        self.money_flow = cl.get('moneyFlow')
        self.over_bt_sl = cl.get('overboughtOversold')
        # relativeStrength: prefer checklist_stocks string; fall back to pgr[3] numeric score
        rs_str = cl.get('relativeStrength')
        pgr_list = data_json.get('pgr') or []
        if rs_str:
            self.relative_strength = rs_str
        elif len(pgr_list) > 3:
            technicals = pgr_list[3].get('Technicals') or []
            rs_score = next((t.get('Rel Strength vs Market') for t in technicals if 'Rel Strength vs Market' in t), None)
            self.relative_strength = str(rs_score) if rs_score is not None else ""
        if check_schema:
            self._check_schema(data_json)

    def _check_schema(self, data_json):
        warnings = []
        pgr_list = data_json.get('pgr') or []
        if not pgr_list:
            warnings.append("'pgr' list missing or empty")
        elif len(pgr_list) <= 5:
            warnings.append(f"'pgr' list shorter than expected (len={len(pgr_list)}, need >=6)")
        else:
            if 'PGR Value' not in pgr_list[0]:
                warnings.append("pgr[0] missing 'PGR Value'")
            if 'Corrected PGR Value' not in pgr_list[5]:
                warnings.append("pgr[5] missing 'Corrected PGR Value'")
        metainfo = data_json.get('metaInfo') or []
        if not metainfo:
            warnings.append("'metaInfo' list missing or empty")
        else:
            for key in ('Last', 'Percentage ', 'Change', 'signals'):
                if key not in metainfo[0]:
                    warnings.append(f"metaInfo[0] missing key '{key}'")
            if not any(k in metainfo[0] for k in ('industry_name', 'etf_group_name', 'industry_logo_name')):
                if not ((metainfo[0].get('etf_data') or {}).get('list_name') or metainfo[0].get('name')):
                    warnings.append("metaInfo[0] missing industry key (industry_name/etf_group_name/industry_logo_name/name)")
        cl = data_json.get('checklist_stocks') or {}
        if not cl:
            warnings.append("'checklist_stocks' missing or empty")
        else:
            for key in ('industry', 'ltTrend', 'moneyFlow', 'overboughtOversold'):
                if key not in cl:
                    warnings.append(f"checklist_stocks missing key '{key}'")
        if warnings:
            print(f"  [SCHEMA WARNING] {self.symbol}: " + "; ".join(warnings))

    def init_from_ohlcv(self, entry: dict):
        """Populate price fields from an Alpha Vantage OHLCV daily entry."""
        close = _to_float(entry.get('4. close'), -1)
        prev_close = _to_float(entry.get('prev_close'), close)
        self.price = close
        self.max_price = _to_float(entry.get('2. high'), close)
        self.change = round(close - prev_close, 4) if prev_close else 0
        self.percentage = round((self.change / prev_close) * 100, 4) if prev_close else 0

    def find_prev_pf(self):
        if self.prevPG is not None:
            return
        if _cache_file_index is None:
            _build_cache_index()
        candidates = (_cache_file_index or {}).get(self.symbol, [])
        today_str = str(self.date)

        # 1. Try Chaikin cache (most recent file before today)
        for path in reversed(candidates):
            fname = os.path.basename(path)
            date_str = fname[len(self.symbol) + 1:-5]
            if date_str < today_str:
                try:
                    prev_date = datetime.date.fromisoformat(date_str)
                except ValueError:
                    continue
                try:
                    with open(path, "r") as f:
                        data_jsn = json.load(f)
                except (json.JSONDecodeError, OSError) as e:
                    print(f"  [CACHE] {self.symbol}: skipping corrupt cache {path}: {e}")
                    continue
                self.prevPG = PowerGauge(self.symbol, prev_date)
                self.prevPG.init_from_json(data_jsn, check_schema=False)
                return

        # 2. Fall back to OHLCV data (Symbol_full/{symbol}_daily.json)
        ohlcv_file = os.path.join(OHLCV_DIR, f"{self.symbol}_daily.json")
        if not os.path.exists(ohlcv_file):
            return
        with open(ohlcv_file) as f:
            ohlcv = json.load(f)
        ts = ohlcv.get('Time Series (Daily)') or {}
        past_dates = sorted((d for d in ts if d < today_str), reverse=True)
        if not past_dates:
            return
        prev_date_str = past_dates[0]
        # Attach previous close so change/percentage are meaningful
        entry = dict(ts[prev_date_str])
        if len(past_dates) > 1:
            entry['prev_close'] = ts[past_dates[1]].get('4. close')
        self.prevPG = PowerGauge(self.symbol, datetime.date.fromisoformat(prev_date_str))
        self.prevPG.init_from_ohlcv(entry)

    def get_prev_same_move_count(self, _depth: int = 0) -> int:
        if _depth > 30:
            return -1 if self.percentage < 0 else 1
        if not self.prevPG:
            self.find_prev_pf()
        if self.prevPG:
            if self.percentage > 0 and self.prevPG.percentage > 0:
                return self.prevPG.get_prev_same_move_count(_depth + 1) + 1
            if self.percentage < 0 and self.prevPG.percentage < 0:
                return self.prevPG.get_prev_same_move_count(_depth + 1) - 1
            return -1 if self.percentage < 0 else 1
        return 0

    def get_prev_same_move_percent(self, _depth: int = 0) -> float:
        if _depth > 30:
            return self.percentage
        if not self.prevPG:
            self.find_prev_pf()
        if self.prevPG:
            if self.percentage > 0 and self.prevPG.percentage > 0:
                return (self.prevPG.get_prev_same_move_percent(_depth + 1) or self.prevPG.percentage) + self.percentage
            if self.percentage < 0 and self.prevPG.percentage < 0:
                return (self.prevPG.get_prev_same_move_percent(_depth + 1) or self.prevPG.percentage) + self.percentage
        return 0

    def get_prev_same_move_price(self, _depth: int = 0) -> float:
        if _depth > 30:
            return self.price
        if not self.prevPG:
            self.find_prev_pf()
        if self.change and self.prevPG and self.prevPG.change:
            if self.change > 0 and self.prevPG.change > 0:
                return self.prevPG.get_prev_same_move_price(_depth + 1) or self.prevPG.price
            if self.change < 0 and self.prevPG.change < 0:
                return self.prevPG.get_prev_same_move_price(_depth + 1) or self.prevPG.price
        return 0

    def get_prev_max_price(self, cur_price):
        if not self.prevPG:
            self.find_prev_pf()
        min_pr = self.get_prev_min_of(deep=3)
        local_max = max(self.max_price, self.get_prev_max_of(deep=3).price)
        if not min_pr.prevPG:
            min_pr.find_prev_pf()
        if min_pr.prevPG:
            if min_pr.price < cur_price:
                return min_pr.prevPG.get_prev_max_price(cur_price)
            return max(min_pr.get_prev_same_move_price() or min_pr.price, local_max)
        return local_max

    def get_prev_min_of(self, deep=3):
        if not self.prevPG:
            self.find_prev_pf()
        if self.prevPG:
            self.max_price = max(self.prevPG.max_price, self.max_price)
            pr = self
            if deep > 0:
                pr = self.prevPG.get_prev_min_of(deep-1)
            if pr.price < self.price:
                return pr
        return self

    def get_prev_max_of(self, deep=3):
        if not self.prevPG:
            self.find_prev_pf()
        if self.prevPG:
            pr = self
            if deep > 0:
                pr = self.prevPG.get_prev_max_of(deep-1)
            if pr.price >= self.price:
                return pr
        return self


def _load_session_from_file() -> dict:
    session_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "session.json")
    if not os.path.exists(session_file):
        txt_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "session.txt")
        if os.path.exists(txt_file):
            with open(txt_file, "r") as f:
                sid = f.read().strip()
                return {
                    "jsessionid": sid,
                    "jwttoken": "",
                    "uuid": _chaikin_uuid()
                }
        return {}
    with open(session_file, "r") as f:
        try:
            return json.load(f)
        except Exception:
            return {}


def _save_session_to_file(session_data: dict):
    session_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "session.json")
    os.makedirs(os.path.dirname(session_file), exist_ok=True)
    with open(session_file, "w") as f:
        json.dump(session_data, f, indent=2)


def _probe_session(session_data: dict) -> str:
    """Probe the Chaikin session, returning 'valid' | 'invalid' | 'unreachable'.

    'unreachable' means a network/proxy fault (or a transient 5xx like Cloudflare's
    503) — NOT proof the token is bad. Callers must keep the existing session and
    skip re-auth in that case, so a proxy blip never nukes a good token or spins up
    a doomed headless browser. Only a real 401/403 (or a missing id) is 'invalid'.
    """
    if not session_data or not session_data.get("jsessionid"):
        return "invalid"
    # New Fastify backend: /api/suggestions/{symbol} is a live, cheap liveness probe
    # (the legacy /CPTRestSecure/* path now 503s for everyone — see chaikin migration).
    test_url = "https://members-backend.chaikinanalytics.com/api/suggestions/AAPL"
    headers = {
        'jsessionid': session_data['jsessionid'],
        'x-session-id': session_data['jsessionid'],
        'uuid': session_data.get('uuid') or _chaikin_uuid(),
        'jwttoken': session_data.get('jwttoken', ''),
        'x-api-key': _CHAIKIN_API_KEY,
        'x-app-id': 'omni',
        'User-Agent': _CHAIKIN_UA
    }
    try:
        r = _get_http_session().get(test_url, headers=headers, timeout=(5, 15))
    except (requests.Timeout, requests.ConnectionError, requests.RequestException):
        return "unreachable"
    if r.status_code == 200:
        return "valid"
    if r.status_code in (401, 403):
        return "invalid"
    if r.status_code >= 500:
        # Upstream/edge transient (e.g. CF 503) — session may well be fine.
        return "unreachable"
    return "invalid"


def _validate_session(session_data: dict) -> bool:
    """Back-compat boolean wrapper: True only when the session is confirmed valid."""
    return _probe_session(session_data) == "valid"


def _jwt_to_session_id(jwt_token: str) -> str:
    url = ("https://members-backend.chaikinanalytics.com/CPTRestSecure/app"
           "/authenticate/getJWTAuthorization?acquireSessionForcibly=Yes"
           f"&jwtToken={jwt_token}")
    headers = {
        'X-Api-Key': _CHAIKIN_API_KEY,
        'X-App-Id': 'omni',
    }
    r = _get_http_session().get(url, headers=headers, timeout=(5, 15))
    if not r.ok:
        raise EnvironmentError(f"JWT exchange failed: HTTP {r.status_code}")
    session_id = r.json().get('sessionId')
    if not session_id:
        raise EnvironmentError(f"No sessionId in JWT exchange response: {r.text[:200]}")
    return session_id


def _load_credentials() -> tuple[str, str]:
    """Load Chaikin credentials from unified config (env vars override config.json)."""
    email, password = CFG.chaikin_email, CFG.chaikin_password
    if not email or not password:
        raise EnvironmentError(
            "Chaikin credentials not found.\n"
            "  Option 1: set env vars CHAIKIN_EMAIL and CHAIKIN_PASSWORD\n"
            "  Option 2: add chaikin.email / chaikin.password to config.json"
        )
    return email, password


def _login_via_browser(headless: bool = False) -> dict:
    if not sync_playwright:
        raise ImportError("Playwright is not installed in the environment.")

    _pg_log.info(f"Opening browser for login (headless={headless})...")
    session_data = [None]

    def on_request(request):
        if 'members-backend.chaikinanalytics.com' in request.url:
            jsid = request.headers.get('jsessionid') or request.headers.get('x-session-id')
            jwt = request.headers.get('jwttoken')
            uuid = request.headers.get('uuid') or _chaikin_uuid()
            if jsid and jsid != 'NULL' and len(jsid) > 10 and jwt:
                session_data[0] = {
                    "jsessionid": jsid,
                    "jwttoken": jwt,
                    "uuid": uuid
                }

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless,
            channel='chrome',
            args=['--disable-blink-features=AutomationControlled'],
        )
        global _CHAIKIN_UA
        _CHAIKIN_UA = f"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{browser.version} Safari/537.36"
        context = browser.new_context(user_agent=_CHAIKIN_UA)
        page = context.new_page()
        if Stealth is not None:
            try:
                Stealth().apply_stealth_sync(page)
            except Exception as e:
                _pg_log.warning(f"Failed to apply playwright-stealth: {e}")
        page.on('request', on_request)

        page.goto('https://members.chaikinanalytics.com/login', wait_until='domcontentloaded', timeout=60000)

        email, password = _load_credentials()
        page.fill('input[name="email"]', email)
        page.fill('input[name="password"]', password)

        # Wait for Turnstile to enable the submit button (auto-verifies or user clicks widget)
        print("Waiting for Turnstile to complete (up to 60s — click the checkbox if it appears)...")
        page.wait_for_selector('button[type="submit"]:not([disabled])', timeout=60000)
        page.click('button[type="submit"]')

        print("Waiting for login to complete (up to 60s)...")
        try:
            page.wait_for_function(
                "window.location.pathname !== '/login'",
                timeout=60000
            )
        except Exception:
            pass

        # Navigate to app.chaikinanalytics.com to fully activate the session
        print("Navigating to app.chaikinanalytics.com to activate the session...")
        try:
            page.goto('https://app.chaikinanalytics.com', timeout=30000)
            page.wait_for_timeout(5000)
        except Exception as e:
            print(f"Warning: Navigation to app.chaikinanalytics.com failed or timed out: {e}")

        browser.close()

    if not session_data[0]:
        raise EnvironmentError(
            "Browser login completed but session ID was not captured. "
            "Fall back to manual session: " + SESSION_FILE
        )

    _save_session_to_file(session_data[0])
    print(f"Session saved to {SESSION_FILE}")
    return session_data[0]


def login(interactive=True) -> dict:
    session_data = _load_session_from_file()
    if session_data:
        print("Loaded session from file, validating...")
        status = _probe_session(session_data)
        if status == "valid":
            print("Session is valid.")
            return session_data
        if status == "unreachable":
            # Can't reach Chaikin (network/proxy/5xx) — the session may well be fine.
            # Don't discard it and don't launch a browser that also can't reach the site.
            print("Chaikin unreachable (network/proxy/5xx) — keeping existing session; skipping browser re-auth.")
            return session_data
        print("Saved session has expired — re-authenticating via browser.")

    # Run headless if we are non-interactive or stdin is not a tty to prevent hanging
    is_tty = sys.stdin and sys.stdin.isatty()
    headless_run = not interactive or not is_tty

    try:
        return _login_via_browser(headless=headless_run)
    except Exception as e:
        print(f"Browser login failed: {e}")
        if not interactive or not sys.stdin or not sys.stdin.isatty():
            raise EnvironmentError(f"Chaikin browser login failed: {e}") from e

    if not interactive or not sys.stdin or not sys.stdin.isatty():
        return {}

    print(SESSION_INSTRUCTIONS.format(session_file=SESSION_FILE))
    try:
        raw = input("Or paste a JSESSIONID here and press Enter (leave blank to abort): ").strip()
    except (EOFError, KeyboardInterrupt):
        raw = ""
    if raw:
        data = {
            "jsessionid": raw,
            "jwttoken": "",
            "uuid": _chaikin_uuid()
        }
        _save_session_to_file(data)
        print(f"Session saved to {SESSION_FILE}")
        return data
    raise EnvironmentError(
        f"No valid session available. Save a JSESSIONID to: {SESSION_FILE}"
    )


# ── New /api/* → legacy getSymbolData schema adapter ───────────────────────────
# Chaikin migrated the data API to /api/suggestions/{symbol}. Its response shape is
# a flat 107-field dict; the rest of this module (init_from_json, _check_schema, the
# on-disk cache, find_prev_pf) still speaks the legacy {pgr[7], metaInfo[1],
# checklist_stocks{}} schema. This adapter converts new→legacy so nothing downstream
# — including the cache format — has to change.

# @doc-sync-start: chaikin_api
# Contract + field mapping documented in plans/chaikin_api.md. If you change the
# endpoint, the header contract, the 7->5 rating maps, or this adapter's shape, update
# that surface in the same commit (enforced by
# scripts/utils/pre_commit_validator.py :: check_feature_doc_sync).
#
# New pgrRating is a 7-level scale (1=Very Bearish … 7=Very Bullish, with Neutral -/+/
# granularity); the legacy code expects the old 5-level rating (1=Be- … 5=Bu+). Both
# name- and integer-based lookups collapse the Neutral -/·/+ band to old 3 (the old
# model had no +/- granularity). 0 = unrated (e.g. leveraged/inverse ETFs, no PGR).
_RATING_INT7_TO_OLD5 = {1: 1, 2: 2, 3: 3, 4: 3, 5: 3, 6: 4, 7: 5}
_RATING_NAME_TO_OLD5 = {
    "very bearish": 1, "bearish": 2,
    "neutral -": 3, "neutral": 3, "neutral +": 3,
    "bullish": 4, "very bullish": 5,
}
# signalInfo dict → legacy 12-char binary string, fixed order (display-only field).
_SIGNAL_KEYS = ("overBoughtSell", "overSoldBuy", "breakdownSell", "breakoutBuy",
                "reversalBuy", "reversalSell", "moneyFlowBuy", "moneyFlowSell",
                "relStrengthBuy", "relStrengthSell", "relStrengthBreakout",
                "relStrengthBreakdown")


def _pgr_rating_old5(int_rating, name) -> int:
    """Resolve the legacy 1-5 PGR rating from the new integer (1-7) or rating name."""
    if isinstance(int_rating, int) and int_rating > 0:
        v = _RATING_INT7_TO_OLD5.get(int_rating, 5 if int_rating > 7 else 0)
        if v:
            return v
    return _RATING_NAME_TO_OLD5.get(str(name or "").strip().lower(), 0)


def _adapt_suggestions_to_legacy(data: dict, symbol: str) -> dict:
    """Map a new /api/suggestions/{symbol} `data` object to the legacy getSymbolData bundle.

    Returns {status, pgr[7], metaInfo[1], checklist_stocks{}}. An unknown ticker still
    returns HTTP 200 but with an empty checklistData and null name — surfaced here as
    status='invalid symbol' so the caller sets price=-1 (matching the old API behavior).
    """
    if not isinstance(data, dict) or (not data.get("name") and not (data.get("checklistData") or {})):
        return {"status": "invalid symbol"}
    cl = data.get("checklistData") or {}
    raw5 = _pgr_rating_old5(data.get("rawPgrRating"), data.get("ratingName"))
    _corr_int = data.get("correctedPgrRating")
    if _corr_int is None:
        _corr_int = data.get("pgrRating")
    corr5 = _pgr_rating_old5(_corr_int, cl.get("pgr") or data.get("ratingName"))

    si = data.get("signalInfo") or {}
    signals = "".join("1" if (si.get(k) or 0) else "0" for k in _SIGNAL_KEYS) if si else "000000000000"

    last = _to_float(data.get("lastPrice"), None) if data.get("lastPrice") is not None else None
    pct = data.get("days1ChangePct")
    if pct is None:
        pct = data.get("latestChangePct")
    chg = data.get("days1Change")
    industry_name = data.get("industry") or data.get("sector") or data.get("name") or ""
    is_etf = bool(data.get("isEtf"))

    pgr_list = [
        {"PGR Value": raw5},
        {"Financials": []},
        {"Earnings": []},
        {"Technicals": []},
        {"Experts": []},
        {"Corrected PGR Value": corr5},
        {"is_etf_symbol": is_etf, "technical_rank": data.get("technicalRank") or 0},
    ]
    meta = {
        "symbol": symbol,
        "Last": last,
        "Percentage ": pct,
        "Change": chg,
        "signals": signals,
        "industry_name": industry_name,
        "etf_group_name": (data.get("name") or "") if is_etf else "",
        "name": data.get("name") or "",
        "marketCap": data.get("marketCap"),
        "is_etf_symbol": is_etf,
        "raw_PGR": data.get("rawPgr"),
    }
    checklist = {
        "symbol": symbol,
        "industry": cl.get("industry"),
        "ltTrend": cl.get("ltTrend"),
        "moneyFlow": cl.get("moneyFlow"),
        "overboughtOversold": cl.get("OBOS"),
        "relativeStrength": cl.get("relativeStrength"),
        "pgr": cl.get("pgr") or data.get("ratingName"),
        "rawPgrRating": raw5,
        "pgrRating": corr5,
        "lastPrice": str(last) if last is not None else None,
        "changePercentage": str(pct) if pct is not None else None,
        "change": str(chg) if chg is not None else None,
        "stockStatus": cl.get("status"),
    }
    return {"status": "ok", "pgr": pgr_list, "metaInfo": [meta], "checklist_stocks": checklist}
# @doc-sync-end: chaikin_api


def get_symbol_data(symbol: str, date, prefer_cache: bool, session_id=None, _allow_reauth: bool = True) -> PowerGauge:
    if not _SYMBOL_RE.match(symbol):
        raise ValueError(f"Invalid symbol format: {symbol!r}")

    session_data = ensure_valid_session()

    # New Fastify backend: a single GET returns the full symbol bundle (PGR + checklist
    # + meta); the legacy getSymbolData/getChecklistStocks pair (and its ?components=…)
    # is gone. _adapt_suggestions_to_legacy() reshapes the response to the old schema.
    url = f"https://members-backend.chaikinanalytics.com/api/suggestions/{symbol}"

    headers = {
        'jsessionid': session_data.get('jsessionid', ''),
        'x-session-id': session_data.get('jsessionid', ''),
        'uuid': session_data.get('uuid') or _chaikin_uuid(),
        'jwttoken': session_data.get('jwttoken', ''),
        'x-api-key': _CHAIKIN_API_KEY,
        'x-app-id': 'omni',
        'User-Agent': _CHAIKIN_UA
    }
    pg = PowerGauge(symbol, date)
    data_jsn = {}

    if date and prefer_cache:
        _base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "Symbol")
        file = os.path.join(_base, symbol, f"{symbol}_{date}.json")
        if not os.path.exists(file):
            file = os.path.join(_base, f"{symbol}_{date}.json")  # flat fallback
        if os.path.exists(file):
            with open(file, "r") as f:
                data_jsn = json.load(f)

    if not data_jsn:
        response = _get_http_session().get(url, headers=headers, timeout=(5, 20))
        if response.ok:
            raw_jsn = response.json()
            new_data = raw_jsn.get("data") if isinstance(raw_jsn, dict) else None
            data_jsn = _adapt_suggestions_to_legacy(new_data, symbol)
            # --- Closing Price Override (Pre-Save Reconciliation) ---
            # Overwrite the Chaikin price fields with the official, settled close from Symbol_full.
            # This guarantees that Chaikin (pg), RapidAPI (Symbol_full), and E*TRADE (live) are 100% synchronized!
            try:
                ohlcv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "Symbol_full", f"{symbol}_daily.json")
                if os.path.exists(ohlcv_path):
                    with open(ohlcv_path) as _f:
                        ohlcv_data = json.load(_f)
                    ohlcv_ts = ohlcv_data.get("Time Series (Daily)", {})
                    cache_date_str = str(date if date else datetime.date.today())
                    if ohlcv_ts and cache_date_str in ohlcv_ts:
                        official_close = float(ohlcv_ts[cache_date_str]["4. close"])
                        if official_close > 0.0:
                            meta_list = data_jsn.get("metaInfo")
                            if isinstance(meta_list, list) and len(meta_list) > 0:
                                meta_list[0]["Last"] = official_close
                            elif isinstance(meta_list, dict):
                                meta_list["Last"] = official_close

                            if "checklist_stocks" in data_jsn:
                                data_jsn["checklist_stocks"]["lastPrice"] = official_close
                            _pg_log.info(f"[Pricing Sync] {symbol}: overrode Chaikin price with settled close ${official_close}")
                    else:
                        _pg_log.debug(f"[Pricing Sync] {symbol}: {cache_date_str} not in OHLCV cache; Chaikin price used as-is.")
            except Exception as e:
                _pg_log.warning(f"Failed to reconcile price: {e}")

            symbol_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "Symbol", symbol)
            os.makedirs(symbol_dir, exist_ok=True)
            cache_date = date if date else datetime.date.today()
            # Safeguard: only persist a genuine "ok" payload. A 200 that the adapter
            # classified as "invalid symbol" (or any non-ok status) is a transient/degraded
            # response — caching it would poison the disk cache and re-serve the symbol as
            # invalid on later cache-preferred reads.
            if data_jsn.get("status") != "ok":
                _pg_log.info(f"[Cache Guard] {symbol}: response status={data_jsn.get('status')!r} (not 'ok'); skipping permanent disk-caching.")
            # Safeguard: Do NOT write/save today's temporary intraday price as today's permanent closing cache if NYSE is currently open!
            elif cache_date == datetime.date.today() and is_nyse_market_open():
                _pg_log.info(f"⚡ [Intraday Volatile] Today is an active trading day and NYSE is open. Skipping permanent disk-caching for {symbol} to force EOD sync.")
            else:
                with open(os.path.join(symbol_dir, f"{symbol}_{cache_date}.json"), "w") as fw:
                    json.dump(data_jsn, fw)

        elif response.status_code in (401, 403):
            if _allow_reauth:
                _pg_log.warning(f"HTTP {response.status_code} for {symbol} — triggering session renewal...")
                global _session_valid_until
                _session_valid_until = 0.0
                fresh = ensure_valid_session()
                if fresh and fresh.get("jsessionid"):
                    return get_symbol_data(symbol, date, prefer_cache=False, session_id=fresh, _allow_reauth=False)
            print(SESSION_INSTRUCTIONS.format(session_file=SESSION_FILE))
            raise EnvironmentError(f"Session rejected (HTTP {response.status_code}). Update {SESSION_FILE}.")
        else:
            print(f"Warning: API error for {symbol} (HTTP {response.status_code}) — row will be skipped")
            pg.price = -1
    if data_jsn:
        pg.init_from_json(data_jsn)
        pg.find_prev_pf()
    return pg


def check_from_file(prefer_cache: bool, date=None):
    if date is None:
        date = datetime.date.today()
    elif isinstance(date, datetime.datetime):
        date = date.date()
    _build_cache_index()
    try:
        session_id = login()
    except EnvironmentError as e:
        try:
            session_abs_path = os.path.abspath(SESSION_FILE)
            send_email(
                subject="ALERT: Chaikin Turnstile Block - Manual Auth Required",
                body=f"Chaikin automated session token renewal failed due to browser login timeout/Turnstile challenge.\n\nError: {e}\n\nActions required:\n1. Log in manually at https://app.chaikinanalytics.com in a regular browser.\n2. Extract JSESSIONID from DevTools request headers.\n3. Save JSESSIONID to {session_abs_path}.\n4. Re-run the daily pipeline."
            )
        except Exception as mail_err:
            _pg_log.warning("Failed to send Turnstile block alert email: %s", mail_err)
        raise
    _pg_log.debug("Session loaded", extra={"jsid_prefix": str(session_id)[:12] + "..."})
    syms_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", "symbols_to_check.txt")
    csv_path  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", f"symbols_to_check_{date}.csv")
    
    if not os.path.exists(syms_path):
        _pg_log.warning(f"Symbols file not found: {syms_path}. Skipping file-based check.")
        return
    
    # ── Phase 1: parse and gather valid rows ──
    valid_entries: list[tuple[str, str]] = []
    with open(syms_path, "r") as f:
        for line in f.readlines():
            split_line = line.strip().split()
            if not split_line:
                continue
            symbol = split_line[-1]
            if not _SYMBOL_RE.match(symbol):
                print(f"  [SKIP] invalid symbol format: {symbol!r}")
                continue
            symbol_line = f"{split_line[0]},{symbol}"
            valid_entries.append((symbol_line, symbol))

    unique_syms = list(dict.fromkeys(sym for _, sym in valid_entries))
    print(f"Fetching {len(unique_syms)} unique symbols ({_FETCH_WORKERS} workers)...")

    # ── Phase 2: parallel fetch ──
    pg_results: dict[str, PowerGauge] = {}
    with ThreadPoolExecutor(max_workers=_FETCH_WORKERS) as pool:
        future_to_sym = {
            pool.submit(get_symbol_data, sym, date, prefer_cache): sym
            for sym in unique_syms
        }
        done = 0
        for future in as_completed(future_to_sym):
            sym = future_to_sym[future]
            done += 1
            try:
                pg_results[sym] = future.result()
            except EnvironmentError:
                pool.shutdown(wait=False, cancel_futures=True)
                raise
            except Exception as e:
                print(f"  [{done}/{len(unique_syms)}] {sym}: fetch error — {e}")
                sentinel = PowerGauge(sym, date)
                sentinel.price = -1
                pg_results[sym] = sentinel
    print(f"Fetch complete ({len(unique_syms)} symbols).")

    # ── Phase 3: serial compute + write ──
    ohlcv_cache: dict = {}
    with open(csv_path, "w") as fw:
        for symbol_line, symbol in valid_entries:
            power_g = pg_results[symbol]

            if symbol not in ohlcv_cache:
                ohlcv_path = os.path.join(OHLCV_DIR, f"{symbol}_daily.json")
                try:
                    with open(ohlcv_path) as _f:
                        ohlcv_cache[symbol] = json.load(_f).get('Time Series (Daily)')
                except FileNotFoundError:
                    ohlcv_cache[symbol] = None
                except (json.JSONDecodeError, OSError) as e:
                    print(f"  [OHLCV] {symbol}: could not load {ohlcv_path}: {e}")
                    ohlcv_cache[symbol] = None
            ohlcv_ts = ohlcv_cache[symbol]

            f_fields = _compute_pgr_fields(power_g, ohlcv_ts=ohlcv_ts)

            prev_change = power_g.prevPG.change if power_g.prevPG else ""
            percentage_delta = 0
            percentage_delta_plus = 0

            if ohlcv_ts and power_g.pgr_value > 3:
                all_dates = sorted(ohlcv_ts.keys())
                date_str = str(date)
                past = [d for d in all_dates if d <= date_str]
                if past:
                    idx = all_dates.index(past[-1])
                    prev_count = _ohlcv_streak_count(ohlcv_ts, all_dates, idx - 1, f_fields['prev_percentage']) if idx >= 1 else 0
                    cur_count  = _ohlcv_streak_count(ohlcv_ts, all_dates, idx,     power_g.percentage)
                    if prev_count < 0 and cur_count > 0:
                        percentage_delta = prev_count
                    elif prev_count > 0 and power_g.percentage < 0:
                        percentage_delta = prev_count
                    elif prev_count > 0 and power_g.percentage > 0:
                        percentage_delta_plus = prev_count + 1
                    else:
                        percentage_delta_plus = prev_count - 1

            msg = f"{symbol_line},{power_g.industry_name},{f_fields['prev_pgr']},{f_fields['pgr']},{power_g.industry_strength}," \
                  f"{round(power_g.price*0.95, 2)},{power_g.price},{f_fields['prev_move_price']}," \
                  f"{f_fields['risk_ratio']},{power_g.signals}," \
                  f"{f_fields['prev_percentage']}%,{power_g.percentage}%,{f_fields['prev_move_perc']}%," \
                  f"${prev_change},${power_g.change}," \
                  f"{f_fields['pgr_delta']},{percentage_delta * (-1)},{percentage_delta_plus}," \
                  f"{power_g.lt_trend},{power_g.money_flow},{power_g.over_bt_sl}"

            print(msg)
            fw.write(f"{msg}\n")


# _week_of_month, _compute_seasonality, _predicted_win_pct, _market_regime,
# _rel_volume_bucket, _short_score, _long_score → moved to scoring.py


def _buying_ratio(power_g: PowerGauge, fields: dict) -> float:
    """
    Composite entry-quality score: -10 (strong sell) to +10 (strong buy).

    Components and weights:
      PGR corrected value  ±2.0   (1=Be- → -2, 5=Bu+ → +2)
      Risk/Reward          -1..+2  (rr=0→-1, rr>=3→+2)
      LT trend             ±1.0   (Weak→+1 recovery play, Strong→-1 already extended)
      Money flow           ±0.75  (Strong/Weak)
      OB/OS zone           -0.25..+1.0  (Optimal→+1, Early→+0.25, Wait→-0.25)
      Industry strength    ±0.5   (Weak→+0.5 recovery, Strong→-0.5 extended)
      PGR delta            +0.25  (any change vs yesterday = interesting)
      Seasonality          ±1.0   (week-of-month 10d avg return buckets)

    setup_ok (col U) gates Stop/Target/R-R writes and flows into the Picks sheet;
    excluded from this score because backtesting showed it is a contrarian indicator
    for raw 10d returns (False=+1.36%, True=+0.48%).
    """
    score = 0.0

    # 1. PGR corrected value (1-5)
    pgr_map = {1: -2.0, 2: -1.0, 3: 0.0, 4: 1.0, 5: 2.0}
    score += pgr_map.get(power_g.pgr_corrected_value, 0.0)

    # 2. Risk/Reward ratio (use raw computed value, not sheet-zeroed value)
    rr = fields.get('risk_ratio', 0)
    if rr >= 3.0:
        score += 2.0
    elif rr >= 2.0:
        score += 1.5
    elif rr >= 1.0:
        score += 1.0
    elif rr >= 0.5:
        score += 0.5
    elif rr > 0:
        score += 0.0
    else:
        score -= 1.0   # no valid stop/target = negative signal

    # 4. Long-term trend
    lt = str(power_g.lt_trend or '').strip()
    lt_map = {'Strong': -1.0, 'Neutral': 0.0, 'Weak': 1.0}
    score += lt_map.get(lt, 0.0)

    # 5. Money flow
    mf = str(power_g.money_flow or '').strip()
    mf_map = {'Strong': 0.75, 'Neutral': 0.0, 'Weak': -0.75}
    score += mf_map.get(mf, 0.0)

    # 6. Overbought/Oversold zone
    ob = str(power_g.over_bt_sl or '').strip()
    ob_map = {'Optimal': 1.0, 'Early': 0.25, 'Neutral': 0.0, 'Wait': -0.25}
    score += ob_map.get(ob, 0.0)

    # 7. Industry strength
    ind = str(power_g.industry_strength or '').strip()
    ind_map = {'Strong': -0.5, 'Weak': 0.5}
    score += ind_map.get(ind, 0.0)

    # 8. PGR delta vs yesterday
    delta = fields.get('pgr_delta', 0)
    score += 0.25 if delta != 0 else 0.0

    # 9. Seasonality
    score += fields.get('seasonality', 0.0)

    return round(max(-10.0, min(10.0, score)), 1)


def _append_ohlcv_entry(symbol: str, date_str: str, power_g: "PowerGauge", ohlcv_full: dict | None) -> None:
    """Append today's closing price from Chaikin into the local OHLCV JSON file.

    Uses the full JSON dict already loaded into ohlcv_cache — zero disk re-reads, zero extra
    API calls. Called once per symbol inside check_from_xls() after power_g is fully populated.
    ohlcv_full is the complete JSON (with Meta Data + Time Series), not just the Time Series slice.
    """
    if ohlcv_full is None:
        return  # file missing; rapidapi.repair_missing() will create it

    ts = ohlcv_full.get("Time Series (Daily)")
    if ts is None:
        return

    if date_str in ts:
        return  # already have an entry for this date

    path = os.path.join(OHLCV_DIR, f"{symbol}_daily.json")
    if not os.path.exists(path):
        return

    close = round(power_g.price, 4)
    high  = round(power_g.max_price, 4) if power_g.max_price and power_g.max_price >= close else close
    # Close-only placeholder: Chaikin/PowerGauge exposes no real low/volume. Mark it
    # provisional so the RapidAPI recovery pass (rapidapi._check_recovery) overwrites it
    # with settled OHLCV and volume/range consumers (MFI, RBR) skip it. The close is real.
    entry = {
        "1. open":   str(close),
        "2. high":   str(high),
        "3. low":    str(close),
        "4. close":  str(close),
        "5. volume": "0",
        "provisional": True,
    }

    try:
        ts[date_str] = entry
        ohlcv_full.setdefault("Meta Data", {})["3. Last Refreshed"] = date_str
        tmp = path + ".tmp"
        with open(tmp, "w") as f:
            json.dump(ohlcv_full, f)
        os.replace(tmp, path)
    except Exception as e:
        print(f"  [OHLCV] {symbol}: could not append today's entry: {e}")


def _compute_pgr_fields(power_g: PowerGauge, ohlcv_ts: dict = None) -> dict:
    pgr_value = _pgr_str(power_g.pgr_value)
    pgr_corrected_value = _pgr_str(power_g.pgr_corrected_value)
    pgr = pgr_corrected_value if pgr_corrected_value == pgr_value else f"{pgr_corrected_value}/{pgr_value}"
    prev_pgr = 0
    prev_percentage = 0
    pgr_delta = 0
    prev_move_perc = 0
    prev_move_price = 0
    stop_price = 0
    risk_ratio = 0

    if power_g.prevPG:
        prev_pgr_v = _pgr_str(power_g.prevPG.pgr_value)
        prev_pgr_cv = _pgr_str(power_g.prevPG.pgr_corrected_value)
        prev_pgr = prev_pgr_cv if prev_pgr_cv == prev_pgr_v else f"{prev_pgr_cv}/{prev_pgr_v}"
        prev_percentage = power_g.prevPG.percentage
        pgr_delta = power_g.pgr_corrected_value - power_g.prevPG.pgr_corrected_value

    # Stop, target, streak, SMA filter — all from OHLCV (O(lookback), no chain traversal)
    setup_ok = None
    if ohlcv_ts:
        all_dates = sorted(ohlcv_ts.keys())
        date_str = str(power_g.date)
        past = [d for d in all_dates if d <= date_str]
        if past:
            idx = all_dates.index(past[-1])

            # Stop = confirmed swing-low support; target = confirmed swing-high
            # resistance — the SAME detectors the dashboard uses (risk_utils), so the
            # sheet and the Research page agree. Levels are computed as-of this bar
            # (series up to idx, no look-ahead).
            past_dates = all_dates[:idx + 1]
            _highs  = [_to_float(ohlcv_ts[d].get('2. high'), 0) for d in past_dates]
            _lows   = [_to_float(ohlcv_ts[d].get('3. low'), 0) for d in past_dates]
            _closes = [_to_float(ohlcv_ts[d].get('4. close'), 0) for d in past_dates]
            _excl = instruments.is_excluded(power_g.symbol)   # TEMPORARY: lev/inverse -> ATR
            stop_price = risk_utils.resolve_stop_detailed(
                power_g.price, highs=_highs, lows=_lows, closes=_closes,
                exclude_swing=_excl)["stop"] or 0
            prev_move_price = risk_utils.resolve_target_detailed(
                power_g.price, highs=_highs, lows=_lows, closes=_closes,
                exclude_swing=_excl)["target"] or 0

            # risk/reward
            if power_g.price > 0 and stop_price and prev_move_price and power_g.price > stop_price:
                risk_ratio = round(
                    (prev_move_price - power_g.price) / (power_g.price - stop_price), 2
                )

            # cumulative same-direction streak percentage
            prev_move_perc = _ohlcv_streak_perc(ohlcv_ts, all_dates, idx, power_g.percentage)

            # entry filter: close > SMA(_TREND_SMA_PERIOD) AND close > close[_DIR_CHECK_DAYS ago]
            sma_w = all_dates[max(0, idx - _TREND_SMA_PERIOD): idx]
            if len(sma_w) >= _TREND_SMA_PERIOD:
                sma = sum(_to_float(ohlcv_ts[d].get('4. close'), 0) for d in sma_w) / len(sma_w)
                trend_ok = power_g.price > sma
            else:
                trend_ok = False
            dir_ok = power_g.price > _to_float(ohlcv_ts[all_dates[idx - _DIR_CHECK_DAYS]].get('4. close'), 0) if idx >= _DIR_CHECK_DAYS else False
            setup_ok = trend_ok and dir_ok

    _date_str = str(power_g.date)
    _pattern_score, _pattern_text = _pattern_summary(ohlcv_ts, _date_str)
    _cs  = _cs_score(ohlcv_ts, _date_str) if ohlcv_ts else 0.0
    _cps, _ = _cp_score(ohlcv_ts, _date_str) if ohlcv_ts else (0.0, [])
    _ms,  _ = _mo_score(ohlcv_ts, _date_str) if ohlcv_ts else (0.0, [])
    _vrec, _vrec_names = _rbr_score(ohlcv_ts, _date_str) if ohlcv_ts else (0.0, [])

    fields = {
        'pgr': pgr,
        'prev_pgr': prev_pgr,
        'prev_percentage': prev_percentage,
        'pgr_delta': pgr_delta,
        'prev_move_perc': prev_move_perc,
        'prev_move_price': prev_move_price,
        'stop_price': stop_price,
        'risk_ratio': risk_ratio,
        'setup_ok': setup_ok,      # True/False/None
        'seasonality':    _compute_seasonality(ohlcv_ts, power_g.date.month, power_g.date.day),
        'rel_vol':        _rel_volume_bucket(ohlcv_ts, _date_str),
        'market_regime':  _market_regime(_date_str),
        'fibonacci':      _fib_score(ohlcv_ts, _date_str),
        'rsi_divergence': _rsi_div_score(ohlcv_ts, _date_str),
        # Chaikin signal fields needed by scoring.py functions
        'ob_os':           str(power_g.over_bt_sl      or '').strip(),
        'money_flow':      str(power_g.money_flow       or '').strip(),
        'lt_trend':        str(power_g.lt_trend         or '').strip(),
        'industry_strength': str(power_g.industry_strength or '').strip(),
        # Pattern recognition fields
        'candlestick_score': _cs,
        'chart_score':       _cps,
        'momentum_score':    _ms,
        'pattern_score':     _pattern_score,
        'pattern_text':      _pattern_text,
        # Rubber-Band Reversal — BULLISH, positively-signed (NOT contrarian)
        'vrecovery_score':   _vrec,
        'vrecovery_text':    ' '.join(_vrec_names),
    }
    fields['buying_ratio'] = _buying_ratio(power_g, fields)
    # Digit-sum close→next-day signal: prior close always available from OHLCV.
    # The open→same-day signal is NOT computed here — it requires today's live
    # open price and is applied in _execute_buys() at buy-decision time only.
    _prev_close = None
    if ohlcv_ts and _date_str:
        all_d = sorted(ohlcv_ts.keys())
        idx_d = next((i for i, d in enumerate(all_d) if d >= _date_str), None)
        if idx_d:  # idx_d=0 means first date — no prior close available
            try:
                _prev_close = float(ohlcv_ts[all_d[idx_d - 1]].get('4. close', 0) or 0)
            except (ValueError, TypeError):
                _prev_close = None
    fields['digit_sum'] = _digit_sum_score(power_g.symbol, close_price=_prev_close)
    fields['short_score']  = _short_score_fn(fields)
    fields['long_score']   = _long_score_fn(fields)
    return fields


def check_from_xls(prefer_cache: bool, date=None, symbols=None):
    """Update Research sheet from PowerGauge data.

    symbols: optional list/set of ticker strings — process only those rows.
             Pass None (default) to process all rows.
    Fetches are parallelised (_FETCH_WORKERS threads); cell writes remain serial.
    """
    if date is None:
        date = datetime.date.today()
    elif isinstance(date, datetime.datetime):
        date = date.date()
    import openpyxl
    _build_cache_index()
    _orig_backup = _backup_xlsx(XLSX_FILE)
    try:
        session_id = login()
    except EnvironmentError as e:
        try:
            session_abs_path = os.path.abspath(SESSION_FILE)
            send_email(
                subject="ALERT: Chaikin Turnstile Block - Manual Auth Required",
                body=f"Chaikin automated session token renewal failed due to browser login timeout/Turnstile challenge.\n\nError: {e}\n\nActions required:\n1. Log in manually at https://app.chaikinanalytics.com in a regular browser.\n2. Extract JSESSIONID from DevTools request headers.\n3. Save JSESSIONID to {session_abs_path}.\n4. Re-run the daily pipeline."
            )
        except Exception as mail_err:
            _pg_log.warning("Failed to send Turnstile block alert email: %s", mail_err)
        raise
    _pg_log.debug("Session loaded", extra={"jsid_prefix": str(session_id)[:12] + "..."})

    # A full run (symbols=None) rebuilds from the ROOT source of truth. A targeted
    # run (symbols=[...]) merges onto the existing OUTPUT so the computed scores of
    # symbols NOT being re-screened are preserved (loading the score-less source
    # would wipe them).
    import os as _os
    base_path = XLSX_FILE if (symbols and _os.path.exists(XLSX_FILE)) else SRC_XLSX
    try:
        wb = openpyxl.load_workbook(base_path)
    except Exception as e:
        alt = SRC_XLSX if base_path == XLSX_FILE else XLSX_FILE
        print(f"  [ERROR] Failed to load {base_path}: {e}")
        print(f"  [INFO] Attempting to load {alt} instead...")
        try:
            wb = openpyxl.load_workbook(alt)
        except Exception:
            print(f"  [FATAL] Both source and output files missing or corrupt.")
            return
    
    ws = wb['Research']
    _write_research_headers(ws)

    filter_set = {s.upper() for s in symbols} if symbols else None

    # ── Phase 1: collect valid (symbol, row) pairs in sheet order ────────────
    valid_rows: list[tuple[str, tuple]] = []
    for row in ws.iter_rows(min_row=2, max_col=26):
        symbol = row[3].value
        if not symbol:
            continue
        symbol = str(symbol).strip()
        if not symbol:
            continue
        if filter_set and symbol.upper() not in filter_set:
            continue
        if not _SYMBOL_RE.match(symbol):
            print(f"  [SKIP] invalid symbol format: {symbol!r}")
            continue
        valid_rows.append((symbol, row))

    total = len(valid_rows)
    unique_syms = list(dict.fromkeys(s for s, _ in valid_rows))
    print(f"Fetching {len(unique_syms)} unique symbols ({_FETCH_WORKERS} workers)...")

    # ── Phase 2: parallel fetch ───────────────────────────────────────────────
    pg_results: dict[str, PowerGauge] = {}
    with ThreadPoolExecutor(max_workers=_FETCH_WORKERS) as pool:
        future_to_sym = {
            pool.submit(get_symbol_data, sym, date, prefer_cache): sym
            for sym in unique_syms
        }
        done = 0
        for future in as_completed(future_to_sym):
            sym = future_to_sym[future]
            done += 1
            try:
                pg_results[sym] = future.result()
            except EnvironmentError:
                pool.shutdown(wait=False, cancel_futures=True)
                raise
            except Exception as e:
                print(f"  [{done}/{len(unique_syms)}] {sym}: fetch error — {e}")
                sentinel = PowerGauge(sym, date)
                sentinel.price = -1
                pg_results[sym] = sentinel
    print(f"Fetch complete ({len(unique_syms)} symbols).")

    # ── Phase 3: serial compute + write ──────────────────────────────────────
    updated = 0
    skipped = 0
    picks_data: list[dict] = []
    ohlcv_cache: dict = {}  # symbol → full JSON dict (Meta Data + Time Series)

    for n, (symbol, row) in enumerate(valid_rows, 1):
        power_g = pg_results[symbol]

        if power_g.price == -1:
            print(f"[{n}/{total}] {symbol}: no market data - row skipped (existing values preserved)")
            skipped += 1
            continue

        # Load full OHLCV JSON — cached per symbol; _append_ohlcv_entry reuses it to avoid re-read
        if symbol not in ohlcv_cache:
            ohlcv_path = os.path.join(OHLCV_DIR, f"{symbol}_daily.json")
            try:
                with open(ohlcv_path) as _f:
                    ohlcv_cache[symbol] = json.load(_f)
            except FileNotFoundError:
                ohlcv_cache[symbol] = None
            except (json.JSONDecodeError, OSError) as e:
                print(f"  [OHLCV] {symbol}: could not load {ohlcv_path}: {e}")
                ohlcv_cache[symbol] = None
        ohlcv_ts = ohlcv_cache[symbol].get('Time Series (Daily)') if ohlcv_cache[symbol] else None

        # --- Closing Price Override (Pre-Computation) ---
        # If we have an official OHLCV close on disk, overwrite power_g.price immediately.
        # This ensures all trend, support/resistance, stop, target, and Reward-to-Risk
        # calculations are computed using the actual correct settled close.
        today_str = date.strftime("%Y-%m-%d")
        if ohlcv_ts and today_str in ohlcv_ts:
            try:
                official_close = float(ohlcv_ts[today_str]['4. close'])
                if official_close > 0.0:
                    power_g.price = official_close
            except (ValueError, KeyError):
                pass

        f = _compute_pgr_fields(power_g, ohlcv_ts=ohlcv_ts)
        setup_ok = f['setup_ok']   # True / False / None

        final_price = power_g.price

        if power_g.industry_name:
            row[4].value = power_g.industry_name
        row[5].value = f['prev_pgr']
        row[6].value = f['pgr']
        row[7].value = power_g.industry_strength
        # row[8] col I: manual price level - preserved
        # J=stop, L=target, M=risk: written for ALL symbols (not just setups) so the
        # sheet carries a stop everywhere. stop_price is the swing-low technical stop;
        # it is 0 only when no recent low sits below price — the dashboard fills those
        # via risk_utils.resolve_stop (ATR -> 8%). See resolve_stop for the full ladder.
        row[9].value  = f['stop_price']                                     # col J
        row[10].value = final_price                                         # col K (Overridden)
        row[11].value = f['prev_move_price']                                # col L
        row[12].value = f['risk_ratio']                                     # col M
        row[13].value = f['prev_move_perc']
        row[14].value = f['prev_percentage']
        row[15].value = power_g.percentage
        # row[16] col Q: notes/category - preserved
        row[17].value = power_g.lt_trend
        row[18].value = power_g.money_flow
        row[19].value = power_g.over_bt_sl
        # col U: entry filter flag (1=valid, 0=filtered out, blank=unknown)
        row[20].value = (1 if setup_ok else 0) if setup_ok is not None else None
        # col V: buying ratio -10..+10
        row[21].value = f['buying_ratio']
        # col W: seasonality score for current month (-1..+1)
        row[22].value = f['seasonality'] if f['seasonality'] != 0.0 else None
        # col X: predicted 10d win% from backtest lookup
        row[23].value = _predicted_win_pct(f['buying_ratio'])
        # col Y: short-term 10d entry score
        row[24].value = f['short_score']
        # col Z: long-term 60d position score
        row[25].value = f['long_score']
        # col AA: pattern recognition summary text
        ws.cell(row[0].row, 27).value = f.get('pattern_text') or None

        # Append today's close from Chaikin into Symbol_full OHLCV JSON (free — no extra API call)
        if power_g.price > 0:
            _append_ohlcv_entry(symbol, today_str, power_g, ohlcv_cache.get(symbol))

        picks_data.append({
            'symbol':   symbol,
            'industry': power_g.industry_name or '',
            'pgr':      f['pgr'],
            'price':    power_g.price,
            'setup':    (1 if setup_ok else 0) if setup_ok is not None else None,
            'br':       f['buying_ratio'],
            'short10':  f['short_score'],
            'long60':   f['long_score'],
            'ob_os':    str(power_g.over_bt_sl or '').strip(),
            'money_fl': str(power_g.money_flow  or '').strip(),
            'lt_trend': str(power_g.lt_trend    or '').strip(),
            'regime':   f['market_regime'],
            'stop':     f['stop_price'],
            'target':   f['prev_move_price'],
        })

        flag = "OK" if setup_ok else ("--" if setup_ok is False else "??")
        print(f"[{n}/{total}] {symbol}: pgr={f['pgr']}, price={power_g.price}, "
              f"stop={f['stop_price']}, target={f['prev_move_price']}, "
              f"rr={f['risk_ratio']}, setup={flag}, br={f['buying_ratio']}, "
              f"s10={f['short_score']}, l60={f['long_score']}")
        updated += 1

    if picks_data:
        _write_picks_sheet(wb, picks_data, date)

    _touched_sheets = {"Research", "Picks"}
    try:
        from aether import etrade as _et
        # Load cached tokens directly to avoid Playwright/MFA interactive prompts
        _cached = _et._load_tokens("production")
        _tok = None
        if _cached:
            _tok = _et.renew_tokens(_cached, "production")
            
        if _tok:
            _lk   = {p["symbol"]: p for p in picks_data}
            _pos  = _et.fetch_positions(_tok, "production")
            _syms = list({p["symbol"] for p in _pos})
            _qts  = _et.fetch_quotes(_tok, _syms, "production")
            _update_short_long_scores(wb, _lk, _qts, _pos, ohlcv_cache)
            _touched_sheets.add("Short_Long")
            print(f"Short_Long sheet synced: {len(_pos)} positions.")
        else:
            print("[E*TRADE] Short_Long skipped (no valid silent token session available).")
    except Exception as _e:
        print(f"[E*TRADE] Short_Long skipped: {_e}")

    if picks_data:
        _update_replacements_sheet(wb, picks_data, date.date() if hasattr(date, "date") else date)
        _touched_sheets.add("Replacements")

    try:
        wb.save(XLSX_FILE)
        _fix_comment_shape_ids(XLSX_FILE,
                               original_xlsx=_orig_backup,
                               touched_sheet_names=_touched_sheets)
        print(f"Research sheet updated ({updated} rows written, {skipped} skipped) -> {XLSX_FILE}")
    except PermissionError:
        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        alt = os.path.join(os.path.dirname(XLSX_FILE), f"investment_pending_{ts}.xlsx")
        wb.save(alt)
        _fix_comment_shape_ids(alt,
                               original_xlsx=_orig_backup,
                               touched_sheet_names=_touched_sheets)
        print(f"ERROR: {XLSX_FILE} is open in another application.")
        print(f"Changes saved to: {alt}")
        print(f"Close Excel and rename/copy that file to state_of_the_day.xlsx")
