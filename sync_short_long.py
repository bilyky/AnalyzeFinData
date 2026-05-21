"""Standalone entry point: sync Short_Long sheet with live E*TRADE positions.

Run:
    python sync_short_long.py              # production account
    python sync_short_long.py --sandbox    # sandbox account

Does NOT require a full check_from_xls run. Uses cached production tokens
(or runs browser auth if not cached). Picks scores come from the last
Research sheet data already in the workbook.
"""

import argparse
import datetime
import openpyxl

import etrade
from excel_output import update_short_long_scores, fix_comment_shape_ids, backup_xlsx
from powergauge import XLSX_FILE


def _read_picks_from_research(wb) -> dict:
    """Build a picks_lookup dict from the current Research sheet data."""
    if "Research" not in wb.sheetnames:
        return {}
    ws     = wb["Research"]
    lookup = {}

    def _g(row, idx, default=None):
        return row[idx] if len(row) > idx else default

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        sym = str(_g(row, 3) or "").strip().upper()   # col D
        if not sym:
            continue
        lookup[sym] = {
            "symbol":   sym,
            "industry": str(_g(row, 4) or "").strip(),    # col E
            "pgr":      _g(row, 6),                        # col G
            "br":       _g(row, 21),                       # col V
            "short10":  _g(row, 24),                       # col Y
            "long60":   _g(row, 25),                       # col Z
            "ob_os":    str(_g(row, 19) or "").strip(),    # col T
            "money_fl": str(_g(row, 18) or "").strip(),    # col S
            "lt_trend": str(_g(row, 17) or "").strip(),    # col R
        }
    return lookup


def main():
    parser = argparse.ArgumentParser(description="Sync Short_Long sheet with E*TRADE")
    parser.add_argument("--sandbox", action="store_true", help="Use sandbox environment")
    args = parser.parse_args()
    env = "sandbox" if args.sandbox else "production"

    print(f"[sync_short_long] env={env}")

    # ── Get tokens ───────────────────────────────────────────────────────────
    tokens = etrade.get_tokens(env)

    # ── Fetch positions + quotes ──────────────────────────────────────────────
    print("Fetching E*TRADE positions...")
    positions = etrade.fetch_positions(tokens, env)
    print(f"  {len(positions)} open positions found.")

    syms   = list({p["symbol"] for p in positions})
    quotes = etrade.fetch_quotes(tokens, syms, env)
    print(f"  {len(quotes)} live quotes fetched.")

    # ── Load workbook + build picks lookup from Research sheet ───────────────
    wb = openpyxl.load_workbook(XLSX_FILE)
    picks_lookup = _read_picks_from_research(wb)
    print(f"  {len(picks_lookup)} symbols in Research sheet for score lookup.")

    # ── Sync Short_Long sheet ────────────────────────────────────────────────
    orig_backup = backup_xlsx(XLSX_FILE)
    update_short_long_scores(wb, picks_lookup, quotes, positions)

    try:
        wb.save(XLSX_FILE)
        fix_comment_shape_ids(XLSX_FILE,
                              original_xlsx=orig_backup,
                              touched_sheet_names={"Short_Long"})
        print(f"Saved -> {XLSX_FILE}")
    except PermissionError:
        import os
        ts  = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        alt = os.path.join(os.path.dirname(XLSX_FILE), f"investment_sl_{ts}.xlsx")
        wb.save(alt)
        fix_comment_shape_ids(alt,
                              original_xlsx=orig_backup,
                              touched_sheet_names={"Short_Long"})
        print(f"File was open — saved to {alt}")


if __name__ == "__main__":
    main()
