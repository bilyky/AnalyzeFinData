"""
Unit tests for update_short_long_scores — specifically the sell/rebuy detection
and buy date handling.  No E*TRADE API calls; uses an in-memory openpyxl workbook.
"""
import sys
import os
import unittest
import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from excel_output import update_short_long_scores


def _make_wb(rows: list[dict]) -> openpyxl.Workbook:
    """Build a minimal Short_Long workbook.

    Each dict in rows: {sym, qty, cost, price, buy_date (date object)}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Short_Long"
    # Header row expected by update_short_long_scores (col B = "Symb")
    ws["B1"] = "Symb"
    for i, r in enumerate(rows, start=3):
        ws.cell(i, 1).value = i - 2        # rank
        ws.cell(i, 2).value = r["sym"]
        ws.cell(i, 3).value = r["qty"]
        ws.cell(i, 4).value = r["cost"]
        ws.cell(i, 5).value = r["price"]
        ws.cell(i, 11).value = r["buy_date"]
    return wb


def _sym_rows(ws) -> dict[str, int]:
    """Return {symbol: row_number} for data rows in the sheet."""
    out = {}
    for row in ws.iter_rows(min_row=3, max_row=50):
        v = row[1].value
        if isinstance(v, str) and v.strip() and v.strip().upper() != "SYMB":
            sym = v.strip().upper()
            if sym not in out:
                out[sym] = row[0].row
    return out


def _buy_date(ws, row_num) -> datetime.date | None:
    val = ws.cell(row_num, 11).value
    if isinstance(val, datetime.datetime):
        return val.date()
    return val if isinstance(val, datetime.date) else None


class TestSellRebuyDetection(unittest.TestCase):

    def test_same_day_rebuy_replaces_row_and_resets_buy_date(self):
        """Symbol sold and re-bought today: old row gone, new row with today's date."""
        old_date = datetime.date(2025, 1, 10)
        new_date = datetime.date.today()

        wb = _make_wb([{"sym": "AAPL", "qty": 10, "cost": 150.0,
                        "price": 175.0, "buy_date": old_date}])
        ws = wb["Short_Long"]

        positions = [{"symbol": "AAPL", "qty": 10, "cost": 182.0,
                      "price": 182.0, "mval": 1820.0, "date_acquired": new_date,
                      "account_last4": "0053"}]

        update_short_long_scores(wb, {}, {}, positions)

        rows = _sym_rows(ws)
        self.assertIn("AAPL", rows, "AAPL should still be in sheet after rebuy")
        self.assertEqual(_buy_date(ws, rows["AAPL"]), new_date,
                         "Buy date should be updated to new acquisition date")

    def test_unchanged_position_preserves_buy_date(self):
        """Symbol held continuously: buy date must not change."""
        old_date = datetime.date(2025, 1, 10)

        wb = _make_wb([{"sym": "MSFT", "qty": 5, "cost": 400.0,
                        "price": 420.0, "buy_date": old_date}])
        ws = wb["Short_Long"]

        positions = [{"symbol": "MSFT", "qty": 5, "cost": 400.0,
                      "price": 425.0, "mval": 2125.0, "date_acquired": old_date,
                      "account_last4": "0053"}]

        update_short_long_scores(wb, {}, {}, positions)

        rows = _sym_rows(ws)
        self.assertIn("MSFT", rows)
        self.assertEqual(_buy_date(ws, rows["MSFT"]), old_date,
                         "Buy date must not change for a held position")

    def test_closed_position_is_removed(self):
        """Symbol no longer in E*TRADE positions → row deleted."""
        wb = _make_wb([{"sym": "GOOG", "qty": 2, "cost": 170.0,
                        "price": 180.0, "buy_date": datetime.date(2025, 3, 1)}])

        update_short_long_scores(wb, {}, {}, positions=[])

        rows = _sym_rows(wb["Short_Long"])
        self.assertNotIn("GOOG", rows, "Closed position should be removed")

    def test_brand_new_position_uses_etrade_date(self):
        """New symbol from E*TRADE uses dateAcquired, not today's date."""
        acq = datetime.date(2026, 5, 19)   # 2 days ago
        wb = _make_wb([])   # empty sheet

        positions = [{"symbol": "NVDA", "qty": 3, "cost": 900.0,
                      "price": 910.0, "mval": 2730.0, "date_acquired": acq,
                      "account_last4": "0053"}]

        update_short_long_scores(wb, {}, {}, positions)

        rows = _sym_rows(wb["Short_Long"])
        self.assertIn("NVDA", rows)
        self.assertEqual(_buy_date(wb["Short_Long"], rows["NVDA"]), acq,
                         "Buy date should match E*TRADE dateAcquired, not today")

    def test_no_etrade_date_falls_back_to_today(self):
        """If E*TRADE returns no dateAcquired, new row gets today's date."""
        today = datetime.date.today()
        wb = _make_wb([])

        positions = [{"symbol": "AMD", "qty": 5, "cost": 120.0,
                      "price": 125.0, "mval": 625.0, "date_acquired": None,
                      "account_last4": "0053"}]

        update_short_long_scores(wb, {}, {}, positions)

        rows = _sym_rows(wb["Short_Long"])
        self.assertIn("AMD", rows)
        self.assertEqual(_buy_date(wb["Short_Long"], rows["AMD"]), today,
                         "Buy date should fall back to today when dateAcquired is absent")


if __name__ == "__main__":
    unittest.main(verbosity=2)
