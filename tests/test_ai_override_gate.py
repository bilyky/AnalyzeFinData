"""
Dedicated unit tests for the AI Second-Opinion Exit Override Gate in ai_portfolio_game.py.
"""
import os
import sys
import unittest
import openpyxl
from unittest import mock

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import ai_portfolio_game as game

class TestAIOverrideGate(unittest.TestCase):
    def setUp(self):
        game._HEAL_ATTEMPTED.clear()
        self.ai_eval_patcher = mock.patch("ai_client.evaluate", return_value="")
        self.mock_ai_eval = self.ai_eval_patcher.start()

    def tearDown(self):
        self.ai_eval_patcher.stop()

    @mock.patch("ai_portfolio_game.is_market_hours", return_value=True)
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    def test_sell_overridden_and_downgraded_to_watch(self, mock_load_wb, mock_save_game, mock_load_game, mock_get_prices, mock_market_hours):
        state = {
            "balance": 5000.0,
            "equity": 10000.0,
            "positions": {
                "ULTA": {
                    "qty": 3,
                    "cost": 469.56,
                    "stop_loss": 400.0,
                    "shadow_verdict": "FLAG-FOR-REVIEW",
                }
            },
            "queued_orders": [],
            "history": []
        }
        mock_load_game.return_value = state
        mock_get_prices.return_value = {"ULTA": 450.0}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        ws.append([1, None, None, "ULTA", "Retail", None, "Bu", None, None, None, 450.0, None, None, None, None, None, None, None, None, None, "0", None, None, 0.65, -5.0, -5.0])
        mock_load_wb.return_value = wb

        game.run_daily_ai_management(force=True, manual_profile="BALANCED")

        self.assertIn("ULTA", state["positions"])
        self.assertEqual(len(state["queued_orders"]), 0)

    @mock.patch("ai_portfolio_game.is_market_hours", return_value=True)
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    def test_sell_overridden_and_downgraded_to_hold(self, mock_load_wb, mock_save_game, mock_load_game, mock_get_prices, mock_market_hours):
        state = {
            "balance": 5000.0,
            "equity": 10000.0,
            "positions": {
                "GMED": {
                    "qty": 20,
                    "cost": 74.80,
                    "stop_loss": 70.0,
                    "shadow_verdict": {"verdict": "HOLD", "note": "Strong structural reasons to hold"},
                }
            },
            "queued_orders": [],
            "history": []
        }
        mock_load_game.return_value = state
        mock_get_prices.return_value = {"GMED": 72.0}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        ws.append([1, None, None, "GMED", "Medical", None, "Bu", None, None, None, 72.0, None, None, None, None, None, None, None, None, None, "0", None, None, 0.65, -3.0, -3.0])
        mock_load_wb.return_value = wb

        game.run_daily_ai_management(force=True, manual_profile="BALANCED")

        self.assertIn("GMED", state["positions"])
        self.assertEqual(len(state["queued_orders"]), 0)

    @mock.patch("ai_portfolio_game.is_market_hours", return_value=True)
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    def test_sell_executes_without_override(self, mock_load_wb, mock_save_game, mock_load_game, mock_get_prices, mock_market_hours):
        """Red-path: a SELL with no AI override verdict must actually execute (position removed)."""
        state = {
            "balance": 5000.0,
            "equity": 10000.0,
            "positions": {
                "TSCO": {
                    "qty": 10,
                    "cost": 100.0,
                    "stop_loss": 80.0,
                    # No shadow_verdict — override gate must not fire
                }
            },
            "queued_orders": [],
            "history": []
        }
        mock_load_game.return_value = state
        # Price is above stop but well below cost with strong negative scores → SELL
        mock_get_prices.return_value = {"TSCO": 90.0}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        # S10=-8.0, L60=-8.0 — deep negative momentum, no winner protection (price < cost)
        ws.append([1, None, None, "TSCO", "Retail", None, "Wk", None, None, None, 90.0, None, None, None, None, None, None, None, None, None, "0", None, None, 0.30, -8.0, -8.0])
        mock_load_wb.return_value = wb

        game.run_daily_ai_management(force=True, manual_profile="BALANCED")

        sold = (
            "TSCO" not in state["positions"]
            or any(t.get("type") == "SELL" and t.get("symbol") == "TSCO" for t in state.get("history", []))
        )
        self.assertTrue(sold, "SELL without AI override should execute — position was not removed.")

if __name__ == "__main__":
    unittest.main()
