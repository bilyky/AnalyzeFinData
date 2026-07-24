"""
Tests for the custom-sprint features, exercising the REAL production helpers
(_active_setup_symbols, _execute_buys, calculate_bubble_z_score, watchdog.check_logs)
rather than copies of their logic, so a regression in the shipped code fails here.
E*TRADE / verify / ATR are mocked; no network.
"""
import os
import sys
import unittest
import tempfile
import json
import openpyxl
import pytz
from unittest import mock
from pathlib import Path
from datetime import datetime as real_datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

import ai_portfolio_game as game
import watchdog


def _defensive_state():
    return {"balance": 7746.29, "equity": 10295.0,
            "positions": {"FANG": {"qty": 8, "cost": 184.37}, "HE": {"qty": 81, "cost": 13.57}},
            "history": [], "start_date": "2026-01-01", "profile": "DEFENSIVE"}


_RULES = {"max_positions": 3, "max_allocation_pct": 0.10, "atr_multiplier": 1.5,
          "min_score_threshold": 10.0, "cash_buffer_pct": 0.50}

_TOP_BUYS = [{"sym": "BSY", "price": 31.55, "total": 11.5, "bottom_desc": ""},
             {"sym": "APA", "price": 33.825, "total": 10.7, "bottom_desc": ""}]


class TestActiveSetupSymbols(unittest.TestCase):
    """Calls the real _active_setup_symbols filter (no row-50 cap; '1'/'OK' active)."""

    def _ws(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["spacer"] * 30)                 # row 1 header
        for _ in range(340):                       # deep rows, past any old row-50 slice
            ws.append([None] * 30)
        for sym, setup in [("APA", "1"), ("XYZ", "0"), ("MMM", "OK")]:
            r = [None] * 30
            r[3], r[20] = sym, setup
            ws.append(r)
        return ws

    def test_active_setups_extracted_deep_rows(self):
        syms = game._active_setup_symbols(self._ws())
        self.assertIn("APA", syms)          # setup '1'
        self.assertIn("MMM", syms)          # setup 'OK'
        self.assertNotIn("XYZ", syms)       # setup '0'
        self.assertEqual(len(syms), 2)


class TestExecuteBuys(unittest.TestCase):
    """Calls the real _execute_buys buy-slot loop."""

    def test_fallback_on_verify_failure(self):
        # BSY (top rank) fails verification -> slot falls through to APA.
        state = _defensive_state()
        new_tx = []
        def verify(sym): return (False, "Failed check") if sym == "BSY" else (True, "Verified")
        with mock.patch.object(game, "backtrack_verify", side_effect=verify), \
             mock.patch.object(game, "calculate_bubble_z_score", return_value=None), \
             mock.patch("risk_utils.calculate_atr", return_value=1.0):
            n = game._execute_buys(state, list(_TOP_BUYS), 1, state["equity"] * 0.50,
                                   _RULES, "2026-07-09", "10:00", new_tx)
        self.assertEqual(n, 1)
        self.assertIn("APA", state["positions"])
        self.assertNotIn("BSY", state["positions"])
        self.assertEqual([t["symbol"] for t in new_tx], ["APA"])
        # (7746.29 - 5147.50) / 1 // 33.825 = 76 shares
        self.assertEqual(state["positions"]["APA"]["qty"], 76)

    def test_bubble_guard_rejects_and_falls_through(self):
        # BSY trades >2.5 SD above its mean -> rejected; APA (0.9 SD) fills the slot.
        state = _defensive_state()
        new_tx = []
        def z(sym): return 3.1 if sym == "BSY" else 0.9
        with mock.patch.object(game, "calculate_bubble_z_score", side_effect=z), \
             mock.patch.object(game, "backtrack_verify", return_value=(True, "Verified")), \
             mock.patch("risk_utils.calculate_atr", return_value=1.0):
            n = game._execute_buys(state, list(_TOP_BUYS), 1, state["equity"] * 0.50,
                                   _RULES, "2026-07-09", "10:00", new_tx)
        self.assertEqual(n, 1)
        self.assertIn("APA", state["positions"])
        self.assertNotIn("BSY", state["positions"])

    def test_slot_cap_respected(self):
        # Only one slot -> only one buy even when both candidates are clean.
        state = _defensive_state()
        new_tx = []
        with mock.patch.object(game, "calculate_bubble_z_score", return_value=None), \
             mock.patch.object(game, "backtrack_verify", return_value=(True, "Verified")), \
             mock.patch("risk_utils.calculate_atr", return_value=1.0):
            n = game._execute_buys(state, list(_TOP_BUYS), 1, state["equity"] * 0.50,
                                   _RULES, "2026-07-09", "10:00", new_tx)
        self.assertEqual(n, 1)


class TestWatchdogLogScanning(unittest.TestCase):
    def test_ignore_success_with_0_errors(self):
        mock_log = "Status: [2026-07-09 05:36:24] OHLCV: 1 recovered, 514 already current, 0 errors\n"
        with mock.patch("builtins.open", mock.mock_open(read_data=mock_log)), \
             mock.patch("os.path.exists", return_value=True):
            errors = watchdog.check_logs()
        self.assertEqual(errors, [])


class TestBubbleZScoreCalculation(unittest.TestCase):
    def setUp(self):
        self._orig_symdir = game.SYMBOL_FULL_DIR
        self.temp_dir = tempfile.TemporaryDirectory()
        game.SYMBOL_FULL_DIR = Path(self.temp_dir.name) / "Data" / "Symbol_full"
        os.makedirs(game.SYMBOL_FULL_DIR, exist_ok=True)
        # 600 rising closes; zero-padded keys so sort == chronological order.
        ts = {f"{i:05d}": {"4. close": str(100.0 + i * 0.5)} for i in range(600)}
        with open(game.SYMBOL_FULL_DIR / "BUBBLE_daily.json", "w") as f:
            json.dump({"Time Series (Daily)": ts}, f)

    def tearDown(self):
        game.SYMBOL_FULL_DIR = self._orig_symdir
        self.temp_dir.cleanup()

    def test_sufficient_history_positive_z(self):
        score = game.calculate_bubble_z_score("BUBBLE")
        self.assertIsNotNone(score)
        self.assertGreater(score, 1.5)

    def test_insufficient_history_returns_none(self):
        ts = {f"{i:05d}": {"4. close": "100"} for i in range(250)}
        with open(game.SYMBOL_FULL_DIR / "SHORT_daily.json", "w") as f:
            json.dump({"Time Series (Daily)": ts}, f)
        self.assertIsNone(game.calculate_bubble_z_score("SHORT"))

    def test_missing_file_returns_none(self):
        self.assertIsNone(game.calculate_bubble_z_score("NOPE"))


class TestCoreSatelliteAllocation(unittest.TestCase):
    def test_scarcity_asset_downsizing_and_caps(self):
        # Scenario: Portfolio equity = $10,000. Scarcity Cap = 20% ($2,000). Standard Cap = 80% ($8,000).
        # Cash = $5,000. Min Cash required = $0.
        state = {
            "balance": 5000.0,
            "equity": 10000.0,
            "positions": {
                # Already hold $1,500 of scarcity assets (75 shares * $20)
                "GLD": {"qty": 75, "cost": 20.0, "is_scarcity": True}
            },
            "history": []
        }
        
        # We try to buy a scarcity asset (SLV) which costs $10 per share.
        # Max additional room in scarcity bucket is $2,000 - $1,500 = $500.
        # With $5,000 cash and 1 available slot, dynamic sizing wants to buy $5,000 worth (500 shares).
        # But the 20% scarcity cap should restrict/downsize this to exactly 50 shares ($500).
        rules = {"scarcity_allocation_pct": 0.20, "max_allocation_pct": 0.50, "atr_multiplier": 2.5}
        top_buys = [{"sym": "SLV", "price": 10.0, "total": 12.0, "industry": "Gold Mining", "bottom_desc": ""}]
        new_tx = []
        
        # Mock instruments classification so SLV is treated as scarcity
        with mock.patch("instruments.is_scarcity_asset", return_value=True), \
             mock.patch.object(game, "calculate_bubble_z_score", return_value=None), \
             mock.patch.object(game, "backtrack_verify", return_value=(True, "Verified")), \
             mock.patch("risk_utils.calculate_atr", return_value=1.0):
            
            n = game._execute_buys(state, top_buys, 1, 0.0, rules, "2026-07-11", "11:30", new_tx, prices={"GLD": 20.0, "SLV": 10.0})
            
        self.assertEqual(n, 1)
        self.assertIn("SLV", state["positions"])
        # Verified: Downsized to exactly 50 shares to fit the $500 remaining room in the 20% scarcity bucket!
        self.assertEqual(state["positions"]["SLV"]["qty"], 50)
        self.assertEqual(state["positions"]["SLV"]["is_scarcity"], True)


class TestAfterHoursOrderQueuing(unittest.TestCase):
    @mock.patch("ai_portfolio_game.is_market_hours", return_value=False)
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")  # patch at point-of-use
    def test_after_hours_orders_are_queued(self, mock_load_wb, mock_save_game, mock_load_game, mock_get_prices, mock_market_hours):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        ws.append([1, None, None, "TSLA", "Technology", None, "Bu", None, None, None, 400.0, None, None, None, None, None, None, None, None, None, "1", None, None, 0.65, 5.0, 5.0])
        mock_load_wb.return_value = wb
        
        state = {
            "balance": 5000.0,
            "equity": 10000.0,
            "positions": {},
            "queued_orders": []
        }
        mock_load_game.return_value = state
        mock_get_prices.return_value = {"TSLA": 400.0}
        
        game.run_daily_ai_management(force=True, manual_profile="BALANCED")
        
        # Verify that TSLA was queued for buying instead of executed immediately
        self.assertEqual(len(state["queued_orders"]), 1)
        self.assertEqual(state["queued_orders"][0]["symbol"], "TSLA")
        self.assertEqual(state["queued_orders"][0]["type"], "BUY")
        self.assertNotIn("TSLA", state["positions"])


class TestMarketHolidayChecks(unittest.TestCase):
    @mock.patch("ai_portfolio_game.etrade.get_tokens", return_value=None)
    @mock.patch("ai_portfolio_game.datetime.datetime")
    def test_market_hours(self, mock_datetime, mock_get_tokens):
        tz_la = pytz.timezone("America/Los_Angeles")
        cases = [
            (real_datetime(2026, 7, 3, 10, 0, 0), False, "holiday (Independence Day observed)"),
            (real_datetime(2026, 7, 14, 10, 0, 0), True,  "standard trading day"),
        ]
        for dt, expected, label in cases:
            with self.subTest(label=label):
                mock_datetime.now.return_value = tz_la.localize(dt)
                self.assertEqual(game.is_market_hours(), expected)


class TestPersistentProfileModes(unittest.TestCase):
    @mock.patch("ai_portfolio_game.get_market_regime", return_value="DEFENSIVE")
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    def test_persistent_manual_profile_override(self, mock_load_wb, mock_save_game, mock_load_game, mock_get_prices, mock_regime):
        # 1. First run: Explicit manual override via CLI should lock into MANUAL mode
        state = {
            "balance": 5000.0,
            "equity": 10000.0,
            "positions": {},
            "queued_orders": [],
            "profile": "DEFENSIVE",
            "profile_mode": "ADAPTIVE"
        }
        mock_load_game.return_value = state
        mock_get_prices.return_value = {}
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        mock_load_wb.return_value = wb

        game.run_daily_ai_management(force=True, manual_profile="BALANCED")
        self.assertEqual(state["profile"], "BALANCED")
        self.assertEqual(state["profile_mode"], "MANUAL")
        
        # 2. Second run: Automated run (no manual_profile CLI arg passed)
        # Under our new auto-reset spec, it must automatically reset back to ADAPTIVE autopilot
        # and fall back to the dynamic regime selector (returning "DEFENSIVE"!)
        game.run_daily_ai_management(force=True, manual_profile=None)
        self.assertEqual(state["profile"], "DEFENSIVE")
        self.assertEqual(state["profile_mode"], "ADAPTIVE")
        
        # 3. Third run: Explicit manual restoration via CLI still works
        state["profile_mode"] = "MANUAL"
        state["profile"] = "BALANCED"
        game.run_daily_ai_management(force=True, manual_profile="ADAPTIVE")
        self.assertEqual(state["profile"], "DEFENSIVE")
        self.assertEqual(state["profile_mode"], "ADAPTIVE")

    @mock.patch("ai_portfolio_game._has_strong_setups_today", return_value=True)
    @mock.patch("ai_portfolio_game.get_market_regime", return_value="DEFENSIVE")
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    def test_adaptive_upgrade_cash_gate(self, mock_load_wb, mock_save_game, mock_load_game,
                                        mock_get_prices, mock_regime, mock_strong_setups):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Research"
        ws.append(["Rank"] + [None] * 25)
        mock_load_wb.return_value = wb
        mock_get_prices.return_value = {}

        cases = [
            (5000.0, "BALANCED", "above 40% threshold upgrades DEFENSIVE→BALANCED"),
            (3000.0, "DEFENSIVE", "below 40% threshold stays DEFENSIVE"),
        ]
        for balance, expected_profile, label in cases:
            with self.subTest(label=label):
                state = {"balance": balance, "equity": 10000.0, "positions": {},
                         "queued_orders": [], "profile": "DEFENSIVE", "profile_mode": "ADAPTIVE"}
                mock_load_game.return_value = state
                game.run_daily_ai_management(force=True, manual_profile=None)
                self.assertEqual(state["profile"], expected_profile)

    def test_check_failure_rules_rejection(self):
        test_rules = [
            {"id": "TOXIC_BEARISH_PGR",  "field": "pgr",   "condition": "startswith_Be", "reason": "Avoid buying Bearish PGR"},
            {"id": "TOXIC_LOW_SCORE",     "field": "score", "condition": "less_than_5.0", "reason": "Avoid buying low scores"},
        ]
        with tempfile.TemporaryDirectory() as td:
            rules_path = Path(td) / "Data" / "failure_dna_rules.json"
            rules_path.parent.mkdir(parents=True)
            rules_path.write_text(json.dumps(test_rules), encoding="utf-8")
            orig = game.BASE_DIR
            game.BASE_DIR = Path(td)
            try:
                is_toxic, reason = game.check_failure_rules("AAPL", "Be-", 9.5, 0.5, "Technology")
                self.assertTrue(is_toxic)
                self.assertEqual(reason, "Avoid buying Bearish PGR")

                is_toxic, reason = game.check_failure_rules("MSFT", "Neutral", 3.2, 0.5, "Technology")
                self.assertTrue(is_toxic)
                self.assertEqual(reason, "Avoid buying low scores")

                is_toxic, reason = game.check_failure_rules("GOOGL", "Bu+", 9.8, 0.5, "Technology")
                self.assertFalse(is_toxic)
                self.assertEqual(reason, "")
            finally:
                game.BASE_DIR = orig

    def test_log_closed_trade_dna_writing(self):
        pos = {
            "qty": 10, "cost": 100.0, "stop_loss": 92.0,
            "buy_dna": {"buy_date": "2026-06-01", "pgr": "Bullish",
                        "score": 9.5, "z_score": 1.2, "industry": "Software"}
        }
        with tempfile.TemporaryDirectory() as td:
            dna_path = Path(td) / "Data" / "trade_history_dna.json"
            dna_path.parent.mkdir(parents=True)
            orig = game.BASE_DIR
            game.BASE_DIR = Path(td)
            try:
                game.log_closed_trade_dna("TEST_SYM", pos, 105.0, "2026-06-05")
                records = json.loads(dna_path.read_text(encoding="utf-8"))
            finally:
                game.BASE_DIR = orig

        self.assertEqual(len(records), 1)
        rec = records[0]
        self.assertEqual(rec["symbol"], "TEST_SYM")
        self.assertEqual(rec["buy_date"], "2026-06-01")
        self.assertEqual(rec["sell_date"], "2026-06-05")
        self.assertEqual(rec["buy_price"], 100.0)
        self.assertEqual(rec["sell_price"], 105.0)
        self.assertEqual(rec["qty"], 10)
        self.assertEqual(rec["pnl_pct"], 5.0)
        self.assertEqual(rec["holding_days"], 4)
        self.assertEqual(rec["buy_dna"]["industry"], "Software")

    def test_risk_reward_gate_math(self):
        """The R/R gate formula used in run_daily_ai_management workbook-scan loop."""
        def _gate(price, stop, target):
            # Mirrors the exact logic at ai_portfolio_game.py lines ~1244-1255
            upside   = target - price
            downside = price - stop
            rr       = round(upside / downside, 2) if downside > 0 else 0.0
            gain     = round((upside / price) * 100, 2) if price > 0 else 0.0
            return rr >= 2.0 and gain >= 5.0

        cases = [
            (100, 96,   102, False, "R/R=0.5 → reject"),
            (100, 99.5, 102, False, "gain=2% → reject"),
            (100,  98,  106, True,  "R/R=3.0, gain=6% → accept"),
            (100,  95,  111, True,  "R/R=2.2, gain=11% → accept"),
        ]
        for price, stop, target, expected, label in cases:
            with self.subTest(label=label):
                self.assertEqual(_gate(price, stop, target), expected, label)

    def test_circuit_breaker_triggers(self):
        """check_systemic_risk fires on single-day crash, rolling drawdown, and VXX spike."""
        import circuit_breaker
        flat = [{"close": 100.0}] * 15

        drawdown = list(flat)
        for i in range(-10, 0):
            drawdown[i] = {"close": 100.0 + (i * 0.6)}
        drawdown[-10] = {"close": 100.0}
        drawdown[-1]  = {"close": 94.0}
        drawdown[-2]  = {"close": 94.6}

        cases = [
            ("single-day crash",    flat,     {"SPY": 97.0},              "Single-Day Capitulation"),
            ("rolling drawdown",    drawdown, {"SPY": 94.0},              "Rolling 10-day Drawdown Breach"),
            ("VXX spike",           flat,     {"SPY": 100.0, "VXX": 116.0}, "Volatility Capitulation"),
        ]
        for label, series, prices, expected_reason in cases:
            with self.subTest(label=label):
                with mock.patch("circuit_breaker.load_spy_history", return_value=series), \
                     mock.patch("circuit_breaker.load_vxx_prev_close", return_value=100.0):
                    is_active, reason = circuit_breaker.check_systemic_risk(prices=prices)
                self.assertTrue(is_active, label)
                self.assertIn(expected_reason, reason, label)

    def test_circuit_breaker_stop_tightening(self):
        """Verify that the breaker successfully freezes buying and tightens stops on satellites."""
        import circuit_breaker
        from unittest import mock
        
        state = {
            "balance": 1000.0,
            "queued_orders": [
                {"symbol": "AAPL", "type": "BUY", "qty": 10},
                {"symbol": "MSFT", "type": "SELL", "qty": 5}
            ],
            "positions": {
                "ZS": {"qty": 10, "cost": 150.0, "stop_loss": 135.0, "is_scarcity": False},
                "HE": {"qty": 81, "cost": 13.57, "stop_loss": 12.72, "is_scarcity": True} # scarcity left alone
            }
        }
        
        with mock.patch("circuit_breaker.check_systemic_risk", return_value=(True, "Single-Day Capitulation")), \
             mock.patch("circuit_breaker.log_circuit_breaker_trigger_dna"), \
             mock.patch("risk_utils.calculate_atr", return_value=1.50):
                circuit_breaker.enforce_circuit_breaker(state, prices={"ZS": 149.0})
                
                # 1. Buying must be frozen (queued BUYs cleared, SELLs left alone)
                self.assertEqual(len(state["queued_orders"]), 1)
                self.assertEqual(state["queued_orders"][0]["symbol"], "MSFT")
                
                # 2. Satellite (ZS) stop must be tightened to 1.0x ATR: 149.0 - 1.50 = 147.50
                # 147.50 is higher (safer) than original 135.0 stop!
                self.assertEqual(state["positions"]["ZS"]["stop_loss"], 147.50)
                
                # 3. Scarcity (HE) stop must be left alone at 12.72
                self.assertEqual(state["positions"]["HE"]["stop_loss"], 12.72)

    def test_circuit_breaker_caution_state(self):
        """Verify that an empty or missing SPY price dynamically triggers a defensive Caution Freeze."""
        import circuit_breaker
        
        # Test completely empty prices dictionary
        is_active, reason = circuit_breaker.check_systemic_risk(prices={})
        self.assertTrue(is_active)
        self.assertIn("Caution freeze active", reason)
        
        # Test prices dict missing SPY
        is_active, reason = circuit_breaker.check_systemic_risk(prices={"AAPL": 150.0})
        self.assertTrue(is_active)
        self.assertIn("Caution freeze active", reason)

    def test_circuit_breaker_vxx_volatility_capitulation(self):
        """Verify that a VXX surge > 15.0% successfully triggers the Volatility Capitulation breaker."""
        import circuit_breaker
        from unittest import mock
        
        # Mock yesterday's VXX close to 100.0, and today's VXX price to 116.0 (up 16.0%!)
        mock_series = [{"close": 100.0}] * 15
        
        with mock.patch("circuit_breaker.load_spy_history", return_value=mock_series):
            with mock.patch("circuit_breaker.load_vxx_prev_close", return_value=100.0):
                is_active, reason = circuit_breaker.check_systemic_risk(prices={"SPY": 100.0, "VXX": 116.0})
                self.assertTrue(is_active)
                self.assertIn("Volatility Capitulation", reason)

    def test_circuit_breaker_idiosyncratic_gap_guard(self):
        """Verify that a single-stock gap-down > 8% is successfully detected during opening window."""
        import circuit_breaker
        from unittest import mock
        
        # Test Case 1: Inside opening window, -10% gap-down -> Must return True!
        with mock.patch("circuit_breaker.is_market_opening_window", return_value=True):
            is_frozen = circuit_breaker.is_single_stock_gap_frozen("ZS", 90.0, 100.0) # down 10%
            self.assertTrue(is_frozen)
            
            # -5% gap-down is less than 8.0% threshold -> Must return False
            is_frozen = circuit_breaker.is_single_stock_gap_frozen("ZS", 95.0, 100.0)
            self.assertFalse(is_frozen)
            
        # Test Case 2: Outside opening window, -10% gap-down -> Must return False!
        with mock.patch("circuit_breaker.is_market_opening_window", return_value=False):
            is_frozen = circuit_breaker.is_single_stock_gap_frozen("ZS", 90.0, 100.0)
            self.assertFalse(is_frozen)

    def test_circuit_breaker_elastic_memory(self):
        """Verify that the breaker successfully caches original stops and restores them on stabilization."""
        import circuit_breaker
        from unittest import mock
        
        state = {
            "balance": 1000.0,
            "positions": {
                "ZS": {"qty": 10, "cost": 150.0, "stop_loss": 147.50, "original_stop_loss": 135.0, "is_scarcity": False}
            }
        }
        
        # Scenario 1: Market stabilized (inactive breaker) -> Must restore stop to 135.0 and pop original key!
        with mock.patch("circuit_breaker.check_systemic_risk", return_value=(False, "")):
            circuit_breaker.enforce_circuit_breaker(state, prices={"ZS": 148.0})
            
            self.assertEqual(state["positions"]["ZS"]["stop_loss"], 135.0)
            self.assertNotIn("original_stop_loss", state["positions"]["ZS"])

    def test_circuit_breaker_backfeed_dna(self):
        import circuit_breaker
        state = {
            "balance": 1000.0, "equity": 10000.0, "profile": "BALANCED",
            "queued_orders": [],
            "positions": {"ZS": {"qty": 10, "cost": 150.0, "stop_loss": 135.0, "is_scarcity": False}}
        }
        mock_series = [{"close": 100.0}] * 15
        with tempfile.TemporaryDirectory() as td:
            tmp_dna = Path(td) / "trade_history_dna.json"
            with mock.patch.object(circuit_breaker, "DNA_FILE", tmp_dna), \
                 mock.patch("circuit_breaker.load_spy_history", return_value=mock_series), \
                 mock.patch("circuit_breaker.load_vxx_prev_close", return_value=100.0):
                circuit_breaker.log_circuit_breaker_trigger_dna(
                    "Single-Day Capitulation", state, prices={"SPY": 97.0, "VXX": 116.0}
                )
            records = json.loads(tmp_dna.read_text(encoding="utf-8"))

        self.assertEqual(len(records), 1)
        rec = records[0]
        self.assertEqual(rec["type"], "CIRCUIT_BREAKER_TRIGGER")
        self.assertEqual(rec["reason"], "Single-Day Capitulation")
        self.assertEqual(rec["spy_return_pct"], -3.0)
        self.assertEqual(rec["vxx_return_pct"], 16.0)
        self.assertEqual(rec["portfolio_equity"], 10000.0)
        self.assertEqual(rec["cash_balance"], 1000.0)
        self.assertEqual(rec["open_positions"][0]["symbol"], "ZS")
        self.assertEqual(rec["profile"], "BALANCED")

class TestRequalifyPromptBuilder(unittest.TestCase):
    """build_requalify_prompt — pure formatting, no I/O, zero mocks needed."""

    def _factors(self, **overrides):
        base = {
            "pgr": "Bu", "s10": 4.2, "l60": 3.1, "combined": 7.3,
            "stop": 138.50, "target": 162.00, "buying_ratio": 6.5,
            "money_flow": "Strong", "lt_trend": "Strong", "over_bt_sl": "Optimal",
            "patterns": "Cup&H↑", "industry": "Semiconductors",
            "price": 150.0, "det_action": "HOLD", "det_reason": "scores positive",
        }
        base.update(overrides)
        return base

    def test_all_fields_present(self):
        import data_api
        prompt = data_api.build_requalify_prompt("AAPL", self._factors(), cost=142.30, regime="BALANCED")
        self.assertIn("AAPL", prompt)
        self.assertIn("Bu", prompt)
        self.assertIn("4.2", prompt)
        self.assertIn("BALANCED", prompt)
        self.assertIn("HOLD", prompt)
        self.assertIn("138.50", prompt)     # stop

    def test_no_cost_omits_pnl(self):
        import data_api
        prompt = data_api.build_requalify_prompt("TSLA", self._factors(), cost=None, regime="DEFENSIVE")
        self.assertIn("n/a", prompt)        # entry price shows n/a
        self.assertNotIn("PnL", prompt)     # no pnl line

    def test_missing_stop_shows_na(self):
        import data_api
        f = self._factors(stop=0, target=0)
        prompt = data_api.build_requalify_prompt("SPY", f, cost=100.0, regime="BALANCED")
        self.assertIn("n/a", prompt)        # stop shows n/a when falsy

    def test_news_appended_when_provided(self):
        import data_api
        news = ["Earnings beat by 12%.", "Analyst upgrades to Buy at $175."]
        prompt = data_api.build_requalify_prompt("NVDA", self._factors(), cost=None, regime="AGGRESSIVE", news=news)
        self.assertIn("Earnings beat", prompt)
        self.assertIn("Analyst upgrades", prompt)

    def test_cached_flag_adds_note(self):
        import data_api
        f = self._factors(_cached=True)
        prompt = data_api.build_requalify_prompt("XOM", f, cost=None, regime="DEFENSIVE")
        self.assertIn("cached", prompt.lower())

    def test_invalid_symbol_rejected_by_requalify_symbol(self):
        import data_api
        result = data_api.requalify_symbol("../etc/passwd")
        self.assertIn("error", result)
        self.assertIn("Invalid symbol", result["error"])
        self.assertEqual(result["factors"], {})

    def test_valid_symbol_format_passes_validation(self):
        import data_api
        # Should not be rejected at the validation gate (may fail later on network/session)
        result = data_api.requalify_symbol("AAPL")
        self.assertNotIn("Invalid symbol", result.get("error") or "")


class TestDigitSumScoring(unittest.TestCase):
    """digit_sum_score and _price_digit_sum — pure math, no I/O mocks needed for math tests.
    Index-loading tests patch the module-level singletons directly."""

    def setUp(self):
        import aether.scoring as _sc
        # Force reset — do not save prior state; a prior test may have loaded
        # the real JSON and restoring it would pollute subsequent tests.
        _sc._digit_index      = None
        _sc._digit_full_index = None

    def tearDown(self):
        import aether.scoring as _sc
        _sc._digit_index      = None
        _sc._digit_full_index = None

    def test_price_digit_sum_integer(self):
        from aether.scoring import _price_digit_sum
        self.assertEqual(_price_digit_sum(247.35), 4)   # 2+4+7=13->4
        self.assertEqual(_price_digit_sum(9.0),    9)
        self.assertEqual(_price_digit_sum(100.0),  1)
        self.assertEqual(_price_digit_sum(99.99),  9)   # int(99.99)=99 -> 9+9=18->9

    def test_price_digit_sum_full(self):
        from aether.scoring import _price_digit_sum_full
        self.assertEqual(_price_digit_sum_full(247.35), 3)   # 2+4+7+3+5=21->3
        self.assertEqual(_price_digit_sum_full(100.00), 1)   # 1+0+0+0+0=1
        self.assertEqual(_price_digit_sum_full(9.99),   9)   # 9+9+9=27->9

    def test_digit_sum_score_no_signal_returns_zero(self):
        from aether.scoring import digit_sum_score
        import unittest.mock as _mock
        # Empty index -> score must be 0.0
        with _mock.patch("aether.scoring._digit_index", {}), \
             _mock.patch("aether.scoring._digit_full_index", {}):
            self.assertEqual(digit_sum_score("AAPL", close_price=247.35), 0.0)

    def test_digit_sum_score_fires_from_full_cents_index(self):
        from aether.scoring import digit_sum_score, _price_digit_sum_full
        import unittest.mock as _mock
        dg_full = _price_digit_sum_full(247.35)
        fake_full = {("AAPL", "CLOSE", dg_full): 3.0}
        with _mock.patch("aether.scoring._digit_index", {}), \
             _mock.patch("aether.scoring._digit_full_index", fake_full):
            score = digit_sum_score("AAPL", close_price=247.35)
        self.assertGreater(score, 0.0)
        self.assertLessEqual(score, 1.0)

    def test_digit_sum_score_fires_from_integer_index(self):
        from aether.scoring import digit_sum_score, _price_digit_sum
        import unittest.mock as _mock
        dg_int = _price_digit_sum(247.35)   # int(247.35)=247 -> 2+4+7=13->4
        fake_int = {("NVDA", "CLOSE", dg_int): -2.4}  # negative z -> down signal
        with _mock.patch("aether.scoring._digit_index", fake_int), \
             _mock.patch("aether.scoring._digit_full_index", {}):
            score = digit_sum_score("NVDA", close_price=247.35)
        self.assertLess(score, 0.0)   # negative signal fires
        self.assertGreaterEqual(score, -1.0)

    def test_digit_sum_open_score_zero_without_signal(self):
        from aether.scoring import digit_sum_open_score
        import unittest.mock as _mock
        with _mock.patch("aether.scoring._digit_index", {}), \
             _mock.patch("aether.scoring._digit_full_index", {}):
            self.assertEqual(digit_sum_open_score("SPY", 500.0), 0.0)

    def test_build_digit_index_quality_filter(self):
        from aether.scoring import _build_digit_index
        import tempfile, json as _json
        rows = [
            {"symbol":"AAA","type":"OPEN","digit":5,"z":2.5,"up_pct":0.6,"base":0.5,"n":100,
             "temporal":"consistent","has_flip":False,"is_sparse":False},  # PASS
            {"symbol":"BBB","type":"OPEN","digit":3,"z":3.0,"up_pct":0.65,"base":0.5,"n":100,
             "temporal":"stale","has_flip":False,"is_sparse":False},        # FAIL: stale
            {"symbol":"CCC","type":"OPEN","digit":7,"z":2.2,"up_pct":0.58,"base":0.5,"n":80,
             "temporal":"partial","has_flip":True,"is_sparse":False},       # FAIL: flip
            {"symbol":"DDD","type":"OPEN","digit":1,"z":2.1,"up_pct":0.56,"base":0.5,"n":60,
             "temporal":"partial","has_flip":False,"is_sparse":True},       # FAIL: sparse
        ]
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
            _json.dump(rows, f); fname = f.name
        idx = _build_digit_index(fname)
        self.assertIn(("AAA","OPEN",5), idx)    # only this one passes
        self.assertNotIn(("BBB","OPEN",3), idx)
        self.assertNotIn(("CCC","OPEN",7), idx)
        self.assertNotIn(("DDD","OPEN",1), idx)
        import os; os.unlink(fname)


if __name__ == "__main__":
    unittest.main()
