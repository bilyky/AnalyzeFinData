"""
Unit tests for update_replacements_sheet — the Rotation ("Replacements") sheet.

Rotation exists to replace *weak* real-account holdings, so these tests pin the
three correctness guarantees added in the badness-gate fix:

  A1  winners (STRONG HOLD / HOLD) are never proposed for sale;
  A2  a held symbol missing from picks_data is logged, not silently dropped;
  A3  the Short_Long reader picks up BOTH real-account tables (T1 + T2), even
      when a blank row sits directly under the T2 header, and the ticker gate
      rejects free-text rows.

No E*TRADE / API calls; uses in-memory openpyxl workbooks.
"""
import sys
import os
import logging
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from workbook_write import update_replacements_sheet


def _make_sl(*tables):
    """Build a Short_Long workbook from one or more account tables.

    Each table is a list of symbol strings; tables are separated by a blank row
    and each gets its own "Symb" header (matching the two real-account layout).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Short_Long"
    r = 1
    for ti, syms in enumerate(tables):
        ws.cell(r, 1).value = f"Account ACCT_{ti}"   # title row (not a header)
        r += 1
        ws.cell(r, 2).value = "Symb"                 # header → opens the table
        r += 1
        for s in syms:
            ws.cell(r, 2).value = s
            r += 1
        r += 1                                        # blank row between tables
    return wb


def _picks(sym, s10, l60, setup=1, pgr="Bu"):
    return {"symbol": sym, "short10": s10, "long60": l60, "setup": setup, "pgr": pgr}


def _sell_symbols(wb) -> list:
    """Symbols in the SELL column (B) of the generated Replacements sheet."""
    ws = wb["Replacements"]
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        v = row[1] if len(row) > 1 else None      # column B = SELL
        if isinstance(v, str) and v.strip():
            out.append(v.strip().upper())
    return out


# Buy candidates: not held, strong scores so they populate the buy side and let
# every eligible sell candidate pair up (n_pairs = min(sells, buys, 30)).
_BUYS = [_picks("CCC", 8, 8), _picks("DDD", 7, 7), _picks("EEE", 6, 6)]


class TestReplacementsRotation(unittest.TestCase):

    def test_winners_are_never_sold(self):
        """A1: STRONG HOLD / HOLD holdings are excluded from the sell side."""
        wb = _make_sl(["WINR", "HELDR", "WEAK"])
        picks = [
            _picks("WINR", 1, 6),    # L60=6 → STRONG HOLD  → must NOT be sold
            _picks("HELDR", 1, 3),   # L60=3 → HOLD         → must NOT be sold
            _picks("WEAK", -3, -6),  # L60=-6 → EXIT        → sell candidate
        ] + _BUYS
        update_replacements_sheet(wb, picks)

        sells = _sell_symbols(wb)
        self.assertIn("WEAK", sells)
        self.assertNotIn("WINR", sells, "STRONG HOLD winner must not be rotated out")
        self.assertNotIn("HELDR", sells, "HOLD winner must not be rotated out")

    def test_all_winners_yields_no_pairs(self):
        """If every holding is a winner, there is nothing to rotate."""
        wb = _make_sl(["WINR", "HELDR"])
        picks = [_picks("WINR", 5, 8), _picks("HELDR", 4, 5)] + _BUYS
        update_replacements_sheet(wb, picks)
        self.assertEqual(_sell_symbols(wb), [])

    def test_unscored_holding_is_logged(self):
        """A2: a held symbol absent from picks_data is logged at WARNING (so it
        reaches the persisted audit log, not just transient stderr), not dropped."""
        wb = _make_sl(["WEAK", "XCLD"])   # XCLD absent from picks below
        picks = [_picks("WEAK", -3, -6)] + _BUYS

        with self.assertLogs("aether.workbook_write", level=logging.WARNING) as cm:
            update_replacements_sheet(wb, picks)

        self.assertEqual(len(cm.records), 1, "exactly one skip line expected")
        rec = cm.records[0]
        self.assertGreaterEqual(rec.levelno, logging.WARNING,
                                "skip must be persisted (>= WARNING), not CONSOLE-only")
        msg = rec.getMessage()
        self.assertIn("XCLD", msg, "unscored holding must be named in the log")
        self.assertIn("not in picks_data", msg)
        # It must not be silently turned into a sell recommendation either.
        self.assertNotIn("XCLD", _sell_symbols(wb))

    def test_reads_both_tables_including_blank_under_t2_header(self):
        """A3: symbols from BOTH real-account tables are read, even with a blank
        row directly under the T2 header — the live Short_Long layout. This is
        the red case: terminating a table on a blank row drops all of T2.
        Prose rows are rejected by the ticker gate.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Short_Long"
        ws.cell(1, 1).value = "Account ACCT_0"
        ws.cell(2, 2).value = "Symb"        # T1 header
        ws.cell(3, 2).value = "AAA"
        # rows 4-5 blank (2-blank gap between tables, matching the live sheet)
        ws.cell(6, 1).value = "Account ACCT_1"
        ws.cell(7, 2).value = "Symb"        # T2 header
        # row 8 blank — directly under the T2 header (the live-sheet gotcha)
        ws.cell(9, 2).value = "BBB"
        ws.cell(10, 2).value = "net liquidation summary"   # prose → ticker gate rejects

        picks = [_picks("AAA", -2, -5), _picks("BBB", -3, -6)] + _BUYS
        update_replacements_sheet(wb, picks)

        sells = _sell_symbols(wb)
        self.assertIn("AAA", sells, "T1 holding must be read")
        self.assertIn("BBB", sells,
                      "T2 holding must be read despite the blank row under its header")


if __name__ == "__main__":
    unittest.main(verbosity=2)
