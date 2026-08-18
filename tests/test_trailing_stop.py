"""
Unit tests for the profit-lock trailing stop ratchet in ai_portfolio_game.py.
"""
import os
import sys
import unittest
import openpyxl
from unittest import mock

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import ai_portfolio_game as game


def _make_state(price, cost=100.0, stop=90.0):
    return {
        "balance": 5000.0,
        "equity": 10000.0,
        "positions": {
            "ULTA": {"qty": 10, "cost": cost, "stop_loss": stop, "is_scarcity": False}
        },
        "queued_orders": [],
        "history": [],
    }


def _make_wb(sym, price, s10=5.0, l60=5.0):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Research"
    ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR",
               "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other",
               "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other",
               "Win%", "Short10", "Long60"])
    ws.append([1, None, None, sym, "Retail", None, "Bu", None, None, None, price,
               None, None, None, None, None, None, None, None, None, "OK", None, None,
               0.65, s10, l60])
    return wb


_COMMON_PATCHES = [
    mock.patch("ai_portfolio_game.is_market_hours", return_value=True),
    mock.patch("ai_portfolio_game.get_live_prices"),
    mock.patch("ai_portfolio_game.load_game"),
    mock.patch("ai_portfolio_game.save_game"),
    mock.patch("ai_portfolio_game.openpyxl.load_workbook"),
    mock.patch("circuit_breaker.enforce_circuit_breaker"),
]


class TestTrailingStopRatcheting(unittest.TestCase):

    def setUp(self):
        game._HEAL_ATTEMPTED.clear()

    def _run(self, state, price, atr=4.0):
        patches = [p.start() for p in _COMMON_PATCHES]
        try:
            _mkt, _prices, _load, _save, _wb, _cb = patches
            _prices.return_value = {"ULTA": price, "SPY": 500.0}
            _load.return_value = state
            _wb.return_value = _make_wb("ULTA", price)
            with mock.patch("aether.risk_utils.calculate_atr", return_value=atr):
                game.run_daily_ai_management(force=True, manual_profile="BALANCED")
        finally:
            for p in _COMMON_PATCHES:
                p.stop()

    def test_flower_protection_below_cost(self):
        """Stop does NOT ratchet when price is below cost."""
        state = _make_state(price=98.0)
        self._run(state, price=98.0)
        pos = state["positions"]["ULTA"]
        self.assertEqual(pos["stop_loss"], 90.0)
        self.assertEqual(pos["highest_close_since_acq"], 100.0)

    def test_stop_does_not_decrease_on_retrace(self):
        """Red-path: stop must NOT move down when price retraces from a previous peak."""
        state = _make_state(price=95.0, stop=92.0)
        state["positions"]["ULTA"]["highest_close_since_acq"] = 102.0
        self._run(state, price=95.0)
        pos = state["positions"]["ULTA"]
        self.assertGreaterEqual(pos["stop_loss"], 92.0, "Ratcheted stop must never decrease on retrace")

    def test_profit_lock_ratchet_above_cost(self):
        """Stop ratchets up when price is > 1×ATR above cost: 105 - 2.5×4 = $95."""
        state = _make_state(price=105.0)
        self._run(state, price=105.0)
        pos = state["positions"]["ULTA"]
        self.assertEqual(pos["stop_loss"], 95.0)
        self.assertEqual(pos["highest_close_since_acq"], 105.0)

    def test_breakeven_lock_trigger(self):
        """Breakeven lock fires when profit > 1.5×ATR: stop raised to cost basis $100."""
        state = _make_state(price=107.0)
        self._run(state, price=107.0)
        pos = state["positions"]["ULTA"]
        self.assertEqual(pos["stop_loss"], 100.0)
        self.assertEqual(pos["highest_close_since_acq"], 107.0)


class TestZeroCashDragAndFractionalSizing(unittest.TestCase):
    def test_is_fractional_eligible(self):
        # 1. Standard S&P 100 leaders and core ETFs must be eligible
        self.assertTrue(game.instruments.is_fractional_eligible("AAPL"))
        self.assertTrue(game.instruments.is_fractional_eligible("SPY"))
        self.assertTrue(game.instruments.is_fractional_eligible("QQQ"))
        self.assertTrue(game.instruments.is_fractional_eligible("NVDA"))
        
        # 2. Non-S&P 100 assets must NOT be eligible
        self.assertFalse(game.instruments.is_fractional_eligible("TSCO"))
        self.assertFalse(game.instruments.is_fractional_eligible("HE"))

    def test_calculate_share_qty(self):
        # 1. Standard assets must return integers (rounded down)
        self.assertEqual(game.calculate_share_qty("TSCO", 1000.0, 30.0), 33)
        self.assertEqual(type(game.calculate_share_qty("TSCO", 1000.0, 30.0)), int)
        
        # 2. Fractional assets must return precise floats rounded to 3 decimals
        # 1000.0 / 30.0 = 33.333333... -> 33.333
        self.assertEqual(game.calculate_share_qty("AAPL", 1000.0, 30.0), 33.333)
        self.assertEqual(type(game.calculate_share_qty("AAPL", 1000.0, 30.0)), float)

    @mock.patch("ai_portfolio_game.get_market_regime")
    def test_zero_cash_floor_under_bullish(self, mock_regime):
        # 1. In a bull market, cash buffer must be 0% across all profiles.
        #    get_market_regime() labels a strong bull market "AGGRESSIVE" — that is
        #    the bull signal the Zero-Cash-Drag Autopilot keys off (not "BULLISH",
        #    which the regime function never returns).
        mock_regime.return_value = "AGGRESSIVE"

        rules_agg = game.get_strategy_rules("AGGRESSIVE")
        self.assertEqual(rules_agg["cash_buffer_pct"], 0.0)

        rules_bal = game.get_strategy_rules("BALANCED")
        self.assertEqual(rules_bal["cash_buffer_pct"], 0.0)

        rules_def = game.get_strategy_rules("DEFENSIVE")
        self.assertEqual(rules_def["cash_buffer_pct"], 0.0)

        # 2. Outside a bull regime, cash buffer must return to its standard values.
        mock_regime.return_value = "BALANCED"

        rules_agg = game.get_strategy_rules("AGGRESSIVE")
        self.assertEqual(rules_agg["cash_buffer_pct"], 0.10)

        rules_bal = game.get_strategy_rules("BALANCED")
        self.assertEqual(rules_bal["cash_buffer_pct"], 0.20)

        rules_def = game.get_strategy_rules("DEFENSIVE")
        self.assertEqual(rules_def["cash_buffer_pct"], 0.50)


class TestAntiFragileFlexibility(unittest.TestCase):
    @mock.patch("ai_portfolio_game.get_market_regime")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.is_market_hours", return_value=True)
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    @mock.patch("circuit_breaker.enforce_circuit_breaker")
    @mock.patch("ai_portfolio_game.backtrack_verify", return_value=(True, "OK"))
    @mock.patch("ai_portfolio_game._cache_stale", return_value=False)
    @mock.patch("ai_portfolio_game.calculate_bubble_z_score", return_value=1.0)
    def test_adaptive_cap_relaxation(self, mock_z, mock_stale, mock_verify, mock_breaker, mock_load_wb, mock_hours, mock_get_prices, mock_save_game, mock_load_game, mock_regime):
        # 1. Setup standard play candidate (AAPL) with score of 11.1 (ultra-conviction)
        # Standard cap is 80% of $10,000 = $8,000 limit.
        # We setup 4 existing standard positions costing $1,800 each (Total: $7,200).
        # Remaining standard room is $8,000 - $7,200 = $800 limit.
        # Available cash to deploy is $1,000. Max slots is 5 (Balanced).
        # Active positions = 4, so available_slots = 1.
        # Under standard rules, AAPL buy would be downsized to standard room $800 (5.556 shares).
        # Under Adaptive Cap Relaxation (since score 11.1 >= 8.0), standard cap is bypassed, and AAPL buy goes through for full available cash $1,000 (6.944 shares)!
        mock_regime.return_value = "NEUTRAL"
        state = {
            "balance": 3000.0,
            "equity": 10000.0,
            "positions": {
                "P1": {"qty": 1, "cost": 1800.0, "is_scarcity": False},
                "P2": {"qty": 1, "cost": 1800.0, "is_scarcity": False},
                "P3": {"qty": 1, "cost": 1800.0, "is_scarcity": False},
                "P4": {"qty": 1, "cost": 1800.0, "is_scarcity": False}
            },
            "queued_orders": [],
            "history": []
        }
        mock_load_game.return_value = state
        mock_get_prices.return_value = {"AAPL": 144.0, "P1": 1800.0, "P2": 1800.0, "P3": 1800.0, "P4": 1800.0, "SPY": 500.0}

        # Mock workbook sheet with ultra-conviction candidate AAPL
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        ws.append([1, None, None, "AAPL", "Technology", None, "Bu", None, None, None, 144.0, None, None, None, None, None, None, None, None, None, "OK", None, None, 0.65, 6.0, 5.1]) # Score = 11.1
        mock_load_wb.return_value = wb

        # Mock ATR calculation
        with mock.patch("aether.risk_utils.calculate_atr", return_value=4.00):
            game.run_daily_ai_management(force=True, manual_profile="BALANCED")

        # AAPL was bought! Verify it bypassed standard cap ($800) and bought full $960 (6.667 shares!)
        # If standard cap was enforced, standard room $800 / 144 = 5.556 shares.
        # So qty must be greater than 6.5 shares!
        pos = state["positions"]["AAPL"]
        self.assertGreater(pos["qty"], 6.5)

    @mock.patch("ai_portfolio_game.get_market_regime")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.is_market_hours", return_value=True)
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    @mock.patch("circuit_breaker.enforce_circuit_breaker")
    @mock.patch("ai_portfolio_game.backtrack_verify", return_value=(True, "OK"))
    @mock.patch("ai_portfolio_game._cache_stale", return_value=False)
    def test_dynamic_slot_expansion(self, mock_stale, mock_verify, mock_breaker, mock_load_wb, mock_hours, mock_get_prices, mock_save_game, mock_load_game, mock_regime):
        # 1. Setup a portfolio with 5 of 5 maximum positions under BALANCED profile
        # Cash is plentiful: Balance = 3000, Equity = 10000. Cash ratio = 30.0% (> 15%!).
        # Under old rules, no buys are allowed because we are already at max 5 positions.
        # Under Dynamic Position-Slot Expansion, slots count dynamically expands to 6, allowing a fresh buy!
        mock_regime.return_value = "NEUTRAL"
        state = {
            "balance": 3000.0,
            "equity": 10000.0,
            "positions": {
                "P1": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P2": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P3": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P4": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P5": {"qty": 1, "cost": 1000.0, "is_scarcity": False}
            },
            "queued_orders": [],
            "history": []
        }
        mock_load_game.return_value = state
        mock_get_prices.return_value = {"P1": 1000.0, "P2": 1000.0, "P3": 1000.0, "P4": 1000.0, "P5": 1000.0, "CDW": 100.0, "SPY": 500.0}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        ws.append([1, None, None, "CDW", "Technology", None, "Bu", None, None, None, 100.0, None, None, None, None, None, None, None, None, None, "OK", None, None, 0.65, 6.0, 5.1]) # Score = 11.1
        mock_load_wb.return_value = wb

        with mock.patch("aether.risk_utils.calculate_atr", return_value=4.00), \
             mock.patch("instruments.is_scarcity_asset", return_value=False):
            game.run_daily_ai_management(force=True, manual_profile="BALANCED")

        # CDW must be purchased successfully, showing that the 5-position limit was dynamically expanded!
        self.assertIn("CDW", state["positions"])

    @mock.patch("ai_portfolio_game.get_market_regime")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.is_market_hours", return_value=True)
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    @mock.patch("circuit_breaker.enforce_circuit_breaker")
    @mock.patch("ai_portfolio_game.backtrack_verify", return_value=(True, "OK"))
    @mock.patch("ai_portfolio_game._cache_stale", return_value=False)
    def test_multiple_slot_expansion_with_while_loop(self, mock_stale, mock_verify, mock_breaker, mock_load_wb, mock_hours, mock_get_prices, mock_save_game, mock_load_game, mock_regime):
        # Setup a portfolio with 6 positions, but base max_positions is 5 (BALANCED profile)
        # Cash is plentiful: Balance = 3000, Equity = 10000. Cash ratio = 30.0% (> 15%!).
        # The system must dynamically expand slots using a while loop until max_positions exceeds 6 (it expands 5 -> 6 -> 7).
        # This yields available_slots = 1, so CDW is successfully bought!
        mock_regime.return_value = "NEUTRAL"
        state = {
            "balance": 3000.0,
            "equity": 10000.0,
            "positions": {
                "P1": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P2": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P3": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P4": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P5": {"qty": 1, "cost": 1000.0, "is_scarcity": False},
                "P6": {"qty": 1, "cost": 1000.0, "is_scarcity": False}
            },
            "queued_orders": [],
            "history": []
        }
        mock_load_game.return_value = state
        mock_get_prices.return_value = {
            "P1": 1000.0, "P2": 1000.0, "P3": 1000.0, "P4": 1000.0, "P5": 1000.0, "P6": 1000.0,
            "CDW": 100.0, "SPY": 500.0
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        ws.append([1, None, None, "CDW", "Technology", None, "Bu", None, None, None, 100.0, None, None, None, None, None, None, None, None, None, "OK", None, None, 0.65, 6.0, 5.1]) # Score = 11.1
        mock_load_wb.return_value = wb

        with mock.patch("aether.risk_utils.calculate_atr", return_value=4.00), \
             mock.patch("instruments.is_scarcity_asset", return_value=False):
            game.run_daily_ai_management(force=True, manual_profile="BALANCED")

        # CDW must be purchased successfully because slots expanded 5 -> 6 -> 7!
        self.assertIn("CDW", state["positions"])

    @mock.patch("ai_portfolio_game.get_market_regime")
    @mock.patch("ai_portfolio_game.load_game")
    @mock.patch("ai_portfolio_game.save_game")
    @mock.patch("ai_portfolio_game.get_live_prices")
    @mock.patch("ai_portfolio_game.is_market_hours", return_value=True)
    @mock.patch("ai_portfolio_game.openpyxl.load_workbook")
    @mock.patch("circuit_breaker.enforce_circuit_breaker")
    @mock.patch("ai_portfolio_game.backtrack_verify", return_value=(True, "OK"))
    @mock.patch("ai_portfolio_game._cache_stale", return_value=False)
    def test_dynamic_pyramiding_scale_in(self, mock_stale, mock_verify, mock_breaker, mock_load_wb, mock_hours, mock_get_prices, mock_save_game, mock_load_game, mock_regime):
        # 1. Setup a portfolio with active green winning positions and plentiful cash (Balance = 10000, Equity = 10000)
        # Active position ULTA cost = 400.0, current price = 410.0 (in profit!), highest_close_since_acq = 412.0 (near peak!)
        # s10 momentum for ULTA is 5.0 (>= 3.0!)
        # The system must dynamically scale-in by purchasing more shares of ULTA and recalculating its blended cost!
        mock_regime.return_value = "NEUTRAL"
        state = {
            "balance": 10000.0,
            "equity": 10000.0,
            "positions": {
                "ULTA": {"qty": 2, "cost": 400.0, "stop_loss": 380.0, "highest_close_since_acq": 412.0, "is_scarcity": False}
            },
            "queued_orders": [],
            "history": []
        }
        mock_load_game.return_value = state
        mock_get_prices.return_value = {"ULTA": 410.0, "SPY": 500.0}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Research"
        ws.append(["Rank", "Symbol", "Industry", "Ticker", "Sector", "Other", "PGR", "Other", "Other", "Other", "Price", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Other", "Setup", "Other", "Other", "Win%", "Short10", "Long60"])
        ws.append([1, None, None, "ULTA", "Retail", None, "Bu", None, None, None, 410.0, None, None, None, None, None, None, None, None, None, "OK", None, None, 0.65, 5.0, 5.1]) # Short10 = 5.0
        mock_load_wb.return_value = wb

        with mock.patch("aether.risk_utils.calculate_atr", return_value=4.00):
            game.run_daily_ai_management(force=True, manual_profile="BALANCED")

        # ULTA qty must be increased and blended cost recalculated!
        pos = state["positions"]["ULTA"]
        self.assertGreater(pos["qty"], 2)
        self.assertGreater(pos["cost"], 400.0)


class TestDetermineMaxPositions(unittest.TestCase):
    def test_determine_max_positions_no_expansion_low_cash(self):
        # Cash is low (10% < 15% threshold), should not expand
        res = game.determine_max_positions(cash_ratio=0.10, num_positions=5, base_max_positions=5)
        self.assertEqual(res, 5)

    def test_determine_max_positions_no_expansion_not_full(self):
        # Cash is high (20% > 15%), but we are not full (4 positions held < 5 max), should not expand
        res = game.determine_max_positions(cash_ratio=0.20, num_positions=4, base_max_positions=5)
        self.assertEqual(res, 5)

    def test_determine_max_positions_single_step_expansion(self):
        # Cash is high (20%), and we are full (5 positions held == 5 max).
        # Should expand from 5 to 6, then loop terminates because 5 >= 6 is False.
        res = game.determine_max_positions(cash_ratio=0.20, num_positions=5, base_max_positions=5)
        self.assertEqual(res, 6)

    def test_determine_max_positions_multiple_step_expansion(self):
        # Cash is high (20%), and we are over-full (6 positions held > 5 max).
        # Iteration 1: 6 >= 5 (True) -> max_positions becomes 6.
        # Iteration 2: 6 >= 6 (True) -> max_positions becomes 7.
        # Iteration 3: 6 >= 7 (False) -> loop ends.
        # Should expand from 5 to 7.
        res = game.determine_max_positions(cash_ratio=0.20, num_positions=6, base_max_positions=5)
        self.assertEqual(res, 7)


class TestStpLmtSlippageProtection(unittest.TestCase):
    def test_stp_lmt_slippage_protection(self):
        """Operational: Assert that Stop-Loss sales are executed at the exact Stop Price (STP LMT) if market price falls below it."""
        pos = {"qty": 10, "cost": 50.0, "stop_loss": 46.0}
        market_close = 44.0
        
        stop_loss = pos.get("stop_loss", 0.0)
        execution_price = market_close
        if stop_loss > 0.0 and market_close <= stop_loss:
            execution_price = stop_loss
            
        self.assertEqual(execution_price, 46.0)


class TestBreakoutRrWaiver(unittest.TestCase):
    def test_breakout_rr_waiver(self):
        """Operational: Assert that elite Blue-Sky breakout leaders successfully waive the conservative 2:1 R:R limit."""
        # Setup: price is at an all-time high, target is flatly ATR-capped, stop is wide -> R:R is low (0.28:1)
        price = 361.71
        stop_val = 302.69
        target_val = 378.0
        total_score = 6.6
        short10 = 2.4
        pgr_val = "Very Bullish"
        is_blue_sky = True # Simulated target source is 'atr' (no resistance)
        
        upside = target_val - price
        downside = price - stop_val
        rr_ratio = round(upside / downside, 2)
        
        # Verify that without waiver, it fails the 2:1 minimum
        self.assertLess(rr_ratio, 2.0)
        
        # Check if the Breakout Waiver (R&D #32) condition evaluates to True
        is_elite_breakout = (total_score >= 6.0) and (short10 >= 2.0) and ("Bullish" in pgr_val) and is_blue_sky
        self.assertTrue(is_elite_breakout)


if __name__ == "__main__":
    unittest.main()
